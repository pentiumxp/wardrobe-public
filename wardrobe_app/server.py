from __future__ import annotations

import csv
import base64
import binascii
import hashlib
import hmac
import io
import json
import mimetypes
import os
import re
import secrets
import shutil
import sqlite3
import socket
import subprocess
import sys
import threading
import time
import unicodedata
import uuid
from datetime import date, datetime, timedelta
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.parse import parse_qs, parse_qsl, quote, urlencode, unquote, urlparse, urlunparse
from urllib.request import Request, urlopen
import xml.etree.ElementTree as ET
import zipfile

mimetypes.add_type("application/manifest+json", ".webmanifest")

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from PIL import ExifTags, Image, ImageFile, ImageOps, UnidentifiedImageError

from wardrobe_app.db import DATA_DIR, DB_PATH, MEDIA_DIR, ROOT_DIR, WEB_DIR, connect, ensure_directories, init_db
from wardrobe_app.excel_import import (
    DESKTOP_DIR,
    _find_daily_update,
    _outfit_rollup,
    _recompute_item_last_worn_on,
    _wear_delta_for_item,
    bootstrap_from_desktop,
)
from wardrobe_app.program_api_uploads import (
    PROGRAM_API_MULTIPART_FILE_PAYLOAD_KEY,
    hashable_program_payload,
    is_multipart_form_data,
    parse_multipart_parts,
    program_item_payload_from_multipart_parts,
)
from wardrobe_app import hermes_plugin, photo_ordering, program_api_sync
from wardrobe_app.item_normalization import normalize_item_brand, normalize_price_currency, normalize_price_text


HOST = os.environ.get("WARDROBE_HOST", "0.0.0.0")
PORT = int(os.environ.get("WARDROBE_PORT", "8765"))
PHOTO_THUMBNAIL_MAX_EDGE = 720
PROGRAM_API_PRIMARY_THUMBNAIL_MAX_EDGE = 768
LIBRARY_ROOT = Path(
    os.environ.get(
        "WARDROBE_LIBRARY_ROOT",
        r"C:\Users\xuxin\SynologyDrive\ChatGPT-Drive\徐欣\奢侈品",
    )
)
DEFAULT_IMPORT_DIR = Path(
    os.environ.get("WARDROBE_CLOSET_DIR", str(LIBRARY_ROOT / "衣橱"))
)
WATCH_IMPORT_DIR = Path(
    os.environ.get("WARDROBE_WATCH_DIR", str(LIBRARY_ROOT / "腕表"))
)
WUPING_CLOSET_DIR = Path(
    os.environ.get(
        "WARDROBE_WUPING_CLOSET_DIR",
        r"C:\Users\xuxin\SynologyDrive\Hermes-吴萍\衣橱",
    )
)
HOST_CLOSET_DIR = Path(
    os.environ.get("WARDROBE_CLOSET_HOST_DIR", str(DEFAULT_IMPORT_DIR))
)
HOST_WATCH_DIR = Path(
    os.environ.get("WARDROBE_WATCH_HOST_DIR", str(WATCH_IMPORT_DIR))
)
HOST_WUPING_CLOSET_DIR = Path(
    os.environ.get("WARDROBE_WUPING_CLOSET_HOST_DIR", str(WUPING_CLOSET_DIR))
)
LOCAL_LIBRARY_ROOT = ROOT_DIR / "baseline_exports"
if "WARDROBE_LIBRARY_ROOT" not in os.environ:
    LIBRARY_ROOT = LOCAL_LIBRARY_ROOT
if "WARDROBE_CLOSET_DIR" not in os.environ:
    DEFAULT_IMPORT_DIR = LIBRARY_ROOT / "衣橱"
if "WARDROBE_WATCH_DIR" not in os.environ:
    WATCH_IMPORT_DIR = LIBRARY_ROOT / "腕表"
if "WARDROBE_WUPING_CLOSET_DIR" not in os.environ:
    if os.name == "nt":
        WUPING_CLOSET_DIR = Path(r"C:\Users\xuxin\SynologyDrive\Hermes-吴萍\衣橱")
    else:
        WUPING_CLOSET_DIR = DEFAULT_IMPORT_DIR
if "WARDROBE_CLOSET_HOST_DIR" not in os.environ:
    HOST_CLOSET_DIR = DEFAULT_IMPORT_DIR
if "WARDROBE_WATCH_HOST_DIR" not in os.environ:
    HOST_WATCH_DIR = WATCH_IMPORT_DIR
if "WARDROBE_WUPING_CLOSET_HOST_DIR" not in os.environ:
    HOST_WUPING_CLOSET_DIR = WUPING_CLOSET_DIR
DRIVE_NOTIFY_QUEUE_DIR = DATA_DIR / "drive-notify-queue"
API_TOKEN_SECRET_DIR = Path(
    os.environ.get("WARDROBE_API_TOKEN_SECRET_DIR", str(DATA_DIR / "api-token-secrets"))
)
WARDROBE_FILE = "\u8863\u6a71.xlsx"
WARDROBE_TEXT_EXPORT_FILE = "\u8863\u6a71.csv"
WARDROBE_CHATGPT_RULES_FILE = "\u8863\u6a71_ChatGPT\u89e3\u6790\u89c4\u5219.md"
WEARCOUNT_CHATGPT_RULES_FILE = "WearCount_ChatGPT\u89e3\u6790\u89c4\u5219.md"
WARDROBE_PROGRAM_API_DOC_FILE = "Hermes_\u8863\u6a71_API\u89c4\u8303.md"
WARDROBE_HERMES_STYLE_RULES_FILE = "Hermes_\u8863\u6a71\u642d\u914d\u89c4\u5219.md"
WEARCOUNT_TEXT_EXPORT_FILE = "WearCount.csv"
WEARCOUNT_NEW_FILE_IMPORT_ENABLED = False
WARDROBE_OWNER_EXPORT_SPECS = [
    ("\u8863\u6a71_\u5f90\u6b23.csv", {"\u5f90\u6b23"}),
    ("\u8863\u6a71_\u5434\u840d.csv", {"\u5434\u840d"}),
]
LOOKS_OWNER_EXPORT_SPECS = [
    ("\u5957\u88c5_\u5f90\u6b23.csv", {"\u5f90\u6b23"}),
    ("\u5957\u88c5_\u5434\u840d.csv", {"\u5434\u840d"}),
]
WEARCOUNT_OWNER_EXPORT_SPECS = [
    ("WearCount_\u5f90\u6b23.csv", {"\u5f90\u6b23"}),
    ("WearCount_\u5434\u840d.csv", {"\u5434\u840d"}),
]
WATCH_FILE = "\u8155\u8868.xlsx"
WATCH_TEXT_EXPORT_FILE = "\u8155\u8868.csv"
LOOKS_FILE = "\u5957\u88c5.xlsx"
WEARCOUNT_EXPORT_FILE = "WearCount.xlsx"
RETIRED_CLOSET_EXPORT_FILES = [
    "\u8863\u6a71.xlsx",
    "\u8863\u6a71_\u5f90\u6b23.xlsx",
    "\u8863\u6a71_\u5434\u840d.xlsx",
    "\u8863\u6a71_\u5434\u840d_\u5bb6\u5ead.xlsx",
    "\u5957\u88c5.xlsx",
    "\u5957\u88c5_\u5f90\u6b23.xlsx",
    "\u5957\u88c5_\u5434\u840d.xlsx",
    "WearCount.xlsx",
    "WearCount_\u5f90\u6b23.xlsx",
    "WearCount_\u5434\u840d.xlsx",
]
RETIRED_WATCH_EXPORT_FILES = [
    "\u8155\u8868.xlsx",
]
IMAGE_INDEX_PATTERN = "\u8863\u6a71\u56fe\u7247\u7d22\u5f15*.xlsx"
WARDROBE_WORD = "\u8863\u6a71"
WATCH_WORD = "\u8155\u8868"
LOOKS_WORD = "\u5957\u88c5"
IMAGE_INDEX_WORD = "\u56fe\u7247\u7d22\u5f15"
FILE_IMPORT_API_ENABLED = False
ITEM_EXPORT_LOCK = threading.Lock()
ITEM_EXPORT_PENDING: set[str] = set()
ITEM_EXPORT_RUNNING = False
AUTH_COOKIE_NAME = "wardrobe_session"
AUTH_USER_HASHES = {
    "\u5f90\u6b23": "pbkdf2_sha256$200000$d8107fe05a45a6fc7315fd33ced88e81$940733d9689ca4733f9cbd17df75b9e13cbd6d07a7e1881340606ea96b458671",
    "\u5434\u840d": "pbkdf2_sha256$200000$17c186233a787b86e22bf125833cbca6$f722418c646ba14f58621b5813850669ab19086e5ac29770c11e17ae3c855b73",
}
AUTH_ADMIN_USERS = {"\u5f90\u6b23"}
CATALOG_MANAGER_USERS = {"\u5f90\u6b23"}
AUTH_SHARED_OWNERS = {"\u5bb6\u5ead"}
AUTH_MAX_ATTEMPTS = 3
AUTH_SESSION_IDLE_SECONDS = int(os.environ.get("WARDROBE_AUTH_IDLE_SECONDS", "43200"))
AUTH_SESSION_ABSOLUTE_SECONDS = int(os.environ.get("WARDROBE_AUTH_ABSOLUTE_SECONDS", "604800"))
AUTH_SESSION_TOUCH_INTERVAL_SECONDS = int(os.environ.get("WARDROBE_AUTH_TOUCH_INTERVAL_SECONDS", "60"))
AUTH_MAX_SESSIONS_PER_USER = max(1, int(os.environ.get("WARDROBE_AUTH_MAX_SESSIONS_PER_USER", "2")))
AUTH_LOGIN_WINDOW_SECONDS = int(os.environ.get("WARDROBE_AUTH_LOGIN_WINDOW_SECONDS", "900"))
AUTH_LOGIN_MAX_PER_IP = int(os.environ.get("WARDROBE_AUTH_LOGIN_MAX_PER_IP", "20"))
AUTH_LOGIN_MAX_PER_USER_IP = int(os.environ.get("WARDROBE_AUTH_LOGIN_MAX_PER_USER_IP", "8"))
AUTH_PASSWORD_MIN_LENGTH = int(os.environ.get("WARDROBE_AUTH_PASSWORD_MIN_LENGTH", "8"))
AUTH_PASSWORD_MAX_LENGTH = int(os.environ.get("WARDROBE_AUTH_PASSWORD_MAX_LENGTH", "24"))
AUTH_PASSWORD_DEFAULT_ITERATIONS = int(os.environ.get("WARDROBE_AUTH_PASSWORD_ITERATIONS", "200000"))
AUTH_PASSWORD_SALT_BYTES = int(os.environ.get("WARDROBE_AUTH_PASSWORD_SALT_BYTES", "16"))
API_TOKEN_PREFIX = os.environ.get("WARDROBE_API_TOKEN_PREFIX", "wd_live_")
API_TOKEN_MIN_LENGTH = 24
API_TOKEN_TOUCH_INTERVAL_SECONDS = int(os.environ.get("WARDROBE_API_TOKEN_TOUCH_INTERVAL_SECONDS", "60"))
AUTH_ALLOWED_ORIGINS = [
    value.strip()
    for value in os.environ.get("WARDROBE_ALLOWED_ORIGINS", "").split(",")
    if value.strip()
]
HERMES_PLUGIN_FRAME_ANCESTORS = [
    value.strip()
    for value in os.environ.get(
        "WARDROBE_HERMES_PLUGIN_FRAME_ANCESTORS",
        "'self' http://127.0.0.1:* http://localhost:*",
    ).split()
    if value.strip()
]
HERMES_PLUGIN_FRAME_ANCESTOR_SETTINGS_KEY = hermes_plugin.FRAME_ANCESTORS_SETTING_KEY
LOGIN_RATE_LOCK = threading.Lock()
LOGIN_RATE_BY_IP: dict[str, list[float]] = {}
LOGIN_RATE_BY_USER_IP: dict[str, list[float]] = {}
UPLOAD_MAX_BYTES = 20 * 1024 * 1024
UPLOAD_ALLOWED_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".heic"}
PROGRAM_API_ITEM_PHOTO_MAX_COUNT = max(
    0,
    int(os.environ.get("WARDROBE_PROGRAM_API_ITEM_PHOTO_MAX_COUNT", "12")),
)
PROGRAM_API_ITEM_PHOTO_MAX_TOTAL_BYTES = max(
    UPLOAD_MAX_BYTES,
    int(os.environ.get("WARDROBE_PROGRAM_API_ITEM_PHOTO_MAX_TOTAL_BYTES", str(80 * 1024 * 1024))),
)
XLSX_NS = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
EXPORT_UID = int(os.environ["WARDROBE_EXPORT_UID"]) if os.environ.get("WARDROBE_EXPORT_UID") else None
EXPORT_GID = int(os.environ["WARDROBE_EXPORT_GID"]) if os.environ.get("WARDROBE_EXPORT_GID") else None
EXIF_GPS_INFO_TAG = next((tag for tag, name in ExifTags.TAGS.items() if name == "GPSInfo"), 34853)
AI_SECRET_FILE = DATA_DIR / "dashscope_api_key.txt"
AI_BASE_URL = os.environ.get("WARDROBE_AI_BASE_URL", "https://dashscope.aliyuncs.com/compatible-mode/v1").rstrip("/")
AI_MODEL = os.environ.get("WARDROBE_AI_MODEL", "qwen3-max").strip() or "qwen3-max"
AI_TIMEOUT_SECONDS = max(5.0, float(os.environ.get("WARDROBE_AI_TIMEOUT_SECONDS", "90")))
AI_TIMEOUT_RETRIES = max(0, int(os.environ.get("WARDROBE_AI_TIMEOUT_RETRIES", "1")))
AI_ROLE_CANDIDATE_LIMIT = max(1, int(os.environ.get("WARDROBE_AI_ROLE_CANDIDATE_LIMIT", "5")))
AI_REVIEW_LOCK = threading.Lock()
AI_REVIEW_PENDING: set[tuple[str, int]] = set()
NATIVE_AI_REMOVED_MESSAGE = "Wardrobe native AI has been removed. Use Hermes Mobile MCP for model-facing wardrobe service."
DEFAULT_AI_PROMPTS = {
    "outfit": (
        "请基于当前历史记录，先判断原搭配在当时场景、温区、材质和颜色条件下是否成立，再输出完整分析。"
        "需要覆盖整体判断、搭配优点、主要问题、单件点评、替换建议、补全建议、腕表建议、轮换优先级提醒。"
        "历史记录分析默认不考虑 Wear count、wear_year、last_worn_on 等轮换字段。"
        "不要因为当前 wear_year、wear_total 或腕表本年佩戴次数较高，就反推原搭配本身不合理。"
        "若原搭配已经成立，替换建议只能作为备选参考，不能为了低出场率强行替换。"
    ),
    "outfit_draft": (
        "请基于当前录入中的今日穿搭草稿进行分析，并尽量给出可直接用于当前表单的角色选择建议。"
        "优先保留已经合理的已选单品，优先补足缺失角色；只有在明显不匹配时才替换已选角色。"
        "推荐时优先考虑本年穿得更少、最近更久未穿、累计穿着更少的单品。"
    ),
    "featured_look": (
        "这次任务只分析当前套装本身，并给出套装评分。"
        "不要提供替换建议、补全建议、腕表建议、优化建议或新的搭配方案。"
        "输出聚焦于整体判断、搭配优点、主要问题、单件点评、套装评分。"
    ),
}


def _web_asset_version() -> str:
    asset_names = ("index.html", "app.js", "styles.css")
    parts: list[str] = []
    for name in asset_names:
        path = WEB_DIR / name
        try:
            stat = path.stat()
            parts.append(f"{name}:{int(stat.st_mtime)}:{stat.st_size}")
        except OSError:
            parts.append(f"{name}:missing")
    return hashlib.sha256("|".join(parts).encode("utf-8")).hexdigest()[:16]


APP_WEB_VERSION = _web_asset_version()
WARDROBE_EDIT_HEADERS = {
    "code": ["货号", "Code"],
    "brand": ["品牌", "Brand"],
    "section": ["Section", "Item"],
    "loc": ["Loc"],
    "owner": ["Owner"],
    "layer_role": ["LayerRole"],
    "outer_type": ["Outer_Type"],
    "scene_tag": ["SceneTag"],
    "relax_index": ["Relax_Index"],
    "temp_min": ["Temp_Min"],
    "temp_max": ["Temp_Max"],
    "standalone_min": ["Standalone_Min"],
    "standalone_max": ["Standalone_Max"],
    "primary_color": ["主色系", "颜色"],
    "secondary_color": ["第二色系"],
    "official_desc": ["官网描述", "官方描述（中国大陆官网）", "官方描述"],
    "price_original": ["原始价格"],
    "price_original_currency": ["原始货币", "原始币种"],
    "price_cny": ["人民币价格"],
    "series": ["系列"],
    "size": ["尺码"],
    "acquired_at": ["入库时间"],
    "official_color_code": ["官方色号"],
    "material": ["材质"],
    "care": ["洗涤方式"],
    "notes": ["说明"],
}
WATCH_EDIT_HEADERS = {
    "section": ["name", "Section", "名称"],
    "code": ["Ref", "Code", "货号"],
    "brand": ["品牌", "Brand"],
    "material": ["材质"],
    "notes": ["机芯", "说明", "Notes"],
    "acquired_at": ["购买日期", "入库时间"],
    "price_original": ["购买价格", "价格", "Price"],
    "price_original_currency": ["原始货币", "原始币种", "Currency"],
    "official_desc": ["中文说明", "说明"],
}
EDIT_DATE_FIELDS = {"acquired_at"}
PRICE_TEXT_FIELDS = {"price_original", "price_cny"}
EDIT_NUMBER_FIELDS = {
    "relax_index",
    "temp_min",
    "temp_max",
    "standalone_min",
    "standalone_max",
    "wear_maintenance",
    "wear_threshold",
}
ITEM_DB_FIELDS = [
    "code", "brand", "section", "loc", "owner", "layer_role", "outer_type", "scene_tag",
    "relax_index", "temp_min", "temp_max", "standalone_min", "standalone_max",
    "wear_maintenance", "wear_threshold",
    "primary_color", "secondary_color", "official_desc", "price_original", "price_original_currency", "price_cny",
    "series", "size", "acquired_at", "official_color_code", "material", "care", "notes",
    "status", "source_sheet",
]
WARDROBE_EXPORT_HEADERS = [
    "序号", "品牌", "Section", "Loc", "保养状态", "Owner", "LayerRole", "Outer_Type", "SceneTag",
    "Relax_Index", "Temp_Min", "Temp_Max", "Standalone_Min", "Standalone_Max",
    "主色系", "第二色系", "官网描述", "原始价格", "原始货币", "人民币价格", "系列", "货号", "尺码",
    "入库时间", "官方色号", "材质", "洗涤方式", "说明",
]
WARDROBE_TEXT_EXPORT_HEADERS = [
    "code",
    "owner",
    "brand",
    "section",
    "loc",
    "status",
    "maintenance_state",
    "layer_role",
    "outer_type",
    "scene_tag",
    "relax_index",
    "temp_min",
    "temp_max",
    "standalone_min",
    "standalone_max",
    "primary_color",
    "secondary_color",
    "material",
    "series",
    "official_color_code",
    "price_original",
    "price_original_currency",
    "price_cny",
    "acquired_at",
    "wear_total",
    "wear_year",
    "last_worn_on",
    "wear_maintenance",
    "wear_threshold",
    "maint_count",
]
WARDROBE_XLSX_EXPORT_FIELDS = {
    "code",
    "brand",
    "section",
    "loc",
    "owner",
    "layer_role",
    "outer_type",
    "scene_tag",
    "relax_index",
    "temp_min",
    "temp_max",
    "standalone_min",
    "standalone_max",
    "primary_color",
    "secondary_color",
    "official_desc",
    "price_original",
    "price_original_currency",
    "price_cny",
    "series",
    "size",
    "acquired_at",
    "official_color_code",
    "material",
    "care",
    "notes",
    "maintenance_state",
}
WARDROBE_AI_EXPORT_FIELDS = {
    "code",
    "owner",
    "brand",
    "section",
    "loc",
    "status",
    "maintenance_state",
    "layer_role",
    "outer_type",
    "scene_tag",
    "relax_index",
    "temp_min",
    "temp_max",
    "standalone_min",
    "standalone_max",
    "primary_color",
    "secondary_color",
    "material",
    "series",
    "official_color_code",
    "price_original",
    "price_original_currency",
    "price_cny",
    "acquired_at",
    "wear_total",
    "wear_year",
    "last_worn_on",
    "wear_maintenance",
    "wear_threshold",
    "maint_count",
}
WATCH_XLSX_EXPORT_FIELDS = {
    "code",
    "brand",
    "status",
    "section",
    "loc",
    "owner",
    "scene_tag",
    "official_desc",
    "price_original",
    "price_original_currency",
    "price_cny",
    "acquired_at",
    "material",
    "notes",
    "maintenance_state",
}
WATCH_AI_EXPORT_FIELDS = {
    "code",
    "owner",
    "brand",
    "section",
    "loc",
    "status",
    "maintenance_state",
    "scene_tag",
    "material",
    "price_original",
    "price_original_currency",
    "price_cny",
    "acquired_at",
    "wear_total",
    "wear_year",
    "last_worn_on",
    "maint_count",
}
WEARCOUNT_EXPORT_FIELDS = {
    "code",
    "section",
    "brand",
    "owner",
    "layer_role",
    "status",
    "wear_maintenance",
    "wear_total",
    "wear_year",
    "maint_count",
    "wear_threshold",
}
ITEM_EXPORT_TASK_ORDER = ["wardrobe_ai", "watch_ai", "looks_ai", "wearcount"]
WARDROBE_TEXT_EXPORT_FIELD_SPECS = [
    ("code", "code"),
    ("owner", "owner"),
    ("brand", "brand"),
    ("section", "section"),
    ("loc", "loc"),
    ("status", "status"),
    ("maintenance_state", "maintenance_state"),
    ("layer_role", "layer_role"),
    ("outer_type", "outer_type"),
    ("scene_tag", "scene_tag"),
    ("relax_index", "relax_index"),
    ("temp_min", "temp_min"),
    ("temp_max", "temp_max"),
    ("standalone_min", "standalone_min"),
    ("standalone_max", "standalone_max"),
    ("primary_color", "primary_color"),
    ("secondary_color", "secondary_color"),
    ("material", "material"),
    ("series", "series"),
    ("official_color_code", "official_color_code"),
    ("price_original", "price_original"),
    ("price_original_currency", "price_original_currency"),
    ("price_cny", "price_cny"),
    ("acquired_at", "acquired_at"),
    ("wear_total", "wear_total"),
    ("wear_year", "wear_year"),
    ("last_worn_on", "last_worn_on"),
    ("wear_maintenance", "wear_maintenance"),
    ("wear_threshold", "wear_threshold"),
    ("maint_count", "maint_count"),
]
WARDROBE_TEXT_EXPORT_FIELD_DESCRIPTIONS = {
    "code": "货号，唯一键。所有单品识别、联表、Program API 历史写入都以它为准。",
    "owner": "归属人。当前实际值以基线为准，默认推荐对象未指定时按徐欣处理。",
    "brand": "品牌。",
    "section": "完整产品名称。禁止截断或自行改写。",
    "loc": "真实地点信息。",
    "status": "当前状态。做主推荐时必须使用 Active；Ordered/ordered 只表示已订购未入库，硬性排除，不参与推荐。",
    "maintenance_state": "保养状态。`激活` 表示正常可穿，`保养中` 表示不应进入正式主案。",
    "layer_role": "衣物基础角色。规范值包括 Inner、Middle、Outer、Bottom、Footwear、Watch、Accessory、Dress、Home、Bespoke；鞋类必须写 Footwear，不再写 Shoes。",
    "outer_type": "当 layer_role=Outer 时使用的外层语义，如 Coat、Jacket、Knit_Outer、Shirt_Outer、Vest_Outer。",
    "scene_tag": "场景语义。当前主值为 Comfort、City、Outdoor、Home、Watch。",
    "relax_index": "视觉松弛度，1-5。",
    "temp_min": "该单品成立的温区下限。",
    "temp_max": "该单品成立的温区上限。",
    "standalone_min": "Inner/Middle 脱外套后单穿成立的温区下限。",
    "standalone_max": "Inner/Middle 脱外套后单穿成立的温区上限。",
    "primary_color": "主色系。色系判断主轴。",
    "secondary_color": "第二色系。只用于桥接与呼应，不应制造新主轴。",
    "material": "材质字段，唯一真相源。正式推荐前应优先读取。",
    "series": "系列信息。",
    "official_color_code": "官方色名/官方颜色编码，用于标准化辅助，不直接替代色系判断。",
    "price_original": "原始入库/购买金额，纯数字字符串。",
    "price_original_currency": "price_original 的原始货币，使用大写币种代码，如 CNY、EUR、USD、HKD。",
    "price_cny": "人民币价格，可用于消费与使用效率分析，纯数字字符串。",
    "acquired_at": "入库/购入日期。",
    "wear_total": "累计总穿着次数。",
    "wear_year": "当年穿着次数。",
    "last_worn_on": "最近一次穿着日期。",
    "wear_maintenance": "保养磨损累计值。",
    "wear_threshold": "保养阈值；大于 0 才有保养判断意义。",
    "maint_count": "已保养次数。",
}
WATCH_TEXT_EXPORT_HEADERS = [
    "code",
    "owner",
    "brand",
    "section",
    "loc",
    "status",
    "maintenance_state",
    "scene_tag",
    "material",
    "price_original",
    "price_original_currency",
    "price_cny",
    "acquired_at",
    "wear_total",
    "wear_year",
    "last_worn_on",
    "maint_count",
]
LOOKS_TEXT_EXPORT_FILE = "套装.csv"
LOOKS_TEXT_EXPORT_HEADERS = [
    "look_id",
    "owner",
    "status",
    "use_case",
    "temp_min",
    "temp_max",
    "relax_center",
    "relax_span",
    "inner_code",
    "middle_code",
    "outer_code",
    "bottom_code",
    "footwear_code",
    "watch_ref",
    "notes",
    "created_at",
]
WEARCOUNT_NEW_DAILY_LOG_FIELDS = [
    ("Date", "必填，北京时间日期，格式 YYYY-MM-DD。"),
    ("City", "建议填写，当天主要城市。"),
    ("Wear_Mode", "建议填写。当前系统已支持独立导入，常用值可写 `normal` 或 `home`。"),
    ("Scene", "必填，真实场景，不要写笼统的 `Normal`。例如 Family_Dinner、Walk、Client。"),
    ("Forecast_Temp", "必填，核心活动时段预报温度口径，建议写 `19-24` 或 `19.0-24.0`。"),
    ("Notes", "选填。"),
    ("Owner", "建议填写。未指定时默认徐欣。"),
]
WEARCOUNT_NEW_ITEMS_FIELDS = [
    ("Code", "必填，正式货号，必须能命中当前基线。"),
    ("Role", "强烈建议填写，允许值：Outer、Middle、Inner、Bottom、Footwear、Accessory、Watch。"),
    ("Has_Base_Layer", "必填，是否有打底/隔离层；接受 1/0、true/false、yes/no、是/否。"),
]
WEARCOUNT_NEW_ROLE_MAPPINGS = [
    ("Outer", "Outer"),
    ("Middle", "Middle"),
    ("Inner", "Inner"),
    ("Bottom", "Bottom"),
    ("Footwear", "Footwear"),
    ("Accessory", "Accessory"),
    ("Watch", "Watch"),
]
WATCH_EXPORT_HEADERS = [
    "Ref", "品牌", "状态", "name", "Loc", "保养状态", "Owner", "SceneTag", "中文说明", "价格", "原始货币", "购买日期", "材质", "机芯",
]
FEATURED_LOOK_EXPORT_HEADERS = [
    "LookID", "Anchor_Type", "Anchor_Code", "Anchor_Section", "Use_Case", "Priority", "Status", "Owner",
    "Temp_Min", "Temp_Max", "SceneTag_Target", "Relax_Center", "Relax_Span", "Notes",
    "Inner_Code", "Mid_Code", "Outer_Code", "Bottom_Code", "Footwear_Code", "Watch_Ref",
]
WEARCOUNT_EXPORT_HEADERS = ["Code", "Item", "Brand", "Wear", "Total", "2026", "Maint", "Thr", "Status"]
WEARCOUNT_TEXT_EXPORT_HEADERS = [
    "role",
    "code",
    "item",
    "brand",
    "owner",
    "wear",
    "total",
    "wear_year",
    "maint",
    "thr",
    "status",
    "maintenance_state",
    "last_worn_on",
]
WEARCOUNT_DAILY_LOG_HEADERS = [
    "Date", "City", "Wear_Mode", "Avg_Relax", "Avg_Temp", "Notes",
    "Inner_Codes", "Middle_Codes", "Outer_Codes", "Bottom_Codes", "Footwear_Codes", "Accessory_Codes", "Watch_Refs",
]


def _local_ipv4_addresses() -> list[str]:
    addresses = {"127.0.0.1"}
    try:
        _, _, host_ips = socket.gethostbyname_ex(socket.gethostname())
        addresses.update(ip for ip in host_ips if "." in ip)
    except Exception:
        pass
    return sorted(addresses)


def _cookie_value(cookie_header: str, key: str) -> str:
    if not cookie_header:
        return ""
    for chunk in cookie_header.split(";"):
        name, _, value = chunk.strip().partition("=")
        if name == key:
            return value.strip()
    return ""


def _pbkdf2_hash(password: str, iterations: int, salt_hex: str) -> str:
    derived = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), bytes.fromhex(salt_hex), iterations)
    return derived.hex()


def _create_password_hash(password: str, iterations: int = AUTH_PASSWORD_DEFAULT_ITERATIONS) -> str:
    salt_hex = os.urandom(max(8, AUTH_PASSWORD_SALT_BYTES)).hex()
    digest_hex = _pbkdf2_hash(password, iterations, salt_hex)
    return f"pbkdf2_sha256${iterations}${salt_hex}${digest_hex}"


def _verify_password_hash(password: str, stored_hash: str) -> bool:
    try:
        algorithm, iteration_text, salt_hex, digest_hex = stored_hash.split("$", 3)
        if algorithm != "pbkdf2_sha256":
            return False
        expected = _pbkdf2_hash(password, int(iteration_text), salt_hex)
        return hmac.compare_digest(expected, digest_hex)
    except Exception:
        return False


def _password_policy_error(password: str) -> str:
    if len(password) < AUTH_PASSWORD_MIN_LENGTH:
        return f"新密码至少 {AUTH_PASSWORD_MIN_LENGTH} 位。"
    if len(password) > AUTH_PASSWORD_MAX_LENGTH:
        return f"新密码最多 {AUTH_PASSWORD_MAX_LENGTH} 位。"
    if any(char.isspace() for char in password):
        return "新密码不能包含空格。"
    checks = (
        (r"[A-Z]", "至少包含 1 个大写字母。"),
        (r"[a-z]", "至少包含 1 个小写字母。"),
        (r"\d", "至少包含 1 个数字。"),
        (r"[^0-9A-Za-z]", "至少包含 1 个特殊字符。"),
    )
    for pattern, message in checks:
        if not re.search(pattern, password):
            return f"新密码需为复杂密码：{message}"
    return ""


def _sync_auth_users(conn: sqlite3.Connection) -> None:
    for username, password_hash in AUTH_USER_HASHES.items():
        conn.execute(
            """
            INSERT INTO auth_users (username, password_hash, enabled, updated_at)
            VALUES (?, ?, 1, CURRENT_TIMESTAMP)
            ON CONFLICT(username) DO UPDATE SET
                enabled = 1,
                updated_at = CURRENT_TIMESTAMP
            """,
            (username, password_hash),
        )
        conn.execute(
            """
            INSERT INTO auth_attempts (username, failed_attempts, locked, updated_at)
            VALUES (?, 0, 0, CURRENT_TIMESTAMP)
            ON CONFLICT(username) DO UPDATE SET
                failed_attempts = 0,
                locked = 0,
                updated_at = CURRENT_TIMESTAMP
            """,
            (username,),
        )
    conn.commit()


def _normalize_authoritative_owners(conn: sqlite3.Connection) -> None:
    return


def _writable_owner_values(username: str) -> list[str]:
    if _user_is_admin(username):
        return []
    normalized = _normalize_edit_value(username)
    return [normalized] if normalized else []


def _owner_read_sql(column: str, username: str) -> tuple[str, list[str]]:
    if _user_is_admin(username):
        return "1 = 1", []
    normalized = _normalize_edit_value(username)
    if not normalized:
        return "1 = 0", []
    return f"COALESCE({column}, '') = ?", [normalized]


def _owner_access_sql(column: str, username: str) -> tuple[str, list[str]]:
    if _user_is_admin(username):
        return "1 = 1", []
    owners = _writable_owner_values(username)
    if not owners:
        return "1 = 0", []
    clauses = [f"COALESCE({column}, '') = ?" for _ in owners]
    return f"({' OR '.join(clauses)})", owners


def _owner_value_viewable(owner: object, username: str) -> bool:
    if _user_is_admin(username):
        return True
    normalized_username = _normalize_edit_value(username)
    if not normalized_username:
        return False
    return _normalize_edit_value(owner) == normalized_username


def _owner_value_allowed(owner: object, username: str) -> bool:
    if _user_is_admin(username):
        return True
    normalized_username = _normalize_edit_value(username)
    if not normalized_username:
        return False
    return _normalize_edit_value(owner) == normalized_username


def _user_is_admin(username: str) -> bool:
    return _normalize_edit_value(username) in {_normalize_edit_value(value) for value in AUTH_ADMIN_USERS}


def _user_can_manage_catalog(username: str) -> bool:
    return _normalize_edit_value(username) in {_normalize_edit_value(value) for value in CATALOG_MANAGER_USERS}


def _app_setting_value(conn: sqlite3.Connection, key: str, default: str = "") -> str:
    row = conn.execute("SELECT value FROM app_settings WHERE key = ?", (key,)).fetchone()
    if row is None:
        return default
    return _normalize_edit_value(row["value"])


def _set_app_setting(conn: sqlite3.Connection, key: str, value: str) -> None:
    conn.execute(
        """
        INSERT INTO app_settings(key, value, updated_at)
        VALUES(?, ?, CURRENT_TIMESTAMP)
        ON CONFLICT(key) DO UPDATE SET
            value = excluded.value,
            updated_at = CURRENT_TIMESTAMP
        """,
        (key, value),
    )


def _ai_prompt_record_type(record_type: str) -> str:
    normalized = _normalize_edit_value(record_type)
    if normalized in {"featured_look", "featured-look"}:
        return "featured_look"
    if normalized in {"outfit_draft", "outfit-draft"}:
        return "outfit_draft"
    return "outfit"


def _ai_prompt_setting_key(record_type: str) -> str:
    return f"{_ai_prompt_record_type(record_type)}_ai_prompt"


def _ai_prompt_setting(conn: sqlite3.Connection, record_type: str) -> str:
    normalized_type = _ai_prompt_record_type(record_type)
    configured = _app_setting_value(conn, _ai_prompt_setting_key(normalized_type), "")
    return configured or DEFAULT_AI_PROMPTS[normalized_type]


def _all_ai_prompt_settings(conn: sqlite3.Connection) -> dict:
    return {
        "outfit": _ai_prompt_setting(conn, "outfit"),
        "outfit_draft": _ai_prompt_setting(conn, "outfit_draft"),
        "featured_look": _ai_prompt_setting(conn, "featured_look"),
    }


def _native_ai_removed_payload() -> dict:
    return {
        "error": "native_ai_removed",
        "message": NATIVE_AI_REMOVED_MESSAGE,
    }


def _session_is_valid(row: sqlite3.Row | None) -> bool:
    if row is None:
        return False
    now = datetime.utcnow()
    for field_name, timeout in (("created_at", AUTH_SESSION_ABSOLUTE_SECONDS), ("last_seen_at", AUTH_SESSION_IDLE_SECONDS)):
        raw_value = str(row[field_name] or "").strip()
        if not raw_value:
            return False
        try:
            seen_at = datetime.strptime(raw_value, "%Y-%m-%d %H:%M:%S")
        except ValueError:
            return False
        if (now - seen_at).total_seconds() > timeout:
            return False
    return True


def _timestamp_older_than(value: object, seconds: int) -> bool:
    if seconds <= 0:
        return True
    normalized = _normalize_edit_value(value)
    if not normalized:
        return True
    try:
        parsed = datetime.strptime(normalized[:19], "%Y-%m-%d %H:%M:%S")
    except ValueError:
        return True
    return (datetime.utcnow() - parsed).total_seconds() >= seconds


def _api_should_touch_token(required_scope: str) -> bool:
    return not _normalize_edit_value(required_scope).endswith(":read")


def _normalize_netloc(value: str) -> str:
    if not value:
        return ""
    parsed = urlparse(value if "://" in value else f"//{value}", scheme="http")
    netloc = parsed.netloc.lower()
    if netloc.endswith(":80"):
        return netloc[:-3]
    if netloc.endswith(":443"):
        return netloc[:-4]
    return netloc


def _prune_attempts(values: list[float], now: float) -> list[float]:
    return [value for value in values if now - value < AUTH_LOGIN_WINDOW_SECONDS]


def _register_login_failure(client_ip: str, username: str) -> None:
    now = time.time()
    ip_key = client_ip or "unknown"
    user_key = f"{ip_key}|{_normalize_edit_value(username)}"
    with LOGIN_RATE_LOCK:
        ip_attempts = _prune_attempts(LOGIN_RATE_BY_IP.get(ip_key, []), now)
        user_attempts = _prune_attempts(LOGIN_RATE_BY_USER_IP.get(user_key, []), now)
        ip_attempts.append(now)
        user_attempts.append(now)
        LOGIN_RATE_BY_IP[ip_key] = ip_attempts
        LOGIN_RATE_BY_USER_IP[user_key] = user_attempts


def _clear_login_failures(client_ip: str, username: str) -> None:
    ip_key = client_ip or "unknown"
    user_key = f"{ip_key}|{_normalize_edit_value(username)}"
    with LOGIN_RATE_LOCK:
        LOGIN_RATE_BY_USER_IP.pop(user_key, None)


def _login_rate_limited(client_ip: str, username: str) -> tuple[bool, int]:
    now = time.time()
    ip_key = client_ip or "unknown"
    user_key = f"{ip_key}|{_normalize_edit_value(username)}"
    with LOGIN_RATE_LOCK:
        ip_attempts = _prune_attempts(LOGIN_RATE_BY_IP.get(ip_key, []), now)
        user_attempts = _prune_attempts(LOGIN_RATE_BY_USER_IP.get(user_key, []), now)
        LOGIN_RATE_BY_IP[ip_key] = ip_attempts
        LOGIN_RATE_BY_USER_IP[user_key] = user_attempts
        retry_after = 0
        if len(ip_attempts) >= AUTH_LOGIN_MAX_PER_IP and ip_attempts:
            retry_after = max(retry_after, int(AUTH_LOGIN_WINDOW_SECONDS - (now - ip_attempts[0])) + 1)
        if len(user_attempts) >= AUTH_LOGIN_MAX_PER_USER_IP and user_attempts:
            retry_after = max(retry_after, int(AUTH_LOGIN_WINDOW_SECONDS - (now - user_attempts[0])) + 1)
        return retry_after > 0, max(retry_after, 1) if retry_after > 0 else 0


def _apply_owner_scope_to_payload(payload: dict, username: str, current_owner: object = None) -> dict:
    scoped = dict(payload)
    current_owner_value = _normalize_edit_value(current_owner)
    if _user_is_admin(username):
        owner_value = _normalize_edit_value(scoped.get("owner"))
        scoped["owner"] = owner_value or current_owner_value or _normalize_edit_value(username)
        return scoped
    normalized_username = _normalize_edit_value(username)
    if current_owner_value:
        if current_owner_value != normalized_username:
            raise PermissionError("forbidden_owner")
        scoped["owner"] = current_owner_value
        return scoped
    if not normalized_username:
        raise PermissionError("forbidden_owner")
    scoped["owner"] = normalized_username
    return scoped


def _apply_item_owner_scope_to_payload(payload: dict, username: str, current_owner: object = None) -> dict:
    scoped = dict(payload)
    if _user_can_manage_catalog(username):
        owner_value = _normalize_edit_value(scoped.get("owner"))
        scoped["owner"] = owner_value or _normalize_edit_value(current_owner) or _normalize_edit_value(username)
        return scoped
    current_owner_value = _normalize_edit_value(current_owner)
    if current_owner_value:
        if current_owner_value != _normalize_edit_value(username):
            raise PermissionError("forbidden_owner")
        scoped["owner"] = current_owner_value
        return scoped
    scoped["owner"] = _normalize_edit_value(username)
    return scoped


def _ensure_option_catalog_table(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS option_catalogs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            option_type TEXT NOT NULL,
            value TEXT NOT NULL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(option_type, value)
        );
        CREATE INDEX IF NOT EXISTS idx_option_catalogs_type_value
        ON option_catalogs(option_type, value);
        """
    )


def _validate_image_upload(file_part: dict) -> str | None:
    filename = str(file_part.get("filename") or "").strip()
    suffix = Path(filename).suffix.lower()
    content = file_part.get("content") or b""
    content_type = str(file_part.get("content_type") or "").lower()
    if suffix not in UPLOAD_ALLOWED_EXTENSIONS:
        return "invalid_file_type"
    if len(content) > UPLOAD_MAX_BYTES:
        return "file_too_large"
    if content_type and not content_type.startswith("image/"):
        return "invalid_file_type"
    return None


def _owner_accounts(conn: sqlite3.Connection) -> list[dict]:
    rows = conn.execute(
        """
        SELECT auth_users.username AS username,
               COALESCE(auth_attempts.failed_attempts, 0) AS failed_attempts,
               COALESCE(auth_attempts.locked, 0) AS locked
        FROM auth_users
        LEFT JOIN auth_attempts ON auth_attempts.username = auth_users.username
        WHERE auth_users.enabled = 1
        ORDER BY auth_users.username
        """
    ).fetchall()
    accounts = []
    for row in rows:
        failed_attempts = int(row["failed_attempts"] or 0)
        accounts.append(
            {
                "username": row["username"],
                "failed_attempts": failed_attempts,
                "remaining_attempts": max(0, AUTH_MAX_ATTEMPTS - failed_attempts),
                "locked": bool(row["locked"]),
            }
        )
    return accounts


def _option_catalog_values(conn: sqlite3.Connection, option_type: str) -> list[str]:
    _ensure_option_catalog_table(conn)
    return [
        str(row["value"] or "").strip()
        for row in conn.execute(
            "SELECT value FROM option_catalogs WHERE option_type = ? ORDER BY value COLLATE NOCASE",
            (option_type,),
        ).fetchall()
        if str(row["value"] or "").strip()
    ]


def _merged_option_values(primary_values: list[str], catalog_values: list[str]) -> list[str]:
    merged: list[str] = []
    seen: set[str] = set()
    for raw in [*primary_values, *catalog_values]:
        value = str(raw or "").strip()
        if not value or value in seen:
            continue
        seen.add(value)
        merged.append(value)
    return merged


def _account_state(conn: sqlite3.Connection, username: str) -> dict:
    row = conn.execute(
        "SELECT failed_attempts, locked FROM auth_attempts WHERE username = ?",
        (username,),
    ).fetchone()
    failed_attempts = int(row["failed_attempts"] or 0) if row else 0
    locked = bool(row["locked"]) if row else False
    return {
        "username": username,
        "failed_attempts": failed_attempts,
        "remaining_attempts": max(0, AUTH_MAX_ATTEMPTS - failed_attempts),
        "locked": locked,
    }


def _account_password_hash(conn: sqlite3.Connection, username: str) -> str:
    row = conn.execute(
        "SELECT password_hash FROM auth_users WHERE username = ? AND enabled = 1",
        (username,),
    ).fetchone()
    return str(row["password_hash"] or "") if row else ""


def _set_account_state(conn: sqlite3.Connection, username: str, failed_attempts: int, locked: bool) -> None:
    conn.execute(
        """
        INSERT INTO auth_attempts (username, failed_attempts, locked, updated_at)
        VALUES (?, ?, ?, CURRENT_TIMESTAMP)
        ON CONFLICT(username) DO UPDATE SET
            failed_attempts = excluded.failed_attempts,
            locked = excluded.locked,
            updated_at = CURRENT_TIMESTAMP
        """,
        (username, failed_attempts, 1 if locked else 0),
    )


def _set_account_password_hash(conn: sqlite3.Connection, username: str, password_hash: str) -> None:
    conn.execute(
        """
        UPDATE auth_users
        SET password_hash = ?, enabled = 1, updated_at = CURRENT_TIMESTAMP
        WHERE username = ?
        """,
        (password_hash, username),
    )


def _create_session(conn: sqlite3.Connection, username: str) -> str:
    session_id = uuid.uuid4().hex
    conn.execute(
        """
        INSERT INTO auth_sessions (session_id, username, created_at, last_seen_at)
        VALUES (?, ?, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
        """,
        (session_id, username),
    )
    _prune_user_sessions(conn, username)
    return session_id


def _prune_user_sessions(conn: sqlite3.Connection, username: str) -> None:
    rows = conn.execute(
        """
        SELECT rowid, session_id, username, created_at, last_seen_at
        FROM auth_sessions
        WHERE username = ?
        ORDER BY last_seen_at DESC, created_at DESC, rowid DESC
        """,
        (username,),
    ).fetchall()
    invalid_session_ids = [row["session_id"] for row in rows if not _session_is_valid(row)]
    if invalid_session_ids:
        conn.executemany(
            "DELETE FROM auth_sessions WHERE session_id = ?",
            [(session_id,) for session_id in invalid_session_ids],
        )
    valid_rows = [row for row in rows if row["session_id"] not in set(invalid_session_ids)]
    excess = valid_rows[AUTH_MAX_SESSIONS_PER_USER:]
    if excess:
        conn.executemany(
            "DELETE FROM auth_sessions WHERE session_id = ?",
            [(row["session_id"],) for row in excess],
        )


def _api_token_hash(token: str) -> str:
    return hashlib.sha256(token.encode("utf-8")).hexdigest()


def _api_token_prefix(token: str) -> str:
    return token[: min(16, len(token))]


def _api_scopes(value: object) -> set[str]:
    try:
        parsed = json.loads(str(value or "[]"))
    except Exception:
        return set()
    if not isinstance(parsed, list):
        return set()
    return {
        _normalize_edit_value(item)
        for item in parsed
        if _normalize_edit_value(item)
    }


def _api_scope_allowed(scopes: set[str], required_scope: str) -> bool:
    if "admin:*" in scopes:
        return True
    if required_scope in scopes:
        return True
    resource, _, action = required_scope.partition(":")
    return bool(resource and action and f"{resource}:*" in scopes)


def _api_owner_token_scope_upgrade_allowed(required_scope: str, scopes: set[str]) -> bool:
    if required_scope not in {"sync:read", "items:write"}:
        return False
    return {"items:read", "history:write"}.issubset(scopes)


def _api_context_has_scope(context: dict, scope: str) -> bool:
    return _api_scope_allowed(set(context.get("scopes") or []), scope)


def _api_context_owner(context: dict) -> str:
    return _normalize_edit_value(context.get("owner"))


def _api_owner_allowed(context: dict, owner: object, any_scope: str = "") -> bool:
    if _api_context_has_scope(context, "admin:*"):
        return True
    if any_scope and _api_context_has_scope(context, any_scope):
        return True
    return _normalize_edit_value(owner) == _api_context_owner(context)


def _api_token_context(conn: sqlite3.Connection, authorization_header: str, required_scope: str) -> tuple[dict | None, dict | None, int]:
    prefix, _, token = _normalize_edit_value(authorization_header).partition(" ")
    if prefix.lower() != "bearer" or len(token) < API_TOKEN_MIN_LENGTH:
        return None, {"error": "invalid_token", "message": "Missing or invalid bearer token."}, 401
    row = conn.execute(
        """
        SELECT *
        FROM api_tokens
        WHERE token_hash = ?
          AND COALESCE(enabled, 1) = 1
        """,
        (_api_token_hash(token),),
    ).fetchone()
    if row is None:
        return None, {"error": "invalid_token", "message": "Token not found or disabled."}, 401
    expires_at = _normalize_edit_value(row["expires_at"])
    if expires_at:
        try:
            expires_dt = datetime.strptime(expires_at[:19], "%Y-%m-%d %H:%M:%S")
            if datetime.utcnow() > expires_dt:
                return None, {"error": "token_expired", "message": "Token expired."}, 401
        except ValueError:
            return None, {"error": "token_invalid_expiry", "message": "Token expiry is invalid."}, 401
    scopes = _api_scopes(row["scopes_json"])
    token_changed = False
    if not _api_scope_allowed(scopes, required_scope):
        if _api_owner_token_scope_upgrade_allowed(required_scope, scopes):
            scopes.add("sync:read")
            scopes.add("items:write")
            conn.execute(
                """
                UPDATE api_tokens
                SET scopes_json = ?
                WHERE id = ?
                """,
                (json.dumps(sorted(scopes), ensure_ascii=False, sort_keys=True), int(row["id"])),
            )
            token_changed = True
        else:
            return None, {"error": "forbidden_scope", "message": f"Missing scope: {required_scope}"}, 403
    if token_changed:
        conn.commit()
    elif _api_should_touch_token(required_scope) and _timestamp_older_than(row["last_used_at"], API_TOKEN_TOUCH_INTERVAL_SECONDS):
        conn.execute("UPDATE api_tokens SET last_used_at = CURRENT_TIMESTAMP WHERE id = ?", (int(row["id"]),))
        conn.commit()
    return {
        "token_id": int(row["id"]),
        "name": _normalize_edit_value(row["name"]),
        "owner": _normalize_edit_value(row["owner"]),
        "scopes": sorted(scopes),
        "token_prefix": _normalize_edit_value(row["token_prefix"]),
    }, None, 200


def _api_access_key_owner_slug(owner: str) -> str:
    normalized = _normalize_edit_value(owner)
    mapping = {
        "徐欣": "xuxin",
        "吴萍": "wuping",
    }
    if normalized in mapping:
        return mapping[normalized]
    ascii_slug = re.sub(r"[^0-9A-Za-z_-]+", "-", normalized).strip("-").lower()
    if ascii_slug:
        return ascii_slug
    return hashlib.sha256(normalized.encode("utf-8")).hexdigest()[:12]


def _api_access_key_secret_path(owner: str) -> Path:
    return API_TOKEN_SECRET_DIR / f"{_api_access_key_owner_slug(owner)}.token"


def _api_doc_token_scopes() -> list[str]:
    return ["history:write", "items:read", "items:write", "sync:read"]


def _api_access_key_row_for_token(conn: sqlite3.Connection, token: str, owner: str) -> sqlite3.Row | None:
    row = conn.execute(
        """
        SELECT *
        FROM api_tokens
        WHERE token_hash = ?
          AND owner = ?
          AND COALESCE(enabled, 1) = 1
        """,
        (_api_token_hash(token), owner),
    ).fetchone()
    if row is None:
        return None
    scopes = _api_scopes(row["scopes_json"])
    required_scopes = set(_api_doc_token_scopes())
    if not required_scopes.issubset(scopes):
        upgraded_scopes = sorted(scopes | required_scopes)
        conn.execute(
            """
            UPDATE api_tokens
            SET scopes_json = ?
            WHERE id = ?
            """,
            (json.dumps(upgraded_scopes, ensure_ascii=False, sort_keys=True), int(row["id"])),
        )
        conn.commit()
        row = conn.execute("SELECT * FROM api_tokens WHERE id = ?", (int(row["id"]),)).fetchone()
    return row


def _read_owner_api_access_key(conn: sqlite3.Connection, owner: str) -> str:
    secret_path = _api_access_key_secret_path(owner)
    try:
        token = secret_path.read_text(encoding="utf-8").strip()
    except OSError:
        return ""
    if not token:
        return ""
    return token if _api_access_key_row_for_token(conn, token, owner) is not None else ""


def _create_owner_api_access_key(conn: sqlite3.Connection, owner: str) -> str:
    token = API_TOKEN_PREFIX + secrets.token_urlsafe(32)
    scopes = _api_doc_token_scopes()
    token_prefix = _api_token_prefix(token)
    conn.execute(
        """
        INSERT INTO api_tokens (
            name, token_prefix, token_hash, owner, scopes_json, enabled
        )
        VALUES (?, ?, ?, ?, ?, 1)
        """,
        (
            f"chatgpt-rules-{_api_access_key_owner_slug(owner)}",
            token_prefix,
            _api_token_hash(token),
            owner,
            json.dumps(scopes, ensure_ascii=False, sort_keys=True),
        ),
    )
    API_TOKEN_SECRET_DIR.mkdir(parents=True, exist_ok=True)
    secret_path = _api_access_key_secret_path(owner)
    secret_path.write_text(token + "\n", encoding="utf-8")
    try:
        os.chmod(secret_path, 0o600)
    except OSError:
        pass
    return token


def _ensure_owner_api_access_key(conn: sqlite3.Connection, owner: str) -> str:
    normalized_owner = _normalize_edit_value(owner)
    if not normalized_owner:
        raise ValueError("owner_required_for_api_access_key")
    token = _read_owner_api_access_key(conn, normalized_owner)
    if token:
        return token
    return _create_owner_api_access_key(conn, normalized_owner)


def _request_external_base_url(handler: BaseHTTPRequestHandler) -> str:
    request_is_secure = getattr(handler, "_request_is_secure", lambda: False)
    proto = handler.headers.get("X-Forwarded-Proto") or handler.headers.get("X-Forwarded-Scheme")
    if not proto:
        proto = "https" if request_is_secure() else "http"
    proto = "https" if str(proto).lower() == "https" else "http"
    host = _normalize_edit_value(handler.headers.get("X-Forwarded-Host") or handler.headers.get("Host"))
    if not host:
        host = f"127.0.0.1:{PORT}"
    return f"{proto}://{host}".rstrip("/")


def _hermes_plugin_frame_ancestors(conn: sqlite3.Connection | None = None) -> list[str]:
    if conn is not None:
        return hermes_plugin.load_frame_ancestors(
            conn,
            HERMES_PLUGIN_FRAME_ANCESTORS,
            setting_key=HERMES_PLUGIN_FRAME_ANCESTOR_SETTINGS_KEY,
        )
    local_conn = connect()
    try:
        return _hermes_plugin_frame_ancestors(local_conn)
    finally:
        local_conn.close()


def _hermes_plugin_manifest_payload(
    base_url: str,
    *,
    frame_ancestors: list[str] | None = None,
    requested_frame_ancestor: str = "",
) -> dict:
    return hermes_plugin.build_plugin_manifest(
        base_url=base_url,
        app_version=APP_WEB_VERSION,
        mcp_version=getattr(__import__("wardrobe_app.wardrobe_mcp", fromlist=["SERVER_VERSION"]), "SERVER_VERSION", ""),
        sync_schema_version=program_api_sync.SYNC_SCHEMA_VERSION,
        resource_names=list(program_api_sync.SYNC_RESOURCE_NAMES),
        frame_ancestors=frame_ancestors or HERMES_PLUGIN_FRAME_ANCESTORS,
        requested_frame_ancestor=requested_frame_ancestor,
    )


def _api_register_hermes_plugin_workspace(conn: sqlite3.Connection, payload: dict) -> dict:
    return hermes_plugin.register_workspace(
        conn,
        payload,
        token_prefix_value=API_TOKEN_PREFIX,
        token_min_length=API_TOKEN_MIN_LENGTH,
        secret_dir=API_TOKEN_SECRET_DIR,
        owner_slug_fn=_api_access_key_owner_slug,
        default_frame_ancestors=HERMES_PLUGIN_FRAME_ANCESTORS,
    )


def _api_register_hermes_plugin_frame_ancestors(conn: sqlite3.Connection, payload: dict) -> dict:
    return hermes_plugin.register_frame_ancestors(
        conn,
        payload,
        HERMES_PLUGIN_FRAME_ANCESTORS,
        setting_key=HERMES_PLUGIN_FRAME_ANCESTOR_SETTINGS_KEY,
    )


def _api_create_hermes_plugin_launch_token(conn: sqlite3.Connection, payload: dict, context: dict) -> dict:
    return hermes_plugin.create_launch_token(conn, payload, context)


def _consume_hermes_plugin_launch(conn: sqlite3.Connection, launch_token: str) -> dict:
    return hermes_plugin.consume_launch_token(conn, launch_token)


def _api_hermes_plugin_session(conn: sqlite3.Connection, session_id: str) -> dict:
    return hermes_plugin.session_appearance(conn, session_id)


def _destroy_session(conn: sqlite3.Connection, cookie_header: str) -> None:
    session_id = _cookie_value(cookie_header, AUTH_COOKIE_NAME)
    if not session_id:
        return
    conn.execute("DELETE FROM auth_sessions WHERE session_id = ?", (session_id,))


def _destroy_session_by_id(conn: sqlite3.Connection, session_id: str) -> None:
    if not session_id:
        return
    conn.execute("DELETE FROM auth_sessions WHERE session_id = ?", (session_id,))


def _session_username_by_id(conn: sqlite3.Connection, session_id: str) -> str:
    if not session_id:
        return ""
    row = conn.execute(
        "SELECT username, created_at, last_seen_at FROM auth_sessions WHERE session_id = ?",
        (session_id,),
    ).fetchone()
    if not _session_is_valid(row):
        conn.execute("DELETE FROM auth_sessions WHERE session_id = ?", (session_id,))
        conn.commit()
        return ""
    if _timestamp_older_than(row["last_seen_at"], AUTH_SESSION_TOUCH_INTERVAL_SECONDS):
        conn.execute(
            "UPDATE auth_sessions SET last_seen_at = CURRENT_TIMESTAMP WHERE session_id = ?",
            (session_id,),
        )
        conn.commit()
    return str(row["username"] or "")


def _session_username(conn: sqlite3.Connection, cookie_header: str) -> str:
    return _session_username_by_id(conn, _cookie_value(cookie_header, AUTH_COOKIE_NAME))


def _extract_date_score(name: str) -> tuple[int, int, int]:
    patterns = [
        r"(20\d{2})[-_]?(\d{2})[-_]?(\d{2})",
        r"(20\d{2})\.(\d{2})\.(\d{2})",
    ]
    for pattern in patterns:
        matched = re.search(pattern, name)
        if matched:
            year, month, day = matched.groups()
            return int(year), int(month), int(day)
    return (0, 0, 0)


def _normalized_import_name(name: str) -> str:
    return re.sub(r"[\s_\-]+", "", str(name or "").lower())


def _pick_best_file(directory: Path, kind: str) -> Path | None:
    if kind == "wearcount_new" and not WEARCOUNT_NEW_FILE_IMPORT_ENABLED:
        return None
    ranked: list[tuple[int, tuple[int, int, int], int, float, Path]] = []
    for path in directory.rglob("*"):
        if not path.is_file() or path.suffix.lower() not in {".xlsx", ".xls", ".xlsm"}:
            continue
        if any(part.lower() == "log" for part in path.parts):
            continue
        name = path.name
        lower_name = name.lower()
        compact_name = _normalized_import_name(name)
        match_score = -10_000
        date_score = _extract_date_score(name)
        has_date = 1 if date_score != (0, 0, 0) else 0
        if kind == "wardrobe":
            if "wearcount" in lower_name or IMAGE_INDEX_WORD in name:
                continue
            if name == WARDROBE_FILE:
                match_score = 7_500
                has_date = 2
            elif WARDROBE_WORD in name:
                match_score = 7_000
        elif kind == "image_index":
            if WARDROBE_WORD in name and IMAGE_INDEX_WORD in name:
                match_score = 9_000
            elif IMAGE_INDEX_WORD in name:
                match_score = 6_000
        elif kind == "wearcount_new":
            if "wearcountnew" in compact_name:
                match_score = 8_500
        elif kind == "watch":
            if name == WATCH_FILE:
                match_score = 8_500
                has_date = 2
            elif WATCH_WORD in name:
                match_score = 7_500
        elif kind == "featured_looks":
            if "wearcount" in lower_name or IMAGE_INDEX_WORD in name:
                continue
            if name == LOOKS_FILE:
                match_score = 8_700
                has_date = 2
            elif LOOKS_WORD in name:
                match_score = 7_200
        if match_score < 0:
            continue
        ranked.append((has_date, date_score, match_score, path.stat().st_mtime, path))
    if not ranked:
        return None
    ranked.sort(key=lambda item: (item[0], item[1], item[2], item[3]), reverse=True)
    return ranked[0][4]


def _import_from_directory(conn: sqlite3.Connection, directory: Path) -> list[dict]:
    summaries: list[dict] = []
    # WearCount_new file import is intentionally disabled. Historical outfit
    # writes must go through POST /api/v1/history/outfits.
    # wear_file = _pick_best_file(directory, "wearcount_new")
    # if wear_file:
    #     summary = import_wearcount_new(conn, wear_file)
    #     summaries.append(summary)
    #     if not summary.get("skipped_duplicate"):
    #         _run_item_export_tasks(conn, {"wardrobe_ai", "watch_ai", "looks_ai", "wearcount"})
    #     try:
    #         wear_file.unlink()
    #         summary["source_removed"] = True
    #     except OSError:
    #         summary["source_removed"] = False
    return summaries


def _pick_directory() -> str:
    script = f"""
[Console]::OutputEncoding = [System.Text.Encoding]::UTF8
Add-Type -AssemblyName System.Windows.Forms
$dialog = New-Object System.Windows.Forms.FolderBrowserDialog
$dialog.Description = '选择导入目录'
$dialog.SelectedPath = '{str(DESKTOP_DIR).replace("'", "''")}'
$dialog.ShowNewFolderButton = $false
if ($dialog.ShowDialog() -eq [System.Windows.Forms.DialogResult]::OK) {{
    Write-Output $dialog.SelectedPath
}}
"""
    try:
        result = subprocess.run(
            ["powershell", "-NoProfile", "-STA", "-Command", script],
            capture_output=True,
            text=True,
            encoding="utf-8",
            timeout=30,
            check=False,
        )
        return (result.stdout or "").strip()
    except Exception:
        return ""


def _pick_directory_native() -> str:
    script = f"""
[Console]::OutputEncoding = [System.Text.Encoding]::UTF8
Add-Type -AssemblyName System.Windows.Forms
[System.Windows.Forms.Application]::EnableVisualStyles()
$selectedPath = ''
$owner = New-Object System.Windows.Forms.Form
$owner.StartPosition = [System.Windows.Forms.FormStartPosition]::CenterScreen
$owner.Size = New-Object System.Drawing.Size(1, 1)
$owner.ShowInTaskbar = $false
$owner.TopMost = $true
$owner.Opacity = 0
$owner.Load.Add({{ $owner.Activate() }})
$owner.Show()
Start-Sleep -Milliseconds 80
try {{
    $shell = New-Object -ComObject Shell.Application
    $folder = $shell.BrowseForFolder($owner.Handle, '选择导入目录', 0, '{str(DESKTOP_DIR).replace("'", "''")}')
    if ($folder -ne $null -and $folder.Self -ne $null) {{
        $selectedPath = $folder.Self.Path
    }}
}} catch {{
    $dialog = New-Object System.Windows.Forms.FolderBrowserDialog
    $dialog.Description = '选择导入目录'
    $dialog.SelectedPath = '{str(DESKTOP_DIR).replace("'", "''")}'
    $dialog.ShowNewFolderButton = $false
    if ($dialog.ShowDialog($owner) -eq [System.Windows.Forms.DialogResult]::OK) {{
        $selectedPath = $dialog.SelectedPath
    }}
}} finally {{
    $owner.Close()
    $owner.Dispose()
}}
if ($selectedPath) {{
    Write-Output $selectedPath
}}
"""
    try:
        result = subprocess.run(
            ["powershell", "-NoProfile", "-STA", "-Command", script],
            capture_output=True,
            text=True,
            encoding="utf-8",
            timeout=30,
            check=False,
        )
        selected = (result.stdout or "").strip()
        if selected:
            return selected
    except Exception:
        pass
    return _pick_directory()


def _import_from_uploaded_parts(conn: sqlite3.Connection, parts: dict[str, dict]) -> tuple[str, list[dict]]:
    folder_name = ""
    temp_path = DATA_DIR / "upload_imports" / uuid.uuid4().hex
    temp_path.mkdir(parents=True, exist_ok=True)
    try:
        for key, part in parts.items():
            if not key.startswith("file_"):
                continue
            suffix = key.split("_", 1)[1]
            relpath_key = f"relpath_{suffix}"
            relpath = ""
            if relpath_key in parts:
                relpath = parts[relpath_key]["content"].decode("utf-8", errors="ignore").strip()
            if relpath and not folder_name:
                folder_name = relpath.split("/")[0]
            file_name = Path(relpath or part.get("filename") or key).name
            (temp_path / file_name).write_bytes(part["content"])
        return folder_name, _import_from_directory(conn, temp_path)
    finally:
        shutil.rmtree(temp_path, ignore_errors=True)


def _item_source_kind(item: sqlite3.Row | dict) -> str:
    return "watch" if _canonical_layer_role(item["layer_role"]) == "Watch" else "wardrobe"


def _baseline_workbook_path(item: sqlite3.Row | dict) -> Path | None:
    kind = _item_source_kind(item)
    directory = WATCH_IMPORT_DIR if kind == "watch" else DEFAULT_IMPORT_DIR
    return _pick_best_file(directory, kind)


def _baseline_workbook_path_for_kind(kind: str) -> Path | None:
    directory = WATCH_IMPORT_DIR if kind == "watch" else DEFAULT_IMPORT_DIR
    return _pick_best_file(directory, kind)


def _baseline_header_map(item: sqlite3.Row | dict) -> dict[str, list[str]]:
    return WATCH_EDIT_HEADERS if _item_source_kind(item) == "watch" else WARDROBE_EDIT_HEADERS


def _baseline_header_map_for_kind(kind: str) -> dict[str, list[str]]:
    return WATCH_EDIT_HEADERS if kind == "watch" else WARDROBE_EDIT_HEADERS


def _normalize_edit_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value}".rstrip("0").rstrip(".")
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def _normalize_date_compare(value: object) -> str:
    text = _normalize_edit_value(value)
    if not text:
        return ""
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    if re.fullmatch(r"\d+(\.\d+)?", text):
        try:
            serial = float(text)
            base = datetime(1899, 12, 30)
            return (base + timedelta(days=serial)).strftime("%Y-%m-%d")
        except Exception:
            return text
    return text


CANONICAL_ITEM_LAYER_ROLES = (
    "Inner",
    "Middle",
    "Outer",
    "Bottom",
    "Footwear",
    "Watch",
    "Accessory",
    "Dress",
    "Home",
    "Bespoke",
)
CANONICAL_OUTFIT_ROLES = ("Inner", "Middle", "Outer", "Bottom", "Footwear", "Accessory", "Watch")
CANONICAL_FEATURED_LOOK_SLOTS = ("inner", "middle", "outer", "bottom", "footwear", "watch")


def _canonical_token(value: object, allowed_values: tuple[str, ...]) -> str:
    normalized = _normalize_edit_value(value)
    if not normalized:
        return ""
    lookup = {entry.lower(): entry for entry in allowed_values}
    return lookup.get(normalized.lower(), "")


def _canonical_layer_role(value: object) -> str:
    return _canonical_token(value, CANONICAL_ITEM_LAYER_ROLES)


def _canonical_outfit_role(value: object) -> str:
    return _canonical_token(value, CANONICAL_OUTFIT_ROLES)


def _canonical_role_key(value: object) -> str:
    role = _canonical_outfit_role(value)
    return role.lower() if role else _normalize_edit_value(value).lower()


def _validate_layer_role(value: object) -> str:
    role = _canonical_layer_role(value)
    if not role:
        raise ValueError("invalid_layer_role")
    return role


def _validate_outfit_role(value: object, code: object = "") -> str:
    role = _canonical_outfit_role(value)
    if not role:
        suffix = f":{_normalize_edit_value(code)}" if _normalize_edit_value(code) else ""
        raise ValueError(f"invalid_outfit_role{suffix}")
    return role


def _parse_export_datetime(value: object) -> datetime | None:
    normalized = _normalize_date_compare(value)
    if not normalized:
        return None
    try:
        return datetime.strptime(normalized, "%Y-%m-%d")
    except ValueError:
        return None


def _normalized_number_value(value: object) -> int | float | None:
    text = _normalize_edit_value(value)
    if not text:
        return None
    number = float(text)
    return int(number) if number.is_integer() else number


def _ai_api_key() -> str:
    return ""


def _ai_role_key(value: object) -> str:
    return _canonical_role_key(value)


def _ai_item_days_since_last_worn(item: sqlite3.Row | dict, today_value: date) -> int | None:
    raw = _normalize_date_compare(item["last_worn_on"] if isinstance(item, sqlite3.Row) else item.get("last_worn_on"))
    if not raw:
        return None
    try:
        last_worn = datetime.strptime(raw, "%Y-%m-%d").date()
    except ValueError:
        return None
    return max(0, (today_value - last_worn).days)


def _ai_item_temperature_match(item: sqlite3.Row | dict, role_key: str, draft_low: float | None, draft_high: float | None) -> bool:
    if draft_low is None and draft_high is None:
        return True
    source = item if isinstance(item, dict) else dict(item)
    if role_key in {"inner", "middle"}:
        item_low = _normalized_number_value(source.get("standalone_min"))
        item_high = _normalized_number_value(source.get("standalone_max"))
    elif role_key in {"outer", "bottom"}:
        item_low = _normalized_number_value(source.get("temp_min"))
        item_high = _normalized_number_value(source.get("temp_max"))
    else:
        return True
    if item_low is None and item_high is None:
        return True
    effective_low = draft_low if draft_low is not None else draft_high
    effective_high = draft_high if draft_high is not None else draft_low
    if effective_low is None or effective_high is None:
        return True
    if item_low is not None and effective_high < item_low:
        return False
    if item_high is not None and effective_low > item_high:
        return False
    return True


def _ai_item_scene_mismatch(item: sqlite3.Row | dict, draft_scene_tag: str) -> int:
    if not draft_scene_tag:
        return 0
    source = item if isinstance(item, dict) else dict(item)
    item_scene_tag = _normalize_edit_value(source.get("scene_tag"))
    if not item_scene_tag or item_scene_tag == draft_scene_tag:
        return 0
    return 1


def _ai_material_text(item: sqlite3.Row | dict | None) -> str:
    if item is None:
        return ""
    source = item if isinstance(item, dict) else dict(item)
    return _normalize_edit_value(source.get("material")).lower()


def _ai_is_refined_material(material_text: str) -> bool:
    text = str(material_text or "").lower()
    if not text:
        return False
    keywords = (
        "cashmere",
        "vicuna",
        "vicuña",
        "silk",
        "suede",
        "leather",
        "crocodile",
        "羊绒",
        "真丝",
        "丝绸",
        "丝缎",
        "麂皮",
        "皮革",
        "鳄鱼皮",
        "小羊皮",
        "鹿皮",
        "驼绒",
    )
    return any(keyword in text for keyword in keywords)


def _ai_is_casual_bottom_material(material_text: str) -> bool:
    text = str(material_text or "").lower()
    if not text:
        return False
    keywords = (
        "cotton",
        "denim",
        "jersey",
        "fleece",
        "棉",
        "牛仔",
        "抓绒",
        "卫衣",
    )
    return any(keyword in text for keyword in keywords)


def _ai_material_mismatch_penalty(item: sqlite3.Row, role_key: str, selected_items: list[dict] | None = None) -> int:
    if role_key != "bottom":
        return 0
    selected_items = selected_items or []
    upper_materials = [
        _ai_material_text(entry)
        for entry in selected_items
        if _ai_role_key(entry.get("role") or entry.get("role_key")) in {"inner", "middle", "outer"}
    ]
    refined_upper = any(_ai_is_refined_material(text) for text in upper_materials if text)
    if not refined_upper:
        return 0
    return 1 if _ai_is_casual_bottom_material(_ai_material_text(item)) else 0


def _ai_item_snapshot(
    item: sqlite3.Row | dict,
    today_value: date,
    role_override: str = "",
    has_base_layer: bool = False,
) -> dict:
    source = item if isinstance(item, dict) else dict(item)
    return {
        "id": int(source.get("id") or 0),
        "role": role_override or _normalize_edit_value(source.get("layer_role")),
        "brand": _normalize_edit_value(source.get("brand")),
        "section": _normalize_edit_value(source.get("section")),
        "code": _normalize_edit_value(source.get("code")),
        "outer_type": _normalize_edit_value(source.get("outer_type")),
        "scene_tag": _normalize_edit_value(source.get("scene_tag")),
        "relax_index": _normalized_number_value(source.get("relax_index")),
        "temp_min": _normalized_number_value(source.get("temp_min")),
        "temp_max": _normalized_number_value(source.get("temp_max")),
        "standalone_min": _normalized_number_value(source.get("standalone_min")),
        "standalone_max": _normalized_number_value(source.get("standalone_max")),
        "primary_color": _normalize_edit_value(source.get("primary_color")),
        "secondary_color": _normalize_edit_value(source.get("secondary_color")),
        "wear_total": _normalized_number_value(source.get("wear_total")) or 0,
        "wear_year": _normalized_number_value(source.get("wear_year")) or 0,
        "last_worn_on": _normalize_date_compare(source.get("last_worn_on")),
        "days_since_last_worn": _ai_item_days_since_last_worn(source, today_value),
        "wear_maintenance": _normalized_number_value(source.get("wear_maintenance")) or 0,
        "wear_threshold": _normalized_number_value(source.get("wear_threshold")) or 0,
        "maintenance_state": int(source.get("maintenance_state") or 0),
        "owner": _normalize_edit_value(source.get("owner")),
        "status": "保养" if _maintenance_state_value(source) == 1 else (_normalize_edit_value(source.get("status")) or "Active"),
        "official_desc": _normalize_edit_value(source.get("official_desc")),
        "notes": _normalize_edit_value(source.get("notes")),
        "material": _normalize_edit_value(source.get("material")),
        "care": _normalize_edit_value(source.get("care")),
        "has_base_layer": 1 if has_base_layer else 0,
    }


AI_ROTATION_CONTEXT_FIELDS = {
    "wear_total",
    "wear_year",
    "last_worn_on",
    "days_since_last_worn",
}


AI_CONTEXT_CANDIDATE_ROLES = ("inner", "middle", "outer", "bottom", "footwear", "watch")


def _strip_ai_rotation_fields_from_item(item: dict) -> dict:
    return {
        key: value
        for key, value in dict(item or {}).items()
        if key not in AI_ROTATION_CONTEXT_FIELDS
    }


def _ai_context_selected_entries(context: dict) -> list[dict]:
    if not isinstance(context, dict):
        return []
    candidates = context.get("candidates")
    selected = candidates.get("selected") if isinstance(candidates, dict) else None
    if selected is None:
        selected = context.get("selected_items")
    return [dict(entry) for entry in (selected or []) if isinstance(entry, dict)]


def _ai_context_candidate_pool(context: dict) -> dict[str, list[dict]]:
    if not isinstance(context, dict):
        return {}
    candidates = context.get("candidates")
    pool = candidates.get("pool") if isinstance(candidates, dict) else None
    if pool is None:
        pool = context.get("candidate_pool")
    normalized_pool: dict[str, list[dict]] = {}
    for role_key, entries in (pool or {}).items():
        normalized_pool[str(role_key)] = [
            dict(entry)
            for entry in (entries or [])
            if isinstance(entry, dict)
        ]
    return normalized_pool


def _strip_ai_rotation_fields_from_candidates(candidates: dict) -> dict:
    cloned = dict(candidates or {})
    context_like = {"candidates": cloned}
    cloned["selected"] = [
        _strip_ai_rotation_fields_from_item(entry)
        for entry in _ai_context_selected_entries(context_like)
    ]
    pool: dict[str, list[dict]] = {}
    for role_key, entries in _ai_context_candidate_pool(context_like).items():
        pool[role_key] = [
            _strip_ai_rotation_fields_from_item(entry)
            for entry in entries
        ]
    cloned["pool"] = pool
    return cloned


def _strip_ai_rotation_fields_from_context(context: dict) -> dict:
    cloned = dict(context or {})
    selected_items = [
        _strip_ai_rotation_fields_from_item(entry)
        for entry in _ai_context_selected_entries(context)
    ]
    candidate_pool = {
        role_key: [
            _strip_ai_rotation_fields_from_item(entry)
            for entry in entries
        ]
        for role_key, entries in _ai_context_candidate_pool(context).items()
    }
    candidate_block = dict(cloned.get("candidates") or {}) if isinstance(cloned.get("candidates"), dict) else {}
    candidate_block["selected"] = selected_items
    candidate_block["pool"] = candidate_pool
    cloned["candidates"] = candidate_block
    cloned.pop("selected_items", None)
    cloned.pop("candidate_pool", None)
    return cloned


def _ai_candidate_sort_key(
    item: sqlite3.Row,
    role_key: str,
    draft_scene_tag: str,
    draft_low: float | None,
    draft_high: float | None,
    today_value: date,
    prefer_low_wear: bool = True,
    selected_items: list[dict] | None = None,
) -> tuple:
    wear_year = _normalized_number_value(item["wear_year"]) or 0
    wear_total = _normalized_number_value(item["wear_total"]) or 0
    days_since = _ai_item_days_since_last_worn(item, today_value)
    days_sort = -(days_since if days_since is not None else 999999)
    base = (
        0 if _ai_item_temperature_match(item, role_key, draft_low, draft_high) else 1,
        _ai_item_scene_mismatch(item, draft_scene_tag),
        _ai_material_mismatch_penalty(item, role_key, selected_items),
    )
    if not prefer_low_wear:
        return (
            *base,
            _normalize_edit_value(item["brand"]),
            _normalize_edit_value(item["section"]),
            _normalize_edit_value(item["code"]),
        )
    return (
        *base,
        wear_year,
        days_sort,
        wear_total,
        _normalize_edit_value(item["brand"]),
        _normalize_edit_value(item["section"]),
        _normalize_edit_value(item["code"]),
    )


def _build_outfit_ai_candidate_pool(
    conn: sqlite3.Connection,
    owner: str,
    selected_item_ids: set[int],
    draft_scene_tag: str,
    draft_low: float | None,
    draft_high: float | None,
    today_value: date,
    prefer_low_wear: bool = True,
    selected_items: list[dict] | None = None,
) -> dict[str, list[dict]]:
    return _build_outfit_ai_candidates(
        conn,
        owner,
        selected_item_ids,
        draft_scene_tag,
        draft_low,
        draft_high,
        today_value,
        prefer_low_wear=prefer_low_wear,
        selected_items=selected_items,
    )["pool"]


def _build_outfit_ai_candidates(
    conn: sqlite3.Connection,
    owner: str,
    selected_item_ids: set[int],
    draft_scene_tag: str,
    draft_low: float | None,
    draft_high: float | None,
    today_value: date,
    prefer_low_wear: bool = True,
    selected_items: list[dict] | None = None,
    include_rotation_fields: bool = True,
) -> dict:
    rows = conn.execute(
        """
        SELECT *
        FROM items
        WHERE COALESCE(owner, '') = ?
          AND COALESCE(status, '') = 'Active'
          AND COALESCE(maintenance_state, 0) = 0
        ORDER BY brand, section, code
        """,
        (owner,),
    ).fetchall()
    pool: dict[str, list[dict]] = {}
    role_counts: dict[str, dict[str, int]] = {}
    for role_key in AI_CONTEXT_CANDIDATE_ROLES:
        role_rows = [
            row
            for row in rows
            if int(row["id"] or 0) not in selected_item_ids
            and _ai_role_key(row["layer_role"]) == role_key
        ]
        role_rows.sort(
            key=lambda row: _ai_candidate_sort_key(
                row,
                role_key,
                draft_scene_tag,
                draft_low,
                draft_high,
                today_value,
                prefer_low_wear=prefer_low_wear,
                selected_items=selected_items,
            )
        )
        snapshots = [_ai_item_snapshot(row, today_value) for row in role_rows[:AI_ROLE_CANDIDATE_LIMIT]]
        if not include_rotation_fields:
            snapshots = [_strip_ai_rotation_fields_from_item(entry) for entry in snapshots]
        pool[role_key] = snapshots
        role_counts[role_key] = {
            "eligible": len(role_rows),
            "returned": len(snapshots),
        }
    selected_snapshots = [dict(entry) for entry in (selected_items or []) if isinstance(entry, dict)]
    if not include_rotation_fields:
        selected_snapshots = [_strip_ai_rotation_fields_from_item(entry) for entry in selected_snapshots]
    screening_flow = [
        "owner",
        "status=Active",
        "status!=Ordered",
        "maintenance_state=0",
        "weather",
        "scene_tag",
        "relax_index",
        "layer_role",
        "outer_type",
        "material",
        "color",
    ]
    if prefer_low_wear and include_rotation_fields:
        screening_flow.append("rotation")
    return {
        "selected": selected_snapshots,
        "pool": pool,
        "screening": {
            "active_only": True,
            "exclude_maintenance_state": 1,
            "exclude_selected_ids": sorted(int(value) for value in selected_item_ids if int(value) > 0),
            "limit_per_role": AI_ROLE_CANDIDATE_LIMIT,
            "role_counts": role_counts,
            "flow": screening_flow,
        },
    }


def _build_outfit_ai_weather(
    temp_low: float | None,
    temp_high: float | None,
) -> dict:
    summary = ""
    if temp_low is None and temp_high is None:
        summary = "unknown"
    elif temp_low is None:
        summary = f"<= {temp_high:g}C"
    elif temp_high is None:
        summary = f">= {temp_low:g}C"
    else:
        summary = f"{temp_low:g} to {temp_high:g}C"
    return {
        "temp_low": temp_low,
        "temp_high": temp_high,
        "summary": summary,
    }


def _ai_truth_source_entries() -> list[dict]:
    return [
        {
            "name": WARDROBE_TEXT_EXPORT_FILE,
            "priority": 1,
            "available": (DEFAULT_IMPORT_DIR / WARDROBE_TEXT_EXPORT_FILE).exists(),
        },
        {
            "name": WATCH_TEXT_EXPORT_FILE,
            "priority": 2,
            "available": (WATCH_IMPORT_DIR / WATCH_TEXT_EXPORT_FILE).exists(),
        },
        {
            "name": WARDROBE_CHATGPT_RULES_FILE,
            "priority": 3,
            "available": (DEFAULT_IMPORT_DIR / WARDROBE_CHATGPT_RULES_FILE).exists(),
        },
    ]


def _build_outfit_ai_rules(record_type: str, prompt_override: str = "") -> dict:
    normalized_type = _ai_prompt_record_type(record_type)
    rules = {
        "truth_priority": _ai_truth_source_entries(),
        "xlsx_policy": "Use xlsx only when baseline edit/create or manual cross-check is required.",
        "watch_policy": "A complete main recommendation must include a watch, otherwise say it cannot be completed.",
        "status_policy": (
            "CSV keeps Ordered/maintenance/other-owner for status queries; "
            "recommendations exclude non-Active, maintenance, wrong owner."
        ),
        "material_priority": "Material coherence is a hard filter and outranks low-wear rotation.",
        "rotation_policy": (
            "Use wear_year, days_since_last_worn, and wear_total only after hard filters pass."
            if normalized_type == "outfit_draft"
            else "Ignore rotation counts when judging saved looks."
        ),
        "selected_item_policy": (
            "Preserve valid selected items first, fill missing roles next, replace only on clear mismatch."
            if normalized_type == "outfit_draft"
            else "Judge the original saved combination first before suggesting alternatives."
        ),
        "analysis_scope": "analysis_only" if normalized_type == "featured_look" else "analysis_and_recommendation",
    }
    if prompt_override:
        rules["prompt_override"] = prompt_override
    return rules


def _build_outfit_ai_context(
    conn: sqlite3.Connection,
    payload: dict,
    resolved_entries: list[dict],
    prompt_override: str = "",
) -> dict:
    owner = _normalize_edit_value(payload.get("owner"))
    draft_scene_tag = _normalize_edit_value(payload.get("scene_tag"))
    draft_low = _normalized_number_value(payload.get("temp_low"))
    draft_high = _normalized_number_value(payload.get("temp_high"))
    today_value = datetime.now().date()
    selected_item_ids = {int(entry["item_id"]) for entry in resolved_entries if int(entry.get("item_id") or 0) > 0}
    prompt_text = _normalize_edit_value(prompt_override)
    selected_items = [
        _ai_item_snapshot(
            entry["item"],
            today_value,
            role_override=_normalize_edit_value(entry.get("role")),
            has_base_layer=bool(entry.get("has_base_layer")),
        )
        for entry in resolved_entries
    ]
    return {
        "meta": {
            "record_type": "outfit_draft",
            "owner": owner,
            "wear_date": _normalize_edit_value(payload.get("wear_date")),
            "inventory_loc": _normalize_edit_value(payload.get("inventory_loc") or payload.get("city")),
            "wear_mode": _normalize_edit_value(payload.get("wear_mode")) or "normal",
            "scene_tag": draft_scene_tag,
            "notes": _normalize_edit_value(payload.get("notes")),
        },
        "weather": _build_outfit_ai_weather(draft_low, draft_high),
        "rules": _build_outfit_ai_rules("outfit_draft", prompt_text),
        "candidates": _build_outfit_ai_candidates(
            conn,
            owner,
            selected_item_ids,
            draft_scene_tag,
            draft_low,
            draft_high,
            today_value,
            prefer_low_wear=True,
            selected_items=selected_items,
            include_rotation_fields=True,
        ),
    }


def _extract_chat_completion_text(payload: dict) -> str:
    choices = payload.get("choices") or []
    if not choices:
        return ""
    message = choices[0].get("message") or {}
    content = message.get("content")
    if isinstance(content, str):
        return content.strip()
    if isinstance(content, list):
        parts: list[str] = []
        for entry in content:
            if isinstance(entry, str):
                parts.append(entry)
                continue
            if isinstance(entry, dict):
                text = entry.get("text")
                if isinstance(text, str):
                    parts.append(text)
        return "\n".join(part for part in parts if part).strip()
    return ""


def _ai_prompt_scalar(value: object, max_length: int = 120) -> str:
    text = _normalize_edit_value(value)
    if not text:
        return "-"
    compact = re.sub(r"\s+", " ", text).strip()
    if len(compact) <= max_length:
        return compact
    return compact[: max_length - 3].rstrip() + "..."


def _ai_prompt_number(value: object) -> str:
    normalized = _normalized_number_value(value)
    if normalized is None:
        return "-"
    if isinstance(normalized, float):
        return f"{normalized:g}"
    return str(normalized)


def _ai_prompt_temp_range(low: object, high: object) -> str:
    normalized_low = _normalized_number_value(low)
    normalized_high = _normalized_number_value(high)
    if normalized_low is None and normalized_high is None:
        return "-"
    if normalized_low is None:
        return f"<= {normalized_high:g}C"
    if normalized_high is None:
        return f">= {normalized_low:g}C"
    return f"{normalized_low:g} to {normalized_high:g}C"


def _ai_prompt_colors(entry: dict) -> str:
    colors = [
        _ai_prompt_scalar(entry.get("primary_color"), max_length=40),
        _ai_prompt_scalar(entry.get("secondary_color"), max_length=40),
    ]
    colors = [color for color in colors if color and color != "-"]
    return " / ".join(colors) if colors else "-"


def _ai_prompt_candidate_line(entry: dict, include_rotation: bool) -> str:
    parts = [
        f"item_id={int(entry.get('id') or 0)}",
        f"role={_ai_prompt_scalar(entry.get('role'), max_length=20)}",
        f"label={_ai_prompt_scalar(entry.get('section') or entry.get('code'), max_length=80)}",
    ]
    brand = _ai_prompt_scalar(entry.get("brand"), max_length=40)
    if brand != "-":
        parts.append(f"brand={brand}")
    scene_tag = _ai_prompt_scalar(entry.get("scene_tag"), max_length=30)
    if scene_tag != "-":
        parts.append(f"scene={scene_tag}")
    relax_index = _ai_prompt_number(entry.get("relax_index"))
    if relax_index != "-":
        parts.append(f"relax={relax_index}")
    temp_range = _ai_prompt_temp_range(entry.get("temp_min"), entry.get("temp_max"))
    if temp_range != "-":
        parts.append(f"temp={temp_range}")
    standalone_range = _ai_prompt_temp_range(entry.get("standalone_min"), entry.get("standalone_max"))
    if standalone_range != "-":
        parts.append(f"standalone={standalone_range}")
    colors = _ai_prompt_colors(entry)
    if colors != "-":
        parts.append(f"colors={colors}")
    material = _ai_prompt_scalar(entry.get("material"), max_length=60)
    if material != "-":
        parts.append(f"material={material}")
    if _boolish_value(entry.get("has_base_layer")):
        parts.append("base_layer=1")
    if int(entry.get("maintenance_state") or 0):
        parts.append("maintenance=1")
    if include_rotation:
        wear_year = _ai_prompt_number(entry.get("wear_year"))
        if wear_year != "-":
            parts.append(f"wear_year={wear_year}")
        days_since = _ai_prompt_number(entry.get("days_since_last_worn"))
        if days_since != "-":
            parts.append(f"days_since_last_worn={days_since}")
        wear_total = _ai_prompt_number(entry.get("wear_total"))
        if wear_total != "-":
            parts.append(f"wear_total={wear_total}")
    return " | ".join(parts)


def _outfit_ai_user_prompt(context: dict) -> str:
    meta = context.get("meta") if isinstance(context, dict) else {}
    weather = context.get("weather") if isinstance(context, dict) else {}
    rules = context.get("rules") if isinstance(context, dict) else {}
    candidates = context.get("candidates") if isinstance(context, dict) else {}
    screening = candidates.get("screening") if isinstance(candidates, dict) else {}
    role_counts = screening.get("role_counts") if isinstance(screening, dict) else {}
    include_rotation = _ai_prompt_record_type(_normalize_edit_value((meta or {}).get("record_type"))) == "outfit_draft"
    pool = _ai_context_candidate_pool(context)
    lines = [
        "META",
        f"record_type: {_ai_prompt_scalar((meta or {}).get('record_type'), max_length=30)}",
        f"owner: {_ai_prompt_scalar((meta or {}).get('owner'), max_length=30)}",
        f"wear_date: {_ai_prompt_scalar((meta or {}).get('wear_date'), max_length=30)}",
        f"inventory_loc: {_ai_prompt_scalar((meta or {}).get('inventory_loc'), max_length=40)}",
        f"wear_mode: {_ai_prompt_scalar((meta or {}).get('wear_mode'), max_length=20)}",
        f"scene_tag: {_ai_prompt_scalar((meta or {}).get('scene_tag'), max_length=40)}",
        f"notes: {_ai_prompt_scalar((meta or {}).get('notes'), max_length=120)}",
        "",
        "WEATHER",
        f"temp_low: {_ai_prompt_number((weather or {}).get('temp_low'))}",
        f"temp_high: {_ai_prompt_number((weather or {}).get('temp_high'))}",
        f"temp_summary: {_ai_prompt_scalar((weather or {}).get('summary'), max_length=40)}",
        "",
        "RULES",
    ]
    truth_priority = [
        entry
        for entry in ((rules or {}).get("truth_priority") or [])
        if isinstance(entry, dict)
    ]
    if truth_priority:
        ordered_truth = " > ".join(_ai_prompt_scalar(entry.get("name"), max_length=60) for entry in truth_priority)
        availability = ", ".join(
            f"{_ai_prompt_scalar(entry.get('name'), max_length=40)}={'yes' if entry.get('available') else 'no'}"
            for entry in truth_priority
        )
        lines.append(f"truth_priority: {ordered_truth}")
        lines.append(f"truth_sources_available: {availability}")
    lines.extend([
        f"xlsx_policy: {_ai_prompt_scalar((rules or {}).get('xlsx_policy'), max_length=120)}",
        f"watch_policy: {_ai_prompt_scalar((rules or {}).get('watch_policy'), max_length=120)}",
        f"status_policy: {_ai_prompt_scalar((rules or {}).get('status_policy'), max_length=120)}",
        f"material_priority: {_ai_prompt_scalar((rules or {}).get('material_priority'), max_length=120)}",
        f"rotation_policy: {_ai_prompt_scalar((rules or {}).get('rotation_policy'), max_length=120)}",
        f"selected_item_policy: {_ai_prompt_scalar((rules or {}).get('selected_item_policy'), max_length=120)}",
        f"analysis_scope: {_ai_prompt_scalar((rules or {}).get('analysis_scope'), max_length=60)}",
    ])
    prompt_override = _ai_prompt_scalar((rules or {}).get("prompt_override"), max_length=160)
    if prompt_override != "-":
        lines.append(f"prompt_override: {prompt_override}")
    flow = screening.get("flow") if isinstance(screening, dict) else []
    if flow:
        lines.append("screening_flow: " + " -> ".join(_ai_prompt_scalar(value, max_length=40) for value in flow))
    lines.extend(["", "CANDIDATES.SELECTED"])
    selected_entries = _ai_context_selected_entries(context)
    if selected_entries:
        lines.extend(f"- {_ai_prompt_candidate_line(entry, include_rotation)}" for entry in selected_entries)
    else:
        lines.append("- none")
    lines.extend(["", "CANDIDATES.POOL"])
    for role_key in AI_CONTEXT_CANDIDATE_ROLES:
        count_info = role_counts.get(role_key) if isinstance(role_counts, dict) else {}
        eligible = count_info.get("eligible") if isinstance(count_info, dict) else None
        returned = count_info.get("returned") if isinstance(count_info, dict) else None
        summary_parts: list[str] = []
        if eligible is not None:
            summary_parts.append(f"eligible={eligible}")
        if returned is not None:
            summary_parts.append(f"returned={returned}")
        header = f"[{role_key}]"
        if summary_parts:
            header += " " + " | ".join(summary_parts)
        lines.append(header)
        role_entries = pool.get(role_key) or []
        if role_entries:
            lines.extend(f"- {_ai_prompt_candidate_line(entry, include_rotation)}" for entry in role_entries)
        else:
            lines.append("- none")
    return "\n".join(lines).strip()


def _outfit_ai_system_prompt(context: dict) -> str:
    meta = context.get("meta") if isinstance(context, dict) else {}
    record_type = _ai_prompt_record_type(_normalize_edit_value((meta or {}).get("record_type")))
    shared_rules = (
        "You are a menswear wardrobe analysis assistant. "
        "Reply only in Simplified Chinese plain text. "
        "Use only the supplied data. Do not invent products that are not present in the input. "
        "When mentioning a specific item, prefer its section name over item code or watch ref; avoid using code unless section is missing. "
        "Do not expose internal implementation words or field names such as ordered, selected_items, candidate_pool, candidates.selected, candidates.pool, role_key, item_id, JSON, schema, raw data, or machine-readable blocks. "
        "wear_year is only a historical count used for rotation priority; this system has no annual wear cap. "
        "Do not say that any item has reached, exceeded, or is close to an annual wear limit. "
        "Material is a hard styling constraint and should be treated as the truth source for texture and refinement. "
        "If the upper body uses refined materials such as cashmere, silk, suede, leather, crocodile, or other luxury textures, do not pair it with obviously casual cotton, denim, jersey, or fleece bottoms unless the scene explicitly requires that contrast and you can justify it. "
        "When material coherence conflicts with low-wear rotation, prioritize material coherence first and rotation second. "
        "If rules.prompt_override is non-empty, treat it as an additional instruction for this record and follow it unless it conflicts with the hard constraints above. "
    )
    if record_type == "featured_look":
        return (
            shared_rules
            + "You will receive one saved featured look plus a candidate pool for the same owner. "
            + "This featured look task is analysis-only. Analyze the current look only. "
            + "Ignore wear-count and rotation fields such as wear_total, wear_year, last_worn_on, and days_since_last_worn. "
            + "Do not provide replacement suggestions, completion suggestions, watch suggestions, optimization suggestions, or any improved alternative plan. "
            + "The response must contain these headings in this exact order: 整体判断、搭配优点、主要问题、单件点评、套装评分。 "
            + "Under 套装评分, you must give one overall score on a 100-point scale and briefly explain the main scoring reasons. "
            + "The analysis must still consider owner, scene, temperature, relax level, color, layering, material coherence, and maintenance state."
        )
    if record_type == "outfit_draft":
        return (
            shared_rules
            + "You will receive a daily outfit draft plus a candidate pool for the same owner. "
            + "For watch recommendations, use only Active watches from the candidate pool as selectable recommendation candidates. "
            + "Focus on helping the current draft become directly fillable in the form. "
            + "The response must contain these headings in this exact order: 整体判断、搭配优点、主要问题、单件点评、替换建议、补全建议、腕表建议、轮换优先级提醒。 "
            + "Recommendations must consider owner, scene, temperature, relax level, color, layering, wear counts, last worn date, and maintenance state. "
            + "When giving a formal or complete recommendation, include a matching watch. If no suitable watch exists, say that a complete main recommendation cannot be formed."
        )
    return (
        shared_rules
        + "You will receive a daily outfit draft or a saved daily outfit, plus a candidate pool for the same owner. "
        + "For watch recommendations, use only Active watches from the candidate pool as selectable recommendation candidates. "
        + "For saved daily outfit analysis, ignore wear-count and rotation fields such as wear_total, wear_year, last_worn_on, and days_since_last_worn. "
        + "For saved daily outfit analysis, first judge whether the original outfit itself was valid under its original scene, temperature, material, color, and layering conditions. "
        + "Do not use current wear counts, current wear_year, or current watch usage counts as a reason to say that the original historical outfit was wrong. "
        + "Do not use low-wear rotation as the basis for replacement suggestions in saved daily outfit evaluation. "
        + "Do not require a middle layer by default. If outer plus inner already works for the scene and temperature, treat the absence of a middle layer as acceptable. "
        + "If the saved outfit already contains a watch and that watch fits the scene, material refinement, and color balance, do not replace it mainly because another watch has been worn less. "
        + "Do not treat lighter sneakers or lighter-toned shoes in a refined city look as an automatic downgrade. Evaluate whether they intentionally add spring lightness versus mature weight, and explain that tradeoff. "
        + "The response must contain these headings in this exact order: 整体判断、搭配优点、主要问题、单件点评、替换建议、补全建议、腕表建议、轮换优先级提醒。 "
        + "Recommendations must consider owner, scene, temperature, relax level, color, layering, material coherence, and maintenance state. "
        + "When giving a formal or complete recommendation, include a matching watch. If no suitable watch exists, say that a complete main recommendation cannot be formed."
    )


def _call_outfit_ai_review(context: dict) -> dict:
    raise RuntimeError(NATIVE_AI_REMOVED_MESSAGE)
    system_prompt = _outfit_ai_system_prompt(context)
    user_prompt = _outfit_ai_user_prompt(context)
    request_payload = {
        "model": AI_MODEL,
        "temperature": 0.2,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
    }
    payload = _execute_ai_request(request_payload)
    analysis = _extract_chat_completion_text(payload)
    if not analysis:
        raise RuntimeError("千问接口返回为空")
    return {
        "analysis": analysis,
        "model": AI_MODEL,
    }


def _is_ai_timeout_error(exc: BaseException | None) -> bool:
    if exc is None:
        return False
    if isinstance(exc, (TimeoutError, socket.timeout)):
        return True
    if isinstance(exc, URLError):
        return _is_ai_timeout_error(exc.reason)
    message = str(exc or "").strip().lower()
    return "timed out" in message or "timeout" in message


def _execute_ai_request(request_payload: dict) -> dict:
    raise RuntimeError(NATIVE_AI_REMOVED_MESSAGE)
    api_key = _ai_api_key()
    if not api_key:
        raise RuntimeError("AI API key 未配置")
    request = Request(
        f"{AI_BASE_URL}/chat/completions",
        data=json.dumps(request_payload, ensure_ascii=False).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    attempts = AI_TIMEOUT_RETRIES + 1
    timeout_exc: BaseException | None = None
    for attempt in range(attempts):
        try:
            with urlopen(request, timeout=AI_TIMEOUT_SECONDS) as response:
                return json.loads(response.read().decode("utf-8"))
        except HTTPError as exc:
            raw = exc.read().decode("utf-8", errors="replace")
            try:
                error_payload = json.loads(raw)
            except Exception:
                error_payload = {}
            message = (
                error_payload.get("error", {}).get("message")
                or error_payload.get("message")
                or raw
                or str(exc)
            )
            raise RuntimeError(f"千问接口调用失败：{message}") from exc
        except Exception as exc:
            if _is_ai_timeout_error(exc):
                timeout_exc = exc
                if attempt + 1 < attempts:
                    time.sleep(0.6)
                    continue
                raise RuntimeError("AI 分析超时，当前模型响应较慢，请稍后重试。") from exc
            if isinstance(exc, URLError):
                raise RuntimeError(f"千问接口连接失败：{exc.reason}") from exc
            raise
    raise RuntimeError("AI 分析超时，当前模型响应较慢，请稍后重试。") from timeout_exc


def _strip_ai_json_fences(text: str) -> str:
    normalized = str(text or "").strip()
    if normalized.startswith("```"):
        normalized = re.sub(r"^```(?:json)?\s*", "", normalized, flags=re.IGNORECASE)
        normalized = re.sub(r"\s*```$", "", normalized)
    return normalized.strip()


def _parse_ai_json_response(text: str) -> dict:
    payload = _parse_ai_json_payload_any(text)
    effective = _extract_effective_ai_payload(payload)
    if isinstance(effective, dict):
        return dict(effective)
    if isinstance(effective, list):
        return {"items": effective}
    raise RuntimeError("AI 结构化结果不是对象")


def _parse_ai_json_payload_any(text: str) -> dict | list:
    normalized = _strip_ai_json_fences(text)
    if not normalized:
        raise RuntimeError("AI structured result is empty")
    try:
        payload = json.loads(normalized)
    except Exception:
        match = re.search(r"(\{[\s\S]*\}|\[[\s\S]*\])", normalized)
        if not match:
            raise RuntimeError("AI structured result is not valid JSON")
        try:
            payload = json.loads(match.group(0))
        except Exception as exc:
            raise RuntimeError("AI structured result is not valid JSON") from exc
    if not isinstance(payload, (dict, list)):
        raise RuntimeError("AI structured result must be a JSON object or array")
    return payload


def _is_effective_ai_payload_value(value: object) -> bool:
    return isinstance(value, (dict, list))


def _effective_ai_payload_fallback(root: dict) -> dict:
    filtered: dict = {}
    for key, value in root.items():
        normalized_key = _normalize_edit_value(key).strip().lower()
        if normalized_key == "meta" or normalized_key.startswith("debug") or normalized_key.startswith("raw"):
            continue
        filtered[key] = value
    return filtered or dict(root)


def _extract_effective_ai_payload(root: dict | list) -> dict | list:
    if isinstance(root, list):
        return root
    for key in ("payload", "data"):
        value = root.get(key)
        if _is_effective_ai_payload_value(value):
            return value
    for key in ("items", "records", "result"):
        value = root.get(key)
        if _is_effective_ai_payload_value(value):
            return value
    return _effective_ai_payload_fallback(root)


def _extract_outfit_ai_structured_payload(root: dict | list) -> dict:
    effective = _extract_effective_ai_payload(root)
    if isinstance(effective, dict):
        structured = dict(effective)
    elif isinstance(effective, list):
        structured = {"selected_items": effective}
    else:
        structured = {}
    if isinstance(root, dict):
        if "analysis" not in structured and isinstance(root.get("analysis"), str):
            structured["analysis"] = root.get("analysis")
        if "selected_items" not in structured and isinstance(root.get("selected_items"), list):
            structured["selected_items"] = root.get("selected_items")
    for alias in ("items", "records"):
        if "selected_items" not in structured and isinstance(structured.get(alias), list):
            structured["selected_items"] = structured.get(alias)
    return structured


def _outfit_ai_output_role_label(value: object) -> str:
    return {
        "inner": "Inner",
        "middle": "Middle",
        "outer": "Outer",
        "bottom": "Bottom",
        "footwear": "Footwear",
        "watch": "Watch",
    }.get(_ai_role_key(value), "")


def _boolish_value(value: object) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    normalized = _normalize_edit_value(value).strip().lower()
    return normalized in {"1", "true", "yes", "y", "是"}


def _ai_available_item_snapshots(context: dict) -> dict[int, dict]:
    available: dict[int, dict] = {}

    def collect(entry: dict | None) -> None:
        if not isinstance(entry, dict):
            return
        item_id = int(entry.get("id") or 0)
        if item_id <= 0 or item_id in available:
            return
        available[item_id] = entry

    for entry in _ai_context_selected_entries(context):
        collect(entry)
    for entries in _ai_context_candidate_pool(context).values():
        for entry in entries:
            collect(entry)
    return available


def _normalize_outfit_ai_selected_items(raw_items: object, context: dict) -> list[dict]:
    if not isinstance(raw_items, list):
        return []
    available = _ai_available_item_snapshots(context)
    normalized_items: list[dict] = []
    seen_roles: set[str] = set()
    for raw_entry in raw_items:
        if not isinstance(raw_entry, dict):
            continue
        role = _outfit_ai_output_role_label(raw_entry.get("role") or raw_entry.get("role_key"))
        if not role or role in seen_roles:
            continue
        try:
            item_id = int(raw_entry.get("item_id") or 0)
        except Exception:
            item_id = 0
        if item_id <= 0:
            continue
        snapshot = available.get(item_id)
        if not snapshot:
            continue
        snapshot_role = _outfit_ai_output_role_label(snapshot.get("role") or snapshot.get("role_key"))
        if snapshot_role != role:
            continue
        has_base_layer = bool(snapshot.get("has_base_layer"))
        if role in {"Inner", "Middle", "Bottom"}:
            if "has_base_layer" in raw_entry:
                has_base_layer = _boolish_value(raw_entry.get("has_base_layer"))
        else:
            has_base_layer = False
        normalized_items.append({
            "role": role,
            "item_id": item_id,
            "has_base_layer": 1 if has_base_layer else 0,
        })
        seen_roles.add(role)
    top_layer_roles = [entry["role"] for entry in normalized_items if entry.get("role") in {"Inner", "Middle"}]
    if len(top_layer_roles) > 1:
        selected_role_order = [
            _outfit_ai_output_role_label(entry.get("role"))
            for entry in _ai_context_selected_entries(context)
            if _outfit_ai_output_role_label(entry.get("role")) in {"Inner", "Middle"}
        ]
        preferred_top_role = next((role for role in selected_role_order if role in top_layer_roles), "")
        if not preferred_top_role:
            preferred_top_role = top_layer_roles[0]
        normalized_items = [
            entry
            for entry in normalized_items
            if entry.get("role") not in {"Inner", "Middle"} or entry.get("role") == preferred_top_role
        ]
    return normalized_items


def _call_outfit_draft_ai_review(context: dict) -> dict:
    raise RuntimeError(NATIVE_AI_REMOVED_MESSAGE)
    system_prompt = (
        _outfit_ai_system_prompt(context)
            + " For this draft-form task, return exactly one JSON object and nothing else. "
            + "The JSON schema is: "
            + "{\"analysis\": string, \"selected_items\": [{\"role\": \"Inner|Middle|Outer|Bottom|Footwear|Watch\", \"item_id\": number, \"has_base_layer\": 0|1}]}. "
            + "analysis must be plain Chinese text for the user only, and must not contain JSON, code fences, machine-readable fields, item_id lists, raw data blocks, or internal English field names. "
            + "selected_items is for the app to fill the current draft form. Use only item_id values that already exist in the input candidates.selected or candidates.pool. "
            + "selected_items must exactly mirror the concrete final recommendation described in analysis. If you recommend a specific item, put it into its matching role slot in selected_items. "
            + "Inner and Middle are mutually exclusive in the final draft recommendation. Choose at most one of them; if Inner is used, omit Middle, and if Middle is used, omit Inner. Outer can coexist with either one. "
            + "Do not recommend or select non-Active items. Items that are Pre-Order, For Sale, Museum, Display, Service Pending, or otherwise not Active must be ignored as recommendation candidates. "
            + "Preserve already selected items when they are acceptable, fill missing roles first, and only replace an existing role when there is a clear mismatch. "
            + "If a role should remain empty, omit that role from selected_items."
        )
    user_prompt = _outfit_ai_user_prompt(context)
    request_payload = {
        "model": AI_MODEL,
        "temperature": 0.1,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
    }
    payload = _execute_ai_request(request_payload)
    raw_text = _extract_chat_completion_text(payload)
    if not raw_text:
        raise RuntimeError("千问接口返回为空")
    structured = _extract_outfit_ai_structured_payload(_parse_ai_json_payload_any(raw_text))
    analysis = _normalize_edit_value(structured.get("analysis"))
    if not analysis:
        raise RuntimeError("AI 分析结果为空")
    return {
        "analysis": analysis,
        "selected_items": _normalize_outfit_ai_selected_items(structured.get("selected_items"), context),
        "model": AI_MODEL,
    }


def _ai_context_code_section_pairs(context: dict) -> list[tuple[str, str]]:
    pairs: list[tuple[str, str]] = []
    seen_codes: set[str] = set()

    def collect(entry: dict | None) -> None:
        if not isinstance(entry, dict):
            return
        code = _normalize_edit_value(entry.get("code"))
        section = _normalize_edit_value(entry.get("section"))
        if not code or not section or code == section or code in seen_codes:
            return
        seen_codes.add(code)
        pairs.append((code, section))

    for entry in _ai_context_selected_entries(context):
        collect(entry)
    for entries in _ai_context_candidate_pool(context).values():
        for entry in entries:
            collect(entry)
    pairs.sort(key=lambda pair: len(pair[0]), reverse=True)
    return pairs


def _replace_ai_codes_with_sections(text: str, context: dict) -> str:
    normalized = str(text or "").strip()
    if not normalized:
        return ""
    for code, section in _ai_context_code_section_pairs(context):
        normalized = normalized.replace(f"`{code}`", f"`{section}`")
        normalized = normalized.replace(code, section)
    normalized = re.sub(
        r"`?(?:ordered|selected_items|candidate_pool|candidates(?:\.(?:selected|pool))?|role_key|item_id|json|schema|raw data)`?",
        "",
        normalized,
        flags=re.IGNORECASE,
    )
    normalized = re.sub(r"[ \t]{2,}", " ", normalized)
    normalized = re.sub(r"\n{3,}", "\n\n", normalized)
    normalized = re.sub(r"(已)?(达到|接近|超过)?年度穿着上限", "本年穿着次数较高", normalized)
    normalized = re.sub(r"(已)?(达到|接近|超过)?年度上限", "本年穿着次数较高", normalized)
    normalized = normalized.replace("本年度穿着上限", "本年穿着次数较高")
    normalized = normalized.replace("年度使用上限", "本年穿着次数较高")
    return normalized


def _set_record_ai_pending(
    conn: sqlite3.Connection,
    table_name: str,
    record_id: int,
) -> None:
    conn.execute(
        f"""
        UPDATE {table_name}
        SET ai_analysis_status = 'pending',
            ai_analysis_error = '',
            updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
        """,
        (record_id,),
    )


def _finish_record_ai_generation(
    conn: sqlite3.Connection,
    table_name: str,
    record_id: int,
    analysis: str,
) -> None:
    conn.execute(
        f"""
        UPDATE {table_name}
        SET ai_analysis = ?,
            ai_analysis_status = '',
            ai_analysis_error = '',
            updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
        """,
        (analysis, record_id),
    )


def _fail_record_ai_generation(
    conn: sqlite3.Connection,
    table_name: str,
    record_id: int,
    error_message: str,
) -> None:
    conn.execute(
        f"""
        UPDATE {table_name}
        SET ai_analysis_status = 'error',
            ai_analysis_error = ?,
            updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
        """,
        (_normalize_edit_value(error_message) or "AI 分析失败。", record_id),
    )


def _schedule_outfit_ai_review_job(outfit_id: int, username: str, prompt_text: str) -> bool:
    job_key = ("outfit", int(outfit_id))
    with AI_REVIEW_LOCK:
        if job_key in AI_REVIEW_PENDING:
            return False
        AI_REVIEW_PENDING.add(job_key)
    threading.Thread(
        target=_run_outfit_ai_review_job,
        args=(int(outfit_id), username, prompt_text),
        daemon=True,
        name=f"outfit-ai-{int(outfit_id)}",
    ).start()
    return True


def _schedule_featured_look_ai_review_job(featured_look_id: int, prompt_text: str) -> bool:
    job_key = ("featured_look", int(featured_look_id))
    with AI_REVIEW_LOCK:
        if job_key in AI_REVIEW_PENDING:
            return False
        AI_REVIEW_PENDING.add(job_key)
    threading.Thread(
        target=_run_featured_look_ai_review_job,
        args=(int(featured_look_id), prompt_text),
        daemon=True,
        name=f"featured-look-ai-{int(featured_look_id)}",
    ).start()
    return True


def _run_outfit_ai_review_job(outfit_id: int, username: str, prompt_text: str) -> None:
    conn = connect()
    try:
        outfit = conn.execute("SELECT * FROM outfits WHERE id = ?", (int(outfit_id),)).fetchone()
        if outfit is None:
            return
        context = _build_saved_outfit_ai_context(conn, outfit, username, prompt_override=prompt_text)
        result = _call_outfit_ai_review(context)
        analysis = _replace_ai_codes_with_sections(result.get("analysis", ""), context)
        _finish_record_ai_generation(conn, "outfits", int(outfit_id), analysis)
        conn.commit()
    except Exception as exc:
        conn.rollback()
        try:
            _fail_record_ai_generation(conn, "outfits", int(outfit_id), str(exc))
            conn.commit()
        except Exception:
            conn.rollback()
        print(f"[outfit-ai] async generation failed for {outfit_id}: {exc}")
    finally:
        conn.close()
        with AI_REVIEW_LOCK:
            AI_REVIEW_PENDING.discard(("outfit", int(outfit_id)))


def _run_featured_look_ai_review_job(featured_look_id: int, prompt_text: str) -> None:
    conn = connect()
    try:
        look = conn.execute("SELECT * FROM featured_looks WHERE id = ?", (int(featured_look_id),)).fetchone()
        if look is None:
            return
        context = _build_saved_featured_look_ai_context(conn, look, prompt_override=prompt_text)
        result = _call_outfit_ai_review(context)
        analysis = _replace_ai_codes_with_sections(result.get("analysis", ""), context)
        _finish_record_ai_generation(conn, "featured_looks", int(featured_look_id), analysis)
        conn.commit()
    except Exception as exc:
        conn.rollback()
        try:
            _fail_record_ai_generation(conn, "featured_looks", int(featured_look_id), str(exc))
            conn.commit()
        except Exception:
            conn.rollback()
        print(f"[featured-look-ai] async generation failed for {featured_look_id}: {exc}")
    finally:
        conn.close()
        with AI_REVIEW_LOCK:
            AI_REVIEW_PENDING.discard(("featured_look", int(featured_look_id)))


def _load_saved_outfit_ai_selected_items(
    conn: sqlite3.Connection,
    outfit_row: sqlite3.Row,
    today_value: date,
) -> tuple[list[dict], set[int]]:
    _, update_item_map = _daily_update_item_map(
        conn,
        _normalize_edit_value(outfit_row["wear_date"]),
        _normalize_edit_value(outfit_row["owner"]),
    )
    item_rows = conn.execute(
        """
        SELECT outfit_items.role, items.*
        FROM outfit_items
        JOIN items ON items.id = outfit_items.item_id
        WHERE outfit_items.outfit_id = ?
        ORDER BY outfit_items.role, items.section
        """,
        (int(outfit_row["id"]),),
    ).fetchall()
    selected_items: list[dict] = []
    selected_item_ids: set[int] = set()
    for item in item_rows:
        item_id = int(item["id"] or 0)
        if item_id > 0:
            selected_item_ids.add(item_id)
        update_row = update_item_map.get(item_id) if item_id > 0 else None
        selected_items.append(
            _ai_item_snapshot(
                item,
                today_value,
                role_override=_normalize_edit_value(item["role"] or item["layer_role"]),
                has_base_layer=bool(update_row.get("has_base_layer")) if update_row is not None else False,
            )
        )
    return selected_items, selected_item_ids


def _load_saved_featured_look_ai_selected_items(
    conn: sqlite3.Connection,
    look_row: sqlite3.Row,
    today_value: date,
) -> tuple[list[dict], set[int]]:
    item_rows = conn.execute(
        """
        SELECT
            featured_look_items.slot,
            featured_look_items.source_code,
            featured_look_items.source_section,
            items.*
        FROM featured_look_items
        LEFT JOIN items ON items.id = featured_look_items.item_id
        WHERE featured_look_items.featured_look_id = ?
        ORDER BY featured_look_items.display_order ASC, featured_look_items.id ASC
        """,
        (int(look_row["id"]),),
    ).fetchall()
    selected_items: list[dict] = []
    selected_item_ids: set[int] = set()
    seen_keys: set[str] = set()
    for entry in item_rows:
        source = dict(entry)
        if not _normalize_edit_value(source.get("section")):
            source["section"] = _normalize_edit_value(source.get("source_section"))
        if not _normalize_edit_value(source.get("code")):
            source["code"] = _normalize_edit_value(source.get("source_code"))
        preferred_role = _normalize_edit_value(source.get("slot"))
        if _ai_role_key(preferred_role) not in AI_CONTEXT_CANDIDATE_ROLES:
            preferred_role = _normalize_edit_value(source.get("role") or source.get("layer_role"))
        item_id = int(source.get("id") or 0)
        if item_id > 0:
            selected_item_ids.add(item_id)
            dedupe_key = f"id:{item_id}"
        else:
            dedupe_key = "src:{code}|{section}".format(
                code=_normalize_edit_value(source.get("source_code") or source.get("code")),
                section=_normalize_edit_value(source.get("source_section") or source.get("section")),
            )
        if dedupe_key in seen_keys:
            continue
        seen_keys.add(dedupe_key)
        selected_items.append(
            _ai_item_snapshot(
                source,
                today_value,
                role_override=preferred_role,
                has_base_layer=False,
            )
        )
    return selected_items, selected_item_ids


def _build_saved_outfit_ai_context(
    conn: sqlite3.Connection,
    outfit_row: sqlite3.Row,
    username: str,
    prompt_override: str = "",
) -> dict:
    outfit = dict(outfit_row)
    today_value = datetime.now().date()
    owner = _normalize_edit_value(outfit.get("owner"))
    draft_scene_tag = _normalize_edit_value(outfit.get("scene_tag"))
    draft_low = _normalized_number_value(outfit.get("temp_low"))
    draft_high = _normalized_number_value(outfit.get("temp_high"))
    selected_items, selected_item_ids = _load_saved_outfit_ai_selected_items(conn, outfit_row, today_value)
    prompt_text = _normalize_edit_value(prompt_override)
    return {
        "meta": {
            "record_type": "outfit",
            "owner": owner,
            "wear_date": _normalize_edit_value(outfit.get("wear_date")),
            "inventory_loc": _normalize_edit_value(outfit.get("inventory_loc") or outfit.get("city")),
            "wear_mode": _normalize_edit_value(outfit.get("wear_mode")) or "normal",
            "scene_tag": draft_scene_tag,
            "notes": _normalize_edit_value(outfit.get("notes")),
        },
        "weather": _build_outfit_ai_weather(draft_low, draft_high),
        "rules": _build_outfit_ai_rules("outfit", prompt_text),
        "candidates": _build_outfit_ai_candidates(
            conn,
            owner,
            selected_item_ids,
            draft_scene_tag,
            draft_low,
            draft_high,
            today_value,
            prefer_low_wear=False,
            selected_items=selected_items,
            include_rotation_fields=False,
        ),
    }


def _build_saved_featured_look_ai_context(
    conn: sqlite3.Connection,
    look_row: sqlite3.Row,
    prompt_override: str = "",
) -> dict:
    look = dict(look_row)
    today_value = datetime.now().date()
    owner = _normalize_edit_value(look.get("owner"))
    draft_scene_tag = _normalize_edit_value(look.get("scene_tag_target"))
    draft_low = _normalized_number_value(look.get("temp_min"))
    draft_high = _normalized_number_value(look.get("temp_max"))
    selected_items, selected_item_ids = _load_saved_featured_look_ai_selected_items(conn, look_row, today_value)
    prompt_text = _normalize_edit_value(prompt_override)
    return {
        "meta": {
            "record_type": "featured_look",
            "owner": owner,
            "look_id": _normalize_edit_value(look.get("look_id")),
            "status": _normalize_edit_value(look.get("status")),
            "use_case": _normalize_edit_value(look.get("use_case")),
            "scene_tag": draft_scene_tag,
            "notes": _normalize_edit_value(look.get("notes")),
        },
        "weather": _build_outfit_ai_weather(draft_low, draft_high),
        "rules": _build_outfit_ai_rules("featured_look", prompt_text),
        "candidates": _build_outfit_ai_candidates(
            conn,
            owner,
            selected_item_ids,
            draft_scene_tag,
            draft_low,
            draft_high,
            today_value,
            prefer_low_wear=False,
            selected_items=selected_items,
            include_rotation_fields=False,
        ),
    }


def _maintenance_state_value(item: sqlite3.Row | dict | None) -> int:
    if item is None:
        return 0
    try:
        value = item["maintenance_state"]  # type: ignore[index]
    except Exception:
        value = 0
    try:
        return 1 if int(value or 0) == 1 else 0
    except Exception:
        return 0


def _maintenance_state_label(item: sqlite3.Row | dict | None) -> str:
    return "保养中" if _maintenance_state_value(item) == 1 else "激活"


def _requested_wardrobe_status(payload: dict | None, existing: sqlite3.Row | dict | None = None) -> str:
    if payload and "status" in payload:
        raw = _normalize_edit_value(payload.get("status"))
    elif existing is not None and _maintenance_state_value(existing) == 1:
        return "maintenance"
    elif existing is not None:
        raw = _normalize_edit_value(existing["status"] if isinstance(existing, sqlite3.Row) else existing.get("status"))
    else:
        raw = ""
    normalized = raw.casefold()
    if normalized in {"maintenance", "保养", "保养中"}:
        return "maintenance"
    if normalized in {"archived", "archive", "归档"}:
        return "archived"
    return "active"


def _apply_wardrobe_status_transition(conn: sqlite3.Connection, item_id: int, requested_status: str) -> None:
    row = conn.execute("SELECT * FROM items WHERE id = ?", (item_id,)).fetchone()
    if row is None or _item_source_kind(row) == "watch":
        return
    requested = requested_status or "active"
    current_loc = _normalize_edit_value(row["loc"])
    previous_loc = _normalize_edit_value(row["maintenance_prev_loc"])
    restored_loc = current_loc or previous_loc or "SH"
    if requested == "maintenance":
        if _maintenance_state_value(row) != 1:
            next_prev_loc = current_loc or previous_loc or "SH"
            conn.execute(
                """
                UPDATE items
                SET status = 'Active',
                    maintenance_state = 1,
                    maintenance_prev_loc = ?,
                    wear_maintenance = 0,
                    maint_count = COALESCE(maint_count, 0) + 1,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (next_prev_loc, item_id),
            )
        else:
            conn.execute(
                """
                UPDATE items
                SET status = 'Active',
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (item_id,),
            )
        return
    next_status = "Archived" if requested == "archived" else "Active"
    if _maintenance_state_value(row) == 1:
        conn.execute(
            """
            UPDATE items
            SET status = ?,
                maintenance_state = 0,
                maintenance_prev_loc = NULL,
                loc = ?,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (next_status, restored_loc, item_id),
        )
        return
    conn.execute(
        """
        UPDATE items
        SET status = ?,
            updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
        """,
        (next_status, item_id),
    )


def _normalize_item_payload(payload: dict, existing: sqlite3.Row | dict | None = None) -> dict:
    data: dict[str, object] = {}
    for field in ITEM_DB_FIELDS:
        raw_value = payload[field] if field in payload else (existing[field] if existing is not None and field in existing.keys() else None)
        if field in EDIT_NUMBER_FIELDS:
            data[field] = _normalized_number_value(raw_value)
            continue
        if field in EDIT_DATE_FIELDS:
            data[field] = _normalize_date_compare(raw_value) or None
            continue
        if field in PRICE_TEXT_FIELDS:
            data[field] = normalize_price_text(raw_value)
            continue
        if field == "price_original_currency":
            raw_price = payload.get("price_original")
            if raw_price is None and existing is not None and "price_original" in existing.keys():
                raw_price = existing["price_original"]
            data[field] = normalize_price_currency(raw_value, raw_price)
            continue
        data[field] = _normalize_edit_value(raw_value)
    data["brand"] = normalize_item_brand(data.get("brand"))
    if not _normalize_edit_value(data.get("code")):
        raise ValueError("code_required")
    if not _normalize_edit_value(data.get("brand")):
        raise ValueError("brand_required")
    if _normalize_edit_value(data.get("price_original")) and not _normalize_edit_value(data.get("price_original_currency")):
        raise ValueError("price_original_currency_required")
    kind = _item_kind_from_values(data)
    if kind == "watch":
        data["layer_role"] = "Watch"
        data["loc"] = _normalize_edit_value(data.get("loc")) or "SH"
        data["outer_type"] = ""
        data["temp_min"] = None
        data["temp_max"] = None
        data["standalone_min"] = None
        data["standalone_max"] = None
        return data
    role = _validate_layer_role(data.get("layer_role"))
    normalized_role = role.lower()
    data["layer_role"] = role
    data["loc"] = _normalize_edit_value(data.get("loc")) or "SH"
    requested_status = _requested_wardrobe_status(data, existing)
    data["status"] = "Archived" if requested_status == "archived" else "Active"
    if normalized_role != "outer":
        data["outer_type"] = ""
    if normalized_role in {"inner", "middle"}:
        data["temp_min"] = None
        data["temp_max"] = None
    elif normalized_role in {"outer", "bottom"}:
        data["standalone_min"] = None
        data["standalone_max"] = None
    else:
        data["temp_min"] = None
        data["temp_max"] = None
        data["standalone_min"] = None
        data["standalone_max"] = None
    return data


def _item_kind_from_values(values: dict) -> str:
    return "watch" if _canonical_layer_role(values.get("layer_role")) == "Watch" else "wardrobe"


def _default_source_sheet(kind: str) -> str:
    return "腕表" if kind == "watch" else "衣橱"


def _apply_export_ownership(path: Path) -> None:
    if EXPORT_UID is None and EXPORT_GID is None:
        return
    uid = EXPORT_UID if EXPORT_UID is not None else -1
    gid = EXPORT_GID if EXPORT_GID is not None else -1
    try:
        os.chown(path, uid, gid)
    except OSError:
        return


def _textual_content_type(content_type: str | None, path: Path) -> str:
    base = (content_type or "application/octet-stream").split(";", 1)[0].strip().lower()
    suffix = path.suffix.lower()
    if base in {"text/html", "text/css", "application/javascript", "text/javascript", "application/json"}:
        return f"{base}; charset=utf-8"
    if suffix in {".html", ".css", ".js", ".json", ".webmanifest"}:
        fallback = {
            ".html": "text/html",
            ".css": "text/css",
            ".js": "application/javascript",
            ".json": "application/json",
            ".webmanifest": "application/manifest+json",
        }[suffix]
        return f"{fallback}; charset=utf-8"
    return content_type or "application/octet-stream"


def _stabilize_export_for_drive(path: Path) -> None:
    now = time.time()
    try:
        os.utime(path, (now, now))
    except OSError:
        return
    try:
        with path.open("r+b") as handle:
            first = handle.read(1)
            if first:
                handle.seek(0)
                handle.write(first)
            handle.flush()
            os.fsync(handle.fileno())
    except OSError:
        return
    try:
        os.utime(path, (now, time.time()))
    except OSError:
        pass
    try:
        dir_fd = os.open(str(path.parent), os.O_RDONLY)
    except OSError:
        return
    try:
        os.fsync(dir_fd)
    except OSError:
        pass
    finally:
        os.close(dir_fd)


def _nudge_export_for_drive(path: Path) -> None:
    temp_name = path.with_name(f"{path.stem}.__sync__{uuid.uuid4().hex}{path.suffix}")
    try:
        path.replace(temp_name)
        _apply_export_ownership(temp_name)
        _stabilize_export_for_drive(temp_name)
        time.sleep(0.35)
        temp_name.replace(path)
    except OSError:
        try:
            temp_name.replace(path)
        except OSError:
            pass
        return
    _apply_export_ownership(path)
    _stabilize_export_for_drive(path)
    time.sleep(0.35)


def _pulse_export_directory(directory: Path) -> None:
    pulse_a = directory / f".drive-pulse-{uuid.uuid4().hex}.tmp"
    pulse_b = directory / f".drive-pulse-{uuid.uuid4().hex}.done"
    try:
        with pulse_a.open("wb") as handle:
            handle.write(f"{time.time():.6f}".encode("utf-8"))
            handle.flush()
            os.fsync(handle.fileno())
        try:
            dir_fd = os.open(str(directory), os.O_RDONLY)
        except OSError:
            dir_fd = None
        if dir_fd is not None:
            try:
                os.fsync(dir_fd)
            except OSError:
                pass
            finally:
                os.close(dir_fd)
        pulse_a.replace(pulse_b)
        try:
            dir_fd = os.open(str(directory), os.O_RDONLY)
        except OSError:
            dir_fd = None
        if dir_fd is not None:
            try:
                os.fsync(dir_fd)
            except OSError:
                pass
            finally:
                os.close(dir_fd)
    except OSError:
        pass
    finally:
        for pulse in (pulse_a, pulse_b):
            try:
                if pulse.exists():
                    pulse.unlink()
            except OSError:
                pass
        try:
            dir_fd = os.open(str(directory), os.O_RDONLY)
        except OSError:
            dir_fd = None
        if dir_fd is not None:
            try:
                os.fsync(dir_fd)
            except OSError:
                pass
            finally:
                os.close(dir_fd)


def _host_visible_export_path(path: Path) -> Path:
    try:
        return HOST_CLOSET_DIR / path.relative_to(DEFAULT_IMPORT_DIR)
    except ValueError:
        pass
    try:
        return HOST_WUPING_CLOSET_DIR / path.relative_to(WUPING_CLOSET_DIR)
    except ValueError:
        pass
    try:
        return HOST_WATCH_DIR / path.relative_to(WATCH_IMPORT_DIR)
    except ValueError:
        pass
    return path


def _owner_specific_export_dir(owners: set[str] | None = None) -> Path:
    normalized_owners = set(_normalized_owner_filters(owners))
    if normalized_owners == {"吴萍"}:
        return WUPING_CLOSET_DIR
    return DEFAULT_IMPORT_DIR


def _rule_doc_target_dirs() -> list[Path]:
    targets: list[Path] = []
    for directory in (DEFAULT_IMPORT_DIR, WUPING_CLOSET_DIR):
        if directory not in targets:
            targets.append(directory)
    return targets


def _rule_doc_target_specs() -> list[tuple[Path, str]]:
    specs: list[tuple[Path, str]] = []
    for directory, owner in ((DEFAULT_IMPORT_DIR, "徐欣"), (WUPING_CLOSET_DIR, "吴萍")):
        if not any(existing_dir == directory for existing_dir, _ in specs):
            specs.append((directory, owner))
    return specs


def _export_markdown_doc_to_targets(
    file_name: str,
    content: str,
    kind: str,
    target_dirs: list[Path],
) -> list[dict]:
    results: list[dict] = []
    for directory in target_dirs:
        target_path = directory / file_name
        temp_path = _export_temp_path(target_path, f"export-{kind}")
        try:
            temp_path.write_text(content, encoding="utf-8")
            _finalize_export_file(temp_path, target_path)
            results.append({
                "kind": kind,
                "path": str(target_path),
            })
        finally:
            temp_path.unlink(missing_ok=True)
    return results


def _export_owner_markdown_docs_to_targets(
    conn: sqlite3.Connection,
    file_name: str,
    render_content,
    kind: str,
) -> list[dict]:
    results: list[dict] = []
    for directory, owner in _rule_doc_target_specs():
        content = render_content(conn, owner)
        target_path = directory / file_name
        temp_path = _export_temp_path(target_path, f"export-{kind}-{_api_access_key_owner_slug(owner)}")
        try:
            temp_path.write_text(content, encoding="utf-8")
            _finalize_export_file(temp_path, target_path)
            results.append({
                "kind": kind,
                "path": str(target_path),
                "owner": owner,
            })
        finally:
            temp_path.unlink(missing_ok=True)
    return results


def _enqueue_drive_notify(path: Path) -> None:
    queue_path = _host_visible_export_path(path)
    try:
        DRIVE_NOTIFY_QUEUE_DIR.mkdir(parents=True, exist_ok=True)
        request_path = DRIVE_NOTIFY_QUEUE_DIR / f"{int(time.time() * 1000)}-{uuid.uuid4().hex}.req"
        request_path.write_text(str(queue_path), encoding="utf-8")
        _stabilize_export_for_drive(request_path)
    except OSError:
        return


def _export_temp_path(target_path: Path, prefix: str) -> Path:
    target_path.parent.mkdir(parents=True, exist_ok=True)
    return target_path.parent / f".{prefix}-{uuid.uuid4().hex}.tmp{target_path.suffix}"


def _finalize_export_file(temp_path: Path, target_path: Path) -> None:
    try:
        temp_path.replace(target_path)
    except PermissionError as exc:
        raise RuntimeError(f"baseline_file_locked:{target_path}") from exc
    _apply_export_ownership(target_path)
    _stabilize_export_for_drive(target_path)
    _nudge_export_for_drive(target_path)
    _pulse_export_directory(target_path.parent)
    _enqueue_drive_notify(target_path)


def _write_export_row(worksheet, row_index: int, values: list[object]) -> None:
    for column_index, value in enumerate(values, start=1):
        cell = worksheet.cell(row=row_index, column=column_index)
        if isinstance(value, datetime):
            cell.value = value
            cell.number_format = "yyyy-mm-dd"
        else:
            cell.value = value


def _display_width(value: object) -> int:
    if value is None:
        return 0
    if isinstance(value, datetime):
        text = value.strftime("%Y-%m-%d")
    else:
        text = str(value)
    width = 0
    for char in text:
        width += 2 if unicodedata.east_asian_width(char) in {"F", "W"} else 1
    return width


def _apply_export_widths(worksheet, headers: list[object], rows: list[list[object]]) -> None:
    column_count = len(headers)
    for column_index in range(1, column_count + 1):
        max_width = _display_width(headers[column_index - 1])
        for row in rows:
            if column_index - 1 < len(row):
                max_width = max(max_width, _display_width(row[column_index - 1]))
        worksheet.column_dimensions[get_column_letter(column_index)].width = max_width + 2


def _normalized_owner_filters(owners: set[str] | None = None) -> list[str]:
    return sorted({
        _normalize_edit_value(owner)
        for owner in (owners or set())
        if _normalize_edit_value(owner)
    })


def _wardrobe_export_rows(conn: sqlite3.Connection, owners: set[str] | None = None) -> list[sqlite3.Row]:
    params: list[object] = []
    where_sql = "COALESCE(layer_role, '') <> 'Watch'"
    normalized_owners = _normalized_owner_filters(owners)
    if normalized_owners:
        where_sql += " AND (" + " OR ".join(["COALESCE(owner, '') = ?" for _ in normalized_owners]) + ")"
        params.extend(normalized_owners)
    rows = conn.execute(
        f"""
        SELECT *
        FROM items
        WHERE {where_sql}
        ORDER BY
            CASE WHEN COALESCE(acquired_at, '') = '' THEN 1 ELSE 0 END,
            acquired_at ASC,
            code ASC
        """,
        params,
    ).fetchall()
    return list(rows)


def _wardrobe_export_groups(conn: sqlite3.Connection, owners: set[str] | None = None) -> list[tuple[str, list[sqlite3.Row]]]:
    def _price_total(rows: list[sqlite3.Row]) -> float:
        total = 0.0
        for row in rows:
            text = normalize_price_text(row["price_cny"])
            if not text:
                continue
            try:
                total += float(text)
            except ValueError:
                continue
        return total

    rows = _wardrobe_export_rows(conn, owners)
    grouped: dict[str, list[sqlite3.Row]] = {}
    for row in rows:
        brand = _normalize_edit_value(row["brand"]) or "未分类"
        grouped.setdefault(brand, []).append(row)
    return sorted(
        grouped.items(),
        key=lambda item: (
            -_price_total(item[1]),
            item[0].casefold(),
        ),
    )


def _watch_export_rows(conn: sqlite3.Connection) -> list[sqlite3.Row]:
    rows = conn.execute(
        """
        SELECT *
        FROM items
        WHERE COALESCE(layer_role, '') = 'Watch'
        ORDER BY
            CASE WHEN COALESCE(acquired_at, '') = '' THEN 1 ELSE 0 END,
            acquired_at ASC,
            code ASC
        """
    ).fetchall()
    return list(rows)


def _featured_look_export_rows(conn: sqlite3.Connection, owners: set[str] | None = None) -> list[sqlite3.Row]:
    params: list[object] = []
    where_sql = "1 = 1"
    normalized_owners = _normalized_owner_filters(owners)
    if normalized_owners:
        where_sql += " AND (" + " OR ".join(["COALESCE(owner, '') = ?" for _ in normalized_owners]) + ")"
        params.extend(normalized_owners)
    rows = conn.execute(
        f"""
        SELECT *
        FROM featured_looks
        WHERE {where_sql}
        ORDER BY
            CASE WHEN status = 'Archived' THEN 1 ELSE 0 END,
            CASE
                WHEN TRIM(COALESCE(priority, '')) GLOB '[0-9]*' THEN CAST(priority AS INTEGER)
                ELSE 999999
            END,
            look_id ASC
        """,
        params,
    ).fetchall()
    return list(rows)


def _featured_look_relax_values(conn: sqlite3.Connection, featured_look_id: int) -> list[float]:
    rows = conn.execute(
        """
        SELECT items.relax_index, items.layer_role
        FROM featured_look_items
        JOIN items ON items.id = featured_look_items.item_id
        WHERE featured_look_items.featured_look_id = ?
          AND items.relax_index IS NOT NULL
        """,
        (featured_look_id,),
    ).fetchall()
    return [
        float(row["relax_index"])
        for row in rows
        if row["relax_index"] is not None and _normalize_edit_value(row["layer_role"]) != "Watch"
    ]


def _featured_look_relax_center(conn: sqlite3.Connection, featured_look_id: int) -> float | None:
    values = _featured_look_relax_values(conn, featured_look_id)
    if not values:
        return None
    return round((max(values) + min(values)) / 2, 2)


def _featured_look_relax_span(conn: sqlite3.Connection, featured_look_id: int) -> float | None:
    values = _featured_look_relax_values(conn, featured_look_id)
    if not values:
        return None
    return round(max(values) - min(values), 2)


def _sync_featured_look_relax_metrics(conn: sqlite3.Connection, featured_look_id: int) -> tuple[float | None, float | None]:
    center = _featured_look_relax_center(conn, featured_look_id)
    span = _featured_look_relax_span(conn, featured_look_id)
    conn.execute(
        "UPDATE featured_looks SET relax_center = ?, relax_span = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
        (center, span, featured_look_id),
    )
    return center, span


def _wearcount_export_sheet_rows(
    conn: sqlite3.Connection,
    layer_role: str,
    owners: set[str] | None = None,
) -> list[sqlite3.Row]:
    normalized_layer_role = _validate_layer_role(layer_role)
    params: list[object] = [normalized_layer_role]
    where_sql = "COALESCE(layer_role, '') = ?"
    normalized_owners = _normalized_owner_filters(owners)
    if normalized_owners:
        where_sql += " AND (" + " OR ".join(["COALESCE(owner, '') = ?" for _ in normalized_owners]) + ")"
        params.extend(normalized_owners)
    rows = conn.execute(
        f"""
        SELECT *
        FROM items
        WHERE {where_sql}
        ORDER BY
            COALESCE(brand, '') ASC,
            COALESCE(section, '') ASC,
            COALESCE(code, '') ASC
        """,
        params,
    ).fetchall()
    return list(rows)


def _wearcount_text_export_rows(
    conn: sqlite3.Connection,
    owners: set[str] | None = None,
) -> list[dict[str, object]]:
    role_sheet_specs = [
        ("Outer", "Outer"),
        ("Middle", "Middle"),
        ("Inner", "Inner"),
        ("Bottom", "Bottom"),
        ("Footwear", "Footwear"),
        ("Accessory", "Accessory"),
        ("Watch", "Watch"),
    ]
    export_rows: list[dict[str, object]] = []
    for layer_role, sheet_name in role_sheet_specs:
        rows = _wearcount_export_sheet_rows(conn, layer_role, owners)
        for row in rows:
            export_rows.append({
                "role": sheet_name,
                "code": _normalize_edit_value(row["code"]),
                "item": _normalize_edit_value(row["section"]),
                "brand": _normalize_edit_value(row["brand"]),
                "owner": _normalize_edit_value(row["owner"]),
                "wear": _normalized_number_value(row["wear_maintenance"]) or 0,
                "total": _normalized_number_value(row["wear_total"]) or 0,
                "wear_year": _normalized_number_value(row["wear_year"]) or 0,
                "maint": _normalized_number_value(row["maint_count"]) or 0,
                "thr": _normalized_number_value(row["wear_threshold"]) or 0,
                "status": _normalize_edit_value(row["status"]),
                "maintenance_state": _maintenance_state_label(row),
                "last_worn_on": _normalize_date_compare(row["last_worn_on"]) or "",
            })
    return export_rows


def _export_wearcount_text_csv(
    conn: sqlite3.Connection,
    owners: set[str] | None = None,
) -> dict:
    target_path = _owner_specific_export_dir(owners) / WEARCOUNT_TEXT_EXPORT_FILE
    temp_path = _export_temp_path(target_path, "export-wearcount-text")
    rows = _wearcount_text_export_rows(conn, owners)
    try:
        with temp_path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(WEARCOUNT_TEXT_EXPORT_HEADERS)
            for row in rows:
                writer.writerow([row.get(field_name, "") for field_name in WEARCOUNT_TEXT_EXPORT_HEADERS])
        _finalize_export_file(temp_path, target_path)
        return {
            "kind": "wearcount_ai_csv",
            "path": str(target_path),
            "rows": len(rows),
            "columns": list(WEARCOUNT_TEXT_EXPORT_HEADERS),
            "owners": _normalized_owner_filters(owners),
        }
    finally:
        temp_path.unlink(missing_ok=True)


def _export_wearcount_workbook(
    conn: sqlite3.Connection,
    file_name: str = WEARCOUNT_EXPORT_FILE,
    owners: set[str] | None = None,
) -> dict:
    target_path = _owner_specific_export_dir(owners) / file_name
    temp_path = _export_temp_path(target_path, "export-wearcount")
    workbook = Workbook()
    role_sheet_specs = [
        ("Outer", "Outer"),
        ("Middle", "Middle"),
        ("Inner", "Inner"),
        ("Bottom", "Bottom"),
        ("Footwear", "Footwear"),
        ("Accessory", "Accessory"),
        ("Watch", "Watch"),
    ]
    daily_log_column_map = {
        "Inner": "Inner_Codes",
        "Middle": "Middle_Codes",
        "Outer": "Outer_Codes",
        "Bottom": "Bottom_Codes",
        "Footwear": "Footwear_Codes",
        "Accessory": "Accessory_Codes",
        "Watch": "Watch_Refs",
    }
    try:
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        total_rows = 0
        for layer_role, sheet_name in role_sheet_specs:
            worksheet = workbook.create_sheet(title=sheet_name)
            _write_export_row(worksheet, 1, WEARCOUNT_EXPORT_HEADERS)
            rows = _wearcount_export_sheet_rows(conn, layer_role, owners)
            export_rows: list[list[object]] = []
            for index, row in enumerate(rows, start=1):
                values = [
                    _normalize_edit_value(row["code"]),
                    _normalize_edit_value(row["section"]),
                    _normalize_edit_value(row["brand"]),
                    _normalized_number_value(row["wear_maintenance"]) or 0,
                    _normalized_number_value(row["wear_total"]) or 0,
                    _normalized_number_value(row["wear_year"]) or 0,
                    _normalized_number_value(row["maint_count"]) or 0,
                    _normalized_number_value(row["wear_threshold"]) or 0,
                    _normalize_edit_value(row["status"]),
                ]
                export_rows.append(values)
                _write_export_row(worksheet, index + 1, values)
            _apply_export_widths(worksheet, WEARCOUNT_EXPORT_HEADERS, export_rows)
            total_rows += len(rows)

        daily_sheet = workbook.create_sheet(title="DAILY_LOG")
        _write_export_row(daily_sheet, 1, WEARCOUNT_DAILY_LOG_HEADERS)
        outfit_params: list[object] = []
        outfit_where = "1 = 1"
        normalized_owners = _normalized_owner_filters(owners)
        if normalized_owners:
            outfit_where += " AND (" + " OR ".join(["COALESCE(owner, '') = ?" for _ in normalized_owners]) + ")"
            outfit_params.extend(normalized_owners)
        outfit_rows = conn.execute(
            f"""
            SELECT *
            FROM outfits
            WHERE {outfit_where}
            ORDER BY wear_date ASC, id ASC
            """,
            outfit_params,
        ).fetchall()
        daily_export_rows: list[list[object]] = []
        for index, outfit in enumerate(outfit_rows, start=1):
            slot_values = {column: "" for column in daily_log_column_map.values()}
            item_rows = conn.execute(
                """
                SELECT outfit_items.role, items.code
                FROM outfit_items
                JOIN items ON items.id = outfit_items.item_id
                WHERE outfit_items.outfit_id = ?
                ORDER BY outfit_items.id ASC, items.code ASC
                """,
                (outfit["id"],),
            ).fetchall()
            grouped_codes: dict[str, list[str]] = {column: [] for column in daily_log_column_map.values()}
            for item_row in item_rows:
                column_name = daily_log_column_map.get(_normalize_edit_value(item_row["role"]))
                code = _normalize_edit_value(item_row["code"])
                if column_name and code:
                    grouped_codes[column_name].append(code)
            for column_name, codes in grouped_codes.items():
                slot_values[column_name] = "|".join(codes)
            values = [
                _normalize_date_compare(outfit["wear_date"]),
                _normalize_edit_value(outfit["inventory_loc"]) or _normalize_edit_value(outfit["city"]),
                _normalize_edit_value(outfit["scene_tag"]) or _normalize_edit_value(outfit["wear_mode"]),
                _normalized_number_value(outfit["avg_relax"]),
                _normalize_edit_value(outfit["avg_temp_label"] or outfit["temp_value"]),
                _normalize_edit_value(outfit["notes"]),
                slot_values["Inner_Codes"],
                slot_values["Middle_Codes"],
                slot_values["Outer_Codes"],
                slot_values["Bottom_Codes"],
                slot_values["Footwear_Codes"],
                slot_values["Accessory_Codes"],
                slot_values["Watch_Refs"],
            ]
            daily_export_rows.append(values)
            _write_export_row(daily_sheet, index + 1, values)
        _apply_export_widths(daily_sheet, WEARCOUNT_DAILY_LOG_HEADERS, daily_export_rows)

        workbook.save(temp_path)
        workbook.close()
        _finalize_export_file(temp_path, target_path)
        return {
            "kind": "wearcount",
            "path": str(target_path),
            "rows": total_rows,
            "logs": len(outfit_rows),
            "owners": _normalized_owner_filters(owners),
        }
    finally:
        workbook.close()
        temp_path.unlink(missing_ok=True)


def _export_wearcount_workbooks(conn: sqlite3.Connection) -> list[dict]:
    results: list[dict] = []
    results.extend(_retire_export_files(DEFAULT_IMPORT_DIR, RETIRED_CLOSET_EXPORT_FILES))
    results.extend(_retire_export_files(WUPING_CLOSET_DIR, RETIRED_CLOSET_EXPORT_FILES))
    results.append(_export_wearcount_text_csv(conn))
    for file_name, owners in WEARCOUNT_OWNER_EXPORT_SPECS:
        results.append(_export_wearcount_owner_text_csv(conn, file_name, owners))
    results.extend(_export_wearcount_chatgpt_rules_docs(conn))
    return results


def _export_single_wardrobe_workbook(conn: sqlite3.Connection, file_name: str, owners: set[str] | None = None) -> dict:
    target_path = _owner_specific_export_dir(owners) / file_name
    temp_path = _export_temp_path(target_path, "export-wardrobe")
    workbook = Workbook()
    try:
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        groups = _wardrobe_export_groups(conn, owners)
        if not groups:
            groups = [("衣橱", [])]
        total_rows = 0
        for brand, rows in groups:
            worksheet = workbook.create_sheet(title=brand[:31] or "衣橱")
            _write_export_row(worksheet, 1, WARDROBE_EXPORT_HEADERS)
            export_rows: list[list[object]] = []
            for index, row in enumerate(rows, start=1):
                values = [
                    index,
                    _normalize_edit_value(row["brand"]),
                    _normalize_edit_value(row["section"]),
                    _normalize_edit_value(row["loc"]),
                    _maintenance_state_label(row),
                    _normalize_edit_value(row["owner"]),
                    _normalize_edit_value(row["layer_role"]),
                    _normalize_edit_value(row["outer_type"]),
                    _normalize_edit_value(row["scene_tag"]),
                    _normalized_number_value(row["relax_index"]),
                    _normalized_number_value(row["temp_min"]),
                    _normalized_number_value(row["temp_max"]),
                    _normalized_number_value(row["standalone_min"]),
                    _normalized_number_value(row["standalone_max"]),
                    _normalize_edit_value(row["primary_color"]),
                    _normalize_edit_value(row["secondary_color"]),
                    _normalize_edit_value(row["official_desc"]),
                    normalize_price_text(row["price_original"]),
                    normalize_price_currency(row["price_original_currency"], row["price_original"]),
                    normalize_price_text(row["price_cny"]),
                    _normalize_edit_value(row["series"]),
                    _normalize_edit_value(row["code"]),
                    _normalize_edit_value(row["size"]),
                    _parse_export_datetime(row["acquired_at"]),
                    _normalize_edit_value(row["official_color_code"]),
                    _normalize_edit_value(row["material"]),
                    _normalize_edit_value(row["care"]),
                    _normalize_edit_value(row["notes"]),
                ]
                export_rows.append(values)
                _write_export_row(worksheet, index + 1, values)
            _apply_export_widths(worksheet, WARDROBE_EXPORT_HEADERS, export_rows)
            total_rows += len(rows)
        workbook.save(temp_path)
        workbook.close()
        _finalize_export_file(temp_path, target_path)
        return {
            "kind": "wardrobe",
            "path": str(target_path),
            "rows": total_rows,
            "sheets": [brand for brand, _ in groups],
            "owners": sorted({
                _normalize_edit_value(owner)
                for owner in (owners or set())
                if _normalize_edit_value(owner)
            }),
        }
    finally:
        workbook.close()
        temp_path.unlink(missing_ok=True)

def _wardrobe_text_export_days_since_last_worn(row: sqlite3.Row, export_on: date) -> int | None:
    last_worn_on = _normalize_date_compare(row["last_worn_on"])
    if not last_worn_on:
        return None
    try:
        last_worn_date = datetime.strptime(last_worn_on, "%Y-%m-%d").date()
    except ValueError:
        return None
    return max(0, (export_on - last_worn_date).days)


def _wardrobe_text_export_maintenance_due(row: sqlite3.Row) -> int:
    wear_threshold = _normalized_number_value(row["wear_threshold"])
    wear_maintenance = _normalized_number_value(row["wear_maintenance"])
    if wear_threshold is None or wear_threshold <= 0:
        return 0
    if wear_maintenance is None:
        return 0
    return 1 if wear_maintenance >= wear_threshold else 0


def _wardrobe_text_export_field_value(
    row: sqlite3.Row,
    field_name: str,
    export_on: date,
    missing_value: str = "",
) -> str:
    if field_name == "maintenance_state":
        return _maintenance_state_label(row)
    if field_name in {"acquired_at", "last_worn_on"}:
        return _normalize_date_compare(row[field_name]) or missing_value
    if field_name in PRICE_TEXT_FIELDS:
        return normalize_price_text(row[field_name]) or missing_value
    if field_name == "days_since_last_worn":
        value = _wardrobe_text_export_days_since_last_worn(row, export_on)
        return str(value) if value is not None else missing_value
    if field_name == "maintenance_due":
        return str(_wardrobe_text_export_maintenance_due(row))
    if field_name in {
        "relax_index",
        "temp_min",
        "temp_max",
        "standalone_min",
        "standalone_max",
        "wear_maintenance",
        "wear_threshold",
        "wear_total",
        "wear_year",
        "maint_count",
    }:
        value = _normalized_number_value(row[field_name])
        return str(value) if value is not None else missing_value
    value = _normalize_edit_value(row[field_name])
    return value or missing_value


def _wardrobe_text_export_row(row: sqlite3.Row, export_on: date) -> list[str]:
    return [
        _wardrobe_text_export_field_value(row, field_name, export_on)
        for field_name in WARDROBE_TEXT_EXPORT_HEADERS
    ]


def _render_program_api_rules_markdown(conn: sqlite3.Connection, owner: str) -> list[str]:
    normalized_owner = _normalize_edit_value(owner)
    access_key = _ensure_owner_api_access_key(conn, normalized_owner)
    external_id_example = f"chatgpt-{_api_access_key_owner_slug(normalized_owner)}-YYYY-MM-DD"
    return [
        "## Program API：接口规范",
        f"- 当前文档 owner：`{normalized_owner}`。",
        "- 本段只定义接口、认证、请求/响应字段和禁用链路；不定义具体搭配审美、材质、颜色、Relax 或场景选择规则。",
        f"- 具体搭配规则由 Hermes 维护的 `{WARDROBE_HERMES_STYLE_RULES_FILE}` 或 owner 目录内同用途文档承担。",
        "- 本项目不再依赖 ChatGPT Drive 同步作为 Hermes Agent 的日常搭配数据通道。",
        "- Wardrobe MCP 负责维护本地缓存文件和缓存元数据；Hermes Agent 只调用 MCP 并消费其返回的资源路径、metadata 与缩略图路径。",
        "- 后续历史穿搭入库只允许走 Program API；不要再生成或投递 `WearCount_new*.xlsx` 文件。",
        "- Access Key 只绑定当前 owner；不得用于写入其他 owner 的历史记录。",
        "- Access Key 当前应具备 `sync:read`、`items:read`、`items:write`、`history:write` scope。",
        "- Access Key：",
        "```text",
        access_key,
        "```",
        "- 客户端必须用上方 Access Key 原文生成 `Authorization` 请求头；不要把占位符、尖括号、`已脱敏`、`REDACTED` 或示例说明当作 token。",
        "- 推荐先用 `dry_run: true` 校验，确认返回 `200` 后再用同一个 payload 改为 `dry_run: false` 正式写入。",
        "",
        "### Hermes 同步接口",
        "- 启动日常搭配任务时，Hermes 先调用 `wardrobe.sync`，由 Wardrobe MCP 请求 manifest；如果返回 `304 Not Modified`，MCP 继续使用已验证的本地分资源缓存。",
        "- 如果 manifest 返回 `200 OK`，MCP 必须逐项比较 `resources[].checksum` 与 `resources[].count`；只请求发生变化的资源端点。",
        "- 全量 bundle 同步端点已停用；不要因为 global `etag` / `data_version` 变化就重新拉取全部单品。",
        "- 同步数据直接由 SQLite 主数据生成，不依赖 CSV 是否导出、Drive 是否同步或共享目录是否刷新。",
        "",
        "#### Manifest",
        "- 内网：`http://192.168.10.99:8765/api/v1/sync/outfit-context/manifest`",
        "- HTTPS：`https://wardrobe-xuxin.synology.me/api/v1/sync/outfit-context/manifest`",
        "",
        "请求头：",
        "```http",
        "Authorization: 运行时填写为 Bearer + 空格 + 上方 Access Key 原文",
        "If-None-Match: \"<本地保存的 etag>\"",
        "```",
        "",
        "有更新时返回：",
        "```json",
        "{",
        f'  "owner": "{normalized_owner}",',
        '  "scope": "outfit_context",',
        '  "schema_version": 4,',
        '  "data_version": "v4-...",',
        '  "etag": "sha256:...",',
        '  "resource_base_endpoint": "/api/v1/sync/outfit-context/resources",',
        '  "resources": [',
        '    {"name": "items", "count": 0, "checksum": "sha256:...", "endpoint": "/api/v1/sync/outfit-context/resources/items"},',
        '    {"name": "wear_counts", "count": 0, "checksum": "sha256:...", "endpoint": "/api/v1/sync/outfit-context/resources/wear_counts"},',
        '    {"name": "featured_looks", "count": 0, "checksum": "sha256:...", "endpoint": "/api/v1/sync/outfit-context/resources/featured_looks"},',
        '    {"name": "wear_history", "count": 0, "checksum": "sha256:...", "endpoint": "/api/v1/sync/outfit-context/resources/wear_history"},',
        '    {"name": "primary_photo_thumbnails", "count": 0, "checksum": "sha256:...", "endpoint": "/api/v1/sync/outfit-context/resources/primary_photo_thumbnails"},',
        '    {"name": "rules", "count": 0, "checksum": "sha256:...", "endpoint": "/api/v1/sync/outfit-context/resources/rules"}',
        '  ]',
        "}",
        "```",
        "",
        "#### Resource",
        "- 内网：`http://192.168.10.99:8765/api/v1/sync/outfit-context/resources/{name}`",
        "- HTTPS：`https://wardrobe-xuxin.synology.me/api/v1/sync/outfit-context/resources/{name}`",
        "- 可用资源：`items`、`wear_counts`、`featured_looks`、`wear_history`、`primary_photo_thumbnails`、`rules`。",
        "- 每个资源端点支持 `If-None-Match: \"<该资源 checksum>\"`；资源未变化时返回 `304 Not Modified`。",
        "- 资源响应包括 `resource`、`checksum`、`count`、`data`，并同时提供同名字段，例如 `items` 或 `wear_counts`。",
        "- `items` 是产品/单品主数据资源，不包含 `wear_total`、`wear_year`、`last_worn_on`、`wear_maintenance`、`maint_count` 等动态穿着字段。",
        "- `items[]` 同时返回 `price_original`、`price_original_currency` 和 `price_cny`；金额统计优先用 `price_cny`，为空时回退 `price_original`。",
        "- `price_original` 与 `price_cny` 的存储/API/MCP 缓存格式统一为纯金额十进制字符串，例如 `61400` 或 `29900.5`；不要写入 `¥`、`￥`、`RMB`、`CNY`、逗号、空格或 `元`，服务端会在入库和读出时规范化这些符号。",
        "- `price_original_currency` 表示 `price_original` 的原始币种，使用大写币种代码，例如 `CNY`、`EUR`、`USD`、`HKD`；写入 `price_original` 时必须提供该字段，或让服务端能从原始金额文本中识别币种；不要把币种混写进金额字段。",
        "- 资源 checksum 包含 sync `schema_version`、资源名和资源版本载荷；当 `items[]` 输出字段变化时必须提升 `schema_version`，让旧缓存失效并重新拉取对应资源。",
        "- `wear_counts` 保存每个 code 的穿着计数、维护计数、`last_worn_on` 与 `maintenance_due`；历史写入通常只会改变 `wear_counts` 与 `wear_history`。",
        "- `items[].primary_photo` 是当前第一张照片 metadata；`primary_photo_thumbnails` 资源供 Wardrobe MCP 按 checksum 增量缓存每件单品第一张安全 JPEG 缩略图。",
        "- 日常版型、穿法、厚薄、颜色辅助判断应优先使用 MCP 返回的本地首图缩略图路径；只有细节/OCR/视觉比对/原图交付需要时才拉 `content_path` 原图。",
        "- 产品没有照片是合法状态：`items[].photo_count=0`、`items[].primary_photo=null`，该产品不会出现在 `primary_photo_thumbnails` 资源里；MCP 必须跳过缩略图缓存并返回无图状态，不能让整次同步失败。",
        "- 单张缩略图下载失败、404、0B 或 MIME 不对时，MCP 只标记该 item 无可用本地缩略图并继续处理其他 item。",
        "- `items[]` 稳定显示字段优先级：`display_name` -> `name` -> `section` -> `brand + official_desc` -> `code`。",
        "- 旧缓存可能没有 `name/display_name`；遇到旧缓存时不得把候选池判为 0，应回退读取 `section`、`brand`、`official_desc`、`code`。",
        "- `items[].recommendation_eligible=false` 的记录可用于状态查询，但默认不得进入正式搭配候选池。",
        "",
        "### 历史穿搭写入接口",
        "- 内网：`http://192.168.10.99:8765/api/v1/history/outfits`",
        "- HTTPS：`https://wardrobe-xuxin.synology.me/api/v1/history/outfits`",
        "",
        "#### 请求头",
        "- `Authorization`：运行时填写为 `Bearer ` + 上方 Access Key 原文。",
        "- `Content-Type: application/json`",
        "- `Idempotency-Key: <稳定唯一键>`，建议使用 `owner + wear_date + source` 组成。",
        "",
        "#### 请求体",
        "```json",
        "{",
        f'  "owner": "{normalized_owner}",',
        '  "wear_date": "YYYY-MM-DD",',
        '  "city": "Shanghai",',
        '  "inventory_loc": "SH",',
        '  "wear_mode": "normal",',
        '  "scene_tag": "City",',
        '  "temp_low": 16,',
        '  "temp_high": 23,',
        '  "temp_value": 20,',
        '  "notes": "当日穿搭说明",',
        '  "source": "chatgpt",',
        f'  "external_id": "{external_id_example}",',
        '  "mode": "create_only",',
        '  "dry_run": true,',
        '  "items": [',
        '    { "code": "单品货号", "role": "Outer", "has_base_layer": false }',
        '  ]',
        "}",
        "```",
        "",
        "#### 字段规则",
        "- `wear_date` 必须是 `YYYY-MM-DD`。",
        "- `temp_low`、`temp_high`、`temp_value` 必须是 `wear_date` 当天、`city` 对应地点的真实天气温度。",
        "- 不得把衣物字段 `temp_min`、`temp_max`、`standalone_min`、`standalone_max` 当作当天气温；这些字段只表示单品适穿温区。",
        "- 无法获得当天真实温度时，温度字段应填 `null` 或省略，不要从衣物候选池推断。",
        "- 通常应满足 `temp_low <= temp_value <= temp_high`；若只知道区间，可只填 `temp_low` 和 `temp_high`。",
        "- `mode` 可选：`create_only`、`upsert`、`replace`；默认使用 `create_only`，避免覆盖已有日期。",
        "- `items[].code` 必须来自当前 owner 的 `items` 同步资源、`GET /api/v1/items`、`衣橱_*.csv` 或 `腕表.csv` 记录。",
        "- `items[].role` 必须使用规范穿搭角色：`Inner`、`Middle`、`Outer`、`Bottom`、`Footwear`、`Accessory`、`Watch`。鞋类统一写 `Footwear`，不得写 `Shoes`；裤装统一写 `Bottom`，不得写 `Pants`。",
        "- `has_base_layer` 仅在内层/中层等需要解释磨损增量时填写；不确定时填 `false`。",
        "- `external_id` 应稳定且唯一；重复请求会按幂等规则返回同一结果。",
        "- 正式写入成功后系统会更新 `outfits / outfit_items / wearcount_daily_updates`，并触发 CSV 导出。",
        "",
        "### 单品查询接口",
        "- 可用 `GET /api/v1/items?limit=500` 查询当前 token 可见单品。",
        "- 可用过滤参数：`q`、`brand`、`status`、`loc`、`layer_role`、`kind=wardrobe|watch`。",
        "- 可用 `GET /api/v1/items/{code}` 查询单个货号。",
        "",
        "### 新产品入库接口",
        "- Hermes 负责根据照片完成识别、归纳和字段结构化；衣橱 API 不做图片识别，只校验并保存 Hermes 提交的结构化结果。",
        "- 内网：`http://192.168.10.99:8765/api/v1/items`",
        "- HTTPS：`https://wardrobe-xuxin.synology.me/api/v1/items`",
        "",
        "#### 请求头",
        "- `Authorization`：运行时填写为 `Bearer ` + 上方 Access Key 原文。",
        "- `Content-Type: application/json`",
        "- `Idempotency-Key: <稳定唯一键>`，建议使用 `owner + code + source` 组成。",
        "",
        "#### 请求体",
        "```json",
        "{",
        '  "mode": "create_only",',
        '  "dry_run": true,',
        '  "source": "hermes-photo-analysis",',
        '  "external_id": "hermes-item-货号或批次号",',
        '  "item": {',
        f'    "owner": "{normalized_owner}",',
        '    "kind": "wardrobe",',
        '    "code": "货号或唯一 Ref",',
        '    "brand": "品牌",',
        '    "section": "产品名称",',
        '    "loc": "SH",',
        '    "layer_role": "Outer",',
        '    "outer_type": "Jacket",',
        '    "scene_tag": "City",',
        '    "relax_index": 3,',
        '    "temp_min": 12,',
        '    "temp_max": 22,',
        '    "standalone_min": null,',
        '    "standalone_max": null,',
        '    "primary_color": "Navy",',
        '    "secondary_color": "Grey",',
        '    "official_desc": "Hermes 图片分析后的描述",',
        '    "price_original": "61400",',
        '    "price_original_currency": "CNY",',
        '    "price_cny": "59900",',
        '    "series": "",',
        '    "size": "",',
        '    "acquired_at": "YYYY-MM-DD",',
        '    "official_color_code": "",',
        '    "material": "材质",',
        '    "care": "",',
        '    "wear_threshold": 30,',
        '    "notes": "识别依据和不确定项",',
        '    "status": "Active"',
        "  },",
        '  "photos": [',
        '    { "file_name": "front.jpg", "content_type": "image/jpeg", "data_base64": "<base64>" }',
        '  ]',
        "}",
        "```",
        "",
        "#### 腕表示例",
        "```json",
        "{",
        '  "mode": "create_only",',
        '  "dry_run": true,',
        '  "item": {',
        f'    "owner": "{normalized_owner}",',
        '    "kind": "watch",',
        '    "code": "腕表 Ref",',
        '    "brand": "品牌（使用已有品牌规范值，例如 Vacheron Constantin）",',
        '    "section": "腕表名称（应包含方位/材质/表盘颜色等区分信息）",',
        '    "primary_color": "Blue",',
        '    "material": "材质",',
        '    "wear_threshold": 30,',
        '    "official_desc": "中文说明",',
        '    "notes": "机芯/尺寸/识别依据",',
        '    "acquired_at": "YYYY-MM-DD"',
        "  }",
        "}",
        "```",
        "",
        "#### 字段规则",
        "- 该接口需要 `items:write` scope；普通 owner key 只能写入 token 绑定 owner，跨 owner 需要 `items:write:any` 或 `admin:*`。",
        "- `dry_run: true` 只校验，不写库、不保存图片、不导出。",
        "- `mode=create_only` 遇到重复 `code` 返回冲突；`mode=upsert` 或 `replace` 会更新同 owner 已有记录。",
        "- `kind=watch` 会强制 `layer_role=Watch`，并写入腕表导出；`kind=wardrobe` 写入衣橱导出。",
        "- `item.layer_role` 是单品主数据角色，必须使用规范值：`Inner`、`Middle`、`Outer`、`Bottom`、`Footwear`、`Watch`、`Accessory`、`Dress`、`Home`、`Bespoke`。鞋类统一写 `Footwear`，不得写 `Shoes`；裤装统一写 `Bottom`，不得写 `Pants`。",
        "- 新产品入库应优先沿用当前 owner 数据里已有的 `brand` 规范值；已知别名如 `江诗丹顿`、`Vacheron Constantin 江诗丹顿` 会规范化为 `Vacheron Constantin`，避免新建重复品牌。",
        "- 腕表 `section` 是 App 和 `items` 同步资源的主显示标题，应由 Hermes 在写入前规范化，包含能区分同系列变体的信息，尤其是方位/限定名、材质和表盘颜色。API 不自动推断或追加盘面颜色；`primary_color`、`secondary_color`、`official_desc` 只作为结构化字段保存。",
        "- `wear_threshold` 是磨损/保养阈值数字字段；Hermes 入库交互里用户说“磨损阈值”“保养阈值”“thr”“maintenance_threshold”等，应映射为 `item.wear_threshold`。未知可省略或填 `null`/`0`；不要把它和当前累计磨损 `wear_maintenance` 混用。",
        "- 图片是可选项。JSON 请求可提交 `photos[]` 的 `data_base64` 或 `data_url`；multipart 请求可直接提交 JPG/JPEG/PNG 等图片文件；已存在产品还可用 `POST /api/v1/items/{code}/photos` 和 `Content-Type: image/jpeg` raw body 补图。单张上限沿用上传限制，默认最多 12 张。",
        "- 推荐 Hermes 稳定流程：先用 JSON `POST /api/v1/items` 完成结构化入库和 readback；再用 `POST /api/v1/items/{code}/photos` 通过 raw image body 或 multipart file part 补传图片。",
        "- 裸 `path`、`file://...` 或本地路径字符串不是上传；HTTP 请求必须携带实际图片字节。若 Hermes 当前工具无法发送文件字节，应报告工具层限制，不能把路径字符串当作附件写库成功。",
        "- `photos[]` 仅用于保存产品照片；Hermes 必须先完成图片分析并提交结构化字段，不能指望衣橱 API 自动识别品牌、材质或颜色。",
        "- 产品照片按 `sort_order ASC, id ASC` 排序；第一张就是客服或自动分析工具应优先拉取的预览/全图。",
        "- `GET /api/v1/items/{code}` 返回 `photos[]`、`primary_photo`、`primary_photo_thumbnail_path` 与 `primary_photo_content_path`；可用 `GET /api/v1/items/{code}/photos/primary/thumbnail` 下载安全 JPEG 缩略图。",
        "- 日常搭配判断应缓存并优先使用首图缩略图；只有细节、OCR、视觉比对或原图交付需要时才用 `GET /api/v1/items/{code}/photos/primary/content` 拉原图。",
        "- 无照片产品不得作为错误处理；`primary_photo` 为 `null` 或缩略图不可用时，Hermes 应继续使用结构化字段完成推荐和状态查询。",
        "- 可用 `POST /api/v1/items/{code}/photos/order` 调整产品照片顺序；传 `{ \"primary_photo_id\": 123 }` 可把某张照片设为第一张，传 `{ \"photo_ids\": [...] }` 可完整重排。",
        "- 照片顺序接口需要 `items:write`；`photo_ids` 必须包含该产品当前全部照片 ID，不能缺失、重复或混入其他产品照片；支持 `dry_run: true` 预览。",
        "- 正式写入成功后系统会更新 SQLite 主数据并触发对应 CSV 导出；manifest 与变化资源的 checksum 会因 SQLite 数据变化而更新。",
    ]


def _render_program_api_reference_lines() -> list[str]:
    return [
        "## Program API 引用",
        f"- Program API 规范由 Codex 生成维护，当前约定文档名为 `{WARDROBE_PROGRAM_API_DOC_FILE}`。",
        f"- 具体搭配规则由 Hermes 维护，当前约定文档名为 `{WARDROBE_HERMES_STYLE_RULES_FILE}` 或 owner 目录内同用途文档。",
        "- 本文件只说明 CSV/解析字段与使用边界，不内嵌 Access Key、endpoint 完整契约或具体搭配规则。",
        "- 历史穿搭写入、manifest/resources 同步、Access Key 解析和禁用文件链路均以 Program API 规范为准。",
    ]


def _render_wardrobe_chatgpt_rules_markdown(conn: sqlite3.Connection, owner: str) -> str:
    lines = [
        "# 衣橱 ChatGPT 解析规则",
        "",
        f"本文件面向 owner：`{_normalize_edit_value(owner)}`。",
        "",
        "本文件用于说明 `衣橱.csv` 的字段含义和推荐使用边界。",
        "",
        "## 1. 基线文件",
        "- `衣橱.csv` 是当前默认的轻量化解析基线。",
        "- 默认不再依赖 `衣橱.xlsx` 参与 ChatGPT/AI 的读取。",
        "- Owner 拆分文件为 `衣橱_徐欣.csv`、`衣橱_吴萍.csv`。",
        "- ChatGPT 解析规则文档会同步写入徐欣目录和吴萍目录。",
        "",
        "## 2. 使用原则",
        "- 搭配、推荐、筛选、统计优先读取 `衣橱.csv`。",
        "- 不需要读取大段官方描述、备注、护理说明等冗余字段。",
        "- 只保留搭配分析有用的结构化字段。",
        "- CSV 中必须保留 Ordered、保养中、其他 Owner 等记录，便于回答状态查询、审计、保养统计和预定统计。",
        "- 硬性要求：这些记录只是保留用于查询；正式推荐、搭配主案和候选池默认必须排除 `status=Ordered/ordered`、保养中以及非目标 Owner。",
        "",
        "## 3. 字段说明",
    ]
    for field_name in WARDROBE_TEXT_EXPORT_HEADERS:
        description = WARDROBE_TEXT_EXPORT_FIELD_DESCRIPTIONS.get(field_name, "")
        lines.append(f"- `{field_name}`: {description}")
    lines.extend(
        [
            "",
            "## 4. 推荐筛选顺序",
            "- 先按 Owner 收窄。",
            "- 再按 `status=Active` 和 `maintenance_state=激活` 过滤。",
            "- 如果 `status` 为 `Ordered`/`ordered`，即使其他字段匹配，也只能用于查询“哪些已预定”，不能进入推荐候选池、主案或备选方案。",
            "- 之后再按地点、场景、温区、松弛指数、层级、二级层级、材质、颜色做搭配。",
            "",
            "## 5. 历史穿搭写入",
            "- `WearCount_new*.xlsx` 文件导入已禁用。",
            "- ChatGPT 不得再生成 `WearCount_new` 文件；历史穿搭必须通过 Program API 入库。",
        ]
    )
    lines.extend([""])
    lines.extend(_render_program_api_reference_lines())
    lines.extend(
        [
            "",
            "## 6. 已禁用的文件链路",
            "- 不再支持通过 `WearCount_new*.xlsx`、`DAILY_LOG`、`ITEMS` sheet 生成历史记录。",
            "- 不要输出 Excel，不要把历史穿搭写入共享目录等待导入。",
            "- 如果需要新增历史记录，只能按上面的 Program API 规则调用接口。",
            "",
            "## 7. 解析要求",
            "- Hermes Agent 在正式分析前，应先调用 Wardrobe MCP 的 `wardrobe.sync`，由 MCP 判断本地缓存是否需要更新。",
            "- manifest 返回 `304 Not Modified` 时 MCP 可以继续使用本地缓存；返回 `200 OK` 时必须按 `resources[].checksum` 只刷新变化资源。",
        ]
    )
    return "\n".join(lines)


def _render_wearcount_chatgpt_rules_markdown(conn: sqlite3.Connection, owner: str) -> str:
    role_sheet_specs = [
        ("Outer", "外套和外层单品"),
        ("Middle", "中层单品"),
        ("Inner", "内层单品"),
        ("Bottom", "裤装"),
        ("Footwear", "鞋类"),
        ("Accessory", "配件"),
        ("Watch", "腕表"),
    ]
    text_field_descriptions = {
        "role": "角色分类，必须使用规范值 Outer / Middle / Inner / Bottom / Footwear / Accessory / Watch。",
        "code": "商品编码或 Ref。",
        "item": "商品名称或 Section。",
        "brand": "品牌。",
        "owner": "Owner。",
        "wear": "磨损指数，对应 wear_maintenance。",
        "total": "总穿着次数，对应 wear_total。",
        "wear_year": "当前年度穿着次数，对应 wear_year。",
        "maint": "保养次数，对应 maint_count。",
        "thr": "磨损阈值，对应 wear_threshold。",
        "status": "状态。",
        "maintenance_state": "保养状态。",
        "last_worn_on": "上次穿着/佩戴日期。",
    }
    lines = [
        "# WearCount ChatGPT 解析规则",
        "",
        f"本文件面向 owner：`{_normalize_edit_value(owner)}`。",
        "",
        "本文件用于说明 `WearCount.csv` 的字段含义和推荐使用边界。",
        "",
        "## 1. 基线文件",
        "- `WearCount.csv` 是当前默认的轻量化 WearCount 基线。",
        "- 默认不再依赖 `WearCount.xlsx` 参与 ChatGPT/AI 的读取。",
        "- Owner 拆分文件为 `WearCount_徐欣.csv`、`WearCount_吴萍.csv`。",
        "- 规则文档会同步写入徐欣目录和吴萍目录。",
        "",
        "## 2. 字段说明",
    ]
    for field_name in WEARCOUNT_TEXT_EXPORT_HEADERS:
        lines.append(f"- `{field_name}`: {text_field_descriptions.get(field_name, '')}")
    lines.extend(["", "## 3. role 含义"])
    for sheet_name, meaning in role_sheet_specs:
        lines.append(f"- `{sheet_name}`: {meaning}")
    lines.extend(
        [
            "",
            "## 4. 使用原则",
            "- ChatGPT/AI 默认读取 `WearCount.csv` 做磨损、保养、累计穿着分析。",
            "- `衣橱.csv` 与 `WearCount.csv` 可以联合使用：前者负责搭配属性，后者负责磨损和穿着统计。",
            "- `wear` 表示当前磨损指数，`thr` 表示阈值，二者需要结合理解保养需求。",
            "- CSV 中必须保留 Ordered、保养中、其他 Owner 等记录，便于回答状态查询、审计、保养统计和预定统计。",
            "- 硬性要求：这些记录只是保留用于查询；正式推荐、搭配主案和候选池默认必须排除 `status=Ordered/ordered`、保养中以及非目标 Owner。",
            "",
            "## 5. 与增量导入的关系",
            "- `WearCount.csv` 不承担 `DAILY_LOG` 的完整回放功能。",
            "- `WearCount_new*.xlsx` 文件导入已禁用。",
            "- 每日新增历史穿搭必须通过 Program API 入库。",
            "- API 写入完成后系统会重新导出最新 `WearCount.csv`。",
        ]
    )
    lines.extend([""])
    lines.extend(_render_program_api_reference_lines())
    return "\n".join(lines)


def _export_wearcount_chatgpt_rules_docs(conn: sqlite3.Connection) -> list[dict]:
    return _export_owner_markdown_docs_to_targets(
        conn,
        WEARCOUNT_CHATGPT_RULES_FILE,
        _render_wearcount_chatgpt_rules_markdown,
        "wearcount_chatgpt_rules",
    )


def _export_wardrobe_text_csv(conn: sqlite3.Connection) -> dict:
    target_path = DEFAULT_IMPORT_DIR / WARDROBE_TEXT_EXPORT_FILE
    temp_path = _export_temp_path(target_path, "export-wardrobe-text")
    rows = _wardrobe_export_rows(conn)
    export_on = datetime.now().date()
    try:
        with temp_path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(WARDROBE_TEXT_EXPORT_HEADERS)
            for row in rows:
                writer.writerow(_wardrobe_text_export_row(row, export_on))
        _finalize_export_file(temp_path, target_path)
        return {
            "kind": "wardrobe_ai_csv",
            "path": str(target_path),
            "rows": len(rows),
            "columns": list(WARDROBE_TEXT_EXPORT_HEADERS),
        }
    finally:
        temp_path.unlink(missing_ok=True)


def _export_wearcount_owner_text_csv(conn: sqlite3.Connection, file_name: str, owners: set[str] | None = None) -> dict:
    target_path = _owner_specific_export_dir(owners) / file_name
    temp_path = _export_temp_path(target_path, "export-wearcount-owner-text")
    rows = _wearcount_text_export_rows(conn, owners)
    try:
        with temp_path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(WEARCOUNT_TEXT_EXPORT_HEADERS)
            for row in rows:
                writer.writerow([row.get(field_name, "") for field_name in WEARCOUNT_TEXT_EXPORT_HEADERS])
        _finalize_export_file(temp_path, target_path)
        return {
            "kind": "wearcount_owner_csv",
            "path": str(target_path),
            "rows": len(rows),
            "columns": list(WEARCOUNT_TEXT_EXPORT_HEADERS),
            "owners": _normalized_owner_filters(owners),
        }
    finally:
        temp_path.unlink(missing_ok=True)


def _export_wardrobe_owner_text_csv(conn: sqlite3.Connection, file_name: str, owners: set[str] | None = None) -> dict:
    target_path = _owner_specific_export_dir(owners) / file_name
    temp_path = _export_temp_path(target_path, "export-wardrobe-owner-text")
    rows = _wardrobe_export_rows(conn, owners)
    export_on = datetime.now().date()
    try:
        with temp_path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(WARDROBE_TEXT_EXPORT_HEADERS)
            for row in rows:
                writer.writerow(_wardrobe_text_export_row(row, export_on))
        _finalize_export_file(temp_path, target_path)
        return {
            "kind": "wardrobe_owner_csv",
            "path": str(target_path),
            "rows": len(rows),
            "columns": list(WARDROBE_TEXT_EXPORT_HEADERS),
            "owners": _normalized_owner_filters(owners),
        }
    finally:
        temp_path.unlink(missing_ok=True)


def _export_wardrobe_chatgpt_rules_docs(conn: sqlite3.Connection) -> list[dict]:
    return _export_owner_markdown_docs_to_targets(
        conn,
        WARDROBE_CHATGPT_RULES_FILE,
        _render_wardrobe_chatgpt_rules_markdown,
        "wardrobe_chatgpt_rules",
    )


def _retire_export_files(directory: Path, file_names: list[str]) -> list[dict]:
    results: list[dict] = []
    for file_name in file_names:
        stale_path = directory / file_name
        if not stale_path.exists():
            continue
        archive_dir = stale_path.parent / "log"
        archive_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        archived_name = f"{stale_path.stem}_retired_{timestamp}{stale_path.suffix}"
        archived_path = archive_dir / archived_name
        try:
            stale_path.replace(archived_path)
        except PermissionError as exc:
            raise RuntimeError(f"baseline_file_locked:{stale_path}") from exc
        _apply_export_ownership(archived_path)
        _stabilize_export_for_drive(archived_path)
        _nudge_export_for_drive(archived_path)
        _pulse_export_directory(archived_path.parent)
        _enqueue_drive_notify(archived_path)
        results.append({
            "kind": "retired_export",
            "path": str(archived_path),
            "replaced": str(stale_path),
        })
    return results


def _export_wardrobe_xlsx_workbooks(conn: sqlite3.Connection) -> list[dict]:
    _retire_export_files(DEFAULT_IMPORT_DIR, RETIRED_CLOSET_EXPORT_FILES)
    results: list[dict] = []
    results.append(_export_single_wardrobe_workbook(conn, WARDROBE_FILE))
    for file_name, owners in WARDROBE_OWNER_EXPORT_SPECS:
        results.append(_export_single_wardrobe_workbook(conn, file_name, owners))
    return results


def _export_wardrobe_workbooks(conn: sqlite3.Connection) -> list[dict]:
    results: list[dict] = []
    results.extend(_retire_export_files(DEFAULT_IMPORT_DIR, RETIRED_CLOSET_EXPORT_FILES))
    results.extend(_retire_export_files(WUPING_CLOSET_DIR, RETIRED_CLOSET_EXPORT_FILES))
    results.append(_export_wardrobe_text_csv(conn))
    for file_name, owners in WARDROBE_OWNER_EXPORT_SPECS:
        results.append(_export_wardrobe_owner_text_csv(conn, file_name, owners))
    results.extend(_export_wardrobe_chatgpt_rules_docs(conn))
    return results


def _export_watch_workbook(conn: sqlite3.Connection) -> dict:
    target_path = WATCH_IMPORT_DIR / WATCH_FILE
    temp_path = _export_temp_path(target_path, "export-watch")
    workbook = Workbook()
    try:
        worksheet = workbook.active
        worksheet.title = "腕表"
        _write_export_row(worksheet, 1, WATCH_EXPORT_HEADERS)
        rows = _watch_export_rows(conn)
        export_rows: list[list[object]] = []
        for index, row in enumerate(rows, start=1):
            values = [
                _normalize_edit_value(row["code"]),
                _normalize_edit_value(row["brand"]),
                _normalize_edit_value(row["status"]),
                _normalize_edit_value(row["section"]),
                _normalize_edit_value(row["loc"]),
                _maintenance_state_label(row),
                _normalize_edit_value(row["owner"]),
                _normalize_edit_value(row["scene_tag"]),
                _normalize_edit_value(row["official_desc"]),
                normalize_price_text(row["price_original"] or row["price_cny"]),
                normalize_price_currency(row["price_original_currency"], row["price_original"] or row["price_cny"]),
                _parse_export_datetime(row["acquired_at"]),
                _normalize_edit_value(row["material"]),
                _normalize_edit_value(row["notes"]),
            ]
            export_rows.append(values)
            _write_export_row(worksheet, index + 1, values)
        _apply_export_widths(worksheet, WATCH_EXPORT_HEADERS, export_rows)
        workbook.save(temp_path)
        workbook.close()
        _finalize_export_file(temp_path, target_path)
        return {"kind": "watch", "path": str(target_path), "rows": len(rows)}
    finally:
        workbook.close()
        temp_path.unlink(missing_ok=True)


def _export_watch_text_csv(conn: sqlite3.Connection) -> dict:
    _retire_export_files(WATCH_IMPORT_DIR, RETIRED_WATCH_EXPORT_FILES)
    target_path = WATCH_IMPORT_DIR / WATCH_TEXT_EXPORT_FILE
    temp_path = _export_temp_path(target_path, "export-watch-text")
    rows = _watch_export_rows(conn)
    try:
        with temp_path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(WATCH_TEXT_EXPORT_HEADERS)
            for row in rows:
                writer.writerow([
                    _normalize_edit_value(row["code"]),
                    _normalize_edit_value(row["owner"]),
                    _normalize_edit_value(row["brand"]),
                    _normalize_edit_value(row["section"]),
                    _normalize_edit_value(row["loc"]),
                    _normalize_edit_value(row["status"]),
                    _maintenance_state_label(row),
                    _normalize_edit_value(row["scene_tag"]),
                    _normalize_edit_value(row["material"]),
                    normalize_price_text(row["price_original"]),
                    normalize_price_currency(row["price_original_currency"], row["price_original"]),
                    normalize_price_text(row["price_cny"]),
                    _normalize_date_compare(row["acquired_at"]),
                    _normalize_edit_value(row["wear_total"]),
                    _normalize_edit_value(row["wear_year"]),
                    _normalize_date_compare(row["last_worn_on"]),
                    _normalize_edit_value(row["maint_count"]),
                ])
        _finalize_export_file(temp_path, target_path)
        return {
            "kind": "watch_ai_csv",
            "path": str(target_path),
            "rows": len(rows),
            "columns": list(WATCH_TEXT_EXPORT_HEADERS),
        }
    finally:
        temp_path.unlink(missing_ok=True)


def _featured_looks_text_export_rows(conn: sqlite3.Connection, owners: set[str] | None = None) -> list[dict[str, object]]:
    rows = _featured_look_export_rows(conn, owners)
    slot_columns = {
        "inner": "inner_code",
        "middle": "middle_code",
        "outer": "outer_code",
        "bottom": "bottom_code",
        "footwear": "footwear_code",
        "watch": "watch_ref",
    }
    export_rows: list[dict[str, object]] = []
    for look in rows:
        slot_rows = conn.execute(
            """
            SELECT slot, source_code
            FROM featured_look_items
            WHERE featured_look_id = ?
            ORDER BY display_order ASC, id ASC
            """,
            (look["id"],),
        ).fetchall()
        slot_values = {column: "" for column in slot_columns.values()}
        for slot_row in slot_rows:
            slot = _normalize_edit_value(slot_row["slot"]).lower()
            column_name = slot_columns.get(slot)
            if column_name:
                slot_values[column_name] = _normalize_edit_value(slot_row["source_code"])
        export_rows.append({
            "look_id": _normalize_edit_value(look["look_id"]),
            "owner": _normalize_edit_value(look["owner"]),
            "status": _normalize_edit_value(look["status"]),
            "use_case": _normalize_edit_value(look["use_case"]),
            "temp_min": _normalized_number_value(look["temp_min"]),
            "temp_max": _normalized_number_value(look["temp_max"]),
            "relax_center": _normalized_number_value(look["relax_center"]),
            "relax_span": _normalized_number_value(look["relax_span"]),
            "inner_code": slot_values["inner_code"],
            "middle_code": slot_values["middle_code"],
            "outer_code": slot_values["outer_code"],
            "bottom_code": slot_values["bottom_code"],
            "footwear_code": slot_values["footwear_code"],
            "watch_ref": slot_values["watch_ref"],
            "notes": _normalize_edit_value(look["notes"]),
            "created_at": _normalize_date_compare(look["created_at"]) or "",
        })
    return export_rows


def _export_featured_looks_text_csv(conn: sqlite3.Connection, file_name: str = LOOKS_TEXT_EXPORT_FILE, owners: set[str] | None = None) -> dict:
    target_path = _owner_specific_export_dir(owners) / file_name
    temp_path = _export_temp_path(target_path, "export-featured-looks-text")
    rows = _featured_looks_text_export_rows(conn, owners)
    try:
        with temp_path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(LOOKS_TEXT_EXPORT_HEADERS)
            for row in rows:
                writer.writerow([row.get(field_name, "") for field_name in LOOKS_TEXT_EXPORT_HEADERS])
        _finalize_export_file(temp_path, target_path)
        return {
            "kind": "featured_looks_csv",
            "path": str(target_path),
            "rows": len(rows),
            "columns": list(LOOKS_TEXT_EXPORT_HEADERS),
            "owners": _normalized_owner_filters(owners),
        }
    finally:
        temp_path.unlink(missing_ok=True)


def _export_featured_looks_workbook(
    conn: sqlite3.Connection,
    file_name: str = LOOKS_FILE,
    owners: set[str] | None = None,
) -> dict:
    target_path = _owner_specific_export_dir(owners) / file_name
    temp_path = _export_temp_path(target_path, "export-featured-looks")
    workbook = Workbook()
    slot_columns = {
        "inner": "Inner_Code",
        "middle": "Mid_Code",
        "outer": "Outer_Code",
        "bottom": "Bottom_Code",
        "footwear": "Footwear_Code",
        "watch": "Watch_Ref",
    }
    try:
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        rows = _featured_look_export_rows(conn, owners)
        sheet_specs = [
            ("LOOK_MASTER", [row for row in rows if _normalize_edit_value(row["status"]) != "Archived"]),
            ("ARCHIVE", [row for row in rows if _normalize_edit_value(row["status"]) == "Archived"]),
        ]
        total_rows = 0
        for sheet_name, look_rows in sheet_specs:
            worksheet = workbook.create_sheet(title=sheet_name)
            _write_export_row(worksheet, 1, FEATURED_LOOK_EXPORT_HEADERS)
            export_rows: list[list[object]] = []
            for index, look in enumerate(look_rows, start=1):
                slot_rows = conn.execute(
                    """
                    SELECT slot, source_code, source_section
                    FROM featured_look_items
                    WHERE featured_look_id = ?
                    ORDER BY display_order ASC, id ASC
                    """,
                    (look["id"],),
                ).fetchall()
                slot_values = {column: "" for column in slot_columns.values()}
                anchor_code = _normalize_edit_value(look["anchor_code"])
                anchor_section = _normalize_edit_value(look["anchor_section"])
                for slot_row in slot_rows:
                    slot = _normalize_edit_value(slot_row["slot"]).lower()
                    source_code = _normalize_edit_value(slot_row["source_code"])
                    source_section = _normalize_edit_value(slot_row["source_section"])
                    if slot == "anchor":
                        if source_code:
                            anchor_code = source_code
                        if source_section:
                            anchor_section = source_section
                        continue
                    column_name = slot_columns.get(slot)
                    if column_name:
                        slot_values[column_name] = source_code
                values = [
                    _normalize_edit_value(look["look_id"]),
                    _normalize_edit_value(look["anchor_type"]),
                    anchor_code,
                    anchor_section,
                    _normalize_edit_value(look["use_case"]),
                    _normalize_edit_value(look["priority"]),
                    _normalize_edit_value(look["status"]),
                    _normalize_edit_value(look["owner"]),
                    _normalized_number_value(look["temp_min"]),
                    _normalized_number_value(look["temp_max"]),
                    _normalize_edit_value(look["scene_tag_target"]),
                    _normalized_number_value(_featured_look_relax_center(conn, int(look["id"]))),
                    _normalized_number_value(_featured_look_relax_span(conn, int(look["id"]))),
                    _normalize_edit_value(look["notes"]),
                    slot_values["Inner_Code"],
                    slot_values["Mid_Code"],
                    slot_values["Outer_Code"],
                    slot_values["Bottom_Code"],
                    slot_values["Footwear_Code"],
                    slot_values["Watch_Ref"],
                ]
                export_rows.append(values)
                _write_export_row(worksheet, index + 1, values)
            _apply_export_widths(worksheet, FEATURED_LOOK_EXPORT_HEADERS, export_rows)
            total_rows += len(look_rows)
        workbook.save(temp_path)
        workbook.close()
        _finalize_export_file(temp_path, target_path)
        return {
            "kind": "featured_looks",
            "path": str(target_path),
            "rows": total_rows,
            "owners": _normalized_owner_filters(owners),
        }
    finally:
        workbook.close()
        temp_path.unlink(missing_ok=True)


def _export_featured_looks_workbooks(conn: sqlite3.Connection) -> list[dict]:
    results: list[dict] = []
    results.extend(_retire_export_files(DEFAULT_IMPORT_DIR, RETIRED_CLOSET_EXPORT_FILES))
    results.extend(_retire_export_files(WUPING_CLOSET_DIR, RETIRED_CLOSET_EXPORT_FILES))
    results.append(_export_featured_looks_text_csv(conn, LOOKS_TEXT_EXPORT_FILE))
    for file_name, owners in LOOKS_OWNER_EXPORT_SPECS:
        results.append(_export_featured_looks_text_csv(conn, file_name, owners))
    return results


def _export_item_baseline(conn: sqlite3.Connection, kind: str) -> list[dict]:
    results = []
    if kind == "watch":
        results.append(_export_watch_text_csv(conn))
    else:
        results.extend(_export_wardrobe_workbooks(conn))
    results.extend(_export_wearcount_workbooks(conn))
    return results


def _run_item_export_tasks(conn: sqlite3.Connection, tasks: set[str]) -> list[dict]:
    results: list[dict] = []
    for task in ITEM_EXPORT_TASK_ORDER:
        if task not in tasks:
            continue
        if task == "wardrobe_ai":
            results.append(_export_wardrobe_text_csv(conn))
            for file_name, owners in WARDROBE_OWNER_EXPORT_SPECS:
                csv_name = Path(file_name).with_suffix(".csv").name
                results.append(_export_wardrobe_owner_text_csv(conn, csv_name, owners))
            results.extend(_export_wardrobe_chatgpt_rules_docs(conn))
        elif task == "watch_ai":
            results.append(_export_watch_text_csv(conn))
        elif task == "looks_ai":
            results.extend(_export_featured_looks_workbooks(conn))
        elif task == "wearcount":
            results.extend(_export_wearcount_workbooks(conn))
    return results


def _item_compare_fields(before: sqlite3.Row | dict | None, after: sqlite3.Row | dict | None) -> set[str]:
    changed: set[str] = set()
    before_keys = set(before.keys()) if before is not None else set()
    after_keys = set(after.keys()) if after is not None else set()
    compare_fields = before_keys | after_keys
    number_fields = set(EDIT_NUMBER_FIELDS) | {"wear_total", "wear_year", "maint_count", "maintenance_state"}
    date_fields = set(EDIT_DATE_FIELDS) | {"last_worn_on"}
    for field in compare_fields:
        before_value = before[field] if before is not None and field in before_keys else None
        after_value = after[field] if after is not None and field in after_keys else None
        if field in date_fields:
            if _normalize_date_compare(before_value) != _normalize_date_compare(after_value):
                changed.add(field)
            continue
        if field in number_fields:
            if _normalize_edit_value(before_value) != _normalize_edit_value(after_value):
                changed.add(field)
            continue
        if _normalize_edit_value(before_value) != _normalize_edit_value(after_value):
            changed.add(field)
    return changed


def _item_export_tasks_for_change(before: sqlite3.Row | dict | None, after: sqlite3.Row | dict | None) -> set[str]:
    tasks: set[str] = set()
    changed_fields = _item_compare_fields(before, after)
    if not changed_fields:
        return tasks
    before_kind = _item_source_kind(before) if before is not None else ""
    after_kind = _item_source_kind(after) if after is not None else ""
    if before_kind != after_kind:
        if before_kind == "watch" or after_kind == "watch":
            tasks.add("watch_ai")
        if before_kind == "wardrobe" or after_kind == "wardrobe":
            tasks.add("wardrobe_ai")
    if "watch_ai" not in tasks and after_kind == "watch" and changed_fields & WATCH_AI_EXPORT_FIELDS:
        tasks.add("watch_ai")
    if "wardrobe_ai" not in tasks and after_kind == "wardrobe" and changed_fields & WARDROBE_AI_EXPORT_FIELDS:
        tasks.add("wardrobe_ai")
    if changed_fields & WEARCOUNT_EXPORT_FIELDS:
        tasks.add("wearcount")
    return tasks


def _wear_aggregate_export_tasks_for_item_ids(conn: sqlite3.Connection, item_ids: list[int] | tuple[int, ...] | set[int]) -> set[str]:
    normalized_ids = sorted({int(value) for value in item_ids if value is not None})
    if not normalized_ids:
        return set()
    placeholders = ",".join("?" for _ in normalized_ids)
    rows = conn.execute(
        f"SELECT id, layer_role FROM items WHERE id IN ({placeholders})",
        normalized_ids,
    ).fetchall()
    tasks = {"wearcount"}
    if any(_item_source_kind(row) == "wardrobe" for row in rows):
        tasks.update({"wardrobe_ai", "looks_ai"})
    if any(_item_source_kind(row) == "watch" for row in rows):
        tasks.add("watch_ai")
    return tasks


def _run_item_export_worker() -> None:
    global ITEM_EXPORT_RUNNING
    while True:
        with ITEM_EXPORT_LOCK:
            if not ITEM_EXPORT_PENDING:
                ITEM_EXPORT_RUNNING = False
                return
            tasks = set(ITEM_EXPORT_PENDING)
            ITEM_EXPORT_PENDING.clear()
        conn = connect()
        try:
            _run_item_export_tasks(conn, tasks)
            conn.commit()
        except Exception as exc:
            conn.rollback()
            print(f"[item-export] async export failed for {sorted(tasks)}: {exc}")
        finally:
            conn.close()


def _schedule_item_export_tasks(tasks: set[str]) -> None:
    global ITEM_EXPORT_RUNNING
    if not tasks:
        return
    should_start = False
    with ITEM_EXPORT_LOCK:
        ITEM_EXPORT_PENDING.update(tasks)
        if not ITEM_EXPORT_RUNNING:
            ITEM_EXPORT_RUNNING = True
            should_start = True
    if should_start:
        threading.Thread(
            target=_run_item_export_worker,
            daemon=True,
            name="item-export-worker",
        ).start()


def _insert_item_record(conn: sqlite3.Connection, payload: dict) -> int:
    values = _normalize_item_payload(payload)
    kind = _new_item_kind(payload)
    values["layer_role"] = "Watch" if kind == "watch" else _normalize_edit_value(values.get("layer_role"))
    values["source_sheet"] = _default_source_sheet(kind)
    values["status"] = _normalize_edit_value(values.get("status")) or "Active"
    cursor = conn.execute(
        """
        INSERT INTO items (
            code, brand, section, loc, owner, layer_role, outer_type, scene_tag,
            relax_index, temp_min, temp_max, standalone_min, standalone_max,
            wear_maintenance, wear_threshold,
            primary_color, secondary_color, official_desc, price_original, price_original_currency, price_cny,
            series, size, acquired_at, official_color_code, material, care, notes,
            status, source_sheet, updated_at
        )
        VALUES (
            :code, :brand, :section, :loc, :owner, :layer_role, :outer_type, :scene_tag,
            :relax_index, :temp_min, :temp_max, :standalone_min, :standalone_max,
            :wear_maintenance, :wear_threshold,
            :primary_color, :secondary_color, :official_desc, :price_original, :price_original_currency, :price_cny,
            :series, :size, :acquired_at, :official_color_code, :material, :care, :notes,
            :status, :source_sheet, CURRENT_TIMESTAMP
        )
        """,
        values,
    )
    return int(cursor.lastrowid)


def _update_item_record(conn: sqlite3.Connection, item: sqlite3.Row, payload: dict) -> int:
    requested_status = _requested_wardrobe_status(payload, item)
    values = _normalize_item_payload(payload, item)
    kind = _new_item_kind(values)
    values["layer_role"] = "Watch" if kind == "watch" else _normalize_edit_value(values.get("layer_role"))
    values["source_sheet"] = _default_source_sheet(kind)
    values["status"] = _normalize_edit_value(values.get("status")) or _normalize_edit_value(item["status"]) or "Active"
    values["item_id"] = int(item["id"])
    conn.execute(
        """
        UPDATE items
        SET code = :code,
            brand = :brand,
            section = :section,
            loc = :loc,
            owner = :owner,
            layer_role = :layer_role,
            outer_type = :outer_type,
            scene_tag = :scene_tag,
            relax_index = :relax_index,
            temp_min = :temp_min,
            temp_max = :temp_max,
            standalone_min = :standalone_min,
            standalone_max = :standalone_max,
            wear_maintenance = :wear_maintenance,
            wear_threshold = :wear_threshold,
            primary_color = :primary_color,
            secondary_color = :secondary_color,
            official_desc = :official_desc,
            price_original = :price_original,
            price_original_currency = :price_original_currency,
            price_cny = :price_cny,
            series = :series,
            size = :size,
            acquired_at = :acquired_at,
            official_color_code = :official_color_code,
            material = :material,
            care = :care,
            notes = :notes,
            status = :status,
            source_sheet = :source_sheet,
            updated_at = CURRENT_TIMESTAMP
        WHERE id = :item_id
        """,
        values,
    )
    if kind != "watch":
        _apply_wardrobe_status_transition(conn, int(item["id"]), requested_status)
    return int(item["id"])


def _send_item_to_maintenance(conn: sqlite3.Connection, item: sqlite3.Row) -> tuple[dict, bool]:
    changed = _maintenance_state_value(item) != 1
    if changed:
        current_loc = _normalize_edit_value(item["loc"])
        previous_loc = _normalize_edit_value(item["maintenance_prev_loc"])
        next_prev_loc = current_loc or previous_loc
        conn.execute(
            """
            UPDATE items
            SET maintenance_state = 1,
                maintenance_prev_loc = ?,
                wear_maintenance = 0,
                maint_count = COALESCE(maint_count, 0) + 1,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (next_prev_loc, int(item["id"])),
        )
    conn.commit()
    updated_row = conn.execute("SELECT * FROM items WHERE id = ?", (int(item["id"]),)).fetchone()
    updated = _item_with_photos(conn, int(item["id"]))
    export_tasks = _item_export_tasks_for_change(item, updated_row)
    _schedule_item_export_tasks(export_tasks)
    return {"item": updated, "exports_pending": bool(export_tasks), "changed": changed}, changed


def _activate_item_from_maintenance(conn: sqlite3.Connection, item: sqlite3.Row) -> tuple[dict, bool]:
    changed = _maintenance_state_value(item) == 1
    if changed:
        previous_loc = _normalize_edit_value(item["maintenance_prev_loc"])
        current_loc = _normalize_edit_value(item["loc"])
        restored_loc = current_loc or previous_loc
        conn.execute(
            """
            UPDATE items
            SET maintenance_state = 0,
                maintenance_prev_loc = NULL,
                loc = ?,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (restored_loc, int(item["id"])),
        )
    conn.commit()
    updated_row = conn.execute("SELECT * FROM items WHERE id = ?", (int(item["id"]),)).fetchone()
    updated = _item_with_photos(conn, int(item["id"]))
    export_tasks = _item_export_tasks_for_change(item, updated_row)
    _schedule_item_export_tasks(export_tasks)
    return {"item": updated, "exports_pending": bool(export_tasks), "changed": changed}, changed


def _item_matches_payload(item: dict | None, payload: dict) -> bool:
    if item is None:
        return False
    for field, expected in payload.items():
        actual = item.get(field)
        if field in EDIT_DATE_FIELDS:
            if _normalize_date_compare(actual) != _normalize_date_compare(expected):
                return False
            continue
        if field in EDIT_NUMBER_FIELDS:
            if _normalize_edit_value(actual) != _normalize_edit_value(expected):
                return False
            continue
        if _normalize_edit_value(actual) != _normalize_edit_value(expected):
            return False
    return True


def _xlsx_normalize_target(target: str) -> str:
    target = target.lstrip("/")
    return target if target.startswith("xl/") else f"xl/{target}"


def _xlsx_column_index(cell_ref: str) -> int:
    letters = "".join(ch for ch in cell_ref if ch.isalpha())
    index = 0
    for ch in letters:
        index = index * 26 + (ord(ch.upper()) - 64)
    return index - 1


def _xlsx_column_label(index: int) -> str:
    index += 1
    result = []
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        result.append(chr(65 + remainder))
    return "".join(reversed(result))


def _xlsx_parse_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []
    root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    values: list[str] = []
    for item in root.findall("main:si", XLSX_NS):
        values.append("".join(node.text or "" for node in item.findall(".//main:t", XLSX_NS)))
    return values


def _xlsx_cell_text(cell: ET.Element, shared_strings: list[str]) -> str:
    cell_type = cell.attrib.get("t")
    value_node = cell.find("main:v", XLSX_NS)
    inline_node = cell.find("main:is", XLSX_NS)
    if cell_type == "s" and value_node is not None and value_node.text:
        try:
            return shared_strings[int(value_node.text)]
        except Exception:
            return value_node.text or ""
    if cell_type == "inlineStr" and inline_node is not None:
        return "".join(node.text or "" for node in inline_node.findall(".//main:t", XLSX_NS))
    if value_node is None:
        return ""
    return value_node.text or ""


def _xlsx_find_sheet_target(archive: zipfile.ZipFile, sheet_name: str) -> str:
    workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
    rels_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    relationships = {rel.attrib["Id"]: _xlsx_normalize_target(rel.attrib["Target"]) for rel in rels_root}
    for sheet in workbook_root.findall("main:sheets/main:sheet", XLSX_NS):
        if sheet.attrib.get("name") != sheet_name:
            continue
        rel_id = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id", "")
        if rel_id in relationships:
            return relationships[rel_id]
    raise ValueError("sheet_not_found")


def _xlsx_parse_date_serial(value: str) -> float:
    parsed = datetime.strptime(value, "%Y-%m-%d")
    base = datetime(1899, 12, 30)
    return float((parsed - base).days)


def _xlsx_set_cell_value(cell: ET.Element, value: str, field_type: str) -> None:
    for child in list(cell):
        cell.remove(child)
    if not value:
        cell.attrib.pop("t", None)
        return
    if field_type == "date":
        cell.attrib.pop("t", None)
        value_node = ET.SubElement(cell, "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v")
        value_node.text = _normalize_edit_value(_xlsx_parse_date_serial(value))
        return
    if field_type == "number":
        cell.attrib.pop("t", None)
        value_node = ET.SubElement(cell, "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v")
        value_node.text = _normalize_edit_value(value)
        return
    cell.attrib["t"] = "inlineStr"
    inline = ET.SubElement(cell, "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}is")
    text_node = ET.SubElement(inline, "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t")
    text_node.text = value


def _xlsx_get_or_create_cell(row: ET.Element, row_index: int, column_index: int) -> ET.Element:
    cell_ref = f"{_xlsx_column_label(column_index)}{row_index}"
    cells = row.findall("main:c", XLSX_NS)
    for cell in cells:
        if cell.attrib.get("r") == cell_ref:
            return cell
    new_cell = ET.Element("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}c", {"r": cell_ref})
    inserted = False
    for idx, existing in enumerate(cells):
        existing_index = _xlsx_column_index(existing.attrib.get("r", "A1"))
        if existing_index > column_index:
            row.insert(idx, new_cell)
            inserted = True
            break
    if not inserted:
        row.append(new_cell)
    return new_cell


def _save_item_to_baseline_workbook(item: sqlite3.Row | dict, payload: dict) -> dict:
    workbook_path = _baseline_workbook_path(item)
    if workbook_path is None or not workbook_path.exists():
        raise FileNotFoundError("baseline workbook not found")
    sheet_name = str(item["source_sheet"] or "").strip()
    if not sheet_name or sheet_name.startswith("_") or sheet_name == "WearCount":
        raise ValueError("item does not map to an editable baseline sheet")

    field_headers = _baseline_header_map(item)
    updates = []
    for field, headers in field_headers.items():
        if field not in payload:
            continue
        field_type = "date" if field in EDIT_DATE_FIELDS else "number" if field in EDIT_NUMBER_FIELDS else "text"
        updates.append(
            {
                "field": field,
                "headers": headers,
                "field_type": field_type,
                "value": _normalize_edit_value(payload.get(field)),
            }
        )
    if not updates:
        raise ValueError("no editable fields provided")

    ET.register_namespace("", "http://schemas.openxmlformats.org/spreadsheetml/2006/main")
    temp_path = DATA_DIR / f"excel-edit-{uuid.uuid4().hex}.xlsx"
    ignored_fields: list[str] = []
    row_index = 0

    try:
        workbook = load_workbook(workbook_path)
        if sheet_name not in workbook.sheetnames:
            raise RuntimeError("sheet_not_found")
        worksheet = workbook[sheet_name]

        column_map: dict[str, int] = {}
        for cell in worksheet[1]:
            header_text = _normalize_edit_value(cell.value)
            if header_text:
                column_map[header_text] = int(cell.column)

        code_column = None
        for candidate in field_headers.get("code", ["货号", "Code", "Ref"]):
            if candidate in column_map:
                code_column = column_map[candidate]
                break
        if code_column is None:
            raise RuntimeError("code_column_not_found")

        target_row = None
        original_code = str(item["code"] or "").strip()
        for row_number in range(2, worksheet.max_row + 1):
            current_code = _normalize_edit_value(worksheet.cell(row=row_number, column=code_column).value)
            if current_code == original_code:
                target_row = row_number
                row_index = row_number
                break
        if target_row is None:
            raise RuntimeError("row_not_found")

        for update in updates:
            target_column = None
            for header_name in update["headers"]:
                if header_name in column_map:
                    target_column = column_map[header_name]
                    break
            if target_column is None:
                ignored_fields.append(update["field"])
                continue
            cell = worksheet.cell(row=target_row, column=target_column)
            value = update["value"]
            if not value:
                cell.value = None
                continue
            if update["field_type"] == "date":
                parsed = datetime.strptime(value, "%Y-%m-%d")
                cell.value = parsed
                if not cell.number_format or cell.number_format == "General":
                    cell.number_format = "yyyy-mm-dd"
                continue
            if update["field_type"] == "number":
                number = float(value)
                cell.value = int(number) if number.is_integer() else number
                continue
            cell.value = value

        workbook.save(temp_path)
        workbook.close()

        temp_path.replace(workbook_path)
        return {
            "saved": True,
            "row_index": row_index,
            "workbook_path": str(workbook_path),
            "sheet_name": sheet_name,
            "ignored_fields": ignored_fields,
        }
    finally:
        temp_path.unlink(missing_ok=True)


def _new_item_kind(payload: dict) -> str:
    kind = _normalize_edit_value(payload.get("kind")).lower()
    if kind in {"watch", "wardrobe"}:
        return kind
    return "watch" if _normalize_edit_value(payload.get("layer_role")) == "Watch" else "wardrobe"


def _resolve_new_item_sheet(workbook, kind: str, payload: dict) -> str:
    visible_sheets = [name for name in workbook.sheetnames if not str(name).startswith("_")]
    if kind == "watch":
        for candidate in ("??", *visible_sheets):
            if candidate in workbook.sheetnames:
                return candidate
        raise RuntimeError("sheet_not_found")

    requested = _normalize_edit_value(payload.get("brand"))
    if not requested:
        raise ValueError("brand_required")
    normalized = requested.casefold()
    for name in visible_sheets:
        if str(name).casefold() == normalized:
            return str(name)
    raise RuntimeError("sheet_not_found")


def _append_item_to_baseline_workbook(payload: dict) -> dict:
    kind = _new_item_kind(payload)
    workbook_path = _baseline_workbook_path_for_kind(kind)
    if workbook_path is None or not workbook_path.exists():
        raise FileNotFoundError("baseline workbook not found")

    field_headers = _baseline_header_map_for_kind(kind)
    code = _normalize_edit_value(payload.get("code"))
    section = _normalize_edit_value(payload.get("section"))
    if not code:
        raise ValueError("code_required")
    if not section:
        raise ValueError("section_required")
    if kind == "wardrobe" and not _normalize_edit_value(payload.get("brand")):
        raise ValueError("brand_required")

    temp_path = DATA_DIR / f"excel-create-{uuid.uuid4().hex}.xlsx"
    workbook = None
    try:
        workbook = load_workbook(workbook_path)
        sheet_name = _resolve_new_item_sheet(workbook, kind, payload)
        worksheet = workbook[sheet_name]

        column_map: dict[str, int] = {}
        for cell in worksheet[1]:
            header_text = _normalize_edit_value(cell.value)
            if header_text:
                column_map[header_text] = int(cell.column)

        code_column = None
        for candidate in field_headers.get("code", ["??", "Code", "Ref"]):
            if candidate in column_map:
                code_column = column_map[candidate]
                break
        if code_column is None:
            raise RuntimeError("code_column_not_found")

        for row_number in range(2, worksheet.max_row + 1):
            current_code = _normalize_edit_value(worksheet.cell(row=row_number, column=code_column).value)
            if current_code == code:
                raise ValueError("duplicate_code")

        row_index = worksheet.max_row + 1
        for field, headers in field_headers.items():
            target_column = None
            for header_name in headers:
                if header_name in column_map:
                    target_column = column_map[header_name]
                    break
            if target_column is None:
                continue
            cell = worksheet.cell(row=row_index, column=target_column)
            value = _normalize_edit_value(payload.get(field))
            if not value:
                cell.value = None
                continue
            if field in EDIT_DATE_FIELDS:
                parsed = datetime.strptime(value, "%Y-%m-%d")
                cell.value = parsed
                if not cell.number_format or cell.number_format == "General":
                    cell.number_format = "yyyy-mm-dd"
                continue
            if field in EDIT_NUMBER_FIELDS:
                number = float(value)
                cell.value = int(number) if number.is_integer() else number
                continue
            cell.value = value

        workbook.save(temp_path)
        workbook.close()
        workbook = None
        temp_path.replace(workbook_path)
        return {
            "saved": True,
            "kind": kind,
            "row_index": row_index,
            "workbook_path": str(workbook_path),
            "sheet_name": sheet_name,
            "code": code,
        }
    finally:
        if workbook is not None:
            workbook.close()
        temp_path.unlink(missing_ok=True)


def _reconcile_imported_item(conn: sqlite3.Connection, old_item_id: int, new_code: str) -> dict | None:
    row = conn.execute("SELECT id FROM items WHERE code = ?", (new_code,)).fetchone()
    if row is None:
        return None
    new_item_id = int(row["id"])
    if new_item_id != old_item_id:
        conn.execute(
            """
            DELETE FROM outfit_items
            WHERE item_id = ?
              AND outfit_id IN (SELECT outfit_id FROM outfit_items WHERE item_id = ?)
            """,
            (old_item_id, new_item_id),
        )
        conn.execute("UPDATE photos SET item_id = ? WHERE item_id = ?", (new_item_id, old_item_id))
        conn.execute("UPDATE outfit_items SET item_id = ? WHERE item_id = ?", (new_item_id, old_item_id))
        conn.execute("UPDATE featured_look_items SET item_id = ? WHERE item_id = ?", (new_item_id, old_item_id))
        conn.execute("DELETE FROM items WHERE id = ?", (old_item_id,))
        conn.commit()
    return _item_with_photos(conn, new_item_id)


def _json_bytes(payload: dict | list) -> bytes:
    return json.dumps(payload, ensure_ascii=False).encode("utf-8")


def _parse_json(handler: BaseHTTPRequestHandler) -> dict:
    length = int(handler.headers.get("Content-Length", "0"))
    raw = handler.rfile.read(length) if length else b"{}"
    return json.loads(raw.decode("utf-8"))


def _parse_multipart(handler: BaseHTTPRequestHandler) -> dict[str, dict]:
    return {
        str(part.get("name") or ""): {
            "filename": part.get("filename"),
            "content": part.get("content") or b"",
            "content_type": part.get("content_type") or "",
        }
        for part in parse_multipart_parts(handler)
    }


def _item_with_photos(conn: sqlite3.Connection, item_id: int) -> dict | None:
    row = conn.execute("SELECT * FROM items WHERE id = ?", (item_id,)).fetchone()
    if row is None:
        return None
    item = dict(row)
    photos = photo_ordering.item_photo_rows(conn, item_id)
    item["photos"] = [
        {
            **dict(photo),
            "content_path": f"/api/photos/{photo['id']}/content",
            "delete_path": f"/api/photos/{photo['id']}",
        }
        for photo in photos
    ]
    return item


def _api_item_photo_content_path(code: str, photo_id: int) -> str:
    return f"/api/v1/items/{quote(_normalize_edit_value(code), safe='')}/photos/{int(photo_id)}/content"


def _api_item_photo_thumbnail_path(code: str, photo_id: int) -> str:
    return f"/api/v1/items/{quote(_normalize_edit_value(code), safe='')}/photos/{int(photo_id)}/thumbnail"


def _api_photo_cache_key(code: str, photo_id: int, checksum: str) -> str:
    safe_code = re.sub(r"[^0-9A-Za-z._-]+", "_", _normalize_edit_value(code))[:96] or "item"
    short_checksum = checksum.removeprefix("sha256:")[:16] or "unknown"
    return f"{safe_code}_{int(photo_id)}_{short_checksum}.jpg"


def _api_photo_source_size(photo: sqlite3.Row | dict) -> int:
    source = dict(photo)
    try:
        size_bytes = int(source.get("size_bytes") or 0)
    except Exception:
        size_bytes = 0
    if size_bytes > 0:
        return size_bytes
    file_name = _normalize_edit_value(source.get("file_name"))
    if not file_name:
        return 0
    try:
        return int((MEDIA_DIR / file_name).stat().st_size)
    except OSError:
        return 0


def _api_photo_source_available(conn: sqlite3.Connection, photo: sqlite3.Row | dict) -> bool:
    source = dict(photo)
    if _api_photo_source_size(source) > 0:
        return True
    file_name = _normalize_edit_value(source.get("file_name"))
    if file_name:
        try:
            if (MEDIA_DIR / file_name).is_file() and (MEDIA_DIR / file_name).stat().st_size > 0:
                return True
        except OSError:
            pass
    photo_id = int(source.get("id") or 0)
    if photo_id <= 0:
        return False
    row = conn.execute("SELECT CASE WHEN data IS NULL THEN 0 ELSE 1 END AS has_blob FROM photos WHERE id = ?", (photo_id,)).fetchone()
    return bool(row and row["has_blob"])


def _api_photo_cache_checksum(photo: sqlite3.Row | dict) -> str:
    source = dict(photo)
    checksum_payload = {
        "id": int(source.get("id") or 0),
        "file_name": _normalize_edit_value(source.get("file_name")),
        "original_name": _normalize_edit_value(source.get("original_name")),
        "mime_type": _normalize_edit_value(source.get("mime_type")),
        "source_tag": _normalize_edit_value(source.get("source_tag")),
        "created_at": _normalize_edit_value(source.get("created_at")),
        "size_bytes": _api_photo_source_size(source),
    }
    return program_api_sync.checksum(checksum_payload)


def _api_primary_photo_row(conn: sqlite3.Connection, item_id: int, include_data: bool = False) -> sqlite3.Row | None:
    data_field = "data" if include_data else "NULL AS data"
    has_blob_field = "CASE WHEN data IS NULL THEN 0 ELSE 1 END AS has_blob" if include_data else "1 AS has_blob"
    size_field = "CASE WHEN data IS NULL THEN 0 ELSE length(data) END AS size_bytes" if include_data else "0 AS size_bytes"
    return conn.execute(
        f"""
        SELECT
            id, item_id, file_name, original_name, sort_order, view_tag, source_tag, created_at, mime_type,
            {has_blob_field},
            {size_field},
            {data_field}
        FROM photos
        WHERE item_id = ?
        ORDER BY sort_order ASC, id ASC
        LIMIT 1
        """,
        (int(item_id),),
    ).fetchone()


def _api_primary_available_photo_row(conn: sqlite3.Connection, item_id: int) -> sqlite3.Row | None:
    for photo in photo_ordering.item_photo_rows(conn, int(item_id)):
        if _api_photo_source_available(conn, photo):
            return photo
    return None


def _api_item_primary_photo_metadata(
    conn: sqlite3.Connection,
    item_id: int,
    code: str,
) -> dict | None:
    photo = _api_primary_available_photo_row(conn, int(item_id))
    if photo is None:
        return None
    source = dict(photo)
    photo_id = int(source["id"])
    checksum = _api_photo_cache_checksum(source)
    return {
        "item_id": int(item_id),
        "code": _normalize_edit_value(code),
        "photo_id": photo_id,
        "checksum": checksum,
        "mime_type": _normalize_edit_value(source.get("mime_type")) or "application/octet-stream",
        "size_bytes": _api_photo_source_size(source),
        "original_name": _normalize_edit_value(source.get("original_name")),
        "thumbnail_mime_type": "image/jpeg",
        "thumbnail_max_edge": PROGRAM_API_PRIMARY_THUMBNAIL_MAX_EDGE,
        "thumbnail_path": _api_item_photo_thumbnail_path(code, photo_id),
        "content_path": _api_item_photo_content_path(code, photo_id),
        "cache_filename": _api_photo_cache_key(code, photo_id, checksum),
    }


def _api_item_photo_payload(photo: sqlite3.Row | dict, code: str) -> dict:
    source = dict(photo)
    photo_id = int(source["id"])
    checksum = _api_photo_cache_checksum(source)
    return {
        "id": photo_id,
        "item_id": int(source.get("item_id") or 0),
        "file_name": _normalize_edit_value(source.get("file_name")),
        "original_name": _normalize_edit_value(source.get("original_name")),
        "sort_order": int(source.get("sort_order") or 0),
        "view_tag": _normalize_edit_value(source.get("view_tag")),
        "source_tag": _normalize_edit_value(source.get("source_tag")),
        "created_at": _normalize_edit_value(source.get("created_at")),
        "mime_type": _normalize_edit_value(source.get("mime_type")),
        "has_blob": bool(source.get("has_blob")),
        "size_bytes": _api_photo_source_size(source),
        "checksum": checksum,
        "content_path": _api_item_photo_content_path(code, photo_id),
        "thumbnail_path": _api_item_photo_thumbnail_path(code, photo_id),
        "thumbnail_mime_type": "image/jpeg",
        "thumbnail_max_edge": PROGRAM_API_PRIMARY_THUMBNAIL_MAX_EDGE,
        "cache_filename": _api_photo_cache_key(code, photo_id, checksum),
    }


def _api_item_detail_payload(conn: sqlite3.Connection, row: sqlite3.Row | dict) -> dict:
    payload = _api_item_payload(row)
    code = _normalize_edit_value(payload.get("code"))
    photos = [
        _api_item_photo_payload(photo, code)
        for photo in photo_ordering.item_photo_rows(conn, int(payload["id"]))
    ]
    payload["photos"] = photos
    payload["photo_count"] = len(photos)
    payload["primary_photo"] = photos[0] if photos else None
    payload["primary_photo_content_path"] = photos[0]["content_path"] if photos else ""
    payload["primary_photo_thumbnail_path"] = photos[0]["thumbnail_path"] if photos else ""
    return payload


def _item_row_authorized(item: sqlite3.Row | dict | None, username: str) -> bool:
    if item is None:
        return False
    return _owner_value_allowed(item["owner"] if isinstance(item, sqlite3.Row) else item.get("owner"), username)


def _item_row_viewable(item: sqlite3.Row | dict | None, username: str) -> bool:
    if item is None:
        return False
    return _owner_value_viewable(item["owner"] if isinstance(item, sqlite3.Row) else item.get("owner"), username)


def _item_owner_scope_where(username: str) -> tuple[str, list[str]]:
    return _owner_read_sql("items.owner", username)


def _outfit_row_authorized(outfit: sqlite3.Row | dict | None, username: str) -> bool:
    if outfit is None:
        return False
    return _owner_value_allowed(outfit["owner"] if isinstance(outfit, sqlite3.Row) else outfit.get("owner"), username)


def _outfit_row_viewable(outfit: sqlite3.Row | dict | None, username: str) -> bool:
    if outfit is None:
        return False
    return _owner_value_viewable(outfit["owner"] if isinstance(outfit, sqlite3.Row) else outfit.get("owner"), username)


def _featured_look_row_authorized(look: sqlite3.Row | dict | None, username: str) -> bool:
    if look is None:
        return False
    return _owner_value_allowed(look["owner"] if isinstance(look, sqlite3.Row) else look.get("owner"), username)


def _featured_look_row_viewable(look: sqlite3.Row | dict | None, username: str) -> bool:
    if look is None:
        return False
    return _owner_value_viewable(look["owner"] if isinstance(look, sqlite3.Row) else look.get("owner"), username)


def _photo_item_authorized(conn: sqlite3.Connection, photo_id: int, username: str) -> bool:
    row = conn.execute(
        """
        SELECT items.owner
        FROM photos
        JOIN items ON items.id = photos.item_id
        WHERE photos.id = ?
        """,
        (photo_id,),
    ).fetchone()
    return _item_row_authorized(row, username)


def _photo_item_viewable(conn: sqlite3.Connection, photo_id: int, username: str) -> bool:
    row = conn.execute(
        """
        SELECT items.owner
        FROM photos
        JOIN items ON items.id = photos.item_id
        WHERE photos.id = ?
        """,
        (photo_id,),
    ).fetchone()
    return _item_row_viewable(row, username)


def _entity_photos(
    conn: sqlite3.Connection,
    table: str,
    foreign_key: str,
    entity_id: int,
    content_prefix: str,
    delete_prefix: str,
) -> list[dict]:
    rows = conn.execute(
        f"""
        SELECT
            id, file_name, original_name, sort_order, source_tag, created_at, mime_type,
            1 AS has_blob
        FROM {table}
        WHERE {foreign_key} = ?
        ORDER BY sort_order ASC, id ASC
        """,
        (entity_id,),
    ).fetchall()
    return [
        {
            **dict(row),
            "content_path": f"{content_prefix}/{row['id']}/content",
            "delete_path": f"{delete_prefix}/{row['id']}",
        }
        for row in rows
    ]


def _backfill_outfit_photo_gps(conn: sqlite3.Connection, photo_id: int, raw: bytes | None) -> tuple[float | None, float | None]:
    latitude, longitude = _extract_image_gps_coordinates(raw)
    conn.execute(
        "UPDATE outfit_photos SET gps_lat = ?, gps_lng = ?, gps_checked = 1 WHERE id = ?",
        (latitude, longitude, photo_id),
    )
    return latitude, longitude


def _outfit_photos(conn: sqlite3.Connection, outfit_id: int) -> list[dict]:
    rows = conn.execute(
        """
        SELECT
            id, file_name, original_name, sort_order, source_tag, created_at, mime_type,
            gps_lat, gps_lng, gps_checked,
            1 AS has_blob
        FROM outfit_photos
        WHERE outfit_id = ?
        ORDER BY sort_order ASC, id ASC
        """,
        (outfit_id,),
    ).fetchall()
    unchecked_ids = [int(row["id"]) for row in rows if not int(row["gps_checked"] or 0)]
    unchecked_payloads: dict[int, bytes | None] = {}
    if unchecked_ids:
        placeholders = ",".join("?" for _ in unchecked_ids)
        payload_rows = conn.execute(
            f"SELECT id, data FROM outfit_photos WHERE id IN ({placeholders})",
            unchecked_ids,
        ).fetchall()
        unchecked_payloads = {int(row["id"]): row["data"] for row in payload_rows}
    photos: list[dict] = []
    updated = False
    for row in rows:
        photo = dict(row)
        latitude = photo.get("gps_lat")
        longitude = photo.get("gps_lng")
        if not int(photo.get("gps_checked") or 0):
            latitude, longitude = _backfill_outfit_photo_gps(
                conn,
                int(photo["id"]),
                unchecked_payloads.get(int(photo["id"])),
            )
            updated = True
        photo["gps_lat"] = latitude
        photo["gps_lng"] = longitude
        photo["location_display"] = _format_photo_location(latitude, longitude)
        photo.pop("gps_checked", None)
        photo["content_path"] = f"/api/outfit-photos/{row['id']}/content"
        photo["delete_path"] = f"/api/outfit-photos/{row['id']}"
        photos.append(photo)
    if updated:
        conn.commit()
    return photos


def _cached_outfit_photo_thumbnail(
    conn: sqlite3.Connection,
    photo_id: int,
    raw: bytes | None,
    mime_type: str,
    cached_raw: bytes | None = None,
    cached_mime: str | None = None,
) -> tuple[bytes, str] | None:
    if cached_raw:
        return cached_raw, cached_mime or "application/octet-stream"
    thumbnail = _thumbnail_bytes_from_payload(raw, mime_type)
    if thumbnail is None:
        return None
    thumb_raw, thumb_mime = thumbnail
    conn.execute(
        "UPDATE outfit_photos SET thumb_data = ?, thumb_mime_type = ? WHERE id = ?",
        (thumb_raw, thumb_mime, photo_id),
    )
    conn.commit()
    return thumb_raw, thumb_mime


def _thumbnail_bytes_from_payload(raw: bytes | None, mime_type: str, max_edge: int = PHOTO_THUMBNAIL_MAX_EDGE) -> tuple[bytes, str] | None:
    if not raw:
        return None
    try:
        with Image.open(io.BytesIO(raw)) as image:
            image = ImageOps.exif_transpose(image)
            if image.mode not in ("RGB", "RGBA"):
                image = image.convert("RGB")
            image.thumbnail((max_edge, max_edge), Image.Resampling.LANCZOS)
            output = io.BytesIO()
            target_format = "JPEG"
            target_mime = "image/jpeg"
            if (mime_type or "").lower() == "image/png" or image.mode == "RGBA":
                target_format = "PNG"
                target_mime = "image/png"
            elif (mime_type or "").lower() == "image/webp":
                target_format = "WEBP"
                target_mime = "image/webp"
            save_kwargs = {"format": target_format}
            if target_format == "JPEG":
                save_kwargs.update({"quality": 82, "optimize": True})
            elif target_format == "WEBP":
                save_kwargs.update({"quality": 82, "method": 6})
            image.save(output, **save_kwargs)
            return output.getvalue(), target_mime
    except (UnidentifiedImageError, OSError, ValueError):
        return None


def _safe_jpeg_thumbnail_bytes(
    raw: bytes | None,
    max_edge: int = PROGRAM_API_PRIMARY_THUMBNAIL_MAX_EDGE,
) -> bytes | None:
    if not raw:
        return None
    previous_truncated_setting = ImageFile.LOAD_TRUNCATED_IMAGES
    try:
        ImageFile.LOAD_TRUNCATED_IMAGES = True
        with Image.open(io.BytesIO(raw)) as image:
            image = ImageOps.exif_transpose(image)
            if image.mode in ("RGBA", "LA"):
                background = Image.new("RGB", image.size, (255, 255, 255))
                alpha = image.getchannel("A")
                background.paste(image.convert("RGBA"), mask=alpha)
                image = background
            elif image.mode != "RGB":
                image = image.convert("RGB")
            image.thumbnail((max_edge, max_edge), Image.Resampling.LANCZOS)
            output = io.BytesIO()
            image.save(output, format="JPEG", quality=84, optimize=True)
            return output.getvalue()
    except (UnidentifiedImageError, OSError, ValueError):
        return None
    finally:
        ImageFile.LOAD_TRUNCATED_IMAGES = previous_truncated_setting


def _exif_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bytes):
        for encoding in ("utf-8", "ascii", "latin1"):
            try:
                return value.decode(encoding).strip()
            except UnicodeDecodeError:
                continue
        return ""
    return str(value).strip()


def _exif_ratio_to_float(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, tuple) and len(value) == 2:
        numerator, denominator = value
        try:
            denominator_float = float(denominator)
            if denominator_float == 0:
                return None
            return float(numerator) / denominator_float
        except (TypeError, ValueError, ZeroDivisionError):
            return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _gps_coordinate_to_decimal(values: object, ref: object) -> float | None:
    if not isinstance(values, (list, tuple)) or len(values) < 3:
        return None
    parts = [_exif_ratio_to_float(values[index]) for index in range(3)]
    if any(part is None for part in parts):
        return None
    degrees, minutes, seconds = parts
    coordinate = degrees + (minutes / 60.0) + (seconds / 3600.0)
    ref_text = _exif_text(ref).upper()
    if ref_text in {"S", "W"}:
        coordinate *= -1
    return round(coordinate, 6)


def _extract_image_gps_coordinates(raw: bytes | None) -> tuple[float | None, float | None]:
    if not raw:
        return None, None
    try:
        with Image.open(io.BytesIO(raw)) as image:
            exif = image.getexif()
            if not exif:
                return None, None
            gps_info = exif.get(EXIF_GPS_INFO_TAG)
            if not isinstance(gps_info, dict) and hasattr(exif, "get_ifd"):
                try:
                    gps_info = exif.get_ifd(EXIF_GPS_INFO_TAG)
                except Exception:
                    gps_info = None
            if hasattr(gps_info, "items"):
                gps_info = dict(gps_info.items())
            if not isinstance(gps_info, dict):
                return None, None
            latitude = _gps_coordinate_to_decimal(gps_info.get(2), gps_info.get(1))
            longitude = _gps_coordinate_to_decimal(gps_info.get(4), gps_info.get(3))
            return latitude, longitude
    except (UnidentifiedImageError, OSError, ValueError, TypeError):
        return None, None


def _format_photo_location(latitude: object, longitude: object) -> str:
    lat_value = _exif_ratio_to_float(latitude)
    lng_value = _exif_ratio_to_float(longitude)
    if lat_value is None or lng_value is None:
        return ""
    return f"{lat_value:.6f}, {lng_value:.6f}"


def _send_photo_payload(handler: "WardrobeHandler", raw: bytes | None, mime_type: str, query: dict[str, list[str]]) -> bool:
    if raw is None:
        return False
    wants_thumb = (query.get("thumb") or [""])[0] in {"1", "true", "yes"}
    if wants_thumb:
        thumbnail = _thumbnail_bytes_from_payload(raw, mime_type)
        if thumbnail is not None:
            thumb_raw, thumb_mime = thumbnail
            handler._send_bytes(thumb_raw, content_type=thumb_mime)
            return True
    handler._send_bytes(raw, content_type=mime_type or "application/octet-stream")
    return True


def _item_photo_raw_bytes(photo: sqlite3.Row | dict) -> tuple[bytes | None, str]:
    source = dict(photo)
    if source.get("data") is not None:
        return source.get("data"), _normalize_edit_value(source.get("mime_type")) or "application/octet-stream"
    file_name = _normalize_edit_value(source.get("file_name"))
    if not file_name:
        return None, _normalize_edit_value(source.get("mime_type")) or "application/octet-stream"
    target = MEDIA_DIR / file_name
    if not target.exists():
        return None, _normalize_edit_value(source.get("mime_type")) or "application/octet-stream"
    try:
        return (
            target.read_bytes(),
            mimetypes.guess_type(target.name)[0]
            or _normalize_edit_value(source.get("mime_type"))
            or "application/octet-stream",
        )
    except OSError:
        return None, _normalize_edit_value(source.get("mime_type")) or "application/octet-stream"


def _send_item_photo_row_payload(
    handler: "WardrobeHandler",
    photo: sqlite3.Row | dict,
    query: dict[str, list[str]],
) -> bool:
    raw, mime_type = _item_photo_raw_bytes(photo)
    if raw is not None:
        return _send_photo_payload(handler, raw, mime_type, query)
    return False


def _send_item_photo_thumbnail_payload(
    handler: "WardrobeHandler",
    photo: sqlite3.Row | dict,
    code: str,
    query: dict[str, list[str]],
) -> bool:
    checksum = _api_photo_cache_checksum(photo)
    if _api_etag_matches(handler.headers.get("If-None-Match", ""), checksum):
        handler._send_empty(
            HTTPStatus.NOT_MODIFIED,
            extra_headers=[("ETag", _api_quote_etag(checksum))],
        )
        return True
    raw, _mime_type = _item_photo_raw_bytes(photo)
    thumbnail = _safe_jpeg_thumbnail_bytes(raw)
    if thumbnail is None:
        return False
    filename = _api_photo_cache_key(code, int(dict(photo)["id"]), checksum)
    handler._send_bytes(
        thumbnail,
        content_type="image/jpeg",
        extra_headers=[
            ("ETag", _api_quote_etag(checksum)),
            ("Content-Disposition", f'inline; filename="{filename}"'),
            ("X-Content-Filename", filename),
            ("X-Thumbnail-Max-Edge", str(PROGRAM_API_PRIMARY_THUMBNAIL_MAX_EDGE)),
        ],
    )
    return True


def _serialize_outfit(conn: sqlite3.Connection, outfit_row: sqlite3.Row, username: str = "") -> dict:
    outfit = dict(outfit_row)
    daily_update, update_item_map = _daily_update_item_map(
        conn,
        _normalize_edit_value(outfit.get("wear_date")),
        _normalize_edit_value(outfit.get("owner")),
    )
    items = conn.execute(
        """
        SELECT outfit_items.role, items.*
        FROM outfit_items
        JOIN items ON items.id = outfit_items.item_id
        WHERE outfit_items.outfit_id = ?
        ORDER BY outfit_items.role, items.section
        """,
        (outfit["id"],),
    ).fetchall()
    visible_items = []
    for item in items:
        if username and not _item_row_viewable(item, username):
            continue
        item_dict = dict(item)
        update_row = update_item_map.get(int(item["id"]))
        if update_row is not None:
            update_dict = dict(update_row)
            item_dict["has_base_layer"] = bool(update_dict.get("has_base_layer"))
            item_dict["wear_delta"] = update_dict.get("wear_delta")
        visible_items.append(item_dict)
    outfit["items"] = visible_items
    outfit["wearcount_linked"] = daily_update is not None
    outfit["daily_update_id"] = int(daily_update["id"]) if daily_update is not None else None
    existing_look = _existing_featured_look_for_outfit(conn, outfit)
    outfit["featured_look_exists"] = existing_look is not None
    outfit["featured_look_id"] = _normalize_edit_value(existing_look["look_id"]) if existing_look is not None else ""
    outfit["photos"] = _outfit_photos(conn, int(outfit["id"]))
    return outfit


def _serialize_outfit_summary(outfit_row: sqlite3.Row) -> dict:
    outfit = dict(outfit_row)
    try:
        outfit["photo_count"] = int(outfit.get("photo_count") or 0)
    except Exception:
        outfit["photo_count"] = 0
    return outfit


def _serialize_featured_look(
    conn: sqlite3.Connection,
    look_row: sqlite3.Row,
    wear_counts: dict[int, int] | None = None,
    username: str = "",
) -> dict:
    look = dict(look_row)
    computed_relax_center = _featured_look_relax_center(conn, int(look["id"]))
    computed_relax_span = _featured_look_relax_span(conn, int(look["id"]))
    stored_relax_center = look.get("relax_center")
    stored_relax_span = look.get("relax_span")
    if computed_relax_center != stored_relax_center or computed_relax_span != stored_relax_span:
        _sync_featured_look_relax_metrics(conn, int(look["id"]))
    look["relax_center"] = computed_relax_center
    look["relax_span"] = computed_relax_span
    look.pop("material_line", None)
    items = conn.execute(
        """
        SELECT
            featured_look_items.slot,
            featured_look_items.source_code,
            featured_look_items.source_section,
            featured_look_items.display_order,
            items.*
        FROM featured_look_items
        LEFT JOIN items ON items.id = featured_look_items.item_id
        WHERE featured_look_items.featured_look_id = ?
        ORDER BY featured_look_items.display_order ASC, featured_look_items.id ASC
        """,
        (look["id"],),
    ).fetchall()
    look["items"] = [
        dict(item)
        for item in items
        if not username or item["id"] is None or _item_row_viewable(item, username)
    ]
    look["photos"] = _entity_photos(
        conn,
        "featured_look_photos",
        "featured_look_id",
        look["id"],
        "/api/featured-look-photos",
        "/api/featured-look-photos",
    )
    if wear_counts is None:
        owner = _normalize_edit_value(look.get("owner"))
        wear_counts = _featured_look_wear_counts(conn, {owner} if owner else None)
    look["wear_count"] = int(wear_counts.get(int(look["id"]), 0))
    return look


def _item_related_outfits(
    conn: sqlite3.Connection,
    item_id: int,
    username: str,
    summary_only: bool = False,
) -> list[dict]:
    item = conn.execute("SELECT * FROM items WHERE id = ?", (item_id,)).fetchone()
    if item is None:
        raise KeyError("item_not_found")
    if username and not _item_row_viewable(item, username):
        raise PermissionError("forbidden")
    outfit_where, outfit_params = _owner_read_sql("outfits.owner", username)
    rows = conn.execute(
        f"""
        SELECT DISTINCT
            outfits.*,
            (SELECT COUNT(*) FROM outfit_photos WHERE outfit_photos.outfit_id = outfits.id) AS photo_count
        FROM outfits
        JOIN outfit_items ON outfit_items.outfit_id = outfits.id
        WHERE outfit_items.item_id = ?
          AND {outfit_where}
        ORDER BY outfits.wear_date DESC
        """,
        [item_id, *outfit_params],
    ).fetchall()
    if summary_only:
        return [_serialize_outfit_summary(row) for row in rows]
    return [_serialize_outfit(conn, row, username) for row in rows]


def _item_related_featured_looks(conn: sqlite3.Connection, item_id: int, username: str) -> list[dict]:
    item = conn.execute("SELECT * FROM items WHERE id = ?", (item_id,)).fetchone()
    if item is None:
        raise KeyError("item_not_found")
    if username and not _item_row_viewable(item, username):
        raise PermissionError("forbidden")
    where_sql, where_params = _owner_read_sql("featured_looks.owner", username)
    rows = conn.execute(
        f"""
        SELECT DISTINCT featured_looks.*
        FROM featured_looks
        JOIN featured_look_items ON featured_look_items.featured_look_id = featured_looks.id
        WHERE featured_look_items.item_id = ?
          AND COALESCE(featured_looks.status, '') <> 'Archived'
          AND {where_sql}
        ORDER BY
            COALESCE(featured_looks.created_at, '') DESC,
            featured_looks.id DESC
        """,
        [item_id, *where_params],
    ).fetchall()
    visible_rows = [row for row in rows if _featured_look_row_viewable(row, username)]
    wear_counts = _featured_look_wear_counts(conn, {
        _normalize_edit_value(row["owner"])
        for row in visible_rows
        if _normalize_edit_value(row["owner"])
    })
    return [_serialize_featured_look(conn, row, wear_counts, username) for row in visible_rows]


def _sync_outfit_items(conn: sqlite3.Connection, outfit_id: int, item_ids: list[int]) -> tuple[float | None, str]:
    wear_date = conn.execute("SELECT wear_date FROM outfits WHERE id = ?", (outfit_id,)).fetchone()["wear_date"]
    conn.execute("DELETE FROM outfit_items WHERE outfit_id = ?", (outfit_id,))
    avg_relax_values = []
    temp_ranges = []
    for item_id in item_ids:
        item = conn.execute("SELECT * FROM items WHERE id = ?", (item_id,)).fetchone()
        if item is None:
            continue
        conn.execute(
            "INSERT OR IGNORE INTO outfit_items (outfit_id, item_id, role) VALUES (?, ?, ?)",
            (outfit_id, item_id, item["layer_role"]),
        )
        conn.execute("UPDATE items SET last_worn_on = ? WHERE id = ?", (wear_date, item_id))
        if item["relax_index"] is not None and item["layer_role"] != "Watch":
            avg_relax_values.append(float(item["relax_index"]))
        if item["temp_min"] is not None or item["temp_max"] is not None:
            temp_ranges.append((item["temp_min"], item["temp_max"]))
    avg_relax = round(sum(avg_relax_values) / len(avg_relax_values), 2) if avg_relax_values else None
    avg_temp_label = ""
    if temp_ranges:
        mins = [value for value, _ in temp_ranges if value is not None]
        maxs = [value for _, value in temp_ranges if value is not None]
        if mins and maxs:
            avg_temp_label = f"{sum(mins) / len(mins):.1f}–{sum(maxs) / len(maxs):.1f}"
    return avg_relax, avg_temp_label


def _daily_update_item_rows(conn: sqlite3.Connection, daily_update_id: int) -> list[sqlite3.Row]:
    return conn.execute(
        "SELECT * FROM wearcount_daily_update_items WHERE daily_update_id = ? ORDER BY id ASC",
        (daily_update_id,),
    ).fetchall()


def _daily_update_item_map(conn: sqlite3.Connection, wear_date: str, owner: str = "") -> tuple[sqlite3.Row | None, dict[int, dict]]:
    daily_update = _find_daily_update(conn, wear_date, owner)
    if daily_update is None:
        return None, {}
    item_map: dict[int, dict] = {}
    for row in _daily_update_item_rows(conn, int(daily_update["id"])):
        item_map[int(row["item_id"])] = dict(row)
    return daily_update, item_map


def _featured_look_slot_from_role(role: str) -> str | None:
    role_value = _canonical_outfit_role(role)
    slot = role_value.lower() if role_value else ""
    return slot if slot in CANONICAL_FEATURED_LOOK_SLOTS else None


def _featured_look_anchor_sort_rank(role: str) -> int:
    slot = _featured_look_slot_from_role(role)
    preferred_order = ("outer", "middle", "inner", "bottom", "footwear", "watch")
    return preferred_order.index(slot) if slot in preferred_order else 999


def _featured_look_signature(
    anchor_item: dict | sqlite3.Row | None,
    slot_entries: dict[str, dict | sqlite3.Row],
) -> tuple[tuple[str, int], ...]:
    signature: list[tuple[str, int]] = []
    if anchor_item is not None:
        anchor_id = int(anchor_item["id"] or 0)
        if anchor_id > 0:
            signature.append(("anchor", anchor_id))
    for slot in CANONICAL_FEATURED_LOOK_SLOTS:
        item = slot_entries.get(slot)
        if item is None:
            continue
        item_id = int(item["id"] or 0)
        if item_id > 0:
            signature.append((slot, item_id))
    signature.sort()
    return tuple(signature)


def _featured_look_signature_for_outfit(conn: sqlite3.Connection, outfit_id: int) -> tuple[tuple[str, int], ...]:
    item_rows = conn.execute(
        """
        SELECT items.id, items.layer_role, outfit_items.role
        FROM outfit_items
        JOIN items ON items.id = outfit_items.item_id
        WHERE outfit_items.outfit_id = ?
        ORDER BY outfit_items.id ASC, items.id ASC
        """,
        (outfit_id,),
    ).fetchall()
    slot_entries: dict[str, sqlite3.Row] = {}
    anchor_item: sqlite3.Row | None = None
    duplicate_slots = set()
    sorted_items = sorted(
        list(item_rows),
        key=lambda item: _featured_look_anchor_sort_rank(item["role"] or item["layer_role"]),
    )
    for item in sorted_items:
        role = _normalize_edit_value(item["role"] or item["layer_role"])
        slot = _featured_look_slot_from_role(role)
        if slot is None:
            continue
        if anchor_item is None:
            anchor_item = item
        if slot in slot_entries:
            duplicate_slots.add(slot)
            continue
        slot_entries[slot] = item
    if duplicate_slots:
        return tuple()
    return _featured_look_signature(anchor_item, slot_entries)


def _featured_look_signature_for_look(conn: sqlite3.Connection, featured_look_id: int) -> tuple[tuple[str, int], ...]:
    item_rows = conn.execute(
        """
        SELECT slot, item_id
        FROM featured_look_items
        WHERE featured_look_id = ?
        ORDER BY slot ASC, item_id ASC
        """,
        (featured_look_id,),
    ).fetchall()
    return tuple(
        sorted(
            (
                _normalize_edit_value(item_row["slot"]).lower(),
                int(item_row["item_id"] or 0),
            )
            for item_row in item_rows
            if int(item_row["item_id"] or 0) > 0
        )
    )


def _featured_look_wear_counts(conn: sqlite3.Connection, owners: set[str] | None = None) -> dict[int, int]:
    params: list[object] = []
    where_sql = "COALESCE(status, '') <> 'Archived'"
    normalized_owners = sorted({
        _normalize_edit_value(owner)
        for owner in (owners or set())
        if _normalize_edit_value(owner)
    })
    if normalized_owners:
        where_sql += " AND (" + " OR ".join(["COALESCE(owner, '') = ?" for _ in normalized_owners]) + ")"
        params.extend(normalized_owners)
    look_rows = conn.execute(
        f"""
        SELECT id, owner
        FROM featured_looks
        WHERE {where_sql}
        ORDER BY id ASC
        """,
        params,
    ).fetchall()
    counts: dict[int, int] = {int(row["id"]): 0 for row in look_rows}
    owner_signatures: dict[str, dict[tuple[tuple[str, int], ...], list[int]]] = {}
    for row in look_rows:
        look_id = int(row["id"] or 0)
        owner = _normalize_edit_value(row["owner"])
        signature = _featured_look_signature_for_look(conn, look_id)
        if look_id <= 0 or not owner or not signature:
            continue
        owner_signatures.setdefault(owner, {}).setdefault(signature, []).append(look_id)
    for owner, signature_map in owner_signatures.items():
        outfit_rows = conn.execute(
            """
            SELECT id
            FROM outfits
            WHERE COALESCE(owner, '') = ?
            ORDER BY wear_date ASC, id ASC
            """,
            (owner,),
        ).fetchall()
        for outfit_row in outfit_rows:
            signature = _featured_look_signature_for_outfit(conn, int(outfit_row["id"] or 0))
            if not signature:
                continue
            for look_id in signature_map.get(signature, []):
                counts[look_id] = counts.get(look_id, 0) + 1
    return counts


def _find_duplicate_featured_look(
    conn: sqlite3.Connection,
    owner: str,
    signature: tuple[tuple[str, int], ...],
) -> sqlite3.Row | None:
    if not signature:
        return None
    rows = conn.execute(
        """
        SELECT *
        FROM featured_looks
        WHERE COALESCE(status, '') <> 'Archived'
          AND COALESCE(owner, '') = ?
        ORDER BY id ASC
        """,
        (owner,),
    ).fetchall()
    for row in rows:
        item_rows = conn.execute(
            """
            SELECT slot, item_id
            FROM featured_look_items
            WHERE featured_look_id = ?
            ORDER BY slot ASC, item_id ASC
            """,
            (int(row["id"]),),
        ).fetchall()
        existing_signature = tuple(
            sorted(
                (
                    _normalize_edit_value(item_row["slot"]).lower(),
                    int(item_row["item_id"] or 0),
                )
                for item_row in item_rows
                if int(item_row["item_id"] or 0) > 0
            )
        )
        if existing_signature == signature:
            return row
    return None


def _existing_featured_look_for_outfit(conn: sqlite3.Connection, outfit_data: dict) -> sqlite3.Row | None:
    owner = _normalize_edit_value(outfit_data.get("owner"))
    if not owner:
        return None
    slot_entries: dict[str, dict] = {}
    anchor_item: dict | None = None
    duplicate_slots = set()
    sorted_items = sorted(
        list(outfit_data.get("items", [])),
        key=lambda item: _featured_look_anchor_sort_rank(item.get("role") or item.get("layer_role")),
    )
    for item in sorted_items:
        role = _normalize_edit_value(item.get("role") or item.get("layer_role"))
        slot = _featured_look_slot_from_role(role)
        if slot is None:
            continue
        if anchor_item is None:
            anchor_item = item
        if slot in slot_entries:
            duplicate_slots.add(slot)
            continue
        slot_entries[slot] = item
    if duplicate_slots:
        return None
    signature = _featured_look_signature(anchor_item, slot_entries)
    return _find_duplicate_featured_look(conn, owner, signature)


def _next_featured_look_id(conn: sqlite3.Connection, wear_date: str) -> str:
    compact_date = re.sub(r"[^0-9]", "", _normalize_edit_value(wear_date)) or datetime.now().strftime("%Y%m%d")
    prefix = f"AUTO-{compact_date}"
    existing = {
        _normalize_edit_value(row["look_id"])
        for row in conn.execute("SELECT look_id FROM featured_looks WHERE look_id LIKE ?", (f"{prefix}%",)).fetchall()
    }
    if prefix not in existing:
        return prefix
    index = 2
    while f"{prefix}-{index}" in existing:
        index += 1
    return f"{prefix}-{index}"


def _subtract_daily_update_effects(conn: sqlite3.Connection, daily_update_id: int) -> list[int]:
    affected_item_ids: list[int] = []
    for row in _daily_update_item_rows(conn, daily_update_id):
        item_id = int(row["item_id"])
        affected_item_ids.append(item_id)
        conn.execute(
            """
            UPDATE items
            SET wear_maintenance = MAX(COALESCE(wear_maintenance, 0) - ?, 0),
                wear_total = MAX(COALESCE(wear_total, 0) - ?, 0),
                wear_year = MAX(COALESCE(wear_year, 0) - ?, 0),
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (
                float(row["wear_delta"] or 0),
                int(row["total_delta"] or 0),
                int(row["year_delta"] or 0),
                item_id,
            ),
        )
    conn.execute("DELETE FROM wearcount_daily_update_items WHERE daily_update_id = ?", (daily_update_id,))
    return affected_item_ids


def _resolve_outfit_item_entries(
    conn: sqlite3.Connection,
    items_payload: list[dict],
    username: str,
    wear_mode: str,
    wear_date: str,
    allow_empty: bool = False,
) -> list[dict]:
    resolved_entries: list[dict] = []
    for raw_entry in items_payload:
        try:
            item_id = int(raw_entry.get("item_id"))
        except Exception:
            continue
        item = conn.execute("SELECT * FROM items WHERE id = ?", (item_id,)).fetchone()
        if item is None:
            raise KeyError("item_not_found")
        if username and not _item_row_authorized(item, username):
            raise PermissionError("forbidden")
        raw_role = raw_entry.get("role") if _normalize_edit_value(raw_entry.get("role")) else item["layer_role"]
        role = _validate_outfit_role(raw_role, item["code"])
        has_base_layer = bool(raw_entry.get("has_base_layer"))
        resolved_entries.append(
            {
                "item_id": item_id,
                "item": item,
                "role": role,
                "has_base_layer": has_base_layer,
                "wear_delta": _wear_delta_for_item(role, has_base_layer, wear_mode),
                "total_delta": 1,
                "year_delta": 1 if wear_date.startswith("2026") else 0,
            }
        )
    if not resolved_entries and not allow_empty:
        raise ValueError("outfit_missing_items")
    return resolved_entries


def _coerce_api_bool(value: object, default: bool = False) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    normalized = _normalize_edit_value(value).lower()
    if normalized in {"1", "true", "yes", "y", "on", "是"}:
        return True
    if normalized in {"0", "false", "no", "n", "off", "否"}:
        return False
    return default


def _api_item_payload(row: sqlite3.Row | dict) -> dict:
    source = dict(row)
    code = _normalize_edit_value(source.get("code"))
    brand = _normalize_edit_value(source.get("brand"))
    section = _normalize_edit_value(source.get("section"))
    official_desc = _normalize_edit_value(source.get("official_desc"))
    name = section or official_desc or code
    if brand and section and not section.lower().startswith(brand.lower()):
        display_name = f"{brand} {section}"
    else:
        display_name = section or " ".join(part for part in [brand, official_desc, code] if part)
    payload = {
        "id": source.get("id"),
        "code": code,
        "owner": _normalize_edit_value(source.get("owner")),
        "kind": "watch" if _canonical_layer_role(source.get("layer_role")) == "Watch" else "wardrobe",
        "brand": brand,
        "name": name,
        "display_name": display_name or name,
        "section": section,
        "loc": _normalize_edit_value(source.get("loc")),
        "status": _normalize_edit_value(source.get("status")) or "Active",
        "maintenance_state": int(source.get("maintenance_state") or 0),
        "layer_role": _canonical_layer_role(source.get("layer_role")) or _normalize_edit_value(source.get("layer_role")),
        "outer_type": _normalize_edit_value(source.get("outer_type")),
        "scene_tag": _normalize_edit_value(source.get("scene_tag")),
        "relax_index": source.get("relax_index"),
        "temp_min": source.get("temp_min"),
        "temp_max": source.get("temp_max"),
        "standalone_min": source.get("standalone_min"),
        "standalone_max": source.get("standalone_max"),
        "primary_color": _normalize_edit_value(source.get("primary_color")),
        "secondary_color": _normalize_edit_value(source.get("secondary_color")),
        "official_desc": official_desc,
        "material": _normalize_edit_value(source.get("material")),
        "care": _normalize_edit_value(source.get("care")),
        "notes": _normalize_edit_value(source.get("notes")),
        "series": _normalize_edit_value(source.get("series")),
        "official_color_code": _normalize_edit_value(source.get("official_color_code")),
        "price_original": normalize_price_text(source.get("price_original")),
        "price_original_currency": normalize_price_currency(source.get("price_original_currency"), source.get("price_original")),
        "price_cny": normalize_price_text(source.get("price_cny")),
        "acquired_at": _normalize_edit_value(source.get("acquired_at")),
        "wear_total": int(source.get("wear_total") or 0),
        "wear_year": int(source.get("wear_year") or 0),
        "wear_maintenance": source.get("wear_maintenance"),
        "wear_threshold": source.get("wear_threshold"),
        "maint_count": int(source.get("maint_count") or 0),
        "last_worn_on": _normalize_edit_value(source.get("last_worn_on")),
        "photo_count": int(source.get("photo_count") or 0),
        "updated_at": _normalize_edit_value(source.get("updated_at")),
    }
    payload["recommendation_eligible"] = _api_item_recommendation_eligible(payload)
    return payload


def _api_item_recommendation_eligible(item: dict) -> bool:
    status_value = _normalize_edit_value(item.get("status")).lower()
    if status_value and status_value != "active":
        return False
    try:
        maintenance_state = int(item.get("maintenance_state") or 0)
    except Exception:
        maintenance_state = 0
    return maintenance_state == 0


def _api_sync_item_payload(row: sqlite3.Row | dict) -> dict:
    payload = _api_item_payload(row)
    for field_name in (
        "wear_total",
        "wear_year",
        "wear_maintenance",
        "maint_count",
        "last_worn_on",
    ):
        payload.pop(field_name, None)
    return payload


def _api_sync_item_payload_with_primary_photo(
    conn: sqlite3.Connection,
    row: sqlite3.Row | dict,
) -> dict:
    payload = _api_sync_item_payload(row)
    payload["primary_photo"] = _api_item_primary_photo_metadata(
        conn,
        int(payload["id"]),
        _normalize_edit_value(payload.get("code")),
    )
    return payload


def _api_sync_primary_photo_payload(conn: sqlite3.Connection, row: sqlite3.Row | dict) -> dict | None:
    source = dict(row)
    code = _normalize_edit_value(source.get("code"))
    photo = _api_primary_photo_row(conn, int(source["id"]))
    if photo is None:
        return None
    metadata = _api_item_primary_photo_metadata(conn, int(source["id"]), code)
    if metadata is None:
        return None
    return {
        **metadata,
        "owner": _normalize_edit_value(source.get("owner")),
        "display_name": _api_item_payload(source).get("display_name"),
        "thumbnail_url": metadata["thumbnail_path"],
    }


def _api_sync_float(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = _normalize_edit_value(value)
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _api_sync_wear_count_payload(row: sqlite3.Row | dict) -> dict:
    source = dict(row)
    wear_maintenance = _api_sync_float(source.get("wear_maintenance"))
    wear_threshold = _api_sync_float(source.get("wear_threshold"))
    maintenance_due = False
    if wear_threshold is not None and wear_threshold > 0 and wear_maintenance is not None:
        maintenance_due = wear_maintenance >= wear_threshold
    return {
        "code": _normalize_edit_value(source.get("code")),
        "owner": _normalize_edit_value(source.get("owner")),
        "wear_total": int(source.get("wear_total") or 0),
        "wear_year": int(source.get("wear_year") or 0),
        "wear_maintenance": wear_maintenance,
        "wear_threshold": wear_threshold,
        "maint_count": int(source.get("maint_count") or 0),
        "last_worn_on": _normalize_edit_value(source.get("last_worn_on")),
        "maintenance_due": maintenance_due,
    }


def _api_sync_rules_payload(owner: str) -> dict:
    normalized_owner = _normalize_edit_value(owner)
    return {
        "owner": normalized_owner,
        "transport": "Hermes Agent uses bearer-token Program API; do not depend on ChatGPT Drive sync.",
        "sync": {
            "manifest": "GET /api/v1/sync/outfit-context/manifest",
            "resource": "GET /api/v1/sync/outfit-context/resources/{name}",
            "resource_names": list(program_api_sync.SYNC_RESOURCE_NAMES),
            "cache_key": "resources[].checksum",
            "no_update_status": 304,
            "bundle": "removed; clients must request only changed resources",
        },
        "hermes_plugin": {
            "manifest": "GET /api/v1/hermes/plugin/manifest",
            "embedded_entry": "/?embed=hermes",
            "workspace_registration": "POST /api/v1/hermes/plugin/workspaces",
            "workspace_launch": "POST /api/v1/hermes/plugin/launch",
            "registration_auth": "Requires owners:write or admin:* bearer scope, or an authenticated same-origin Wardrobe admin session.",
            "owner_binding": "Hermes Mobile generates the workspace Access Key and sends it once. Wardrobe stores token_hash and owner/workspace metadata, and never returns the raw key.",
        },
        "history_write": {
            "endpoint": "POST /api/v1/history/outfits",
            "dry_run_first": True,
            "idempotency": "Use Idempotency-Key or stable external_id.",
        },
        "item_query": {
            "list": "GET /api/v1/items?limit=500",
            "detail": "GET /api/v1/items/{code}",
        },
        "item_write": {
            "endpoint": "POST /api/v1/items",
            "scope": "items:write",
            "owner": "Token owner is the default item owner; items:write:any is required to target another owner.",
            "dry_run_first": True,
            "idempotency": "Use Idempotency-Key or stable external_id.",
            "photos": "Optional photos[] may include JSON base64/data_url image payloads, multipart image file parts, or a follow-up raw image body request to POST /api/v1/items/{code}/photos. Path/file:// strings are not uploads.",
            "photo_endpoint": "POST /api/v1/items/{code}/photos supports multipart image file parts and raw image body with Content-Type image/jpeg or image/png; use it after structured item write/readback.",
            "photo_order_endpoint": "POST /api/v1/items/{code}/photos/order accepts primary_photo_id or a complete photo_ids array to define the first/preview photo.",
            "primary_photo_endpoint": "GET /api/v1/items/{code}/photos/primary/content returns the first ordered product image bytes.",
            "primary_thumbnail_endpoint": "GET /api/v1/items/{code}/photos/primary/thumbnail returns a safe JPEG thumbnail for the first ordered product image.",
        },
        "photo_cache": {
            "resource": "primary_photo_thumbnails",
            "strategy": "Cache only first-photo thumbnails by checksum; fetch originals on demand for detail/OCR/deliverables.",
            "local_path_pattern": ".hermes-cache/photos/{cache_filename}",
            "thumbnail_mime_type": "image/jpeg",
            "thumbnail_max_edge": PROGRAM_API_PRIMARY_THUMBNAIL_MAX_EDGE,
            "no_photo_item_behavior": "Items without photos are valid: keep the item, set primary_photo to null, skip thumbnail caching, and continue sync.",
            "thumbnail_failure_behavior": "Treat one thumbnail 404/0B/MIME/generation failure as an item-level missing visual, not a whole-sync failure.",
        },
        "item_schema": {
            "stable_name_fields": ["display_name", "name", "section", "brand", "official_desc", "code"],
            "display_name": "Human-readable label; clients should prefer this for recommendation output.",
            "name": "Backward-compatible alias generated from section/official_desc/code.",
            "section": "Original wardrobe item title from the SQLite master data.",
            "primary_photo": "First ordered product photo metadata; use primary_photo_thumbnails resource to cache the safe local thumbnail.",
            "wear_counts": "Wear totals, yearly counts, last_worn_on, and maintenance counters are a separate sync resource.",
        },
        "recommendation_filter": {
            "default_owner": normalized_owner,
            "include_for_status_queries": ["Ordered", "maintenance", "other owner records if explicitly visible"],
            "exclude_from_recommendation_candidates": [
                "status=Ordered/ordered",
                "maintenance_state != 0",
                "non-target owner",
            ],
        },
        "temperature": {
            "daily_fields": ["temp_low", "temp_high", "temp_value"],
            "rule": "Daily temperatures must come from actual weather for wear_date and city.",
            "do_not_infer_from_item_fields": [
                "temp_min",
                "temp_max",
                "standalone_min",
                "standalone_max",
            ],
        },
        "wearcount_new": "disabled",
    }


def _api_sync_featured_look_payload(conn: sqlite3.Connection, look_row: sqlite3.Row) -> dict:
    look = dict(look_row)
    item_rows = conn.execute(
        """
        SELECT
            featured_look_items.slot,
            featured_look_items.source_code,
            featured_look_items.source_section,
            featured_look_items.display_order,
            items.code,
            items.owner,
            items.brand,
            items.section,
            items.layer_role,
            items.status,
            items.maintenance_state
        FROM featured_look_items
        LEFT JOIN items ON items.id = featured_look_items.item_id
        WHERE featured_look_items.featured_look_id = ?
        ORDER BY featured_look_items.display_order ASC, featured_look_items.id ASC
        """,
        (int(look["id"]),),
    ).fetchall()
    status_value = _normalize_edit_value(look.get("status")) or "Active"
    return {
        "id": int(look["id"]),
        "look_id": _normalize_edit_value(look.get("look_id")),
        "owner": _normalize_edit_value(look.get("owner")),
        "status": status_value,
        "use_case": _normalize_edit_value(look.get("use_case")),
        "priority": _normalize_edit_value(look.get("priority")),
        "temp_min": look.get("temp_min"),
        "temp_max": look.get("temp_max"),
        "scene_tag_target": _normalize_edit_value(look.get("scene_tag_target")),
        "relax_center": look.get("relax_center"),
        "relax_span": look.get("relax_span"),
        "notes": _normalize_edit_value(look.get("notes")),
        "source_sheet": _normalize_edit_value(look.get("source_sheet")),
        "updated_at": _normalize_edit_value(look.get("updated_at")),
        "recommendation_eligible": status_value.lower() not in {"archived", "disabled"},
        "items": [
            {
                "slot": _normalize_edit_value(item.get("slot")),
                "source_code": _normalize_edit_value(item.get("source_code")),
                "source_section": _normalize_edit_value(item.get("source_section")),
                "code": _normalize_edit_value(item.get("code")),
                "owner": _normalize_edit_value(item.get("owner")),
                "brand": _normalize_edit_value(item.get("brand")),
                "section": _normalize_edit_value(item.get("section")),
                "layer_role": _normalize_edit_value(item.get("layer_role")),
                "status": _normalize_edit_value(item.get("status")) or "Active",
                "maintenance_state": int(item.get("maintenance_state") or 0),
            }
            for item in (dict(row) for row in item_rows)
        ],
    }


def _api_sync_history_payload(conn: sqlite3.Connection, outfit_row: sqlite3.Row) -> dict:
    outfit = dict(outfit_row)
    item_rows = conn.execute(
        """
        SELECT outfit_items.role, items.code, items.brand, items.section, items.layer_role
        FROM outfit_items
        JOIN items ON items.id = outfit_items.item_id
        WHERE outfit_items.outfit_id = ?
        ORDER BY outfit_items.id ASC, items.code ASC
        """,
        (int(outfit["id"]),),
    ).fetchall()
    return {
        "id": int(outfit["id"]),
        "wear_date": _normalize_date_compare(outfit.get("wear_date")) or "",
        "owner": _normalize_edit_value(outfit.get("owner")),
        "city": _normalize_edit_value(outfit.get("city")),
        "inventory_loc": _normalize_edit_value(outfit.get("inventory_loc")),
        "wear_mode": _normalize_edit_value(outfit.get("wear_mode")),
        "scene_tag": _normalize_edit_value(outfit.get("scene_tag")),
        "temp_value": outfit.get("temp_value"),
        "temp_low": outfit.get("temp_low"),
        "temp_high": outfit.get("temp_high"),
        "avg_relax": outfit.get("avg_relax"),
        "avg_temp_label": _normalize_edit_value(outfit.get("avg_temp_label")),
        "notes": _normalize_edit_value(outfit.get("notes")),
        "created_at": _normalize_edit_value(outfit.get("created_at")),
        "updated_at": _normalize_edit_value(outfit.get("updated_at")),
        "items": [
            {
                "role": _normalize_edit_value(item.get("role")) or _normalize_edit_value(item.get("layer_role")),
                "code": _normalize_edit_value(item.get("code")),
                "brand": _normalize_edit_value(item.get("brand")),
                "section": _normalize_edit_value(item.get("section")),
                "layer_role": _normalize_edit_value(item.get("layer_role")),
            }
            for item in (dict(row) for row in item_rows)
        ],
    }


def _api_sync_owner_from_query(context: dict, query: dict[str, list[str]]) -> str:
    requested_owner = _normalize_edit_value((query.get("owner") or [""])[0])
    if requested_owner:
        if not _api_owner_allowed(context, requested_owner, "sync:read:any"):
            raise PermissionError("forbidden_owner")
        return requested_owner
    owner = _api_context_owner(context)
    if not owner:
        raise PermissionError("token_owner_required")
    return owner


def _api_sync_items_payload(conn: sqlite3.Connection, owner: str) -> list[dict]:
    normalized_owner = _normalize_edit_value(owner)
    item_rows = conn.execute(
        """
        SELECT items.*,
               (SELECT COUNT(*) FROM photos WHERE photos.item_id = items.id) AS photo_count
        FROM items
        WHERE COALESCE(owner, '') = ?
        ORDER BY COALESCE(layer_role, ''), COALESCE(brand, ''), COALESCE(section, ''), COALESCE(code, '')
        """,
        (normalized_owner,),
    ).fetchall()
    return [_api_sync_item_payload_with_primary_photo(conn, row) for row in item_rows]


def _api_sync_primary_photo_thumbnails_payload(conn: sqlite3.Connection, owner: str) -> list[dict]:
    normalized_owner = _normalize_edit_value(owner)
    item_rows = conn.execute(
        """
        SELECT items.*,
               (SELECT COUNT(*) FROM photos WHERE photos.item_id = items.id) AS photo_count
        FROM items
        WHERE COALESCE(owner, '') = ?
        ORDER BY COALESCE(layer_role, ''), COALESCE(brand, ''), COALESCE(section, ''), COALESCE(code, '')
        """,
        (normalized_owner,),
    ).fetchall()
    payload = []
    for row in item_rows:
        item_payload = _api_sync_primary_photo_payload(conn, row)
        if item_payload is not None:
            payload.append(item_payload)
    return payload


def _api_sync_wear_counts_payload(conn: sqlite3.Connection, owner: str) -> list[dict]:
    normalized_owner = _normalize_edit_value(owner)
    item_rows = conn.execute(
        """
        SELECT code, owner, wear_total, wear_year, wear_maintenance, wear_threshold, maint_count, last_worn_on
        FROM items
        WHERE COALESCE(owner, '') = ?
        ORDER BY COALESCE(code, '')
        """,
        (normalized_owner,),
    ).fetchall()
    return [_api_sync_wear_count_payload(row) for row in item_rows]


def _api_sync_featured_looks_payload(conn: sqlite3.Connection, owner: str) -> list[dict]:
    normalized_owner = _normalize_edit_value(owner)
    look_rows = conn.execute(
        """
        SELECT *
        FROM featured_looks
        WHERE COALESCE(owner, '') = ?
          AND COALESCE(status, '') <> 'Archived'
        ORDER BY
            CASE
                WHEN TRIM(COALESCE(priority, '')) GLOB '[0-9]*' THEN CAST(priority AS INTEGER)
                ELSE 999999
            END,
            COALESCE(updated_at, created_at, '') DESC,
            look_id ASC
        """,
        (normalized_owner,),
    ).fetchall()
    return [_api_sync_featured_look_payload(conn, row) for row in look_rows]


def _api_sync_wear_history_payload(conn: sqlite3.Connection, owner: str) -> list[dict]:
    normalized_owner = _normalize_edit_value(owner)
    outfit_rows = conn.execute(
        """
        SELECT *
        FROM outfits
        WHERE COALESCE(owner, '') = ?
        ORDER BY wear_date DESC, id DESC
        """,
        (normalized_owner,),
    ).fetchall()
    return [_api_sync_history_payload(conn, row) for row in outfit_rows]


def _api_sync_resource_payload(conn: sqlite3.Connection, owner: str, resource_name: str) -> object:
    normalized_owner = _normalize_edit_value(owner)
    if resource_name == "items":
        return _api_sync_items_payload(conn, normalized_owner)
    if resource_name == "wear_counts":
        return _api_sync_wear_counts_payload(conn, normalized_owner)
    if resource_name == "featured_looks":
        return _api_sync_featured_looks_payload(conn, normalized_owner)
    if resource_name == "wear_history":
        return _api_sync_wear_history_payload(conn, normalized_owner)
    if resource_name == "primary_photo_thumbnails":
        return _api_sync_primary_photo_thumbnails_payload(conn, normalized_owner)
    if resource_name == "rules":
        return _api_sync_rules_payload(normalized_owner)
    raise ValueError("unknown_sync_resource")


def _api_rows_as_dicts(rows: list[sqlite3.Row]) -> list[dict]:
    return [dict(row) for row in rows]


def _api_sync_resource_version_payload(conn: sqlite3.Connection, owner: str, resource_name: str) -> object:
    normalized_owner = _normalize_edit_value(owner)
    if resource_name == "items":
        return _api_rows_as_dicts(
            conn.execute(
                """
                WITH photo_rollup AS (
                    SELECT
                        item_id,
                        COUNT(*) AS photo_count,
                        COALESCE(SUM(id * COALESCE(sort_order, 0)), 0) AS photo_order_hash,
                        COALESCE(MAX(COALESCE(created_at, '')), '') AS max_photo_created_at
                    FROM photos
                    GROUP BY item_id
                )
                SELECT
                    items.id,
                    items.code,
                    COALESCE(items.updated_at, '') AS updated_at,
                    COALESCE(photo_rollup.photo_count, 0) AS photo_count,
                    COALESCE(photo_rollup.photo_order_hash, 0) AS photo_order_hash,
                    COALESCE(photo_rollup.max_photo_created_at, '') AS max_photo_created_at,
                    (SELECT p.id FROM photos p WHERE p.item_id = items.id ORDER BY p.sort_order ASC, p.id ASC LIMIT 1) AS primary_photo_id,
                    COALESCE((SELECT p.file_name FROM photos p WHERE p.item_id = items.id ORDER BY p.sort_order ASC, p.id ASC LIMIT 1), '') AS primary_photo_file_name,
                    COALESCE((SELECT p.original_name FROM photos p WHERE p.item_id = items.id ORDER BY p.sort_order ASC, p.id ASC LIMIT 1), '') AS primary_photo_original_name,
                    COALESCE((SELECT p.mime_type FROM photos p WHERE p.item_id = items.id ORDER BY p.sort_order ASC, p.id ASC LIMIT 1), '') AS primary_photo_mime_type,
                    COALESCE((SELECT p.source_tag FROM photos p WHERE p.item_id = items.id ORDER BY p.sort_order ASC, p.id ASC LIMIT 1), '') AS primary_photo_source_tag,
                    COALESCE((SELECT p.created_at FROM photos p WHERE p.item_id = items.id ORDER BY p.sort_order ASC, p.id ASC LIMIT 1), '') AS primary_photo_created_at
                FROM items
                LEFT JOIN photo_rollup ON photo_rollup.item_id = items.id
                WHERE COALESCE(owner, '') = ?
                ORDER BY COALESCE(layer_role, ''), COALESCE(brand, ''), COALESCE(section, ''), COALESCE(code, '')
                """,
                (normalized_owner,),
            ).fetchall()
        )
    if resource_name == "wear_counts":
        return _api_rows_as_dicts(
            conn.execute(
                """
                SELECT code, owner, wear_total, wear_year, wear_maintenance, wear_threshold, maint_count, last_worn_on
                FROM items
                WHERE COALESCE(owner, '') = ?
                ORDER BY COALESCE(code, '')
                """,
                (normalized_owner,),
            ).fetchall()
        )
    if resource_name == "featured_looks":
        return _api_rows_as_dicts(
            conn.execute(
                """
                SELECT
                    featured_looks.id,
                    featured_looks.look_id,
                    COALESCE(featured_looks.updated_at, '') AS updated_at,
                    COALESCE(featured_looks.status, '') AS status,
                    (SELECT COUNT(*) FROM featured_look_items WHERE featured_look_items.featured_look_id = featured_looks.id) AS item_count,
                    (SELECT COUNT(*) FROM featured_look_photos WHERE featured_look_photos.featured_look_id = featured_looks.id) AS photo_count
                FROM featured_looks
                WHERE COALESCE(owner, '') = ?
                  AND COALESCE(status, '') <> 'Archived'
                ORDER BY
                    CASE
                        WHEN TRIM(COALESCE(priority, '')) GLOB '[0-9]*' THEN CAST(priority AS INTEGER)
                        ELSE 999999
                    END,
                    COALESCE(updated_at, created_at, '') DESC,
                    look_id ASC
                """,
                (normalized_owner,),
            ).fetchall()
        )
    if resource_name == "wear_history":
        return _api_rows_as_dicts(
            conn.execute(
                """
                SELECT
                    outfits.id,
                    outfits.wear_date,
                    COALESCE(outfits.updated_at, '') AS updated_at,
                    (SELECT COUNT(*) FROM outfit_items WHERE outfit_items.outfit_id = outfits.id) AS item_count,
                    (SELECT COUNT(*) FROM outfit_photos WHERE outfit_photos.outfit_id = outfits.id) AS photo_count
                FROM outfits
                WHERE COALESCE(owner, '') = ?
                ORDER BY wear_date DESC, id DESC
                """,
                (normalized_owner,),
            ).fetchall()
        )
    if resource_name == "primary_photo_thumbnails":
        return _api_rows_as_dicts(
            conn.execute(
                """
                SELECT
                    items.id AS item_id,
                    items.code,
                    photos.id AS photo_id,
                    COALESCE(photos.file_name, '') AS file_name,
                    COALESCE(photos.original_name, '') AS original_name,
                    COALESCE(photos.sort_order, 0) AS sort_order,
                    COALESCE(photos.mime_type, '') AS mime_type,
                    COALESCE(photos.source_tag, '') AS source_tag,
                    COALESCE(photos.created_at, '') AS created_at
                FROM items
                JOIN photos ON photos.id = (
                    SELECT p.id
                    FROM photos p
                    WHERE p.item_id = items.id
                    ORDER BY COALESCE(p.sort_order, 0) ASC, p.id ASC
                    LIMIT 1
                )
                WHERE COALESCE(items.owner, '') = ?
                ORDER BY COALESCE(items.layer_role, ''), COALESCE(items.brand, ''), COALESCE(items.section, ''), COALESCE(items.code, '')
                """,
                (normalized_owner,),
            ).fetchall()
        )
    if resource_name == "rules":
        return _api_sync_rules_payload(normalized_owner)
    raise ValueError("unknown_sync_resource")


def _api_sync_outfit_context_manifest(conn: sqlite3.Connection, owner: str) -> dict:
    normalized_owner = _normalize_edit_value(owner)
    resources = [
        program_api_sync.resource_manifest(
            resource_name,
            _api_sync_resource_version_payload(conn, normalized_owner, resource_name),
        )
        for resource_name in program_api_sync.SYNC_RESOURCE_NAMES
    ]
    return program_api_sync.build_manifest(normalized_owner, resources)


def _api_sync_outfit_context_resource(conn: sqlite3.Connection, owner: str, resource_name: str) -> dict:
    normalized_owner = _normalize_edit_value(owner)
    if resource_name not in program_api_sync.SYNC_RESOURCE_NAMES:
        raise ValueError("unknown_sync_resource")
    payload = _api_sync_resource_payload(conn, normalized_owner, resource_name)
    version_payload = _api_sync_resource_version_payload(conn, normalized_owner, resource_name)
    return program_api_sync.build_resource_envelope(normalized_owner, resource_name, payload, checksum_payload=version_payload)


def _api_quote_etag(etag: str) -> str:
    return f'"{etag.replace(chr(34), "")}"'


def _api_etag_matches(header_value: str, etag: str) -> bool:
    normalized_etag = etag.strip()
    for raw_part in str(header_value or "").split(","):
        candidate = raw_part.strip()
        if candidate.startswith("W/"):
            candidate = candidate[2:].strip()
        candidate = candidate.strip('"')
        if candidate == normalized_etag or candidate == "*":
            return True
    return False


def _api_resolve_outfit_item_entries(
    conn: sqlite3.Connection,
    items_payload: list[dict],
    context: dict,
    target_owner: str,
    wear_mode: str,
    wear_date: str,
) -> list[dict]:
    resolved_entries: list[dict] = []
    seen_item_ids: set[int] = set()
    for raw_entry in items_payload:
        if not isinstance(raw_entry, dict):
            raise ValueError("invalid_item_entry")
        code = _normalize_edit_value(raw_entry.get("code"))
        item_id_value = raw_entry.get("item_id")
        if code:
            item = conn.execute("SELECT * FROM items WHERE code = ?", (code,)).fetchone()
        elif item_id_value is not None:
            try:
                item_id = int(item_id_value)
            except Exception as exc:
                raise ValueError("invalid_item_id") from exc
            item = conn.execute("SELECT * FROM items WHERE id = ?", (item_id,)).fetchone()
        else:
            raise ValueError("item_code_required")
        if item is None:
            raise ValueError(f"item_not_found:{code or item_id_value}")
        item_owner = _normalize_edit_value(item["owner"])
        target_owner = _normalize_edit_value(target_owner)
        if not (
            _api_owner_allowed(context, item_owner, "items:write:any")
            or (
                target_owner
                and item_owner == target_owner
                and _api_context_has_scope(context, "history:write:any")
            )
        ):
            raise PermissionError("item_owner_forbidden")
        item_id = int(item["id"])
        if item_id in seen_item_ids:
            continue
        seen_item_ids.add(item_id)
        raw_role = raw_entry.get("role") if _normalize_edit_value(raw_entry.get("role")) else item["layer_role"]
        role = _validate_outfit_role(raw_role, item["code"])
        has_base_layer = _coerce_api_bool(raw_entry.get("has_base_layer"), False)
        resolved_entries.append(
            {
                "item_id": item_id,
                "item": item,
                "role": role,
                "has_base_layer": has_base_layer,
                "wear_delta": _wear_delta_for_item(role, has_base_layer, wear_mode),
                "total_delta": 1,
                "year_delta": 1 if wear_date.startswith("2026") else 0,
            }
        )
    if not resolved_entries:
        raise ValueError("outfit_missing_items")
    return resolved_entries


def _api_request_hash(payload: dict | list) -> str:
    return hashlib.sha256(
        json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
    ).hexdigest()


def _api_idempotency_response(
    conn: sqlite3.Connection,
    context: dict,
    idempotency_key: str,
    request_hash: str,
) -> tuple[dict | None, int]:
    if not idempotency_key:
        return None, 0
    row = conn.execute(
        """
        SELECT request_hash, response_json, status_code
        FROM api_idempotency_keys
        WHERE token_id = ?
          AND idempotency_key = ?
        """,
        (int(context["token_id"]), idempotency_key),
    ).fetchone()
    if row is None:
        return None, 0
    if _normalize_edit_value(row["request_hash"]) != request_hash:
        return {"error": "idempotency_conflict", "message": "Idempotency key was used with a different payload."}, 409
    try:
        return json.loads(row["response_json"] or "{}"), int(row["status_code"] or 200)
    except Exception:
        return {"error": "idempotency_replay_failed"}, 500


def _store_api_idempotency_response(
    conn: sqlite3.Connection,
    context: dict,
    idempotency_key: str,
    request_hash: str,
    response: dict,
    status_code: int,
) -> None:
    if not idempotency_key:
        return
    conn.execute(
        """
        INSERT INTO api_idempotency_keys (
            token_id, idempotency_key, request_hash, response_json, status_code, updated_at
        )
        VALUES (?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
        ON CONFLICT(token_id, idempotency_key) DO UPDATE SET
            response_json = excluded.response_json,
            status_code = excluded.status_code,
            updated_at = CURRENT_TIMESTAMP
        """,
        (
            int(context["token_id"]),
            idempotency_key,
            request_hash,
            json.dumps(response, ensure_ascii=False, sort_keys=True),
            status_code,
        ),
    )


def _api_item_write_owner(context: dict, item_payload: dict) -> str:
    requested_owner = _normalize_edit_value(item_payload.get("owner"))
    token_owner = _api_context_owner(context)
    if _api_context_has_scope(context, "items:write:any"):
        owner = requested_owner or token_owner
    else:
        if requested_owner and requested_owner != token_owner:
            raise PermissionError("forbidden_owner")
        owner = token_owner
    if not owner:
        raise PermissionError("token_owner_required")
    return owner


def _api_apply_item_payload_aliases(payload: dict) -> dict:
    data = dict(payload)

    def copy_first_empty(target: str, *aliases: str) -> None:
        if _normalize_edit_value(data.get(target)):
            return
        for alias in aliases:
            value = data.get(alias)
            if _normalize_edit_value(value):
                data[target] = value
                return

    copy_first_empty("code", "ref", "reference", "model_ref", "sku")
    copy_first_empty("section", "name", "title", "product_name")
    copy_first_empty("loc", "inventory_loc")
    copy_first_empty("official_desc", "official_description", "description", "chinese_desc")
    copy_first_empty("price_original", "original_price", "list_price", "retail_price", "原始价格")
    copy_first_empty("price_original_currency", "original_currency", "original_price_currency", "currency", "currency_code", "原始货币", "原始币种", "币种")
    copy_first_empty("price_cny", "cny_price", "actual_price", "paid_price", "purchase_price", "amount", "price", "人民币价格", "实际价格", "购买价格", "价格", "Price")
    copy_first_empty("acquired_at", "purchase_date", "bought_at")
    copy_first_empty("primary_color", "main_color", "color")
    copy_first_empty("secondary_color", "accent_color")
    copy_first_empty("scene_tag", "scene", "use_case")
    copy_first_empty(
        "wear_threshold",
        "wearThreshold",
        "wear_count_threshold",
        "maintenance_threshold",
        "maintenanceThreshold",
        "service_threshold",
        "threshold",
        "thr",
        "wear_thr",
        "磨损阈值",
        "保养阈值",
        "维护阈值",
    )

    raw_kind = _normalize_edit_value(data.get("kind") or data.get("item_type") or data.get("type") or data.get("category")).lower()
    if raw_kind in {"watch", "腕表"}:
        data["kind"] = "watch"
        data["layer_role"] = "Watch"
        if not _normalize_edit_value(data.get("scene_tag")):
            data["scene_tag"] = "Watch"
    elif raw_kind in {"wardrobe", "clothing", "clothes", "garment", "衣物", "衣橱"}:
        data["kind"] = "wardrobe"
    return data


def _api_item_payload_from_request(payload: dict) -> dict:
    nested_item = payload.get("item")
    if isinstance(nested_item, dict):
        item_payload = dict(nested_item)
        for field in ("kind", "owner", "source", "external_id"):
            if field not in item_payload and field in payload:
                item_payload[field] = payload[field]
        return item_payload
    return dict(payload)


def _api_parse_item_request_payload(handler: BaseHTTPRequestHandler) -> dict:
    if is_multipart_form_data(handler.headers.get("Content-Type", "")):
        return program_item_payload_from_multipart_parts(parse_multipart_parts(handler))
    payload = _parse_json(handler)
    if not isinstance(payload, dict):
        raise ValueError("JSON body must be an object.")
    return payload


def _api_query_value(query: dict[str, list[str]], name: str) -> str:
    values = query.get(name) or []
    return _normalize_edit_value(values[0] if values else "")


def _api_parse_item_photo_request_payload(handler: BaseHTTPRequestHandler, query: dict[str, list[str]]) -> dict:
    content_type = handler.headers.get("Content-Type", "")
    media_type = content_type.split(";", 1)[0].strip().lower()
    if media_type.startswith("image/") or media_type == "application/octet-stream":
        length = int(handler.headers.get("Content-Length", "0"))
        content = handler.rfile.read(length) if length else b""
        filename = (
            _normalize_edit_value(handler.headers.get("X-Filename"))
            or _normalize_edit_value(handler.headers.get("X-File-Name"))
            or _api_query_value(query, "filename")
            or f"photo-{uuid.uuid4().hex[:8]}{_api_photo_suffix(media_type)}"
        )
        upload_content_type = media_type
        if upload_content_type == "application/octet-stream":
            guessed_type = mimetypes.guess_type(filename)[0] or ""
            upload_content_type = guessed_type if guessed_type.startswith("image/") else ""
        if not upload_content_type.startswith("image/"):
            raise ValueError("raw_photo_content_type_required")
        payload = {
            PROGRAM_API_MULTIPART_FILE_PAYLOAD_KEY: [
                {
                    "name": "raw",
                    "filename": filename,
                    "content_type": upload_content_type,
                    "content": content,
                }
            ]
        }
        for field, header_name in (
            ("dry_run", "X-Dry-Run"),
            ("replace_photos", "X-Replace-Photos"),
            ("external_id", "X-External-Id"),
        ):
            value = _api_query_value(query, field) or _normalize_edit_value(handler.headers.get(header_name))
            if value:
                payload[field] = value
        return payload
    return _api_parse_item_request_payload(handler)


def _api_photo_suffix(content_type: str) -> str:
    normalized = content_type.split(";", 1)[0].strip().lower()
    return {
        "image/jpeg": ".jpg",
        "image/jpg": ".jpg",
        "image/png": ".png",
        "image/webp": ".webp",
        "image/gif": ".gif",
        "image/heic": ".heic",
        "image/heif": ".heic",
    }.get(normalized, ".jpg")


def _api_decode_item_photo(photo: object, index: int) -> dict:
    if isinstance(photo, str):
        source = {"data_url": photo}
    elif isinstance(photo, dict):
        source = photo
    else:
        raise ValueError(f"photos[{index}]_invalid")

    filename = _normalize_edit_value(
        source.get("file_name")
        or source.get("filename")
        or source.get("name")
        or source.get("original_name")
    )
    content_type = _normalize_edit_value(source.get("content_type") or source.get("mime_type"))
    raw_b64 = _normalize_edit_value(
        source.get("data_base64")
        or source.get("base64")
        or source.get("content_base64")
    )
    data_url = _normalize_edit_value(source.get("data_url"))
    if data_url:
        match = re.match(r"^data:([^;,]+);base64,(.+)$", data_url, flags=re.IGNORECASE | re.DOTALL)
        if not match:
            raise ValueError(f"photos[{index}]_invalid_data_url")
        content_type = content_type or match.group(1)
        raw_b64 = match.group(2)
    if not raw_b64:
        raise ValueError(f"photos[{index}]_missing_data")
    try:
        content = base64.b64decode(re.sub(r"\s+", "", raw_b64), validate=True)
    except (binascii.Error, ValueError) as exc:
        raise ValueError(f"photos[{index}]_invalid_base64") from exc
    content_type = content_type or "image/jpeg"
    if not filename:
        filename = f"photo-{index}{_api_photo_suffix(content_type)}"
    else:
        suffix = Path(filename).suffix
        if not suffix:
            filename = f"{filename}{_api_photo_suffix(content_type)}"
    file_part = {
        "filename": Path(filename).name,
        "content_type": content_type,
        "content": content,
    }
    upload_error = _validate_image_upload(file_part)
    if upload_error:
        raise ValueError(f"photos[{index}]_{upload_error}")
    return file_part


def _api_decode_multipart_item_photo(photo: object, index: int) -> dict:
    if not isinstance(photo, dict):
        raise ValueError(f"photos[{index}]_invalid")
    filename = _normalize_edit_value(photo.get("filename"))
    content = photo.get("content") or b""
    if isinstance(content, str):
        content = content.encode("utf-8")
    if not filename:
        raise ValueError(f"photos[{index}]_missing_filename")
    if not content:
        raise ValueError(f"photos[{index}]_missing_data")
    file_part = {
        "filename": Path(filename).name,
        "content_type": _normalize_edit_value(photo.get("content_type")) or "image/jpeg",
        "content": content,
    }
    upload_error = _validate_image_upload(file_part)
    if upload_error:
        raise ValueError(f"photos[{index}]_{upload_error}")
    return file_part


def _api_decode_item_photos(payload: dict, item_payload: dict) -> list[dict]:
    raw_photos = payload.get("photos")
    if raw_photos is None:
        raw_photos = item_payload.get("photos")
    multipart_photos = payload.get(PROGRAM_API_MULTIPART_FILE_PAYLOAD_KEY)
    if raw_photos is None and not multipart_photos:
        return []
    if raw_photos is not None and not isinstance(raw_photos, list):
        raise ValueError("photos_must_be_array")
    if multipart_photos is not None and not isinstance(multipart_photos, list):
        raise ValueError("multipart_photos_must_be_array")
    raw_photo_count = len(raw_photos or [])
    multipart_photo_count = len(multipart_photos or [])
    if raw_photo_count + multipart_photo_count > PROGRAM_API_ITEM_PHOTO_MAX_COUNT:
        raise ValueError("too_many_photos")
    photos: list[dict] = []
    total_bytes = 0
    index = 1
    for photo in raw_photos or []:
        decoded = _api_decode_item_photo(photo, index)
        total_bytes += len(decoded.get("content") or b"")
        if total_bytes > PROGRAM_API_ITEM_PHOTO_MAX_TOTAL_BYTES:
            raise ValueError("photos_total_too_large")
        photos.append(decoded)
        index += 1
    for photo in multipart_photos or []:
        decoded = _api_decode_multipart_item_photo(photo, index)
        total_bytes += len(decoded.get("content") or b"")
        if total_bytes > PROGRAM_API_ITEM_PHOTO_MAX_TOTAL_BYTES:
            raise ValueError("photos_total_too_large")
        photos.append(decoded)
        index += 1
    return photos


def _api_insert_item_photos(conn: sqlite3.Connection, item_id: int, code: str, photos: list[dict]) -> int:
    if not photos:
        return 0
    next_order = conn.execute(
        "SELECT COALESCE(MAX(sort_order), 0) + 1 AS next_order FROM photos WHERE item_id = ?",
        (item_id,),
    ).fetchone()["next_order"]
    saved = 0
    safe_code = re.sub(r"[^0-9A-Za-z._-]+", "_", code)[:80] or "item"
    for offset, photo in enumerate(photos):
        original_name = _normalize_edit_value(photo.get("filename")) or f"photo-{offset + 1}.jpg"
        suffix = Path(original_name).suffix.lower() or _api_photo_suffix(_normalize_edit_value(photo.get("content_type")))
        safe_name = f"{safe_code}_{uuid.uuid4().hex[:8]}{suffix}"
        conn.execute(
            """
            INSERT INTO photos (item_id, file_name, original_name, sort_order, source_tag, mime_type, data)
            VALUES (?, ?, ?, ?, 'api', ?, ?)
            """,
            (
                item_id,
                safe_name,
                original_name,
                int(next_order) + offset,
                _normalize_edit_value(photo.get("content_type")) or "application/octet-stream",
                photo.get("content") or b"",
            ),
        )
        saved += 1
    return saved


def _api_prepare_item_write_payload(context: dict, payload: dict) -> tuple[dict, list[dict]]:
    item_payload = _api_item_payload_from_request(payload)
    item_payload = _api_apply_item_payload_aliases(item_payload)
    item_payload["owner"] = _api_item_write_owner(context, item_payload)
    photos = _api_decode_item_photos(payload, item_payload)
    return item_payload, photos


def _api_item_write_preview(conn: sqlite3.Connection, context: dict, payload: dict) -> tuple[dict, int]:
    item_payload, photos = _api_prepare_item_write_payload(context, payload)
    mode = _normalize_edit_value(payload.get("mode")).lower() or "create_only"
    if mode not in {"create_only", "upsert", "replace"}:
        raise ValueError("invalid_mode")
    code = _normalize_edit_value(item_payload.get("code"))
    existing = conn.execute("SELECT * FROM items WHERE code = ?", (code,)).fetchone() if code else None
    if existing is not None and not _api_owner_allowed(context, existing["owner"], "items:write:any"):
        raise PermissionError("forbidden_owner")
    if existing is not None and mode == "create_only":
        return {
            "error": "duplicate_code",
            "dry_run": True,
            "owner": _normalize_edit_value(existing["owner"]),
            "item_id": int(existing["id"]),
            "code": _normalize_edit_value(existing["code"]),
        }, 409
    if existing is not None:
        normalized = _normalize_item_payload(item_payload, existing)
        after = dict(existing)
        after.update(normalized)
        action = "updated"
        item_id = int(existing["id"])
    else:
        normalized = _normalize_item_payload(item_payload)
        after = dict(normalized)
        action = "created"
        item_id = None
    kind = _new_item_kind(after)
    if kind == "watch":
        after["layer_role"] = "Watch"
        after["source_sheet"] = _default_source_sheet("watch")
    else:
        after["source_sheet"] = _default_source_sheet("wardrobe")
    export_tasks = _item_export_tasks_for_change(existing, after)
    return {
        "saved": False,
        "dry_run": True,
        "action": action,
        "owner": _normalize_edit_value(after.get("owner")),
        "kind": kind,
        "item_id": item_id,
        "code": _normalize_edit_value(after.get("code")),
        "photos_received": len(photos),
        "exports_pending": bool(export_tasks),
        "export_tasks": sorted(export_tasks),
    }, 200


def _api_write_item(
    conn: sqlite3.Connection,
    context: dict,
    payload: dict,
) -> tuple[dict, int]:
    item_payload, photos = _api_prepare_item_write_payload(context, payload)
    mode = _normalize_edit_value(payload.get("mode")).lower() or "create_only"
    if mode not in {"create_only", "upsert", "replace"}:
        raise ValueError("invalid_mode")
    code = _normalize_edit_value(item_payload.get("code"))
    existing = conn.execute("SELECT * FROM items WHERE code = ?", (code,)).fetchone() if code else None
    if existing is not None and not _api_owner_allowed(context, existing["owner"], "items:write:any"):
        raise PermissionError("forbidden_owner")
    if existing is not None and mode == "create_only":
        raise RuntimeError("duplicate_code")

    if existing is not None:
        item_id = _update_item_record(conn, existing, item_payload)
        if _coerce_api_bool(payload.get("replace_photos"), False):
            conn.execute("DELETE FROM photos WHERE item_id = ?", (item_id,))
        action = "updated"
        status_code = 200
        before = existing
    else:
        item_id = _insert_item_record(conn, item_payload)
        action = "created"
        status_code = 201
        before = None
    photos_saved = _api_insert_item_photos(conn, item_id, code or _normalize_edit_value(item_payload.get("code")), photos)
    after_row = conn.execute(
        """
        SELECT items.*,
               (SELECT COUNT(*) FROM photos WHERE photos.item_id = items.id) AS photo_count
        FROM items
        WHERE id = ?
        """,
        (item_id,),
    ).fetchone()
    export_tasks = _item_export_tasks_for_change(before, after_row)
    export_results = _run_item_export_tasks(conn, export_tasks)
    item_detail = _api_item_detail_payload(conn, after_row) if after_row is not None else _item_with_photos(conn, item_id)
    response = {
        "saved": True,
        "action": action,
        "owner": _normalize_edit_value(after_row["owner"] if after_row is not None else item_payload.get("owner")),
        "kind": _item_source_kind(after_row or item_payload),
        "item_id": item_id,
        "code": _normalize_edit_value(after_row["code"] if after_row is not None else item_payload.get("code")),
        "photos_saved": photos_saved,
        "exports_pending": bool(export_tasks),
        "export_tasks": sorted(export_tasks),
        "exports": export_results,
        "item": _api_item_payload(after_row) if after_row is not None else item_detail,
        "item_detail": item_detail,
    }
    return response, status_code


def _api_find_item_for_photo_write(conn: sqlite3.Connection, context: dict, code: str) -> sqlite3.Row:
    normalized_code = _normalize_edit_value(code)
    if not normalized_code:
        raise ValueError("missing_item_code")
    item = conn.execute("SELECT * FROM items WHERE code = ?", (normalized_code,)).fetchone()
    if item is None:
        raise KeyError("item_not_found")
    if not _api_owner_allowed(context, item["owner"], "items:write:any"):
        raise PermissionError("forbidden_owner")
    return item


def _api_find_item_for_photo_read(conn: sqlite3.Connection, context: dict, code: str) -> sqlite3.Row:
    normalized_code = _normalize_edit_value(code)
    if not normalized_code:
        raise ValueError("missing_item_code")
    item = conn.execute("SELECT * FROM items WHERE code = ?", (normalized_code,)).fetchone()
    if item is None:
        raise KeyError("item_not_found")
    if not _api_owner_allowed(context, item["owner"], "items:read:any"):
        raise PermissionError("forbidden_owner")
    return item


def _api_int_value(value: object, field_name: str) -> int:
    normalized = _normalize_edit_value(value)
    if not normalized:
        raise ValueError(f"{field_name}_required")
    try:
        return int(normalized)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{field_name}_invalid") from exc


def _api_int_list(value: object, field_name: str) -> list[int]:
    if isinstance(value, str):
        raw_values: object = [part.strip() for part in value.split(",") if part.strip()]
    else:
        raw_values = value
    if not isinstance(raw_values, list):
        raise ValueError(f"{field_name}_must_be_array")
    return [_api_int_value(item, field_name) for item in raw_values]


def _api_photo_order_request(payload: dict) -> tuple[list[int] | None, int | None]:
    raw_photo_ids = (
        payload.get("photo_ids")
        if "photo_ids" in payload
        else payload.get("ordered_photo_ids")
        if "ordered_photo_ids" in payload
        else payload.get("order")
    )
    raw_first_photo_id = (
        payload.get("first_photo_id")
        or payload.get("primary_photo_id")
        or payload.get("photo_id")
    )
    photo_ids = _api_int_list(raw_photo_ids, "photo_ids") if raw_photo_ids is not None else None
    first_photo_id = _api_int_value(raw_first_photo_id, "first_photo_id") if raw_first_photo_id is not None else None
    if photo_ids is not None and first_photo_id is not None:
        raise ValueError("choose_photo_ids_or_first_photo_id")
    return photo_ids, first_photo_id


def _item_photo_order_result(
    conn: sqlite3.Connection,
    item: sqlite3.Row,
    next_photo_ids: list[int],
    *,
    saved: bool,
    dry_run: bool = False,
) -> dict:
    item_id = int(item["id"])
    item_detail = _api_item_detail_payload(conn, item)
    if dry_run:
        photo_by_id = {int(photo["id"]): photo for photo in item_detail.get("photos") or []}
        item_detail["photos"] = [photo_by_id[photo_id] for photo_id in next_photo_ids if photo_id in photo_by_id]
        for sort_order, photo in enumerate(item_detail["photos"], start=1):
            photo["sort_order"] = sort_order
        item_detail["primary_photo"] = item_detail["photos"][0] if item_detail["photos"] else None
        item_detail["primary_photo_content_path"] = (
            item_detail["primary_photo"]["content_path"] if item_detail["primary_photo"] else ""
        )
    return {
        "saved": saved,
        "dry_run": dry_run,
        "action": "photos_reordered",
        "owner": _normalize_edit_value(item["owner"]),
        "item_id": item_id,
        "code": _normalize_edit_value(item["code"]),
        "photo_ids": next_photo_ids,
        "primary_photo_id": next_photo_ids[0] if next_photo_ids else None,
        "item_detail": item_detail,
    }


def _api_item_photo_order_preview(
    conn: sqlite3.Connection,
    context: dict,
    code: str,
    payload: dict,
) -> tuple[dict, int]:
    item = _api_find_item_for_photo_write(conn, context, code)
    photo_ids, first_photo_id = _api_photo_order_request(payload)
    current_ids = [int(photo["id"]) for photo in photo_ordering.item_photo_rows(conn, int(item["id"]))]
    next_photo_ids = photo_ordering.ordered_photo_ids(
        current_ids,
        photo_ids=photo_ids,
        first_photo_id=first_photo_id,
    )
    return _item_photo_order_result(conn, item, next_photo_ids, saved=False, dry_run=True), 200


def _api_reorder_item_photos(
    conn: sqlite3.Connection,
    context: dict,
    code: str,
    payload: dict,
) -> tuple[dict, int]:
    item = _api_find_item_for_photo_write(conn, context, code)
    photo_ids, first_photo_id = _api_photo_order_request(payload)
    current_ids = [int(photo["id"]) for photo in photo_ordering.item_photo_rows(conn, int(item["id"]))]
    next_photo_ids = photo_ordering.ordered_photo_ids(
        current_ids,
        photo_ids=photo_ids,
        first_photo_id=first_photo_id,
    )
    photo_ordering.apply_item_photo_order(conn, int(item["id"]), next_photo_ids)
    return _item_photo_order_result(conn, item, next_photo_ids, saved=True), 200


def _api_item_photo_media_target(path: str) -> tuple[str, int | None, str] | None:
    prefix = "/api/v1/items/"
    if not path.startswith(prefix):
        return None
    body = path[len(prefix):]
    for variant in ("content", "thumbnail"):
        primary_suffix = f"/photos/primary/{variant}"
        if body.endswith(primary_suffix):
            return body[:-len(primary_suffix)], None, variant
    match = re.match(r"^(.+)/photos/([0-9]+)/(content|thumbnail)$", body)
    if not match:
        return None
    return match.group(1), int(match.group(2)), match.group(3)


def _api_item_photo_content_row(
    conn: sqlite3.Connection,
    item_id: int,
    photo_id: int | None,
) -> sqlite3.Row | None:
    if photo_id is None:
        return conn.execute(
            """
            SELECT photos.*,
                   CASE WHEN data IS NULL THEN 0 ELSE length(data) END AS size_bytes
            FROM photos
            WHERE item_id = ?
            ORDER BY sort_order ASC, id ASC
            LIMIT 1
            """,
            (int(item_id),),
        ).fetchone()
    return conn.execute(
        """
        SELECT photos.*,
               CASE WHEN data IS NULL THEN 0 ELSE length(data) END AS size_bytes
        FROM photos
        WHERE item_id = ? AND id = ?
        """,
        (int(item_id), int(photo_id)),
    ).fetchone()


def _api_item_photo_write_preview(
    conn: sqlite3.Connection,
    context: dict,
    code: str,
    payload: dict,
) -> tuple[dict, int]:
    item = _api_find_item_for_photo_write(conn, context, code)
    photos = _api_decode_item_photos(payload, {})
    if not photos:
        raise ValueError("photos_required")
    action = "photos_replaced" if _coerce_api_bool(payload.get("replace_photos"), False) else "photos_appended"
    return {
        "saved": False,
        "dry_run": True,
        "action": action,
        "owner": _normalize_edit_value(item["owner"]),
        "item_id": int(item["id"]),
        "code": _normalize_edit_value(item["code"]),
        "photos_received": len(photos),
    }, 200


def _api_write_item_photos(
    conn: sqlite3.Connection,
    context: dict,
    code: str,
    payload: dict,
) -> tuple[dict, int]:
    item = _api_find_item_for_photo_write(conn, context, code)
    photos = _api_decode_item_photos(payload, {})
    if not photos:
        raise ValueError("photos_required")
    replace_photos = _coerce_api_bool(payload.get("replace_photos"), False)
    if replace_photos:
        conn.execute("DELETE FROM photos WHERE item_id = ?", (int(item["id"]),))
    photos_saved = _api_insert_item_photos(conn, int(item["id"]), _normalize_edit_value(item["code"]), photos)
    item_row = conn.execute(
        """
        SELECT items.*,
               (SELECT COUNT(*) FROM photos WHERE photos.item_id = items.id) AS photo_count
        FROM items
        WHERE id = ?
        """,
        (int(item["id"]),),
    ).fetchone()
    item_detail = _api_item_detail_payload(conn, item_row or item)
    return {
        "saved": True,
        "action": "photos_replaced" if replace_photos else "photos_appended",
        "owner": _normalize_edit_value(item["owner"]),
        "item_id": int(item["id"]),
        "code": _normalize_edit_value(item["code"]),
        "photos_saved": photos_saved,
        "item_detail": item_detail,
    }, 200


def _resolved_outfit_entry_signature(resolved_entries: list[dict]) -> list[tuple[str, int, int]]:
    signature: list[tuple[str, int, int]] = []
    for entry in resolved_entries:
        item_id = int(entry.get("item_id") or 0)
        role = _canonical_outfit_role(entry.get("role"))
        if item_id <= 0 or not role:
            continue
        signature.append((role, item_id, 1 if bool(entry.get("has_base_layer")) else 0))
    signature.sort()
    return signature


def _serialized_outfit_item_signature(outfit_data: dict) -> list[tuple[str, int, int]]:
    signature: list[tuple[str, int, int]] = []
    for item in outfit_data.get("items", []):
        item_id = int(item.get("id") or 0)
        role = _canonical_outfit_role(item.get("role") or item.get("layer_role"))
        if item_id <= 0 or not role:
            continue
        signature.append((role, item_id, 1 if bool(item.get("has_base_layer")) else 0))
    signature.sort()
    return signature


def _outfit_daily_update_payload_hash(
    wear_date: str,
    city: str,
    wear_mode: str,
    notes: str,
    owner: str,
    resolved_entries: list[dict],
) -> str:
    payload_for_hash = {
        "wear_date": wear_date,
        "city": city,
        "wear_mode": wear_mode,
        "notes": notes,
        "owner": owner,
        "items": [
            {
                "item_id": int(entry["item_id"]),
                "role": entry["role"],
                "has_base_layer": 1 if entry["has_base_layer"] else 0,
            }
            for entry in resolved_entries
        ],
    }
    return hashlib.sha256(
        json.dumps(payload_for_hash, ensure_ascii=False, sort_keys=True).encode("utf-8")
    ).hexdigest()


def _apply_outfit_daily_update_entries(
    conn: sqlite3.Connection,
    daily_update_id: int,
    wear_date: str,
    resolved_entries: list[dict],
) -> list[int]:
    affected_item_ids: list[int] = []
    year_delta = 1 if wear_date.startswith("2026") else 0
    for entry in resolved_entries:
        item_id = int(entry["item_id"])
        affected_item_ids.append(item_id)
        conn.execute(
            """
            INSERT INTO wearcount_daily_update_items (
                daily_update_id, item_id, code, role, has_base_layer, wear_delta, total_delta, year_delta
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                int(daily_update_id),
                item_id,
                _normalize_edit_value(entry["item"]["code"]),
                entry["role"],
                1 if entry["has_base_layer"] else 0,
                float(entry["wear_delta"]),
                int(entry["total_delta"]),
                year_delta,
            ),
        )
        conn.execute(
            """
            UPDATE items
            SET wear_maintenance = COALESCE(wear_maintenance, 0) + ?,
                wear_total = COALESCE(wear_total, 0) + ?,
                wear_year = COALESCE(wear_year, 0) + ?,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (
                float(entry["wear_delta"]),
                int(entry["total_delta"]),
                year_delta,
                item_id,
            ),
        )
    return affected_item_ids


def _replace_outfit_items_and_rollup(
    conn: sqlite3.Connection,
    outfit_id: int,
    payload: dict,
    resolved_entries: list[dict],
) -> sqlite3.Row:
    wear_date = _normalize_edit_value(payload.get("wear_date"))
    city = _normalize_edit_value(payload.get("city"))
    wear_mode = _normalize_edit_value(payload.get("wear_mode")) or "normal"
    notes = _normalize_edit_value(payload.get("notes"))
    owner = _normalize_edit_value(payload.get("owner")) or "徐欣"
    conn.execute("DELETE FROM outfit_items WHERE outfit_id = ?", (outfit_id,))
    for entry in resolved_entries:
        conn.execute(
            "INSERT OR IGNORE INTO outfit_items (outfit_id, item_id, role) VALUES (?, ?, ?)",
            (outfit_id, int(entry["item_id"]), entry["role"]),
        )
    avg_relax, avg_temp_label = _outfit_rollup(conn, [int(entry["item_id"]) for entry in resolved_entries])
    conn.execute(
        """
        UPDATE outfits
        SET wear_date = ?,
            city = ?,
            inventory_loc = ?,
            owner = ?,
            wear_mode = ?,
            scene_tag = ?,
            temp_value = ?,
            temp_low = ?,
            temp_high = ?,
            avg_relax = ?,
            avg_temp_label = ?,
            notes = ?,
            updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
        """,
        (
            wear_date,
            city,
            _normalize_edit_value(payload.get("inventory_loc")),
            owner,
            wear_mode,
            _normalize_edit_value(payload.get("scene_tag")),
            payload.get("temp_value"),
            payload.get("temp_low"),
            payload.get("temp_high"),
            avg_relax,
            avg_temp_label,
            notes,
            outfit_id,
        ),
    )
    return conn.execute("SELECT * FROM outfits WHERE id = ?", (outfit_id,)).fetchone()


def _create_outfit_daily_update(
    conn: sqlite3.Connection,
    payload: dict,
    resolved_entries: list[dict],
) -> tuple[sqlite3.Row, list[int]]:
    wear_date = _normalize_edit_value(payload.get("wear_date"))
    if not wear_date:
        raise ValueError("outfit_missing_wear_date")
    city = _normalize_edit_value(payload.get("city"))
    wear_mode = _normalize_edit_value(payload.get("wear_mode")) or "normal"
    notes = _normalize_edit_value(payload.get("notes"))
    owner = _normalize_edit_value(payload.get("owner")) or "徐欣"
    source_path = _normalize_edit_value(payload.get("source_path")) or "manual_edit"
    existing_update = _find_daily_update(conn, wear_date, owner)
    payload_hash = _outfit_daily_update_payload_hash(wear_date, city, wear_mode, notes, owner, resolved_entries)
    affected_item_ids: list[int] = []
    if existing_update is not None:
        affected_item_ids.extend(_subtract_daily_update_effects(conn, int(existing_update["id"])))
        conn.execute(
            """
            UPDATE wearcount_daily_updates
            SET city = ?,
                wear_mode = ?,
                notes = ?,
                source_path = ?,
                payload_hash = ?,
                owner = ?,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (
                city,
                wear_mode,
                notes,
                source_path,
                payload_hash,
                owner,
                int(existing_update["id"]),
            ),
        )
        daily_update_id = int(existing_update["id"])
    else:
        cursor = conn.execute(
            """
            INSERT INTO wearcount_daily_updates (wear_date, city, wear_mode, notes, source_path, payload_hash, owner)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (wear_date, city, wear_mode, notes, source_path, payload_hash, owner),
        )
        daily_update_id = int(cursor.lastrowid)
    affected_item_ids.extend(_apply_outfit_daily_update_entries(conn, daily_update_id, wear_date, resolved_entries))
    conn.execute(
        """
        INSERT INTO outfits (
            wear_date, city, inventory_loc, owner, wear_mode, scene_tag,
            temp_value, temp_low, temp_high, notes, updated_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
        ON CONFLICT(wear_date, owner) DO UPDATE SET
            city = excluded.city,
            inventory_loc = excluded.inventory_loc,
            wear_mode = excluded.wear_mode,
            scene_tag = excluded.scene_tag,
            temp_value = excluded.temp_value,
            temp_low = excluded.temp_low,
            temp_high = excluded.temp_high,
            notes = excluded.notes,
            updated_at = CURRENT_TIMESTAMP
        """,
        (
            wear_date,
            city,
            _normalize_edit_value(payload.get("inventory_loc")),
            owner,
            wear_mode,
            _normalize_edit_value(payload.get("scene_tag")),
            payload.get("temp_value"),
            payload.get("temp_low"),
            payload.get("temp_high"),
            notes,
        ),
    )
    outfit_row = conn.execute(
        "SELECT * FROM outfits WHERE wear_date = ? AND COALESCE(owner, '') = ? ORDER BY id DESC LIMIT 1",
        (wear_date, owner),
    ).fetchone()
    saved_outfit = _replace_outfit_items_and_rollup(conn, int(outfit_row["id"]), payload, resolved_entries)
    _recompute_item_last_worn_on(conn, affected_item_ids)
    return saved_outfit, sorted({int(value) for value in affected_item_ids if value is not None})


def _outfit_metadata_only_edit(
    conn: sqlite3.Connection,
    outfit: sqlite3.Row,
    payload: dict,
    resolved_entries: list[dict],
    username: str,
) -> sqlite3.Row | None:
    wear_date = _normalize_edit_value(payload.get("wear_date")) or _normalize_edit_value(outfit["wear_date"])
    if wear_date != _normalize_edit_value(outfit["wear_date"]):
        return None
    serialized = _serialize_outfit(conn, outfit, username)
    if _serialized_outfit_item_signature(serialized) != _resolved_outfit_entry_signature(resolved_entries):
        return None
    city = _normalize_edit_value(payload.get("city")) or _normalize_edit_value(outfit["city"])
    inventory_loc = _normalize_edit_value(payload.get("inventory_loc")) or _normalize_edit_value(outfit["inventory_loc"])
    owner = _normalize_edit_value(payload.get("owner")) or _normalize_edit_value(outfit["owner"]) or "徐欣"
    wear_mode = _normalize_edit_value(payload.get("wear_mode")) or _normalize_edit_value(outfit["wear_mode"]) or "normal"
    scene_tag = _normalize_edit_value(payload.get("scene_tag")) or _normalize_edit_value(outfit["scene_tag"])
    temp_value = payload.get("temp_value")
    temp_low = payload.get("temp_low")
    temp_high = payload.get("temp_high")
    notes = _normalize_edit_value(payload.get("notes")) or _normalize_edit_value(outfit["notes"])
    conn.execute(
        """
        UPDATE outfits
        SET city = ?,
            inventory_loc = ?,
            owner = ?,
            wear_mode = ?,
            scene_tag = ?,
            temp_value = ?,
            temp_low = ?,
            temp_high = ?,
            notes = ?,
            updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
        """,
        (
            city,
            inventory_loc,
            owner,
            wear_mode,
            scene_tag,
            temp_value,
            temp_low,
            temp_high,
            notes,
            int(outfit["id"]),
        ),
    )
    return conn.execute("SELECT * FROM outfits WHERE id = ?", (int(outfit["id"]),)).fetchone()


def _latest_outfit_id(conn: sqlite3.Connection, owner: str = "") -> int | None:
    normalized_owner = _normalize_edit_value(owner)
    if normalized_owner:
        row = conn.execute(
            "SELECT id FROM outfits WHERE COALESCE(owner, '') = ? ORDER BY wear_date DESC, id DESC LIMIT 1",
            (normalized_owner,),
        ).fetchone()
    else:
        row = conn.execute("SELECT id FROM outfits ORDER BY wear_date DESC, id DESC LIMIT 1").fetchone()
    return int(row["id"]) if row is not None else None


def _ensure_latest_outfit_editable(conn: sqlite3.Connection, outfit_id: int, owner: str = "") -> None:
    latest_outfit_id = _latest_outfit_id(conn, owner)
    if latest_outfit_id is None or int(outfit_id) != int(latest_outfit_id):
        raise RuntimeError("outfit_only_latest_record_editable")


def _upsert_outfit_daily_update(
    conn: sqlite3.Connection,
    outfit_id: int,
    original_wear_date: str,
    payload: dict,
    resolved_entries: list[dict],
    username: str,
    outfit_row: sqlite3.Row,
) -> tuple[sqlite3.Row, list[int]]:
    wear_date = _normalize_edit_value(payload.get("wear_date"))
    if not wear_date:
        raise ValueError("outfit_missing_wear_date")
    city = _normalize_edit_value(payload.get("city"))
    wear_mode = _normalize_edit_value(payload.get("wear_mode")) or "normal"
    notes = _normalize_edit_value(payload.get("notes"))
    owner = _normalize_edit_value(payload.get("owner")) or "徐欣"
    source_path = _normalize_edit_value(payload.get("source_path")) or "manual_edit"
    original_owner = _normalize_edit_value(outfit_row["owner"]) or owner
    existing_update = _find_daily_update(conn, original_wear_date, original_owner)
    if existing_update is None:
        metadata_only_saved = _outfit_metadata_only_edit(conn, outfit_row, payload, resolved_entries, username)
        if metadata_only_saved is not None:
            return metadata_only_saved, []
        raise RuntimeError("outfit_missing_wearcount_basis")
    conflicting_outfit = conn.execute(
        "SELECT id FROM outfits WHERE wear_date = ? AND COALESCE(owner, '') = ? AND id <> ?",
        (wear_date, owner, outfit_id),
    ).fetchone()
    if conflicting_outfit is not None:
        raise RuntimeError("outfit_date_conflict")
    conflicting_update = conn.execute(
        "SELECT id FROM wearcount_daily_updates WHERE wear_date = ? AND COALESCE(owner, '') = ? AND id <> ?",
        (wear_date, owner, int(existing_update["id"])),
    ).fetchone()
    if conflicting_update is not None:
        raise RuntimeError("outfit_date_conflict")
    affected_item_ids = _subtract_daily_update_effects(conn, int(existing_update["id"]))
    payload_hash = _outfit_daily_update_payload_hash(wear_date, city, wear_mode, notes, owner, resolved_entries)
    conn.execute(
        """
        UPDATE wearcount_daily_updates
        SET wear_date = ?,
            city = ?,
            wear_mode = ?,
            notes = ?,
            source_path = ?,
            payload_hash = ?,
            owner = ?,
            updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
        """,
        (
            wear_date,
            city,
            wear_mode,
            notes,
            source_path,
            payload_hash,
            owner,
            int(existing_update["id"]),
        ),
    )
    affected_item_ids.extend(
        _apply_outfit_daily_update_entries(conn, int(existing_update["id"]), wear_date, resolved_entries)
    )
    saved_outfit = _replace_outfit_items_and_rollup(conn, outfit_id, payload, resolved_entries)
    _recompute_item_last_worn_on(conn, affected_item_ids)
    return saved_outfit, sorted({int(value) for value in affected_item_ids if value is not None})


def _delete_outfit_with_wearcount(conn: sqlite3.Connection, outfit: sqlite3.Row) -> list[int]:
    daily_update = _find_daily_update(
        conn,
        _normalize_edit_value(outfit["wear_date"]),
        _normalize_edit_value(outfit["owner"]),
    )
    if daily_update is None:
        raise RuntimeError("outfit_missing_wearcount_basis")
    affected_item_ids = _subtract_daily_update_effects(conn, int(daily_update["id"]))
    conn.execute("DELETE FROM wearcount_daily_updates WHERE id = ?", (int(daily_update["id"]),))
    conn.execute("DELETE FROM outfits WHERE id = ?", (int(outfit["id"]),))
    _recompute_item_last_worn_on(conn, affected_item_ids)
    return affected_item_ids


def _create_featured_look_from_outfit(conn: sqlite3.Connection, outfit: sqlite3.Row, username: str) -> dict:
    serialized = _serialize_outfit(conn, outfit, username)
    slot_entries: dict[str, dict] = {}
    duplicate_roles: list[str] = []
    anchor_item: dict | None = None
    sorted_items = sorted(
        serialized.get("items", []),
        key=lambda item: _featured_look_anchor_sort_rank(item.get("role") or item.get("layer_role")),
    )
    for item in sorted_items:
        role = _normalize_edit_value(item.get("role") or item.get("layer_role"))
        slot = _featured_look_slot_from_role(role)
        if anchor_item is None and slot is not None:
            anchor_item = item
        if slot is None:
            continue
        if slot in slot_entries:
            duplicate_roles.append(role or slot)
            continue
        slot_entries[slot] = item
    if duplicate_roles:
        raise RuntimeError(f"featured_look_duplicate_slots:{','.join(sorted(set(duplicate_roles)))}")
    owner = _normalize_edit_value(outfit["owner"]) or "寰愭"
    signature = _featured_look_signature(anchor_item, slot_entries)
    duplicate_look = _find_duplicate_featured_look(conn, owner, signature)
    if duplicate_look is not None:
        raise RuntimeError(f"featured_look_duplicate_existing:{_normalize_edit_value(duplicate_look['look_id'])}")
    look_id = _next_featured_look_id(conn, _normalize_edit_value(outfit["wear_date"]))
    cursor = conn.execute(
        """
        INSERT INTO featured_looks (
            look_id, anchor_type, anchor_code, anchor_section, use_case, priority, status, owner,
            temp_min, temp_max, scene_tag_target, relax_center, relax_span, material_line, notes, source_sheet, updated_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
        """,
        (
            look_id,
            _featured_look_slot_from_role(_normalize_edit_value(anchor_item.get("role") or anchor_item.get("layer_role"))) if anchor_item else "",
            _normalize_edit_value(anchor_item.get("code")) if anchor_item else "",
            _normalize_edit_value(anchor_item.get("section")) if anchor_item else "",
            _normalize_edit_value(outfit["scene_tag"]) or "历史记录",
            "50",
            "Active",
            _normalize_edit_value(outfit["owner"]) or "徐欣",
            outfit["temp_low"],
            outfit["temp_high"],
            _normalize_edit_value(outfit["scene_tag"]),
            outfit["avg_relax"],
            None,
            "",
            _normalize_edit_value(outfit["notes"]),
            "Manual",
        ),
    )
    featured_look_id = int(cursor.lastrowid)
    display_order = 1
    if anchor_item is not None:
        conn.execute(
            """
            INSERT INTO featured_look_items (featured_look_id, item_id, slot, source_code, source_section, display_order)
            VALUES (?, ?, 'anchor', ?, ?, ?)
            """,
            (
                featured_look_id,
                int(anchor_item["id"]),
                _normalize_edit_value(anchor_item.get("code")),
                _normalize_edit_value(anchor_item.get("section")),
                display_order,
            ),
        )
        display_order += 1
    for slot in CANONICAL_FEATURED_LOOK_SLOTS:
        item = slot_entries.get(slot)
        if item is None:
            continue
        conn.execute(
            """
            INSERT INTO featured_look_items (featured_look_id, item_id, slot, source_code, source_section, display_order)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                featured_look_id,
                int(item["id"]),
                slot,
                _normalize_edit_value(item.get("code")),
                _normalize_edit_value(item.get("section")),
                display_order,
            ),
        )
        display_order += 1
    outfit_photos = conn.execute(
        """
        SELECT file_name, original_name, sort_order, source_tag, mime_type, data
        FROM outfit_photos
        WHERE outfit_id = ?
        ORDER BY sort_order, id
        """,
        (int(outfit["id"]),),
    ).fetchall()
    for index, photo in enumerate(outfit_photos, start=1):
        conn.execute(
            """
            INSERT INTO featured_look_photos (
                featured_look_id, file_name, original_name, sort_order, source_tag, mime_type, data
            )
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                featured_look_id,
                _normalize_edit_value(photo["file_name"]) or f"look_{featured_look_id}_{index}.bin",
                _normalize_edit_value(photo["original_name"]),
                index,
                _normalize_edit_value(photo["source_tag"]) or "copied_from_outfit",
                _normalize_edit_value(photo["mime_type"]),
                photo["data"],
            ),
    )
    _sync_featured_look_relax_metrics(conn, featured_look_id)
    _export_featured_looks_workbooks(conn)
    saved = conn.execute("SELECT * FROM featured_looks WHERE id = ?", (featured_look_id,)).fetchone()
    return _serialize_featured_look(conn, saved)


def _sync_outfit_photo_into_existing_featured_look(
    conn: sqlite3.Connection,
    outfit: sqlite3.Row,
    photo_row: sqlite3.Row,
    username: str,
) -> bool:
    serialized = _serialize_outfit(conn, outfit, username)
    existing_look = _existing_featured_look_for_outfit(conn, serialized)
    if existing_look is None:
        return False
    existing_photos = conn.execute(
        """
        SELECT original_name, mime_type, data
        FROM featured_look_photos
        WHERE featured_look_id = ?
        """,
        (int(existing_look["id"]),),
    ).fetchall()
    original_name = _normalize_edit_value(photo_row["original_name"])
    mime_type = _normalize_edit_value(photo_row["mime_type"])
    data = photo_row["data"]
    for existing_photo in existing_photos:
        if (
            _normalize_edit_value(existing_photo["original_name"]) == original_name
            and _normalize_edit_value(existing_photo["mime_type"]) == mime_type
            and existing_photo["data"] == data
        ):
            return False
    next_order = conn.execute(
        """
        SELECT COALESCE(MAX(sort_order), 0) + 1 AS next_order
        FROM featured_look_photos
        WHERE featured_look_id = ?
        """,
        (int(existing_look["id"]),),
    ).fetchone()["next_order"]
    suffix = Path(_normalize_edit_value(photo_row["file_name"]) or _normalize_edit_value(photo_row["original_name"])).suffix.lower() or ".bin"
    safe_name = f"look_{int(existing_look['id'])}_{uuid.uuid4().hex[:8]}{suffix}"
    conn.execute(
        """
        INSERT INTO featured_look_photos (
            featured_look_id, file_name, original_name, sort_order, source_tag, mime_type, data
        )
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (
            int(existing_look["id"]),
            safe_name,
            original_name,
            next_order,
            "copied_from_outfit",
            mime_type,
            data,
        ),
    )
    return True


def _update_featured_look_record(conn: sqlite3.Connection, look: sqlite3.Row, payload: dict) -> dict:
    values = {
        "id": int(look["id"]),
        "look_id": _normalize_edit_value(payload.get("look_id")) or _normalize_edit_value(look["look_id"]),
        "use_case": _normalize_edit_value(payload.get("use_case")) or _normalize_edit_value(look["use_case"]),
        "priority": _normalize_edit_value(look["priority"]),
        "status": _normalize_edit_value(payload.get("status")) or _normalize_edit_value(look["status"]) or "Active",
        "owner": _normalize_edit_value(payload.get("owner")) or _normalize_edit_value(look["owner"]) or "徐欣",
        "temp_min": payload.get("temp_min"),
        "temp_max": payload.get("temp_max"),
        "scene_tag_target": _normalize_edit_value(payload.get("scene_tag_target")) or _normalize_edit_value(look["scene_tag_target"]),
        "relax_center": _featured_look_relax_center(conn, int(look["id"])),
        "relax_span": _featured_look_relax_span(conn, int(look["id"])),
        "notes": _normalize_edit_value(payload.get("notes")) or _normalize_edit_value(look["notes"]),
    }
    conn.execute(
        """
        UPDATE featured_looks
        SET look_id = :look_id,
            use_case = :use_case,
            priority = :priority,
            status = :status,
            owner = :owner,
            temp_min = :temp_min,
            temp_max = :temp_max,
            scene_tag_target = :scene_tag_target,
            relax_center = :relax_center,
            relax_span = :relax_span,
            notes = :notes,
            updated_at = CURRENT_TIMESTAMP
        WHERE id = :id
        """,
        values,
    )
    _sync_featured_look_relax_metrics(conn, int(look["id"]))
    _export_featured_looks_workbooks(conn)
    saved = conn.execute("SELECT * FROM featured_looks WHERE id = ?", (int(look["id"]),)).fetchone()
    return _serialize_featured_look(conn, saved)


class WardrobeHandler(BaseHTTPRequestHandler):
    server_version = "WardrobeApp/0.1"

    def _send_common_security_headers(self) -> None:
        self.send_header("Cache-Control", "no-store")
        self.send_header("X-Content-Type-Options", "nosniff")
        parsed = urlparse(getattr(self, "path", "") or "")
        is_hermes_embed = (parse_qs(parsed.query).get("embed") or [""])[0] == "hermes"
        if is_hermes_embed:
            try:
                frame_ancestors = _hermes_plugin_frame_ancestors()
            except Exception:
                frame_ancestors = hermes_plugin.normalize_frame_ancestors(HERMES_PLUGIN_FRAME_ANCESTORS)
            self.send_header("Content-Security-Policy", f"frame-ancestors {' '.join(frame_ancestors)}")
        else:
            self.send_header("X-Frame-Options", "DENY")
        self.send_header("Referrer-Policy", "same-origin")

    def _send_json(
        self,
        payload: dict | list,
        status: int = 200,
        extra_headers: list[tuple[str, str]] | None = None,
    ) -> None:
        body = _json_bytes(payload)
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self._send_common_security_headers()
        for header, value in [*(extra_headers or []), *self._consume_pending_headers()]:
            self.send_header(header, value)
        self.end_headers()
        self.wfile.write(body)

    def _send_empty(
        self,
        status: int,
        extra_headers: list[tuple[str, str]] | None = None,
    ) -> None:
        self.send_response(status)
        self._send_common_security_headers()
        for header, value in [*(extra_headers or []), *self._consume_pending_headers()]:
            self.send_header(header, value)
        self.end_headers()

    def _send_file(self, path: Path) -> None:
        if not path.exists() or not path.is_file():
            self.send_error(HTTPStatus.NOT_FOUND)
            return
        content_type, _ = mimetypes.guess_type(str(path))
        raw = path.read_bytes()
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", _textual_content_type(content_type, path))
        self.send_header("Content-Length", str(len(raw)))
        self._send_common_security_headers()
        for header, value in self._consume_pending_headers():
            self.send_header(header, value)
        self.end_headers()
        self.wfile.write(raw)

    def _send_bytes(
        self,
        raw: bytes,
        content_type: str = "application/octet-stream",
        extra_headers: list[tuple[str, str]] | None = None,
    ) -> None:
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(raw)))
        self._send_common_security_headers()
        for header, value in [*(extra_headers or []), *self._consume_pending_headers()]:
            self.send_header(header, value)
        self.end_headers()
        self.wfile.write(raw)

    def _db(self) -> sqlite3.Connection:
        return connect()

    def _public_api_path(self, path: str) -> bool:
        return path in {"/api/auth/status", "/api/auth/login", "/api/auth/logout", "/api/app-version"}

    def _client_ip(self) -> str:
        forwarded_for = self.headers.get("CF-Connecting-IP") or self.headers.get("X-Forwarded-For", "")
        if forwarded_for:
            return forwarded_for.split(",")[0].strip()
        return self.client_address[0] if self.client_address else ""

    def _request_is_secure(self) -> bool:
        forwarded_proto = (self.headers.get("X-Forwarded-Proto") or self.headers.get("X-Forwarded-Scheme") or "").lower()
        if forwarded_proto == "https":
            return True
        cf_visitor = self.headers.get("CF-Visitor", "")
        return '"scheme":"https"' in cf_visitor.lower()

    def _session_cookie_header(self, session_id: str, max_age: int | None = None) -> tuple[str, str]:
        if self._request_is_secure():
            parts = [f"{AUTH_COOKIE_NAME}={session_id}", "Path=/", "HttpOnly", "SameSite=None", "Secure"]
        else:
            parts = [f"{AUTH_COOKIE_NAME}={session_id}", "Path=/", "HttpOnly", "SameSite=Lax"]
        if max_age is not None:
            parts.append(f"Max-Age={max_age}")
        return ("Set-Cookie", "; ".join(parts))

    def _queue_session_refresh(self, session_id: str) -> None:
        if not session_id:
            return
        self._pending_auth_headers = [self._session_cookie_header(session_id, AUTH_SESSION_IDLE_SECONDS)]

    def _request_session_id(self) -> str:
        header_session = _normalize_edit_value(self.headers.get("X-Wardrobe-Session"))
        if header_session:
            return header_session
        parsed = urlparse(getattr(self, "path", "") or "")
        query_session = _normalize_edit_value((parse_qs(parsed.query).get("plugin_session") or [""])[0])
        if query_session:
            return query_session
        return _cookie_value(self.headers.get("Cookie", ""), AUTH_COOKIE_NAME)

    def _request_cookie_session_id(self) -> str:
        return _cookie_value(self.headers.get("Cookie", ""), AUTH_COOKIE_NAME)

    def _consume_pending_headers(self) -> list[tuple[str, str]]:
        pending = list(getattr(self, "_pending_auth_headers", []) or [])
        self._pending_auth_headers = []
        return pending

    def _allowed_request_hosts(self) -> set[str]:
        hosts = {_normalize_netloc(self.headers.get("Host", ""))}
        hosts.update(_normalize_netloc(value) for value in AUTH_ALLOWED_ORIGINS)
        return {host for host in hosts if host}

    def _verify_request_origin(self) -> bool:
        origin = self.headers.get("Origin", "").strip()
        referer = self.headers.get("Referer", "").strip()
        candidate = origin or referer
        if not candidate:
            self._send_json({"error": "forbidden_origin"}, status=403)
            return False
        if _normalize_netloc(candidate) not in self._allowed_request_hosts():
            self._send_json({"error": "forbidden_origin"}, status=403)
            return False
        return True

    def _authenticated_username(self) -> str:
        conn = self._db()
        try:
            return _session_username_by_id(conn, self._request_session_id())
        finally:
            conn.close()

    def _authorize_api(self, path: str) -> bool:
        if self._public_api_path(path):
            return True
        username = self._authenticated_username()
        if username:
            self._queue_session_refresh(self._request_cookie_session_id())
            return True
        self._send_json({"error": "unauthorized"}, status=401)
        return False

    def _authorize_stateful_api(self, path: str) -> bool:
        if not self._authorize_api(path):
            return False
        return self._verify_request_origin()

    def _program_api_context(self, conn: sqlite3.Connection, required_scope: str) -> dict | None:
        context, error_payload, status = _api_token_context(
            conn,
            self.headers.get("Authorization", ""),
            required_scope,
        )
        if error_payload is not None:
            self._send_json(error_payload, status=status)
            return None
        return context

    def _owner_registration_context(self, conn: sqlite3.Connection) -> dict | None:
        authorization = self.headers.get("Authorization", "")
        if _normalize_edit_value(authorization):
            return self._program_api_context(conn, "owners:write")
        username = _session_username_by_id(conn, self._request_session_id())
        if username and _user_is_admin(username):
            if not self._verify_request_origin():
                return None
            return {"owner": username, "scopes": ["admin:*"], "token_id": None, "source": "admin_session"}
        self._send_json({"error": "invalid_token", "message": "Missing owner registration credential."}, status=401)
        return None

    def _handle_program_api_get(self, path: str, query: dict[str, list[str]]) -> None:
        if path == "/api/v1/hermes/plugin/manifest":
            conn = self._db()
            try:
                requested_origin = _normalize_edit_value(
                    (query.get("origin") or query.get("hermes_origin") or query.get("frame_ancestor") or [""])[0]
                )
                frame_ancestors = _hermes_plugin_frame_ancestors(conn)
                self._send_json(
                    _hermes_plugin_manifest_payload(
                        _request_external_base_url(self),
                        frame_ancestors=frame_ancestors,
                        requested_frame_ancestor=requested_origin,
                    )
                )
            except hermes_plugin.HermesPluginError as exc:
                self._send_json({"error": exc.code, "message": exc.message}, status=400)
            finally:
                conn.close()
            return
        if path == "/api/v1/hermes/plugin/session":
            conn = self._db()
            try:
                session_id = self._request_session_id()
                username = _session_username_by_id(conn, session_id)
                if not username:
                    self._send_json({"error": "unauthorized"}, status=401)
                    return
                self._send_json(_api_hermes_plugin_session(conn, session_id))
            except hermes_plugin.HermesPluginError as exc:
                self._send_json({"error": exc.code, "message": exc.message}, status=400)
            finally:
                conn.close()
            return
        conn = self._db()
        try:
            sync_resource_prefix = f"{program_api_sync.SYNC_RESOURCE_ENDPOINT_PREFIX}/"
            if (
                path == "/api/v1/sync/outfit-context/manifest"
                or path == "/api/v1/sync/outfit-context/bundle"
                or path.startswith(sync_resource_prefix)
            ):
                context = self._program_api_context(conn, "sync:read")
                if context is None:
                    return
                try:
                    owner = _api_sync_owner_from_query(context, query)
                except PermissionError as exc:
                    self._send_json({"error": "forbidden_owner", "message": str(exc)}, status=403)
                    return
                if path == "/api/v1/sync/outfit-context/bundle":
                    self._send_json(
                        {
                            "error": "bundle_sync_removed",
                            "message": "Use /api/v1/sync/outfit-context/manifest and /api/v1/sync/outfit-context/resources/{name}.",
                        },
                        status=HTTPStatus.GONE,
                    )
                    return
                if path == "/api/v1/sync/outfit-context/manifest":
                    manifest = _api_sync_outfit_context_manifest(conn, owner)
                    etag_header = ("ETag", _api_quote_etag(manifest["etag"]))
                    if _api_etag_matches(self.headers.get("If-None-Match", ""), manifest["etag"]):
                        self._send_empty(HTTPStatus.NOT_MODIFIED, extra_headers=[etag_header])
                        return
                    self._send_json(manifest, extra_headers=[etag_header])
                    return
                resource_name = _normalize_edit_value(unquote(path.removeprefix(sync_resource_prefix)))
                if not resource_name or "/" in resource_name:
                    self._send_json({"error": "invalid_sync_resource"}, status=400)
                    return
                try:
                    resource_payload = _api_sync_outfit_context_resource(conn, owner, resource_name)
                except ValueError:
                    self._send_json({"error": "unknown_sync_resource", "resource": resource_name}, status=404)
                    return
                etag_header = ("ETag", _api_quote_etag(resource_payload["checksum"]))
                if _api_etag_matches(self.headers.get("If-None-Match", ""), resource_payload["checksum"]):
                    self._send_empty(HTTPStatus.NOT_MODIFIED, extra_headers=[etag_header])
                    return
                self._send_json(resource_payload, extra_headers=[etag_header])
                return
            photo_media_target = _api_item_photo_media_target(path)
            if photo_media_target is not None:
                context = self._program_api_context(conn, "items:read")
                if context is None:
                    return
                code, photo_id, media_variant = photo_media_target
                try:
                    item = _api_find_item_for_photo_read(conn, context, code)
                except ValueError as exc:
                    self._send_json({"error": "invalid_payload", "message": str(exc)}, status=400)
                    return
                except KeyError:
                    self._send_json({"error": "item_not_found"}, status=404)
                    return
                except PermissionError:
                    self._send_json({"error": "forbidden_owner"}, status=403)
                    return
                photo = _api_item_photo_content_row(conn, int(item["id"]), photo_id)
                if photo is None:
                    self._send_json({"error": "photo_not_found"}, status=404)
                    return
                if media_variant == "thumbnail":
                    sent = _send_item_photo_thumbnail_payload(self, photo, _normalize_edit_value(item["code"]), query)
                else:
                    sent = _send_item_photo_row_payload(self, photo, query)
                if not sent:
                    self._send_json({"error": "photo_not_found"}, status=404)
                return
            if path == "/api/v1/items" or path.startswith("/api/v1/items/"):
                context = self._program_api_context(conn, "items:read")
                if context is None:
                    return
                if path.startswith("/api/v1/items/"):
                    code = _normalize_edit_value(path.removeprefix("/api/v1/items/"))
                    if not code:
                        self._send_json({"error": "missing_item_code"}, status=400)
                        return
                    row = conn.execute(
                        """
                        SELECT items.*,
                               (SELECT COUNT(*) FROM photos WHERE photos.item_id = items.id) AS photo_count
                        FROM items
                        WHERE code = ?
                        """,
                        (code,),
                    ).fetchone()
                    if row is None:
                        self._send_json({"error": "item_not_found"}, status=404)
                        return
                    if not _api_owner_allowed(context, row["owner"], "items:read:any"):
                        self._send_json({"error": "forbidden_owner"}, status=403)
                        return
                    self._send_json({"item": _api_item_detail_payload(conn, row)})
                    return
                where: list[str] = []
                params: list[object] = []
                requested_owner = _normalize_edit_value((query.get("owner") or [""])[0])
                if requested_owner:
                    if not _api_owner_allowed(context, requested_owner, "items:read:any"):
                        self._send_json({"error": "forbidden_owner"}, status=403)
                        return
                    where.append("COALESCE(owner, '') = ?")
                    params.append(requested_owner)
                elif not _api_context_has_scope(context, "items:read:any"):
                    where.append("COALESCE(owner, '') = ?")
                    params.append(_api_context_owner(context))
                search = _normalize_edit_value((query.get("q") or [""])[0])
                brand = _normalize_edit_value((query.get("brand") or [""])[0])
                status_value = _normalize_edit_value((query.get("status") or [""])[0])
                loc = _normalize_edit_value((query.get("loc") or [""])[0])
                layer_role = _normalize_edit_value((query.get("layer_role") or [""])[0])
                if layer_role:
                    layer_role = _canonical_layer_role(layer_role)
                    if not layer_role:
                        self._send_json({"error": "invalid_layer_role"}, status=400)
                        return
                kind = _normalize_edit_value((query.get("kind") or [""])[0]).lower()
                if search:
                    where.append("(code LIKE ? OR section LIKE ? OR brand LIKE ? OR material LIKE ?)")
                    like_value = f"%{search}%"
                    params.extend([like_value, like_value, like_value, like_value])
                if brand:
                    where.append("brand = ?")
                    params.append(brand)
                if status_value:
                    where.append("status = ?")
                    params.append(status_value)
                if loc:
                    where.append("loc = ?")
                    params.append(loc)
                if layer_role:
                    where.append("layer_role = ?")
                    params.append(layer_role)
                if kind == "watch":
                    where.append("layer_role = 'Watch'")
                elif kind == "wardrobe":
                    where.append("COALESCE(layer_role, '') <> 'Watch'")
                try:
                    limit = min(max(int((query.get("limit") or ["200"])[0]), 1), 500)
                except Exception:
                    limit = 200
                where_sql = " AND ".join(where) if where else "1 = 1"
                rows = conn.execute(
                    f"""
                    SELECT items.*,
                           (SELECT COUNT(*) FROM photos WHERE photos.item_id = items.id) AS photo_count
                    FROM items
                    WHERE {where_sql}
                    ORDER BY COALESCE(owner, ''), COALESCE(layer_role, ''), COALESCE(brand, ''), COALESCE(section, '')
                    LIMIT ?
                    """,
                    [*params, limit],
                ).fetchall()
                self._send_json({
                    "items": [_api_item_payload(row) for row in rows],
                    "count": len(rows),
                    "limit": limit,
                    "owner": requested_owner or (_api_context_owner(context) if not _api_context_has_scope(context, "items:read:any") else ""),
                })
                return
            self._send_json({"error": "not_found"}, status=404)
        finally:
            conn.close()

    def _handle_program_api_post(self, path: str, query: dict[str, list[str]] | None = None) -> None:
        query = query or {}
        conn = self._db()
        try:
            if path == "/api/v1/hermes/plugin/frame-ancestors":
                context = self._owner_registration_context(conn)
                if context is None:
                    return
                payload = _parse_json(self)
                try:
                    response = _api_register_hermes_plugin_frame_ancestors(conn, payload)
                except hermes_plugin.HermesPluginError as exc:
                    self._send_json({"error": exc.code, "message": exc.message}, status=400)
                    return
                conn.commit()
                self._send_json(response, status=201)
                return
            if path == "/api/v1/hermes/plugin/launch":
                context = self._program_api_context(conn, "sync:read")
                if context is None:
                    return
                payload = _parse_json(self)
                try:
                    response = _api_create_hermes_plugin_launch_token(conn, payload, context)
                except hermes_plugin.HermesPluginError as exc:
                    status = 403 if exc.code == "workspace_token_mismatch" else 400
                    if exc.code == "workspace_not_registered":
                        status = 404
                    self._send_json({"error": exc.code, "message": exc.message}, status=status)
                    return
                conn.commit()
                self._send_json(response, status=201)
                return
            if path == "/api/v1/hermes/plugin/workspaces":
                context = self._owner_registration_context(conn)
                if context is None:
                    return
                payload = _parse_json(self)
                try:
                    response = _api_register_hermes_plugin_workspace(conn, payload)
                except hermes_plugin.HermesPluginError as exc:
                    status = 409 if exc.code in {"access_key_already_registered", "owner_key_exists"} else 400
                    self._send_json({"error": exc.code, "message": exc.message}, status=status)
                    return
                conn.commit()
                self._send_json(response, status=201)
                return
            if path.startswith("/api/v1/items/") and path.endswith("/photos/order"):
                context = self._program_api_context(conn, "items:write")
                if context is None:
                    return
                code = _normalize_edit_value(path[len("/api/v1/items/") : -len("/photos/order")])
                payload = _parse_json(self)
                if not isinstance(payload, dict):
                    self._send_json({"error": "invalid_payload", "message": "JSON body must be an object."}, status=400)
                    return
                dry_run = _coerce_api_bool(payload.get("dry_run"), False)
                if dry_run:
                    response, status_code = _api_item_photo_order_preview(conn, context, code, payload)
                    self._send_json(response, status=status_code)
                    return
                response, status_code = _api_reorder_item_photos(conn, context, code, payload)
                conn.commit()
                self._send_json(response, status=status_code)
                return
            if path.startswith("/api/v1/items/") and path.endswith("/photos"):
                context = self._program_api_context(conn, "items:write")
                if context is None:
                    return
                code = _normalize_edit_value(path[len("/api/v1/items/") : -len("/photos")])
                payload = _api_parse_item_photo_request_payload(self, query)
                dry_run = _coerce_api_bool(payload.get("dry_run"), False)
                if dry_run:
                    response, status_code = _api_item_photo_write_preview(conn, context, code, payload)
                    self._send_json(response, status=status_code)
                    return
                request_hash = _api_request_hash(hashable_program_payload(payload))
                idempotency_key = _normalize_edit_value(self.headers.get("Idempotency-Key")) or _normalize_edit_value(payload.get("external_id"))
                replay_payload, replay_status = _api_idempotency_response(conn, context, idempotency_key, request_hash)
                if replay_payload is not None:
                    self._send_json(replay_payload, status=replay_status)
                    return
                response, status_code = _api_write_item_photos(conn, context, code, payload)
                _store_api_idempotency_response(conn, context, idempotency_key, request_hash, response, status_code)
                conn.commit()
                self._send_json(response, status=status_code)
                return
            if path == "/api/v1/items":
                context = self._program_api_context(conn, "items:write")
                if context is None:
                    return
                payload = _api_parse_item_request_payload(self)
                dry_run = _coerce_api_bool(payload.get("dry_run"), False)
                if dry_run:
                    response, status_code = _api_item_write_preview(conn, context, payload)
                    self._send_json(response, status=status_code)
                    return
                request_hash = _api_request_hash(hashable_program_payload(payload))
                idempotency_key = (
                    _normalize_edit_value(self.headers.get("Idempotency-Key"))
                    or _normalize_edit_value(payload.get("external_id"))
                    or _normalize_edit_value((payload.get("item") or {}).get("external_id") if isinstance(payload.get("item"), dict) else "")
                )
                replay_payload, replay_status = _api_idempotency_response(conn, context, idempotency_key, request_hash)
                if replay_payload is not None:
                    self._send_json(replay_payload, status=replay_status)
                    return
                response, status_code = _api_write_item(conn, context, payload)
                _store_api_idempotency_response(conn, context, idempotency_key, request_hash, response, status_code)
                conn.commit()
                self._send_json(response, status=status_code)
                return
            if path == "/api/v1/history/outfits":
                context = self._program_api_context(conn, "history:write")
                if context is None:
                    return
                payload = _parse_json(self)
                if not isinstance(payload, dict):
                    self._send_json({"error": "invalid_payload", "message": "JSON body must be an object."}, status=400)
                    return
                dry_run = _coerce_api_bool(payload.get("dry_run"), False)
                request_hash = _api_request_hash(payload)
                idempotency_key = ""
                if not dry_run:
                    idempotency_key = _normalize_edit_value(self.headers.get("Idempotency-Key")) or _normalize_edit_value(payload.get("external_id"))
                    replay_payload, replay_status = _api_idempotency_response(conn, context, idempotency_key, request_hash)
                    if replay_payload is not None:
                        self._send_json(replay_payload, status=replay_status)
                        return
                owner = _api_context_owner(context)
                if _api_context_has_scope(context, "history:write:any"):
                    owner = _normalize_edit_value(payload.get("owner")) or owner
                if not owner:
                    self._send_json({"error": "token_owner_required"}, status=403)
                    return
                wear_date = _normalize_date_compare(payload.get("wear_date"))
                if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", wear_date or ""):
                    self._send_json({"error": "invalid_wear_date"}, status=400)
                    return
                mode = _normalize_edit_value(payload.get("mode")).lower() or "create_only"
                if mode not in {"create_only", "replace", "upsert"}:
                    self._send_json({"error": "invalid_mode"}, status=400)
                    return
                source = _normalize_edit_value(payload.get("source")) or "api"
                external_id = _normalize_edit_value(payload.get("external_id"))
                source_path = f"api:{source}:{external_id}" if external_id else f"api:{source}"
                scoped_payload = dict(payload)
                scoped_payload["owner"] = owner
                scoped_payload["wear_date"] = wear_date
                scoped_payload["source_path"] = source_path
                wear_mode = _normalize_edit_value(scoped_payload.get("wear_mode")) or "normal"
                resolved_entries = _api_resolve_outfit_item_entries(
                    conn,
                    list(scoped_payload.get("items") or []),
                    context,
                    owner,
                    wear_mode,
                    wear_date,
                )
                existing_outfit = conn.execute(
                    "SELECT * FROM outfits WHERE wear_date = ? AND COALESCE(owner, '') = ? ORDER BY id DESC LIMIT 1",
                    (wear_date, owner),
                ).fetchone()
                if existing_outfit is not None and mode == "create_only":
                    self._send_json(
                        {
                            "error": "existing_outfit",
                            "dry_run": dry_run,
                            "owner": owner,
                            "outfit_id": int(existing_outfit["id"]),
                            "wear_date": wear_date,
                        },
                        status=409,
                    )
                    return
                if dry_run:
                    affected_item_ids = sorted({int(entry["item_id"]) for entry in resolved_entries})
                    export_tasks = _wear_aggregate_export_tasks_for_item_ids(conn, affected_item_ids)
                    self._send_json(
                        {
                            "saved": False,
                            "dry_run": True,
                            "action": "replaced" if existing_outfit is not None else "created",
                            "owner": owner,
                            "outfit_id": int(existing_outfit["id"]) if existing_outfit is not None else None,
                            "wear_date": wear_date,
                            "wearcount_linked": False,
                            "affected_item_ids": affected_item_ids,
                            "exports_pending": bool(export_tasks),
                            "export_tasks": sorted(export_tasks),
                        }
                    )
                    return
                if existing_outfit is not None:
                    saved, affected_item_ids = _upsert_outfit_daily_update(
                        conn,
                        int(existing_outfit["id"]),
                        _normalize_edit_value(existing_outfit["wear_date"]),
                        scoped_payload,
                        resolved_entries,
                        owner,
                        existing_outfit,
                    )
                    status_code = 200
                    action = "replaced"
                else:
                    saved, affected_item_ids = _create_outfit_daily_update(conn, scoped_payload, resolved_entries)
                    status_code = 201
                    action = "created"
                export_tasks = _wear_aggregate_export_tasks_for_item_ids(conn, affected_item_ids)
                response = {
                    "saved": True,
                    "action": action,
                    "owner": owner,
                    "outfit_id": int(saved["id"]),
                    "wear_date": _normalize_edit_value(saved["wear_date"]),
                    "wearcount_linked": True,
                    "affected_item_ids": affected_item_ids,
                    "exports_pending": bool(export_tasks),
                    "export_tasks": sorted(export_tasks),
                }
                _store_api_idempotency_response(conn, context, idempotency_key, request_hash, response, status_code)
                conn.commit()
                _schedule_item_export_tasks(export_tasks)
                self._send_json(response, status=status_code)
                return
            self._send_json({"error": "not_found"}, status=404)
        except PermissionError as exc:
            conn.rollback()
            self._send_json({"error": "forbidden", "message": str(exc)}, status=403)
        except (KeyError, ValueError) as exc:
            conn.rollback()
            self._send_json({"error": "invalid_payload", "message": str(exc)}, status=400)
        except RuntimeError as exc:
            conn.rollback()
            self._send_json({"error": "conflict", "message": str(exc)}, status=409)
        except Exception as exc:
            conn.rollback()
            self._send_json({"error": "db_or_export_failed", "message": str(exc)}, status=500)
        finally:
            conn.close()

    def log_message(self, format: str, *args) -> None:
        sys.stdout.write("%s - - [%s] %s\n" % (self.address_string(), self.log_date_time_string(), format % args))

    def _redirect(
        self,
        location: str,
        status: int = HTTPStatus.FOUND,
        extra_headers: list[tuple[str, str]] | None = None,
    ) -> None:
        self.send_response(status)
        self.send_header("Location", location)
        self.send_header("Cache-Control", "no-store, no-cache, must-revalidate, max-age=0")
        self.send_header("Pragma", "no-cache")
        self.send_header("Expires", "0")
        for header, value in extra_headers or []:
            self.send_header(header, value)
        self.end_headers()

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        path = unquote(parsed.path)
        if path == "/api/app-version":
            self._send_json({"version": APP_WEB_VERSION})
            return
        if path == "/api/auth/status":
            conn = self._db()
            try:
                cookie_session_id = self._request_cookie_session_id()
                request_session_id = self._request_session_id()
                username = _session_username_by_id(conn, request_session_id)
                extra_headers = []
                if username and cookie_session_id:
                    extra_headers.append(
                        self._session_cookie_header(
                            cookie_session_id,
                            AUTH_SESSION_IDLE_SECONDS,
                        )
                    )
                self._send_json(
                    {
                        "authenticated": bool(username),
                        "username": username,
                        "is_admin": _user_is_admin(username),
                        "accounts": _owner_accounts(conn),
                    },
                    extra_headers=extra_headers,
                )
            finally:
                conn.close()
            return
        if path.startswith("/api/"):
            if path.startswith("/api/v1/"):
                self._handle_program_api_get(path, parse_qs(parsed.query))
                return
            if not self._authorize_api(path):
                return
            self._handle_api_get(path, parse_qs(parsed.query))
            return
        if path.startswith("/media/"):
            self._send_file(MEDIA_DIR / path.removeprefix("/media/"))
            return
        if path == "/" or path == "/index.html":
            query_items = parse_qsl(parsed.query, keep_blank_values=True)
            query_map = dict(query_items)
            launch_token = _normalize_edit_value(query_map.get("launch"))
            if launch_token:
                conn = self._db()
                try:
                    launch = _consume_hermes_plugin_launch(conn, launch_token)
                    session_id = _create_session(conn, launch["owner"])
                    appearance = hermes_plugin.bind_session_appearance(conn, session_id, launch)
                    conn.commit()
                    filtered = [
                        (key, value)
                        for key, value in query_items
                        if key not in {"launch", "pluginTheme", "pluginFontSize"}
                    ]
                    if not any(key == "embed" and value == "hermes" for key, value in filtered):
                        filtered.append(("embed", "hermes"))
                    if any(key == "embed" and value == "hermes" for key, value in filtered):
                        filtered.append(("plugin_session", session_id))
                        filtered.append(("pluginTheme", appearance["theme"]))
                        filtered.append(("pluginFontSize", appearance["fontSize"]))
                    redirect_target = urlunparse(("", "", path, "", urlencode(filtered), ""))
                    self._redirect(
                        redirect_target,
                        status=HTTPStatus.FOUND,
                        extra_headers=[self._session_cookie_header(session_id, AUTH_SESSION_IDLE_SECONDS)],
                    )
                except hermes_plugin.HermesPluginError as exc:
                    conn.rollback()
                    self._send_json({"error": exc.code, "message": exc.message}, status=401)
                finally:
                    conn.close()
                return
            current_version = dict(query_items).get("_appv", "")
            if current_version != APP_WEB_VERSION:
                filtered = [(key, value) for key, value in query_items if key != "_appv"]
                filtered.append(("_appv", APP_WEB_VERSION))
                redirect_target = urlunparse(("", "", path, "", urlencode(filtered), ""))
                self._redirect(redirect_target, status=HTTPStatus.FOUND)
                return
            self._send_file(WEB_DIR / "index.html")
            return
        static_path = WEB_DIR / path.lstrip("/")
        if static_path.exists():
            self._send_file(static_path)
            return
        self.send_error(HTTPStatus.NOT_FOUND)

    def _handle_api_get(self, path: str, query: dict[str, list[str]]) -> None:
        conn = self._db()
        try:
            username = self._authenticated_username() if not self._public_api_path(path) else ""
            if path.startswith("/api/outfit-photos/") and path.endswith("/content"):
                photo_id = int(path.split("/")[-2])
                if username:
                    outfit = conn.execute(
                        """
                        SELECT outfits.*
                        FROM outfit_photos
                        JOIN outfits ON outfits.id = outfit_photos.outfit_id
                        WHERE outfit_photos.id = ?
                        """,
                        (photo_id,),
                    ).fetchone()
                    if not _outfit_row_viewable(outfit, username):
                        self._send_json({"error": "forbidden"}, status=403)
                        return
                photo = conn.execute(
                    "SELECT mime_type, data, thumb_data, thumb_mime_type FROM outfit_photos WHERE id = ?",
                    (photo_id,),
                ).fetchone()
                if photo is None or photo["data"] is None:
                    self._send_json({"error": "photo_not_found"}, status=404)
                    return
                wants_thumb = (query.get("thumb") or [""])[0] in {"1", "true", "yes"}
                if wants_thumb:
                    thumbnail = _cached_outfit_photo_thumbnail(
                        conn,
                        photo_id,
                        photo["data"],
                        photo["mime_type"] or "application/octet-stream",
                        cached_raw=photo["thumb_data"],
                        cached_mime=photo["thumb_mime_type"],
                    )
                    if thumbnail is not None:
                        thumb_raw, thumb_mime = thumbnail
                        self._send_bytes(thumb_raw, content_type=thumb_mime)
                        return
                _send_photo_payload(self, photo["data"], photo["mime_type"] or "application/octet-stream", query)
                return
            if path.startswith("/api/featured-look-photos/") and path.endswith("/content"):
                photo_id = int(path.split("/")[-2])
                if username:
                    look = conn.execute(
                        """
                        SELECT featured_looks.*
                        FROM featured_look_photos
                        JOIN featured_looks ON featured_looks.id = featured_look_photos.featured_look_id
                        WHERE featured_look_photos.id = ?
                        """,
                        (photo_id,),
                    ).fetchone()
                    if not _featured_look_row_viewable(look, username):
                        self._send_json({"error": "forbidden"}, status=403)
                        return
                photo = conn.execute(
                    "SELECT mime_type, data FROM featured_look_photos WHERE id = ?",
                    (photo_id,),
                ).fetchone()
                if photo is None or photo["data"] is None:
                    self._send_json({"error": "photo_not_found"}, status=404)
                    return
                _send_photo_payload(self, photo["data"], photo["mime_type"] or "application/octet-stream", query)
                return
            if path == "/api/auth/status":
                cookie_session_id = self._request_cookie_session_id()
                request_session_id = self._request_session_id()
                username = _session_username_by_id(conn, request_session_id)
                extra_headers = []
                if username and cookie_session_id:
                    extra_headers.append(
                        self._session_cookie_header(
                            cookie_session_id,
                            AUTH_SESSION_IDLE_SECONDS,
                        )
                    )
                self._send_json(
                    {
                        "authenticated": bool(username),
                        "username": username,
                        "is_admin": _user_is_admin(username),
                        "accounts": _owner_accounts(conn),
                    },
                    extra_headers=extra_headers,
                )
                return
            if path.startswith("/api/photos/") and path.endswith("/content"):
                photo_id = int(path.split("/")[-2])
                if username and not _photo_item_viewable(conn, photo_id, username):
                    self._send_json({"error": "forbidden"}, status=403)
                    return
                photo = conn.execute(
                    "SELECT mime_type, data, file_name FROM photos WHERE id = ?",
                    (photo_id,),
                ).fetchone()
                if photo is None:
                    self._send_json({"error": "photo_not_found"}, status=404)
                    return
                if photo["data"] is not None:
                    _send_photo_payload(self, photo["data"], photo["mime_type"] or "application/octet-stream", query)
                    return
                target = MEDIA_DIR / photo["file_name"]
                if target.exists():
                    if (query.get("thumb") or [""])[0] in {"1", "true", "yes"}:
                        try:
                            thumbnail = _thumbnail_bytes_from_payload(target.read_bytes(), mimetypes.guess_type(target.name)[0] or "application/octet-stream")
                        except OSError:
                            thumbnail = None
                        if thumbnail is not None:
                            thumb_raw, thumb_mime = thumbnail
                            self._send_bytes(thumb_raw, content_type=thumb_mime)
                            return
                    self._send_file(target)
                    return
                self._send_json({"error": "photo_not_found"}, status=404)
                return
            if path == "/api/bootstrap-status":
                item_where, item_params = _item_owner_scope_where(username)
                imports = conn.execute(
                    "SELECT import_type, source_path, summary_json, created_at FROM imports ORDER BY id DESC LIMIT 10"
                ).fetchall()
                item_count = conn.execute(f"SELECT COUNT(*) AS count FROM items WHERE {item_where}", item_params).fetchone()["count"]
                try:
                    db_size_bytes = DB_PATH.stat().st_size
                except OSError:
                    db_size_bytes = 0
                self._send_json(
                        {
                            "item_count": item_count,
                            "imports": [dict(row) for row in imports],
                            "default_import_dir": str(DEFAULT_IMPORT_DIR),
                            "db_size_bytes": db_size_bytes,
                        }
                    )
                return
            if path == "/api/items":
                owner_where, owner_params = _item_owner_scope_where(username)
                where = [owner_where]
                params: list = list(owner_params)
                search = (query.get("q") or [""])[0].strip()
                brand = (query.get("brand") or [""])[0].strip()
                owner = (query.get("owner") or [""])[0].strip()
                loc = (query.get("loc") or [""])[0].strip()
                layer_role = (query.get("layer_role") or [""])[0].strip()
                if layer_role:
                    layer_role = _canonical_layer_role(layer_role)
                    if not layer_role:
                        self._send_json({"error": "invalid_layer_role"}, status=400)
                        return
                if search:
                    where.append("(code LIKE ? OR section LIKE ? OR brand LIKE ? OR material LIKE ?)")
                    like_value = f"%{search}%"
                    params.extend([like_value, like_value, like_value, like_value])
                if brand:
                    where.append("brand = ?")
                    params.append(brand)
                if owner:
                    where.append("owner = ?")
                    params.append(owner)
                if loc:
                    where.append("loc = ?")
                    params.append(loc)
                if layer_role:
                    where.append("layer_role = ?")
                    params.append(layer_role)
                rows = conn.execute(
                    f"""
                    WITH photo_rollup AS (
                        SELECT item_id, COUNT(*) AS photo_count
                        FROM photos
                        GROUP BY item_id
                    )
                    SELECT items.*,
                           COALESCE(photo_rollup.photo_count, 0) AS photo_count
                    FROM items
                    LEFT JOIN photo_rollup ON photo_rollup.item_id = items.id
                    WHERE {' AND '.join(where)}
                    ORDER BY COALESCE(owner, ''), COALESCE(layer_role, ''), COALESCE(wear_total, 0) ASC, section ASC
                    """,
                    params,
                ).fetchall()
                self._send_json([dict(row) for row in rows])
                return
            if path.startswith("/api/items/") and path.endswith("/outfits"):
                item_id = int(path.split("/")[3])
                summary_only = (query.get("summary") or [""])[0].strip().lower() in {"1", "true", "yes"}
                try:
                    self._send_json(_item_related_outfits(conn, item_id, username, summary_only=summary_only))
                except KeyError:
                    self._send_json({"error": "item_not_found"}, status=404)
                except PermissionError:
                    self._send_json({"error": "forbidden"}, status=403)
                return
            if path.startswith("/api/items/") and path.endswith("/featured-looks"):
                item_id = int(path.split("/")[3])
                try:
                    self._send_json(_item_related_featured_looks(conn, item_id, username))
                except KeyError:
                    self._send_json({"error": "item_not_found"}, status=404)
                except PermissionError:
                    self._send_json({"error": "forbidden"}, status=403)
                return
            if path.startswith("/api/items/"):
                item_id = int(path.split("/")[-1])
                item = _item_with_photos(conn, item_id)
                if item is None:
                    self._send_json({"error": "item_not_found"}, status=404)
                    return
                if username and not _item_row_viewable(item, username):
                    self._send_json({"error": "forbidden"}, status=403)
                    return
                self._send_json(item)
                return
            if path.startswith("/api/outfits/") and path.count("/") == 3:
                outfit_id = int(path.split("/")[3])
                outfit = conn.execute("SELECT * FROM outfits WHERE id = ?", (outfit_id,)).fetchone()
                if outfit is None:
                    self._send_json({"error": "outfit_not_found"}, status=404)
                    return
                if username and not _outfit_row_viewable(outfit, username):
                    self._send_json({"error": "forbidden"}, status=403)
                    return
                self._send_json(_serialize_outfit(conn, outfit, username))
                return
            if path == "/api/outfits":
                outfit_where, outfit_params = _owner_read_sql("owner", username)
                summary_only = (query.get("summary") or [""])[0].strip().lower() in {"1", "true", "yes"}
                rows = conn.execute(
                    f"""
                    SELECT
                        outfits.*,
                        (SELECT COUNT(*) FROM outfit_photos WHERE outfit_photos.outfit_id = outfits.id) AS photo_count
                    FROM outfits
                    WHERE {outfit_where}
                    ORDER BY wear_date DESC
                    """,
                    outfit_params,
                ).fetchall()
                if summary_only:
                    self._send_json([_serialize_outfit_summary(row) for row in rows])
                else:
                    self._send_json([_serialize_outfit(conn, row, username) for row in rows])
                return
            if path == "/api/featured-looks":
                where_sql, where_params = _owner_read_sql("owner", username)
                rows = conn.execute(
                    f"""
                    SELECT *
                    FROM featured_looks
                    WHERE COALESCE(status, '') <> 'Archived'
                      AND {where_sql}
                    ORDER BY
                        COALESCE(created_at, '') DESC,
                        id DESC
                    """,
                    where_params,
                ).fetchall()
                visible_rows = [row for row in rows if _featured_look_row_viewable(row, username)]
                wear_counts = _featured_look_wear_counts(conn, {
                    _normalize_edit_value(row["owner"])
                    for row in visible_rows
                    if _normalize_edit_value(row["owner"])
                })
                self._send_json([_serialize_featured_look(conn, row, wear_counts, username) for row in visible_rows])
                return
            if path == "/api/meta/options":
                owner_where, owner_params = _item_owner_scope_where(username)
                actual_brands = [
                    row["brand"]
                    for row in conn.execute(
                        f"""
                        SELECT DISTINCT brand
                        FROM items
                        WHERE {owner_where}
                          AND brand <> ''
                        ORDER BY brand
                        """,
                        owner_params,
                    )
                ]
                actual_owners = [
                    row["owner"]
                    for row in conn.execute(
                        f"SELECT DISTINCT owner FROM items WHERE {owner_where} AND owner <> '' ORDER BY owner",
                        owner_params,
                    )
                ]
                actual_locs = [
                    row["loc"]
                    for row in conn.execute(
                        f"SELECT DISTINCT loc FROM items WHERE {owner_where} AND loc <> '' ORDER BY loc",
                        owner_params,
                    )
                ]
                clothing_brands = [
                    row["brand"]
                    for row in conn.execute(
                        f"""
                        SELECT DISTINCT brand
                        FROM items
                        WHERE {owner_where}
                          AND brand <> ''
                          AND COALESCE(layer_role, '') <> 'Watch'
                        ORDER BY brand
                        """,
                        owner_params,
                    )
                ]
                watch_brands = [
                    row["brand"]
                    for row in conn.execute(
                        f"""
                        SELECT DISTINCT brand
                        FROM items
                        WHERE {owner_where}
                          AND brand <> ''
                          AND COALESCE(layer_role, '') = 'Watch'
                        ORDER BY brand
                        """,
                        owner_params,
                    )
                ]
                catalog_owners = _merged_option_values(actual_owners, _option_catalog_values(conn, "owner"))
                if not _user_can_manage_catalog(username):
                    normalized_username = _normalize_edit_value(username)
                    catalog_owners = [
                        value
                        for value in catalog_owners
                        if _normalize_edit_value(value) == normalized_username
                    ]
                    if normalized_username and not catalog_owners:
                        catalog_owners = [normalized_username]
                result = {
                    "brands": actual_brands,
                    "owners": actual_owners,
                    "locs": actual_locs,
                    "roles": [row["layer_role"] for row in conn.execute(f"SELECT DISTINCT layer_role FROM items WHERE {owner_where} AND layer_role <> '' ORDER BY layer_role", owner_params)],
                    "relax_levels": [
                        row["relax_index"]
                        for row in conn.execute(
                            f"SELECT DISTINCT relax_index FROM items WHERE {owner_where} AND relax_index IS NOT NULL ORDER BY relax_index",
                            owner_params,
                        )
                    ],
                    "scene_tags": [
                        row["scene_tag"]
                        for row in conn.execute(
                            f"SELECT DISTINCT scene_tag FROM items WHERE {owner_where} AND scene_tag <> '' AND scene_tag <> 'Watch' ORDER BY scene_tag",
                            owner_params,
                        )
                    ],
                    "catalog_owners": catalog_owners,
                    "catalog_wardrobe_brands": _merged_option_values(clothing_brands, _option_catalog_values(conn, "wardrobe_brand")),
                    "catalog_watch_brands": _merged_option_values(watch_brands, _option_catalog_values(conn, "watch_brand")),
                }
                self._send_json(result)
                return
            self._send_json({"error": "not_found"}, status=404)
        except Exception as error:
            self._send_json(
                {
                    "error": "api_get_failed",
                    "message": str(error) or error.__class__.__name__,
                    "path": path,
                },
                status=500,
            )
        finally:
            conn.close()

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        path = unquote(parsed.path)
        if path == "/api/auth/login":
            if not self._verify_request_origin():
                return
            conn = self._db()
            try:
                payload = _parse_json(self)
                username = str(payload.get("username", "")).strip()
                password = str(payload.get("password", ""))
                client_ip = self._client_ip()
                rate_limited, retry_after = _login_rate_limited(client_ip, username)
                if rate_limited:
                    self._send_json(
                        {
                            "authenticated": False,
                            "error": "rate_limited",
                            "retry_after": retry_after,
                            "accounts": _owner_accounts(conn),
                        },
                        status=429,
                    )
                    return
                password_hash = _account_password_hash(conn, username)
                if not password_hash:
                    _register_login_failure(client_ip, username)
                    self._send_json(
                        {
                            "authenticated": False,
                            "error": "invalid_username",
                            "accounts": _owner_accounts(conn),
                        },
                        status=400,
                    )
                    return
                account = _account_state(conn, username)
                if account["locked"]:
                    self._send_json(
                        {
                            "authenticated": False,
                            "error": "locked",
                            **account,
                            "accounts": _owner_accounts(conn),
                        },
                        status=423,
                    )
                    return
                if not _verify_password_hash(password, password_hash):
                    _register_login_failure(client_ip, username)
                    failed_attempts = account["failed_attempts"] + 1
                    locked = failed_attempts >= AUTH_MAX_ATTEMPTS
                    _set_account_state(conn, username, failed_attempts, locked)
                    conn.commit()
                    updated = _account_state(conn, username)
                    self._send_json(
                        {
                            "authenticated": False,
                            "error": "invalid_password",
                            **updated,
                            "accounts": _owner_accounts(conn),
                        },
                        status=423 if locked else 401,
                    )
                    return
                _clear_login_failures(client_ip, username)
                _set_account_state(conn, username, 0, False)
                session_id = _create_session(conn, username)
                conn.commit()
                self._send_json(
                    {
                        "authenticated": True,
                        "username": username,
                        "is_admin": _user_is_admin(username),
                        "accounts": _owner_accounts(conn),
                    },
                    extra_headers=[
                        self._session_cookie_header(session_id, AUTH_SESSION_IDLE_SECONDS)
                    ],
                )
            finally:
                conn.close()
            return
        if path == "/api/auth/logout":
            if not self._verify_request_origin():
                return
            conn = self._db()
            try:
                _destroy_session_by_id(conn, self._request_session_id())
                conn.commit()
                self._send_json(
                    {"logged_out": True},
                    extra_headers=[
                        self._session_cookie_header("", 0)
                    ],
                )
            finally:
                conn.close()
            return
        if path.startswith("/api/v1/"):
            self._handle_program_api_post(path, parse_qs(parsed.query))
            return
        if path.startswith("/api/") and not self._authorize_stateful_api(path):
            return
        conn = self._db()
        try:
            username = self._authenticated_username()
            if path == "/api/auth/change-password":
                payload = _parse_json(self)
                current_password = str(payload.get("current_password", ""))
                new_password = str(payload.get("new_password", ""))
                confirm_password = str(payload.get("confirm_password", ""))
                if not current_password:
                    self._send_json({"error": "missing_current_password", "message": "请输入当前密码。"}, status=400)
                    return
                if not new_password:
                    self._send_json({"error": "missing_new_password", "message": "请输入新密码。"}, status=400)
                    return
                if new_password != confirm_password:
                    self._send_json({"error": "password_mismatch", "message": "两次输入的新密码不一致。"}, status=400)
                    return
                if current_password == new_password:
                    self._send_json({"error": "password_unchanged", "message": "新密码不能与当前密码相同。"}, status=400)
                    return
                password_rule_error = _password_policy_error(new_password)
                if password_rule_error:
                    self._send_json({"error": "invalid_new_password", "message": password_rule_error}, status=400)
                    return
                current_hash = _account_password_hash(conn, username)
                if not current_hash or not _verify_password_hash(current_password, current_hash):
                    self._send_json({"error": "invalid_current_password", "message": "当前密码不正确。"}, status=400)
                    return
                new_hash = _create_password_hash(new_password)
                _set_account_password_hash(conn, username, new_hash)
                _set_account_state(conn, username, 0, False)
                _clear_login_failures(self._client_ip(), username)
                conn.commit()
                self._send_json({"updated": True, "username": username})
                return
            if path == "/api/meta/catalog":
                if not _user_can_manage_catalog(username):
                    self._send_json({"error": "forbidden"}, status=403)
                    return
                _ensure_option_catalog_table(conn)
                payload = _parse_json(self)
                option_type = str(payload.get("option_type", "") or "").strip()
                value = str(payload.get("value", "") or "").strip()
                if option_type not in {"owner", "wardrobe_brand", "watch_brand"}:
                    self._send_json({"error": "invalid_option_type"}, status=400)
                    return
                if not value:
                    self._send_json({"error": "empty_option_value"}, status=400)
                    return
                conn.execute(
                    "INSERT OR IGNORE INTO option_catalogs(option_type, value) VALUES(?, ?)",
                    (option_type, value),
                )
                conn.commit()
                self._send_json({"saved": True, "option_type": option_type, "value": value}, status=201)
                return
            if path == "/api/ai-prompts":
                self._send_json(_native_ai_removed_payload(), status=410)
                return
            if path == "/api/items":
                payload = _parse_json(self)
                try:
                    payload = _apply_item_owner_scope_to_payload(payload, username)
                except PermissionError:
                    self._send_json({"error": "forbidden"}, status=403)
                    return
                try:
                    payload = _normalize_item_payload(payload)
                    cursor = conn.execute(
                        """
                        INSERT INTO items (
                            code, brand, section, loc, owner, layer_role, outer_type, scene_tag,
                            relax_index, temp_min, temp_max, standalone_min, standalone_max,
                            primary_color, secondary_color, official_desc, price_original, price_original_currency, price_cny,
                            series, size, acquired_at, official_color_code, material, care, notes,
                            status, wear_total, wear_maintenance, wear_year, maint_count, wear_threshold
                        )
                        VALUES (
                            :code, :brand, :section, :loc, :owner, :layer_role, :outer_type, :scene_tag,
                            :relax_index, :temp_min, :temp_max, :standalone_min, :standalone_max,
                            :primary_color, :secondary_color, :official_desc, :price_original, :price_original_currency, :price_cny,
                            :series, :size, :acquired_at, :official_color_code, :material, :care, :notes,
                            COALESCE(:status, 'Active'), COALESCE(:wear_total, 0), COALESCE(:wear_maintenance, 0),
                            COALESCE(:wear_year, 0), COALESCE(:maint_count, 0), COALESCE(:wear_threshold, 0)
                        )
                        """,
                        payload,
                    )
                    conn.commit()
                except ValueError as exc:
                    conn.rollback()
                    self._send_json({"error": "invalid_item", "message": str(exc)}, status=400)
                    return
                except sqlite3.IntegrityError:
                    conn.rollback()
                    self._send_json({"error": "invalid_item", "message": "duplicate_code"}, status=409)
                    return
                except Exception as exc:
                    conn.rollback()
                    self._send_json({"error": "db_or_export_failed", "message": str(exc)}, status=500)
                    return
                self._send_json(_item_with_photos(conn, int(cursor.lastrowid)), status=201)
                return
            if path == "/api/pick-directory":
                selected = _pick_directory_native()
                self._send_json({"directory": selected})
                return
            if path == "/api/import-files":
                self._send_json(
                    {
                        "error": "file_import_disabled",
                        "message": "File import is disabled. Use POST /api/v1/history/outfits.",
                    },
                    status=410,
                )
                return
            if path == "/api/bootstrap":
                self._send_json(
                    {
                        "error": "file_import_disabled",
                        "message": "File import is disabled. Use POST /api/v1/history/outfits.",
                    },
                    status=410,
                )
                return
            if path == "/api/items/baseline-create":
                payload = _parse_json(self)
                try:
                    payload = _apply_item_owner_scope_to_payload(payload, username)
                except PermissionError:
                    self._send_json({"error": "forbidden"}, status=403)
                    return
                kind = _new_item_kind(payload)
                try:
                    item_id = _insert_item_record(conn, payload)
                    export_results = _export_item_baseline(conn, kind)
                    conn.commit()
                except ValueError as exc:
                    conn.rollback()
                    self._send_json({"error": "invalid_baseline_create", "message": str(exc)}, status=400)
                    return
                except sqlite3.IntegrityError:
                    conn.rollback()
                    self._send_json({"error": "invalid_baseline_create", "message": "duplicate_code"}, status=409)
                    return
                except RuntimeError as exc:
                    conn.rollback()
                    message = str(exc)
                    status = 423 if message.startswith("baseline_file_locked:") else 500
                    self._send_json({"error": "db_or_export_failed", "message": message}, status=status)
                    return
                except Exception as exc:
                    conn.rollback()
                    self._send_json({"error": "db_or_export_failed", "message": str(exc)}, status=500)
                    return
                saved_item = _item_with_photos(conn, item_id)
                self._send_json({**saved_item, "create_result": {"saved": True, "mode": "db_export"}, "exports": export_results}, status=201)
                return
            if path.startswith("/api/items/") and path.endswith("/baseline-save"):
                item_id = int(path.split("/")[3])
                item = conn.execute("SELECT * FROM items WHERE id = ?", (item_id,)).fetchone()
                if item is None:
                    self._send_json({"error": "item_not_found"}, status=404)
                    return
                if not _item_row_authorized(item, username):
                    self._send_json({"error": "forbidden"}, status=403)
                    return
                payload = _parse_json(self)
                try:
                    payload = _apply_item_owner_scope_to_payload(payload, username, item["owner"])
                except PermissionError:
                    self._send_json({"error": "forbidden"}, status=403)
                    return
                try:
                    _update_item_record(conn, item, payload)
                    conn.commit()
                except ValueError as exc:
                    conn.rollback()
                    self._send_json({"error": "invalid_baseline_edit", "message": str(exc)}, status=400)
                    return
                except sqlite3.IntegrityError:
                    conn.rollback()
                    self._send_json({"error": "invalid_baseline_edit", "message": "duplicate_code"}, status=409)
                    return
                except RuntimeError as exc:
                    conn.rollback()
                    message = str(exc)
                    status = 423 if message.startswith("baseline_file_locked:") else 500
                    self._send_json({"error": "db_or_export_failed", "message": message}, status=status)
                    return
                except Exception as exc:
                    conn.rollback()
                    self._send_json({"error": "db_or_export_failed", "message": str(exc)}, status=500)
                    return
                updated_row = conn.execute("SELECT * FROM items WHERE id = ?", (item_id,)).fetchone()
                saved_item = _item_with_photos(conn, item_id)
                export_tasks = _item_export_tasks_for_change(item, updated_row)
                _schedule_item_export_tasks(export_tasks)
                self._send_json({
                    **saved_item,
                    "save_result": {"saved": True, "mode": "db_async_export" if export_tasks else "db_only"},
                    "exports_pending": bool(export_tasks),
                })
                return
            if path == "/api/ai/outfit-review":
                self._send_json(_native_ai_removed_payload(), status=410)
                return
            if path.startswith("/api/outfits/") and path.endswith("/ai-review"):
                self._send_json(_native_ai_removed_payload(), status=410)
                return
            if path.startswith("/api/featured-looks/") and path.endswith("/ai-review"):
                self._send_json(_native_ai_removed_payload(), status=410)
                return
            if path.startswith("/api/outfits/") and path.endswith("/ai-analysis"):
                self._send_json(_native_ai_removed_payload(), status=410)
                return
            if path.startswith("/api/featured-looks/") and path.endswith("/ai-analysis"):
                self._send_json(_native_ai_removed_payload(), status=410)
                return
            if path == "/api/outfits":
                payload = _parse_json(self)
                requested_owner = _normalize_edit_value(payload.get("owner")) or _normalize_edit_value(username)
                existing_outfit = conn.execute(
                    "SELECT * FROM outfits WHERE wear_date = ? AND COALESCE(owner, '') = ? ORDER BY id DESC LIMIT 1",
                    (_normalize_edit_value(payload.get("wear_date")), requested_owner),
                ).fetchone()
                if existing_outfit is not None and not _outfit_row_authorized(existing_outfit, username):
                    self._send_json({"error": "forbidden"}, status=403)
                    return
                try:
                    payload = _apply_owner_scope_to_payload(payload, username, existing_outfit["owner"] if existing_outfit is not None else None)
                    resolved_entries = _resolve_outfit_item_entries(
                        conn,
                        list(payload.get("items") or []),
                        username,
                        _normalize_edit_value(payload.get("wear_mode")) or "normal",
                        _normalize_edit_value(payload.get("wear_date")),
                    )
                    if existing_outfit is not None:
                        _ensure_latest_outfit_editable(conn, int(existing_outfit["id"]), existing_outfit["owner"])
                        saved, affected_item_ids = _upsert_outfit_daily_update(
                            conn,
                            int(existing_outfit["id"]),
                            _normalize_edit_value(existing_outfit["wear_date"]),
                            payload,
                            resolved_entries,
                            username,
                            existing_outfit,
                        )
                        status_code = 200
                    else:
                        saved, affected_item_ids = _create_outfit_daily_update(conn, payload, resolved_entries)
                        status_code = 201
                    export_tasks = _wear_aggregate_export_tasks_for_item_ids(conn, affected_item_ids)
                    conn.commit()
                    _schedule_item_export_tasks(export_tasks)
                except PermissionError:
                    conn.rollback()
                    self._send_json({"error": "forbidden"}, status=403)
                    return
                except (KeyError, ValueError) as exc:
                    conn.rollback()
                    self._send_json({"error": "invalid_outfit_create", "message": str(exc)}, status=400)
                    return
                except RuntimeError as exc:
                    conn.rollback()
                    self._send_json({"error": "invalid_outfit_create", "message": str(exc)}, status=409)
                    return
                except Exception as exc:
                    conn.rollback()
                    self._send_json({"error": "db_or_export_failed", "message": str(exc)}, status=500)
                    return
                self._send_json({
                    "outfit": _serialize_outfit(conn, saved, username),
                    "exports_pending": bool(export_tasks),
                    "export_tasks": sorted(export_tasks),
                }, status=status_code)
                return
            if path.startswith("/api/outfits/") and path.endswith("/save-as-featured-look"):
                outfit_id = int(path.split("/")[3])
                outfit = conn.execute("SELECT * FROM outfits WHERE id = ?", (outfit_id,)).fetchone()
                if outfit is None:
                    self._send_json({"error": "outfit_not_found"}, status=404)
                    return
                if not _outfit_row_authorized(outfit, username):
                    self._send_json({"error": "forbidden"}, status=403)
                    return
                try:
                    look = _create_featured_look_from_outfit(conn, outfit, username)
                    conn.commit()
                except RuntimeError as exc:
                    conn.rollback()
                    self._send_json({"error": "invalid_featured_look_create", "message": str(exc)}, status=409)
                    return
                except Exception as exc:
                    conn.rollback()
                    self._send_json({"error": "db_or_export_failed", "message": str(exc)}, status=500)
                    return
                self._send_json(look, status=201)
                return
            if path.startswith("/api/outfits/") and path.endswith("/save"):
                outfit_id = int(path.split("/")[3])
                outfit = conn.execute("SELECT * FROM outfits WHERE id = ?", (outfit_id,)).fetchone()
                if outfit is None:
                    self._send_json({"error": "outfit_not_found"}, status=404)
                    return
                if not _outfit_row_authorized(outfit, username):
                    self._send_json({"error": "forbidden"}, status=403)
                    return
                try:
                    _ensure_latest_outfit_editable(conn, outfit_id, outfit["owner"])
                except RuntimeError as exc:
                    self._send_json({"error": "outfit_edit_not_allowed", "message": str(exc)}, status=409)
                    return
                payload = _parse_json(self)
                try:
                    payload = _apply_owner_scope_to_payload(payload, username, outfit["owner"])
                    resolved_entries = _resolve_outfit_item_entries(
                        conn,
                        list(payload.get("items") or []),
                        username,
                        _normalize_edit_value(payload.get("wear_mode")) or _normalize_edit_value(outfit["wear_mode"]) or "normal",
                        _normalize_edit_value(payload.get("wear_date")) or _normalize_edit_value(outfit["wear_date"]),
                    )
                    updated, affected_item_ids = _upsert_outfit_daily_update(
                        conn,
                        outfit_id,
                        _normalize_edit_value(outfit["wear_date"]),
                        payload,
                        resolved_entries,
                        username,
                        outfit,
                    )
                    export_tasks = _wear_aggregate_export_tasks_for_item_ids(conn, affected_item_ids)
                    conn.commit()
                    _schedule_item_export_tasks(export_tasks)
                except PermissionError:
                    conn.rollback()
                    self._send_json({"error": "forbidden"}, status=403)
                    return
                except (KeyError, ValueError) as exc:
                    conn.rollback()
                    self._send_json({"error": "invalid_outfit_edit", "message": str(exc)}, status=400)
                    return
                except RuntimeError as exc:
                    conn.rollback()
                    self._send_json({"error": "invalid_outfit_edit", "message": str(exc)}, status=409)
                    return
                except Exception as exc:
                    conn.rollback()
                    self._send_json({"error": "db_or_export_failed", "message": str(exc)}, status=500)
                    return
                self._send_json({
                    "outfit": _serialize_outfit(conn, updated, username),
                    "exports_pending": bool(export_tasks),
                    "export_tasks": sorted(export_tasks),
                })
                return
            if path.startswith("/api/featured-looks/") and path.endswith("/save"):
                featured_look_id = int(path.split("/")[3])
                look = conn.execute("SELECT * FROM featured_looks WHERE id = ?", (featured_look_id,)).fetchone()
                if look is None:
                    self._send_json({"error": "featured_look_not_found"}, status=404)
                    return
                if not _featured_look_row_authorized(look, username):
                    self._send_json({"error": "forbidden"}, status=403)
                    return
                payload = _parse_json(self)
                try:
                    payload = _apply_owner_scope_to_payload(payload, username, look["owner"])
                    saved = _update_featured_look_record(conn, look, payload)
                    conn.commit()
                except PermissionError:
                    conn.rollback()
                    self._send_json({"error": "forbidden"}, status=403)
                    return
                except sqlite3.IntegrityError:
                    conn.rollback()
                    self._send_json({"error": "invalid_featured_look_edit", "message": "duplicate_look_id"}, status=409)
                    return
                except Exception as exc:
                    conn.rollback()
                    self._send_json({"error": "db_or_export_failed", "message": str(exc)}, status=500)
                    return
                self._send_json(saved)
                return
            if path.startswith("/api/items/") and path.endswith("/photos/order"):
                item_id = int(path.split("/")[3])
                item = conn.execute("SELECT * FROM items WHERE id = ?", (item_id,)).fetchone()
                if item is None:
                    self._send_json({"error": "item_not_found"}, status=404)
                    return
                if not _item_row_authorized(item, username):
                    self._send_json({"error": "forbidden"}, status=403)
                    return
                payload = _parse_json(self)
                if not isinstance(payload, dict):
                    self._send_json({"error": "invalid_payload", "message": "JSON body must be an object."}, status=400)
                    return
                try:
                    photo_ids, first_photo_id = _api_photo_order_request(payload)
                    current_ids = [int(photo["id"]) for photo in photo_ordering.item_photo_rows(conn, item_id)]
                    next_photo_ids = photo_ordering.ordered_photo_ids(
                        current_ids,
                        photo_ids=photo_ids,
                        first_photo_id=first_photo_id,
                    )
                    photo_ordering.apply_item_photo_order(conn, item_id, next_photo_ids)
                    conn.commit()
                except ValueError as exc:
                    conn.rollback()
                    self._send_json({"error": "invalid_photo_order", "message": str(exc)}, status=400)
                    return
                self._send_json(_item_with_photos(conn, item_id))
                return
            if path.startswith("/api/items/") and path.endswith("/photos"):
                item_id = int(path.split("/")[3])
                item = conn.execute("SELECT * FROM items WHERE id = ?", (item_id,)).fetchone()
                if item is None:
                    self._send_json({"error": "item_not_found"}, status=404)
                    return
                if not _item_row_authorized(item, username):
                    self._send_json({"error": "forbidden"}, status=403)
                    return
                parts = _parse_multipart(self)
                file_part = parts.get("file")
                if not file_part or not file_part.get("filename"):
                    self._send_json({"error": "file_missing"}, status=400)
                    return
                upload_error = _validate_image_upload(file_part)
                if upload_error:
                    self._send_json({"error": upload_error}, status=400 if upload_error == "invalid_file_type" else 413)
                    return
                suffix = Path(file_part["filename"]).suffix.lower() or ".bin"
                safe_name = f"{item['code']}_{uuid.uuid4().hex[:8]}{suffix}"
                next_order = conn.execute(
                    "SELECT COALESCE(MAX(sort_order), 0) + 1 AS next_order FROM photos WHERE item_id = ?",
                    (item_id,),
                ).fetchone()["next_order"]
                conn.execute(
                    """
                    INSERT INTO photos (item_id, file_name, original_name, sort_order, source_tag, mime_type, data)
                    VALUES (?, ?, ?, ?, 'upload', ?, ?)
                    """,
                    (
                        item_id,
                        safe_name,
                        file_part["filename"],
                        next_order,
                        file_part.get("content_type", "application/octet-stream"),
                        file_part["content"],
                    ),
                )
                conn.commit()
                self._send_json(_item_with_photos(conn, item_id), status=201)
                return
            if path.startswith("/api/items/") and path.endswith("/maintenance"):
                item_id = int(path.split("/")[3])
                item = conn.execute("SELECT * FROM items WHERE id = ?", (item_id,)).fetchone()
                if item is None:
                    self._send_json({"error": "item_not_found"}, status=404)
                    return
                if not _item_row_authorized(item, username):
                    self._send_json({"error": "forbidden"}, status=403)
                    return
                try:
                    result, _changed = _send_item_to_maintenance(conn, item)
                except Exception as exc:
                    conn.rollback()
                    self._send_json({"error": "db_or_export_failed", "message": str(exc)}, status=500)
                    return
                self._send_json(result)
                return
            if path.startswith("/api/items/") and path.endswith("/activate"):
                item_id = int(path.split("/")[3])
                item = conn.execute("SELECT * FROM items WHERE id = ?", (item_id,)).fetchone()
                if item is None:
                    self._send_json({"error": "item_not_found"}, status=404)
                    return
                if not _item_row_authorized(item, username):
                    self._send_json({"error": "forbidden"}, status=403)
                    return
                try:
                    result, _changed = _activate_item_from_maintenance(conn, item)
                except Exception as exc:
                    conn.rollback()
                    self._send_json({"error": "db_or_export_failed", "message": str(exc)}, status=500)
                    return
                self._send_json(result)
                return
            if path.startswith("/api/outfits/") and path.endswith("/photos"):
                outfit_id = int(path.split("/")[3])
                outfit = conn.execute("SELECT * FROM outfits WHERE id = ?", (outfit_id,)).fetchone()
                if outfit is None:
                    self._send_json({"error": "outfit_not_found"}, status=404)
                    return
                if not _outfit_row_authorized(outfit, username):
                    self._send_json({"error": "forbidden"}, status=403)
                    return
                parts = _parse_multipart(self)
                file_part = parts.get("file")
                if not file_part or not file_part.get("filename"):
                    self._send_json({"error": "file_missing"}, status=400)
                    return
                upload_error = _validate_image_upload(file_part)
                if upload_error:
                    self._send_json({"error": upload_error}, status=400 if upload_error == "invalid_file_type" else 413)
                    return
                suffix = Path(file_part["filename"]).suffix.lower() or ".bin"
                safe_name = f"outfit_{outfit_id}_{uuid.uuid4().hex[:8]}{suffix}"
                next_order = conn.execute(
                    "SELECT COALESCE(MAX(sort_order), 0) + 1 AS next_order FROM outfit_photos WHERE outfit_id = ?",
                    (outfit_id,),
                ).fetchone()["next_order"]
                gps_lat, gps_lng = _extract_image_gps_coordinates(file_part["content"])
                thumbnail = _thumbnail_bytes_from_payload(
                    file_part["content"],
                    file_part.get("content_type", "application/octet-stream"),
                )
                thumb_raw = thumbnail[0] if thumbnail is not None else None
                thumb_mime = thumbnail[1] if thumbnail is not None else None
                conn.execute(
                    """
                    INSERT INTO outfit_photos (
                        outfit_id, file_name, original_name, sort_order, source_tag, mime_type, data,
                        gps_lat, gps_lng, gps_checked, thumb_mime_type, thumb_data
                    )
                    VALUES (?, ?, ?, ?, 'upload', ?, ?, ?, ?, 1, ?, ?)
                    """,
                    (
                        outfit_id,
                        safe_name,
                        file_part["filename"],
                        next_order,
                        file_part.get("content_type", "application/octet-stream"),
                        file_part["content"],
                        gps_lat,
                        gps_lng,
                        thumb_mime,
                        thumb_raw,
                    ),
                )
                photo_row = conn.execute(
                    "SELECT * FROM outfit_photos WHERE id = last_insert_rowid()"
                ).fetchone()
                if photo_row is not None:
                    _sync_outfit_photo_into_existing_featured_look(conn, outfit, photo_row, username)
                conn.commit()
                outfit_row = conn.execute("SELECT * FROM outfits WHERE id = ?", (outfit_id,)).fetchone()
                self._send_json(_serialize_outfit(conn, outfit_row), status=201)
                return
            if path.startswith("/api/featured-looks/") and path.endswith("/photos"):
                featured_look_id = int(path.split("/")[3])
                look = conn.execute("SELECT * FROM featured_looks WHERE id = ?", (featured_look_id,)).fetchone()
                if look is None:
                    self._send_json({"error": "featured_look_not_found"}, status=404)
                    return
                if not _featured_look_row_authorized(look, username):
                    self._send_json({"error": "forbidden"}, status=403)
                    return
                parts = _parse_multipart(self)
                file_part = parts.get("file")
                if not file_part or not file_part.get("filename"):
                    self._send_json({"error": "file_missing"}, status=400)
                    return
                upload_error = _validate_image_upload(file_part)
                if upload_error:
                    self._send_json({"error": upload_error}, status=400 if upload_error == "invalid_file_type" else 413)
                    return
                suffix = Path(file_part["filename"]).suffix.lower() or ".bin"
                safe_name = f"look_{featured_look_id}_{uuid.uuid4().hex[:8]}{suffix}"
                next_order = conn.execute(
                    """
                    SELECT COALESCE(MAX(sort_order), 0) + 1 AS next_order
                    FROM featured_look_photos
                    WHERE featured_look_id = ?
                    """,
                    (featured_look_id,),
                ).fetchone()["next_order"]
                conn.execute(
                    """
                    INSERT INTO featured_look_photos (featured_look_id, file_name, original_name, sort_order, source_tag, mime_type, data)
                    VALUES (?, ?, ?, ?, 'upload', ?, ?)
                    """,
                    (
                        featured_look_id,
                        safe_name,
                        file_part["filename"],
                        next_order,
                        file_part.get("content_type", "application/octet-stream"),
                        file_part["content"],
                    ),
                )
                conn.commit()
                look_row = conn.execute("SELECT * FROM featured_looks WHERE id = ?", (featured_look_id,)).fetchone()
                self._send_json(_serialize_featured_look(conn, look_row), status=201)
                return
            self._send_json({"error": "not_found"}, status=404)
        finally:
            conn.close()

    def do_PUT(self) -> None:
        parsed = urlparse(self.path)
        path = unquote(parsed.path)
        if path.startswith("/api/") and not self._authorize_stateful_api(path):
            return
        conn = self._db()
        try:
            username = self._authenticated_username()
            if path.startswith("/api/items/"):
                item_id = int(path.split("/")[-1])
                item = conn.execute("SELECT * FROM items WHERE id = ?", (item_id,)).fetchone()
                if item is None:
                    self._send_json({"error": "item_not_found"}, status=404)
                    return
                if not _item_row_authorized(item, username):
                    self._send_json({"error": "forbidden"}, status=403)
                    return
                payload = _parse_json(self)
                try:
                    payload = _apply_item_owner_scope_to_payload(payload, username, item["owner"])
                except PermissionError:
                    self._send_json({"error": "forbidden"}, status=403)
                    return
                try:
                    payload = _normalize_item_payload(payload, item)
                    payload["id"] = item_id
                    conn.execute(
                        """
                        UPDATE items
                        SET code=:code,
                            brand=:brand,
                            section=:section,
                            loc=:loc,
                            owner=:owner,
                            layer_role=:layer_role,
                            outer_type=:outer_type,
                            scene_tag=:scene_tag,
                            relax_index=:relax_index,
                            temp_min=:temp_min,
                            temp_max=:temp_max,
                            standalone_min=:standalone_min,
                            standalone_max=:standalone_max,
                            primary_color=:primary_color,
                            secondary_color=:secondary_color,
                            official_desc=:official_desc,
                            price_original=:price_original,
                            price_original_currency=:price_original_currency,
                            price_cny=:price_cny,
                            series=:series,
                            size=:size,
                            acquired_at=:acquired_at,
                            official_color_code=:official_color_code,
                            material=:material,
                            care=:care,
                            notes=:notes,
                            status=:status,
                            wear_total=:wear_total,
                            wear_maintenance=:wear_maintenance,
                            wear_year=:wear_year,
                            maint_count=:maint_count,
                            wear_threshold=:wear_threshold,
                            updated_at=CURRENT_TIMESTAMP
                        WHERE id=:id
                        """,
                        payload,
                    )
                    conn.commit()
                except ValueError as exc:
                    conn.rollback()
                    self._send_json({"error": "invalid_item_edit", "message": str(exc)}, status=400)
                    return
                except sqlite3.IntegrityError:
                    conn.rollback()
                    self._send_json({"error": "invalid_item_edit", "message": "duplicate_code"}, status=409)
                    return
                except Exception as exc:
                    conn.rollback()
                    self._send_json({"error": "db_or_export_failed", "message": str(exc)}, status=500)
                    return
                self._send_json(_item_with_photos(conn, item_id))
                return
            self._send_json({"error": "not_found"}, status=404)
        finally:
            conn.close()

    def do_DELETE(self) -> None:
        parsed = urlparse(self.path)
        path = unquote(parsed.path)
        if path.startswith("/api/") and not self._authorize_stateful_api(path):
            return
        conn = self._db()
        try:
            username = self._authenticated_username()
            if path.startswith("/api/outfit-photos/"):
                photo_id = int(path.split("/")[-1])
                photo = conn.execute("SELECT * FROM outfit_photos WHERE id = ?", (photo_id,)).fetchone()
                if photo is None:
                    self._send_json({"error": "photo_not_found"}, status=404)
                    return
                outfit = conn.execute("SELECT * FROM outfits WHERE id = ?", (photo["outfit_id"],)).fetchone()
                if not _outfit_row_authorized(outfit, username):
                    self._send_json({"error": "forbidden"}, status=403)
                    return
                conn.execute("DELETE FROM outfit_photos WHERE id = ?", (photo_id,))
                conn.commit()
                self._send_json({"deleted": True})
                return
            if path.startswith("/api/featured-look-photos/"):
                photo_id = int(path.split("/")[-1])
                photo = conn.execute("SELECT * FROM featured_look_photos WHERE id = ?", (photo_id,)).fetchone()
                if photo is None:
                    self._send_json({"error": "photo_not_found"}, status=404)
                    return
                look = conn.execute("SELECT * FROM featured_looks WHERE id = ?", (photo["featured_look_id"],)).fetchone()
                if not _featured_look_row_authorized(look, username):
                    self._send_json({"error": "forbidden"}, status=403)
                    return
                conn.execute("DELETE FROM featured_look_photos WHERE id = ?", (photo_id,))
                conn.commit()
                self._send_json({"deleted": True})
                return
            if path.startswith("/api/photos/"):
                photo_id = int(path.split("/")[-1])
                photo = conn.execute("SELECT * FROM photos WHERE id = ?", (photo_id,)).fetchone()
                if photo is None:
                    self._send_json({"error": "photo_not_found"}, status=404)
                    return
                if not _photo_item_authorized(conn, photo_id, username):
                    self._send_json({"error": "forbidden"}, status=403)
                    return
                conn.execute("DELETE FROM photos WHERE id = ?", (photo_id,))
                conn.commit()
                target = MEDIA_DIR / photo["file_name"]
                if target.exists() and photo["source_tag"] == "upload":
                    target.unlink()
                self._send_json({"deleted": True})
                return
            if path.startswith("/api/items/"):
                item_id = int(path.split("/")[-1])
                photos = conn.execute("SELECT * FROM photos WHERE item_id = ?", (item_id,)).fetchall()
                item = conn.execute("SELECT * FROM items WHERE id = ?", (item_id,)).fetchone()
                if item is None:
                    self._send_json({"error": "item_not_found"}, status=404)
                    return
                if not _item_row_authorized(item, username):
                    self._send_json({"error": "forbidden"}, status=403)
                    return
                conn.execute("DELETE FROM items WHERE id = ?", (item_id,))
                try:
                    export_results = _export_item_baseline(conn, _item_source_kind(item))
                except Exception as exc:
                    conn.rollback()
                    self._send_json({"error": "db_or_export_failed", "message": str(exc)}, status=500)
                    return
                conn.commit()
                for photo in photos:
                    target = MEDIA_DIR / photo["file_name"]
                    if target.exists() and photo["source_tag"] == "upload":
                        target.unlink()
                self._send_json({"deleted": True, "exports": export_results})
                return
            if path.startswith("/api/featured-looks/"):
                featured_look_id = int(path.split("/")[-1])
                look = conn.execute("SELECT * FROM featured_looks WHERE id = ?", (featured_look_id,)).fetchone()
                if look is None:
                    self._send_json({"error": "featured_look_not_found"}, status=404)
                    return
                if not _featured_look_row_authorized(look, username):
                    self._send_json({"error": "forbidden"}, status=403)
                    return
                photos = conn.execute(
                    "SELECT * FROM featured_look_photos WHERE featured_look_id = ?",
                    (featured_look_id,),
                ).fetchall()
                conn.execute("DELETE FROM featured_look_items WHERE featured_look_id = ?", (featured_look_id,))
                conn.execute("DELETE FROM featured_look_photos WHERE featured_look_id = ?", (featured_look_id,))
                conn.execute("DELETE FROM featured_looks WHERE id = ?", (featured_look_id,))
                try:
                    export_results = _export_featured_looks_workbooks(conn)
                except Exception as exc:
                    conn.rollback()
                    self._send_json({"error": "db_or_export_failed", "message": str(exc)}, status=500)
                    return
                conn.commit()
                for photo in photos:
                    target = MEDIA_DIR / photo["file_name"]
                    if target.exists() and photo["source_tag"] == "upload":
                        target.unlink()
                self._send_json({"deleted": True, "exports": export_results})
                return
            if path.startswith("/api/outfits/"):
                outfit_id = int(path.split("/")[-1])
                outfit = conn.execute("SELECT * FROM outfits WHERE id = ?", (outfit_id,)).fetchone()
                if outfit is None:
                    self._send_json({"error": "outfit_not_found"}, status=404)
                    return
                if not _outfit_row_authorized(outfit, username):
                    self._send_json({"error": "forbidden"}, status=403)
                    return
                try:
                    _ensure_latest_outfit_editable(conn, outfit_id, outfit["owner"])
                    affected_item_ids = _delete_outfit_with_wearcount(conn, outfit)
                    export_tasks = _wear_aggregate_export_tasks_for_item_ids(conn, affected_item_ids)
                    conn.commit()
                    _schedule_item_export_tasks(export_tasks)
                except RuntimeError as exc:
                    conn.rollback()
                    self._send_json({"error": "invalid_outfit_delete", "message": str(exc)}, status=409)
                    return
                except Exception as exc:
                    conn.rollback()
                    self._send_json({"error": "db_or_export_failed", "message": str(exc)}, status=500)
                    return
                self._send_json({
                    "deleted": True,
                    "exports_pending": bool(export_tasks),
                    "export_tasks": sorted(export_tasks),
                })
                return
            self._send_json({"error": "not_found"}, status=404)
        finally:
            conn.close()


def run() -> None:
    ensure_directories()
    conn = connect()
    init_db(conn)
    _sync_auth_users(conn)
    summaries = bootstrap_from_desktop(conn)
    conn.close()
    if summaries:
        print("Bootstrap import completed:")
        for summary in summaries:
          print(json.dumps(summary, ensure_ascii=False))
    httpd = ThreadingHTTPServer((HOST, PORT), WardrobeHandler)
    print(f"Wardrobe app running at http://{HOST}:{PORT}")
    for ip in _local_ipv4_addresses():
        print(f"Access URL: http://{ip}:{PORT}")
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        httpd.server_close()
